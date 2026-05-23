"""
speculative_engine.py — Engine for the "Maybe Unsafe Financial Freedom" speculative tool.

Scores 12 dimensions across two categories:

  Auto-scored (FMP + yfinance + free public APIs):
    1.  Price Momentum        — RSI + price vs 50d/200d MA
    2.  Volume Signal         — recent volume surge vs 30d baseline
    3.  Short Interest        — short float %, days to cover (squeeze fuel)
    4.  Analyst Revisions     — estimate revision direction + upgrade/downgrade balance
    5.  Float Size            — small float/mktcap = explosive potential
    6.  Insider Buying        — open-market purchases last 90 days
    7.  Downside Floor        — net cash position, debt load
    8.  Options Activity      — put/call ratio via yfinance
    9.  Technical Setup       — MACD crossover, 52-week high proximity, EMA stack alignment
    10. Social / Trend Momentum — Google Trends search spike + StockTwits sentiment

  Manual (user-supplied):
    11. Narrative Theme       — dropdown selection + strength rating
    12. Catalyst Quality      — user-described catalyst + timing tier

Scoring: HIGH = 10 pts · MOD = 5 pts · LOW = 0 pts  (max 120)

Verdicts (75% / ~54% / 45% / 30% of 120):
    ≥ 90 → Moonshot Conviction
    ≥ 65 → Strong Speculative
    ≥ 54 → Speculative Play
    ≥ 36 → High Risk
     < 36 → Pass
"""

import datetime
import io

import requests as _req

# ── Constants ──────────────────────────────────────────────────────────────────

TIER_PTS = {"HIGH": 10, "MOD": 5, "LOW": 0}

POLYGON_API_KEY = ""   # set from env in server.py

NARRATIVE_THEMES = [
    "AI / Machine Learning",
    "Defence & Aerospace",
    "Biotech Catalyst",
    "GLP-1 / Weight Loss",
    "Energy Transition",
    "Nuclear / SMR",
    "Crypto-Adjacent",
    "Turnaround / Restructuring",
    "Supply Chain Re-shoring",
    "Space / Deep Tech",
    "Other",
]

VERDICTS = [
    (90, "Moonshot Conviction"),
    (65, "Strong Speculative"),
    (54, "Speculative Play"),
    (36, "High Risk"),
    (0,  "Pass"),
]

MAX_SCORE = 120

# ── Sector-calibrated scenario multiples ──────────────────────────────────────
# Each bucket: (bear_factor, default_bull_factor) applied to current EV/Revenue.
# Bear factor reflects narrative-fails de-rating risk in that sector; bull factor
# reflects realistic re-rating ceiling on success.  Tighter bear/bull ranges in
# stable sectors (defence), wider in binary-catalyst sectors (biotech, crypto).
SECTOR_SCENARIOS = {
    "ai_ml":              (0.55, 2.10),  # rich multiples, sharp de-rating on disappointment
    "defense_aerospace":  (0.70, 1.55),  # stable contract revenue, contained re-rating
    "biotech_catalyst":   (0.35, 3.00),  # binary outcomes — phase 3 / FDA / readouts
    "glp1_weight_loss":   (0.55, 2.10),  # demand-driven, regulatory/supply risk
    "energy_transition":  (0.60, 1.85),  # policy-sensitive, slower re-rating
    "nuclear_smr":        (0.50, 2.50),  # nascent commercial deployment
    "crypto_adjacent":    (0.40, 2.80),  # high beta to BTC/ETH price action
    "turnaround":         (0.40, 2.30),  # execution-binary, often levered
    "supply_chain":       (0.65, 1.70),  # mature industries, capex-driven
    "space_deeptech":     (0.45, 2.60),  # contract / mission cadence dependent
    "default":            (0.65, 1.80),
}

THEME_TO_BUCKET = {
    "AI / Machine Learning":      "ai_ml",
    "Defence & Aerospace":        "defense_aerospace",
    "Biotech Catalyst":           "biotech_catalyst",
    "GLP-1 / Weight Loss":        "glp1_weight_loss",
    "Energy Transition":          "energy_transition",
    "Nuclear / SMR":              "nuclear_smr",
    "Crypto-Adjacent":            "crypto_adjacent",
    "Turnaround / Restructuring": "turnaround",
    "Supply Chain Re-shoring":    "supply_chain",
    "Space / Deep Tech":          "space_deeptech",
}


def _verdict(score: float) -> str:
    for threshold, label in VERDICTS:
        if score >= threshold:
            return label
    return "Pass"


# ── Mock data (used when FMP quota is exhausted or MOCK_MODE=True) ─────────────

def _mock_data(ticker: str) -> dict:
    """Return plausible dummy signals so the full pipeline renders without API calls."""
    return {
        "ticker":         ticker,
        "mock":           True,
        "current_price":  42.50,
        "market_cap":     3_800_000_000,
        "float_shares":   85_000_000,
        "shares_out":     92_000_000,
        "net_debt":       -120_000_000,          # net cash
        "trailing_rev":   680_000_000,
        "fwd_rev_1yr":    820_000_000,
        "fwd_rev_2yr":    1_010_000_000,
        "current_ev":     3_680_000_000,
        # Momentum
        "rsi_14":         61.4,
        "price_vs_50ma":  0.07,                  # +7% above 50d MA
        "price_vs_200ma": 0.22,                  # +22% above 200d MA
        # Volume
        "vol_10d_avg":    4_200_000,
        "vol_50d_avg":    2_800_000,
        # Short interest
        "short_float_pct": 18.5,
        "days_to_cover":   5.2,
        # Analyst revisions
        "upgrades_90d":    4,
        "downgrades_90d":  1,
        "est_rev_pct":     0.12,                 # +12% estimate revision
        # Insider buying (last 90d open-market)
        "insider_buys":    3,
        "insider_buy_usd": 820_000,
        "insider_sells":   0,
        # Options
        "put_call_ratio":  0.42,
        "call_oi":         45_000,
        "put_oi":          18_900,
        # Balance sheet
        "cash":            380_000_000,
        "total_debt":      260_000_000,
        "ebitda_ttm":      95_000_000,
        # Social / trend momentum
        "trend_ratio":      2.4,
        "best_theme_ratio": 3.1,
        "best_theme_kw":    "AI stocks",
        "all_gt_ratios":    {"MOCK": 2.4, "AI stocks": 3.1, "HBM memory": 2.8},
        "trend_values":     [30,28,32,35,40,38,42,45,55,60,72,80,85,72],
        # yfinance news sentiment
        "news_count":        8,
        "news_7d":           5,
        "news_bullish":      5,
        "news_bearish":      1,
        "news_bullish_pct":  0.83,
        # Reddit sentiment
        "reddit_mentions":   38,
        "reddit_24h":        14,
        "reddit_bullish":    24,
        "reddit_bearish":    6,
        "reddit_bullish_pct": 0.80,
        # Technical setup
        "macd":            1.42,          # MACD line
        "macd_signal":     0.88,          # signal line
        "macd_hist":       0.54,          # histogram (positive = bullish)
        "ema_21":          40.10,
        "ema_50":          38.60,
        "high_52w":        46.80,
        "low_52w":         22.30,
        "pct_from_52w_high": -0.093,      # -9.3% from 52w high
        # Parabolic precursor block (mock illustrates a fully-firing setup)
        "bb_upper":        45.20,
        "bb_lower":        39.80,
        "bb_mid":          42.50,
        "bb_width_pct":    0.127,         # 12.7% — moderate
        "bb_pos_pct":      0.62,
        "bb_squeeze":      True,
        "bb_expansion":    True,
        "obv_slope_pct":   0.18,          # +18% relative — strong accumulation
        "obv_rising":      True,
        "obv_divergence":  True,
        "up_vol_share":    0.68,
        "acc_dist_tier":   "accumulation",
        "roc_20d":         0.084,         # +8.4% over 20 days
        "roc_accelerating": True,
        "atr_14":          1.32,
        "atr_pct_price":   0.031,
        "company_name":    f"{ticker} Corp.",
        "sector":          "Technology",
        "profile":         {},
    }


# ── Polygon helpers ────────────────────────────────────────────────────────────

def _fetch_polygon_ohlcv(ticker: str, polygon_key: str) -> list | None:
    """Fetch ~500 days of daily OHLCV bars from Polygon, sorted ascending.

    Returns a list of dicts with keys: o, h, l, c, v, t (timestamp ms).
    Returns None if the key is missing or the request fails.
    """
    if not polygon_key:
        return None
    try:
        from datetime import date, timedelta
        end   = date.today().strftime("%Y-%m-%d")
        start = (date.today() - timedelta(days=720)).strftime("%Y-%m-%d")
        url   = (f"https://api.polygon.io/v2/aggs/ticker/{ticker}"
                 f"/range/1/day/{start}/{end}")
        r = _req.get(url, params={"adjusted": "true", "sort": "asc",
                                  "limit": 500, "apiKey": polygon_key},
                     timeout=15)
        return r.json().get("results") or None
    except Exception:
        return None


def _fetch_polygon_ticker_details(ticker: str, polygon_key: str) -> dict:
    """Fetch ticker reference data from Polygon (market cap, shares outstanding)."""
    if not polygon_key:
        return {}
    try:
        r = _req.get(f"https://api.polygon.io/v3/reference/tickers/{ticker}",
                     params={"apiKey": polygon_key}, timeout=10)
        return r.json().get("results") or {}
    except Exception:
        return {}


def _fetch_polygon_news(ticker: str, polygon_key: str) -> dict:
    """Fetch recent news from Polygon with built-in per-ticker sentiment.

    Polygon's insights array carries a sentiment field ('positive' / 'negative' /
    'neutral') pre-scored by their NLP model — cleaner than keyword matching.
    """
    if not polygon_key:
        return {"news_count": None, "news_bullish_pct": None,
                "news_error": "No Polygon API key"}
    try:
        r = _req.get("https://api.polygon.io/v2/reference/news",
                     params={"ticker": ticker, "limit": 20, "apiKey": polygon_key},
                     timeout=10)
        articles = r.json().get("results") or []
        if not articles:
            return {"news_count": 0, "news_7d": 0, "news_bullish": 0,
                    "news_bearish": 0, "news_bullish_pct": None}

        cutoff_7d = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(days=7)
        bullish = bearish = recent_7d = 0
        for article in articles:
            insights      = article.get("insights") or []
            ticker_insight = next((i for i in insights if i.get("ticker") == ticker), None)
            if ticker_insight:
                s = ticker_insight.get("sentiment", "neutral")
                if s == "positive":
                    bullish += 1
                elif s == "negative":
                    bearish += 1
            pub = article.get("published_utc", "")
            try:
                ts = datetime.datetime.fromisoformat(pub.replace("Z", "+00:00"))
                if ts > cutoff_7d:
                    recent_7d += 1
            except Exception:
                pass

        total_scored = bullish + bearish
        return {
            "news_count":       len(articles),
            "news_7d":          recent_7d,
            "news_bullish":     bullish,
            "news_bearish":     bearish,
            "news_bullish_pct": (bullish / total_scored) if total_scored else None,
        }
    except Exception as e:
        return {"news_count": None, "news_bullish_pct": None, "news_error": str(e)}


# ── Sentiment keyword sets (used by Reddit scorer) ────────────────────────────

_BULL_WORDS = frozenset([
    "beat", "beats", "upgrade", "upgrades", "surge", "surges", "soar", "soars",
    "strong", "buy", "outperform", "record", "win", "wins", "breakthrough",
    "deal", "contract", "raises", "raised", "rally", "rallies", "bullish",
    "growth", "profit", "gains", "exceeded", "exceeds", "upside", "expansion",
    "accelerat", "positive", "boost", "boosts", "momentum", "milestone",
])
_BEAR_WORDS = frozenset([
    "miss", "misses", "downgrade", "downgrades", "fall", "falls", "drop", "drops",
    "weak", "sell", "underperform", "concern", "concerns", "warn", "warns",
    "warning", "cut", "cuts", "loss", "losses", "decline", "disappoint",
    "disappoints", "bearish", "negative", "shortfall", "slowdown", "risk",
    "risks", "lawsuit", "probe", "investigation", "recall", "layoff", "layoffs",
])


def _text_sentiment(text: str) -> int:
    """Return +1 (bullish), -1 (bearish), 0 (neutral) for a piece of text."""
    lower = text.lower()
    bull  = any(w in lower for w in _BULL_WORDS)
    bear  = any(w in lower for w in _BEAR_WORDS)
    if bull and not bear:
        return 1
    if bear and not bull:
        return -1
    return 0


def _fetch_google_trends(ticker: str, company_name: str = "",
                         custom_terms: list | None = None) -> dict:
    """Fetch Google Trends search-interest data for the ticker and optional custom keywords.

    Uses pytrends (unofficial Google Trends API — no key required).
    Compares recent 30-day average search interest vs prior 60-day baseline.
    A ratio >1 means rising interest; >2 means a significant spike.

    custom_terms: optional list of sector-theme keywords (e.g. ['HBM memory', 'AI chips'])
                  compared in addition to the ticker itself.

    Returns dict with:
        trend_ratio     — recent/baseline ratio for the ticker keyword
        trend_kw        — keyword used (ticker or company name, whichever had more data)
        best_theme_ratio — highest ratio across ticker + custom_terms
        best_theme_kw   — keyword that achieved best_theme_ratio
        all_gt_ratios   — {keyword: ratio} for all terms attempted
        trend_error     — error string if unavailable (None when successful)
    """
    try:
        from pytrends.request import TrendReq
        import time as _time

        pytrends = TrendReq(hl="en-US", tz=360, timeout=(10, 30), retries=1, backoff_factor=0.5)

        def _ratio_for_kw(kw: str) -> float | None:
            """Return recent/baseline ratio for a single keyword. Returns None on failure."""
            try:
                pytrends.build_payload([kw], cat=0, timeframe="today 3-m", geo="")
                df = pytrends.interest_over_time()
                if df is None or df.empty or kw not in df.columns:
                    return None
                series = df[kw].dropna()
                if len(series) < 10:
                    return None
                # Last ~4 data points ≈ most recent ~4 weeks; everything before = baseline
                recent   = float(series.iloc[-4:].mean())
                baseline = float(series.iloc[:-4].mean())
                if baseline <= 0:
                    return None
                return round(recent / baseline, 2)
            except Exception:
                return None

        # Primary: try ticker symbol first, fall back to company name
        primary_kw    = ticker
        primary_ratio = _ratio_for_kw(ticker)

        if primary_ratio is None and company_name and company_name != ticker:
            short_name = company_name.split()[0]  # just first word (e.g. "Apple" not "Apple Inc.")
            r = _ratio_for_kw(short_name)
            if r is not None:
                primary_ratio = r
                primary_kw    = short_name

        all_ratios: dict = {}
        if primary_ratio is not None:
            all_ratios[primary_kw] = primary_ratio

        best_ratio = primary_ratio or 0.0
        best_kw    = primary_kw

        # Custom sector-theme keywords
        terms_to_check = (custom_terms or [])[:4]  # cap at 4 to stay within rate limits
        for term in terms_to_check:
            if not term.strip():
                continue
            _time.sleep(0.6)   # polite delay — Google Trends throttles aggressive scrapers
            r = _ratio_for_kw(term)
            if r is not None:
                all_ratios[term] = r
                if r > best_ratio:
                    best_ratio = r
                    best_kw    = term

        return {
            "trend_ratio":      primary_ratio,
            "trend_kw":         primary_kw,
            "best_theme_ratio": best_ratio if best_ratio > 0 else primary_ratio,
            "best_theme_kw":    best_kw,
            "all_gt_ratios":    all_ratios,
            "trend_error":      None,
        }

    except ImportError:
        return {
            "trend_ratio":   None, "trend_error": "pytrends not installed (pip install pytrends)",
            "trend_kw":      ticker, "best_theme_ratio": None, "best_theme_kw": ticker,
            "all_gt_ratios": {},
        }
    except Exception as e:
        return {
            "trend_ratio":   None, "trend_error": str(e),
            "trend_kw":      ticker, "best_theme_ratio": None, "best_theme_kw": ticker,
            "all_gt_ratios": {},
        }


def _fetch_reddit_sentiment(ticker: str) -> dict:
    """Search r/wallstreetbets, r/stocks, r/investing for ticker mentions (last 7 days).

    Scores each post via keyword sentiment on title + body text, and uses the
    Reddit upvote_ratio as a secondary signal (≥0.75 = bullish crowd, <0.35 = bearish).

    Requires env vars: REDDIT_CLIENT_ID, REDDIT_CLIENT_SECRET.
    Optional: REDDIT_USER_AGENT (defaults to 'InvestmentResearch/1.0').
    Returns None fields gracefully when credentials are absent.
    """
    import os
    client_id     = os.environ.get("REDDIT_CLIENT_ID", "")
    client_secret = os.environ.get("REDDIT_CLIENT_SECRET", "")
    if not client_id or not client_secret:
        return {
            "reddit_mentions": None, "reddit_bullish_pct": None,
            "reddit_error": "Missing REDDIT_CLIENT_ID / REDDIT_CLIENT_SECRET env vars",
        }

    try:
        import praw
        user_agent = os.environ.get("REDDIT_USER_AGENT", "InvestmentResearch/1.0")
        reddit = praw.Reddit(
            client_id=client_id,
            client_secret=client_secret,
            user_agent=user_agent,
        )

        subreddits = ["wallstreetbets", "stocks", "investing", "StockMarket"]
        cutoff_24h = datetime.datetime.now(datetime.timezone.utc) - datetime.timedelta(hours=24)

        all_posts: list = []
        for sub_name in subreddits:
            try:
                results = reddit.subreddit(sub_name).search(
                    ticker, time_filter="week", limit=15, sort="new"
                )
                all_posts.extend(list(results))
            except Exception:
                pass

        if not all_posts:
            return {"reddit_mentions": 0, "reddit_24h": 0,
                    "reddit_bullish": 0, "reddit_bearish": 0, "reddit_bullish_pct": None}

        bullish = bearish = recent_24h = 0
        for post in all_posts:
            text          = (post.title or "") + " " + (getattr(post, "selftext", "") or "")
            kw_sentiment  = _text_sentiment(text)
            upvote_ratio  = float(getattr(post, "upvote_ratio", 0.5) or 0.5)
            # Bullish if keyword OR crowd upvote signal; bearish if keyword AND downvoted
            if kw_sentiment > 0 or upvote_ratio >= 0.75:
                bullish += 1
            elif kw_sentiment < 0 or upvote_ratio < 0.35:
                bearish += 1
            try:
                ts = datetime.datetime.fromtimestamp(post.created_utc, tz=datetime.timezone.utc)
                if ts > cutoff_24h:
                    recent_24h += 1
            except Exception:
                pass

        total_scored  = bullish + bearish
        bullish_pct   = (bullish / total_scored) if total_scored > 0 else None

        return {
            "reddit_mentions":    len(all_posts),
            "reddit_24h":         recent_24h,
            "reddit_bullish":     bullish,
            "reddit_bearish":     bearish,
            "reddit_bullish_pct": bullish_pct,
        }
    except Exception as e:
        return {"reddit_mentions": None, "reddit_bullish_pct": None, "reddit_error": str(e)}


# ── Live data fetch ────────────────────────────────────────────────────────────

def fetch_speculative_data(ticker: str, api_key: str, polygon_key: str = "",
                           mock: bool = False, narrative_theme: str = "",
                           custom_terms: list[str] | None = None) -> dict:
    """Fetch all signals needed to score the speculative scorecard.

    Falls back to mock data if `mock=True` or on any network failure, so the
    pipeline always has something to work with.
    """
    if mock:
        return _mock_data(ticker)

    data: dict = {"ticker": ticker, "mock": False}

    base = "https://financialmodelingprep.com/stable"

    def _get(path, params=None):
        try:
            p = {"apikey": api_key}
            if params:
                p.update(params)
            r = _req.get(f"{base}/{path}", params=p, timeout=10)
            return r.json()
        except Exception:
            return None

    # ── Profile (FMP — name, sector, price) ──────────────────────────────────
    prof_raw = _get(f"profile?symbol={ticker}")
    profile  = (prof_raw[0] if isinstance(prof_raw, list) and prof_raw else prof_raw or {})
    data["profile"]       = profile
    data["company_name"]  = profile.get("companyName") or ticker
    data["sector"]        = profile.get("sector") or "Unknown"
    data["current_price"] = float(profile.get("price") or 0) or None

    # ── Market cap / shares — Polygon ticker details (more reliable than FMP) ─
    poly_details          = _fetch_polygon_ticker_details(ticker, polygon_key)
    data["market_cap"]    = float(poly_details.get("market_cap") or 0) or None
    data["shares_out"]    = float(poly_details.get("weighted_shares_outstanding") or 0) or None
    data["float_shares"]  = float(poly_details.get("share_class_shares_outstanding") or 0) or None
    # Fallback to FMP profile if Polygon didn't return market cap
    if not data["market_cap"]:
        data["market_cap"]  = float(profile.get("mktCap") or 0) or None
    if not data["shares_out"]:
        data["shares_out"]  = float(profile.get("sharesOutstanding") or 0) or None

    # ── Balance sheet (latest) ────────────────────────────────────────────────
    bs_raw = _get(f"balance-sheet-statement?symbol={ticker}&limit=1")
    bs     = (bs_raw[0] if isinstance(bs_raw, list) and bs_raw else {})
    cash       = float(bs.get("cashAndCashEquivalents") or 0)
    total_debt = float(bs.get("totalDebt") or 0)
    data["cash"]       = cash
    data["total_debt"] = total_debt
    data["net_debt"]   = total_debt - cash       # positive = net debt, negative = net cash

    # ── Income statement (latest, for trailing revenue + EBITDA) ─────────────
    is_raw = _get(f"income-statement?symbol={ticker}&limit=1")
    is_    = (is_raw[0] if isinstance(is_raw, list) and is_raw else {})
    data["trailing_rev"] = float(is_.get("revenue") or 0) or None
    data["ebitda_ttm"]   = float(is_.get("ebitda") or 0) or None

    # ── EV ────────────────────────────────────────────────────────────────────
    mktcap   = data["market_cap"] or 0
    net_debt = data["net_debt"]   or 0
    data["current_ev"] = mktcap + net_debt

    # ── Analyst estimates (fwd revenue) ───────────────────────────────────────
    ae_raw = _get(f"analyst-estimates?symbol={ticker}&period=annual&limit=3")
    fwd_revs = []
    if isinstance(ae_raw, list):
        today_yr = datetime.date.today().year
        for e in sorted(ae_raw, key=lambda x: x.get("date", "")):
            yr = int(str(e.get("date", "0"))[:4])
            if yr >= today_yr:
                rev = float(e.get("estimatedRevenueAvg") or e.get("revenueAvg") or 0)
                if rev:
                    fwd_revs.append(rev)
    data["fwd_rev_1yr"] = fwd_revs[0] if len(fwd_revs) > 0 else None
    data["fwd_rev_2yr"] = fwd_revs[1] if len(fwd_revs) > 1 else None

    # ── OHLCV + all technical indicators — Polygon + pandas-ta ────────────────
    # Polygon returns full bars (O/H/L/C/V), sorted ascending. pandas-ta then
    # computes RSI/MACD/EMA locally — no extra API calls needed.
    data["price_vs_50ma"]  = None
    data["price_vs_200ma"] = None
    data["vol_10d_avg"]    = None
    data["vol_50d_avg"]    = None
    data["rsi_14"]         = None
    data["macd"]           = data["macd_signal"] = data["macd_hist"] = None
    data["ema_21"]         = data["ema_50"]       = None
    data["high_52w"]       = data["low_52w"]      = data["pct_from_52w_high"] = None

    poly_bars = _fetch_polygon_ohlcv(ticker, polygon_key)
    if poly_bars and len(poly_bars) >= 50:
        closes_asc = [float(b["c"]) for b in poly_bars]
        highs_asc  = [float(b["h"]) for b in poly_bars]
        lows_asc   = [float(b["l"]) for b in poly_bars]
        vols_asc   = [float(b["v"]) for b in poly_bars]

        # Descending lists for simple slice-based MA / volume calcs
        closes_desc = list(reversed(closes_asc))
        vols_desc   = list(reversed(vols_asc))

        if len(closes_desc) >= 200:
            cur   = closes_desc[0]
            ma50  = sum(closes_desc[:50])  / 50
            ma200 = sum(closes_desc[:200]) / 200
            data["price_vs_50ma"]  = (cur / ma50  - 1) if ma50  else None
            data["price_vs_200ma"] = (cur / ma200 - 1) if ma200 else None

        if len(vols_desc) >= 50:
            data["vol_10d_avg"] = sum(vols_desc[:10]) / 10
            data["vol_50d_avg"] = sum(vols_desc[:50]) / 50

        # 52-week high / low — use actual H/L bars from Polygon (not just close)
        year_bars = min(len(poly_bars), 252)
        year_highs = highs_asc[-year_bars:]
        year_lows  = lows_asc[-year_bars:]
        data["high_52w"] = max(year_highs)
        data["low_52w"]  = min(year_lows)
        if data["high_52w"] > 0 and closes_desc:
            data["pct_from_52w_high"] = (closes_desc[0] / data["high_52w"]) - 1

        # ta library — RSI/MACD/EMA computed from Polygon closes, no extra API calls
        try:
            import pandas as pd
            import ta as ta_lib

            close_s = pd.Series(closes_asc)

            rsi_v = ta_lib.momentum.RSIIndicator(close=close_s, window=14).rsi().iloc[-1]
            data["rsi_14"] = float(rsi_v) if pd.notna(rsi_v) else None

            macd_obj = ta_lib.trend.MACD(close=close_s)
            macd_v   = macd_obj.macd().iloc[-1]
            hist_v   = macd_obj.macd_diff().iloc[-1]
            sig_v    = macd_obj.macd_signal().iloc[-1]
            data["macd"]        = float(macd_v) if pd.notna(macd_v) else None
            data["macd_hist"]   = float(hist_v) if pd.notna(hist_v) else None
            data["macd_signal"] = float(sig_v)  if pd.notna(sig_v)  else None

            for period, key in [(21, "ema_21"), (50, "ema_50")]:
                try:
                    ema_v = ta_lib.trend.EMAIndicator(close=close_s, window=period).ema_indicator().iloc[-1]
                    data[key] = float(ema_v) if pd.notna(ema_v) else None
                except Exception:
                    pass
                # Manual fallback if library returns NaN
                if data[key] is None and len(closes_asc) >= period:
                    alpha = 2.0 / (period + 1)
                    ema = sum(closes_asc[:period]) / period
                    for v in closes_asc[period:]:
                        ema = alpha * v + (1 - alpha) * ema
                    data[key] = ema
        except Exception as e:
            data["ta_error"] = str(e)

        # ── Parabolic-move precursor indicators ──────────────────────────────
        # All computed from the Polygon OHLCV bars already in memory — no extra calls.
        # These detect the "coiling spring" + "smart-money accumulation" pattern
        # that historically precedes parabolic moves.
        data["bb_upper"]        = data["bb_lower"]      = data["bb_mid"]        = None
        data["bb_width_pct"]    = data["bb_pos_pct"]    = None
        data["bb_squeeze"]      = data["bb_expansion"]  = None
        data["obv_slope_pct"]   = data["obv_rising"]    = data["obv_divergence"] = None
        data["up_vol_share"]    = data["acc_dist_tier"] = None
        data["roc_20d"]         = data["roc_accelerating"] = None
        data["atr_14"]          = data["atr_pct_price"] = None
        try:
            import pandas as _pd
            import numpy as _np

            close_s = _pd.Series(closes_asc)
            high_s  = _pd.Series(highs_asc)
            low_s   = _pd.Series(lows_asc)
            vol_s   = _pd.Series(vols_asc)

            # Bollinger Bands (20-period SMA, 2 std dev). Width as % of mid band
            # is the cleanest cross-stock comparable metric (volatility-normalised).
            if len(close_s) >= 60:
                bb_mid_s   = close_s.rolling(20).mean()
                bb_std_s   = close_s.rolling(20).std()
                bb_up_s    = bb_mid_s + 2 * bb_std_s
                bb_dn_s    = bb_mid_s - 2 * bb_std_s
                bb_width_s = ((bb_up_s - bb_dn_s) / bb_mid_s).replace([_np.inf, -_np.inf], _np.nan)

                cur_mid   = float(bb_mid_s.iloc[-1])
                cur_up    = float(bb_up_s.iloc[-1])
                cur_dn    = float(bb_dn_s.iloc[-1])
                cur_width = float(bb_width_s.iloc[-1])
                cur_close = float(close_s.iloc[-1])

                data["bb_mid"]       = cur_mid
                data["bb_upper"]     = cur_up
                data["bb_lower"]     = cur_dn
                data["bb_width_pct"] = cur_width
                if cur_up > cur_dn:
                    data["bb_pos_pct"] = (cur_close - cur_dn) / (cur_up - cur_dn)

                # Squeeze: current BB width sits in the lowest 25% of the last 60 bars
                width_hist  = bb_width_s.dropna().iloc[-60:]
                if len(width_hist) >= 40:
                    p25 = float(width_hist.quantile(0.25))
                    data["bb_squeeze"] = bool(cur_width <= p25)
                    # Expansion: width was in lowest 25% within last 10 bars AND is
                    # now expanding (width up >15% from its 10-bar low). The
                    # classic coiled-spring breakout signature.
                    recent_widths = width_hist.iloc[-10:]
                    recent_min    = float(recent_widths.min())
                    was_squeezed_recently = recent_min <= p25
                    expansion = (cur_width > recent_min * 1.15) and was_squeezed_recently
                    data["bb_expansion"] = bool(expansion)

            # OBV (On-Balance Volume) — cumulative signed volume. A rising OBV with
            # flat price = institutional accumulation; classic pre-parabolic tell.
            if len(close_s) >= 30:
                price_diff = close_s.diff().fillna(0)
                signed_vol = vol_s.where(price_diff > 0, -vol_s.where(price_diff < 0, 0))
                obv_s = signed_vol.cumsum()

                obv_now   = float(obv_s.iloc[-1])
                obv_10ago = float(obv_s.iloc[-11]) if len(obv_s) >= 11 else float(obv_s.iloc[0])
                obv_denom = max(abs(obv_10ago), abs(obv_s.iloc[-30:]).max(), 1.0)
                obv_slope = (obv_now - obv_10ago) / obv_denom
                data["obv_slope_pct"] = float(obv_slope)
                data["obv_rising"]    = bool(obv_slope > 0.02)

                # Divergence: OBV up >5% relative to recent range, price flat or down
                if len(close_s) >= 11:
                    price_10d_change = float(close_s.iloc[-1] / close_s.iloc[-11] - 1)
                    data["obv_divergence"] = bool(
                        obv_slope > 0.05 and price_10d_change < 0.02
                    )

            # Accumulation/Distribution — share of last 20 days' volume that
            # occurred on up-close days. >0.60 = accumulation, <0.40 = distribution.
            if len(close_s) >= 21:
                last20_close = close_s.iloc[-20:]
                last20_vol   = vol_s.iloc[-20:]
                last20_prev  = close_s.iloc[-21:-1].values
                up_mask      = last20_close.values > last20_prev
                up_vol       = float(last20_vol[up_mask].sum())
                tot_vol      = float(last20_vol.sum())
                if tot_vol > 0:
                    share = up_vol / tot_vol
                    data["up_vol_share"] = share
                    if share >= 0.60:
                        data["acc_dist_tier"] = "accumulation"
                    elif share <= 0.40:
                        data["acc_dist_tier"] = "distribution"
                    else:
                        data["acc_dist_tier"] = "neutral"

            # 20-day Rate of Change + acceleration check
            if len(close_s) >= 26:
                roc_now   = float(close_s.iloc[-1]  / close_s.iloc[-21] - 1)
                roc_5ago  = float(close_s.iloc[-6]  / close_s.iloc[-26] - 1)
                data["roc_20d"]            = roc_now
                data["roc_accelerating"]   = bool(roc_now > roc_5ago and roc_now > 0)

            # ATR-14 (Wilder's smoothing approximated by simple mean of TR)
            if len(close_s) >= 15:
                prev_close = close_s.shift(1)
                tr = _pd.concat([
                    high_s - low_s,
                    (high_s - prev_close).abs(),
                    (low_s  - prev_close).abs(),
                ], axis=1).max(axis=1)
                atr14 = float(tr.iloc[-14:].mean())
                data["atr_14"] = atr14
                if cur_close := float(close_s.iloc[-1]):
                    data["atr_pct_price"] = atr14 / cur_close
        except Exception as e:
            data["parabolic_ta_error"] = str(e)

    # ── Short interest ────────────────────────────────────────────────────────
    si_raw = _get(f"short-interest?symbol={ticker}")
    data["short_float_pct"] = None
    data["days_to_cover"]   = None
    if isinstance(si_raw, list) and si_raw:
        si = si_raw[0]
        data["short_float_pct"] = float(si.get("shortPercentOfFloat") or si.get("shortPercent") or 0) or None
        if data["short_float_pct"]:
            data["short_float_pct"] *= 100   # convert 0.18 → 18.0
        data["days_to_cover"] = float(si.get("daysToCover") or si.get("shortRatio") or 0) or None

    # ── Upgrades / downgrades (last 90 days) ──────────────────────────────────
    cutoff = (datetime.date.today() - datetime.timedelta(days=90)).isoformat()
    ud_raw = _get(f"upgrades-downgrades?symbol={ticker}&limit=50")
    ups = downs = 0
    ud_list = ud_raw if isinstance(ud_raw, list) else (ud_raw.get("data") or [] if isinstance(ud_raw, dict) else [])
    if not ud_list and isinstance(ud_raw, dict):
        data["ud_error"] = ud_raw.get("message") or str(ud_raw)
    for e in ud_list:
        if (e.get("publishedDate") or e.get("date") or "") < cutoff:
            continue
        action = (e.get("action") or e.get("newGrade") or e.get("gradingCompany") or "").lower()
        grade  = (e.get("newGrade") or "").lower()
        if any(w in action or w in grade for w in ("upgrade", "buy", "outperform", "overweight", "positive")):
            ups += 1
        elif any(w in action or w in grade for w in ("downgrade", "sell", "underperform", "underweight", "negative")):
            downs += 1
    data["upgrades_90d"]   = ups
    data["downgrades_90d"] = downs

    # ── Analyst estimate revision (compare latest to 3mo-ago estimate) ────────
    data["est_rev_pct"] = None
    if isinstance(ae_raw, list) and len(ae_raw) >= 2:
        try:
            new_est = float(ae_raw[0].get("estimatedRevenueAvg") or 0)
            old_est = float(ae_raw[1].get("estimatedRevenueAvg") or 0)
            if old_est:
                data["est_rev_pct"] = (new_est - old_est) / abs(old_est)
        except Exception:
            pass

    # ── Insider trading (last 90 days, open-market only) ──────────────────────
    cutoff_ins = (datetime.date.today() - datetime.timedelta(days=90)).isoformat()
    ins_raw = _get(f"insider-trading?symbol={ticker}&limit=50")
    buys = sells = 0
    buy_usd = 0.0
    if isinstance(ins_raw, list):
        for e in ins_raw:
            if (e.get("transactionDate") or "") < cutoff_ins:
                continue
            ttype = e.get("transactionType") or ""
            if ttype == "P-Purchase":
                buys += 1
                buy_usd += float(e.get("value") or 0)
            elif ttype == "S-Sale":
                sells += 1
    data["insider_buys"]    = buys
    data["insider_buy_usd"] = buy_usd
    data["insider_sells"]   = sells

    # ── Options activity (yfinance — no free alternative for options) ─────────
    data["put_call_ratio"] = None
    data["call_oi"]        = None
    data["put_oi"]         = None
    try:
        import yfinance as yf
        yt   = yf.Ticker(ticker)
        exps = yt.options
        if exps:
            chain   = yt.option_chain(exps[0])
            call_oi = chain.calls["openInterest"].sum()
            put_oi  = chain.puts["openInterest"].sum()
            data["call_oi"]        = int(call_oi)
            data["put_oi"]         = int(put_oi)
            data["put_call_ratio"] = (put_oi / call_oi) if call_oi else None
    except Exception as e:
        data["options_error"] = str(e)

    # ── News sentiment — Polygon NLP (replaces yfinance keyword scraping) ─────
    pn = _fetch_polygon_news(ticker, polygon_key)
    data["news_count"]       = pn.get("news_count")
    data["news_7d"]          = pn.get("news_7d")
    data["news_bullish"]     = pn.get("news_bullish")
    data["news_bearish"]     = pn.get("news_bearish")
    data["news_bullish_pct"] = pn.get("news_bullish_pct")

    # ── Reddit sentiment (PRAW — requires REDDIT_CLIENT_ID + REDDIT_CLIENT_SECRET) ─
    rd = _fetch_reddit_sentiment(ticker)
    data["reddit_mentions"]    = rd.get("reddit_mentions")
    data["reddit_24h"]         = rd.get("reddit_24h")
    data["reddit_bullish"]     = rd.get("reddit_bullish")
    data["reddit_bearish"]     = rd.get("reddit_bearish")
    data["reddit_bullish_pct"] = rd.get("reddit_bullish_pct")

    # ── Google Trends (pytrends — no API key required) ────────────────────────
    gt = _fetch_google_trends(
        ticker,
        company_name = data.get("company_name", ""),
        custom_terms = custom_terms or [],
    )
    data["trend_ratio"]      = gt.get("trend_ratio")
    data["trend_kw"]         = gt.get("trend_kw")
    data["best_theme_ratio"] = gt.get("best_theme_ratio")
    data["best_theme_kw"]    = gt.get("best_theme_kw")
    data["all_gt_ratios"]    = gt.get("all_gt_ratios")
    data["trend_error"]      = gt.get("trend_error")

    return data


# ── Scoring functions ──────────────────────────────────────────────────────────

def _score_momentum(data: dict) -> tuple[str, int, str]:
    rsi   = data.get("rsi_14")
    vs50  = data.get("price_vs_50ma")
    vs200 = data.get("price_vs_200ma")

    if rsi is None:
        return "MOD", 5, "RSI unavailable — defaulting to neutral"

    above_50  = vs50  is not None and vs50  > 0
    above_200 = vs200 is not None and vs200 > 0

    # Ideal speculative zone: RSI 50-75, above both MAs (trending, not extended)
    if 50 <= rsi <= 75 and above_50 and above_200:
        return "HIGH", 10, f"RSI {rsi:.0f} in momentum zone; +{vs50*100:.0f}% vs 50d MA, +{vs200*100:.0f}% vs 200d MA"
    if rsi > 75 and above_50 and above_200:
        return "MOD", 5, f"RSI {rsi:.0f} extended (overbought risk) — strong trend but pullback possible"
    if 40 <= rsi < 50 and (above_50 or above_200):
        return "MOD", 5, f"RSI {rsi:.0f} recovering — early-stage setup, watch for breakout above 50"
    if rsi > 50 and (above_50 or above_200):
        return "MOD", 5, f"RSI {rsi:.0f} positive but only above one MA — partial momentum"
    return "LOW", 0, f"RSI {rsi:.0f} — no clear upward momentum, price below key moving averages"


def _score_volume(data: dict) -> tuple[str, int, str]:
    v10 = data.get("vol_10d_avg")
    v50 = data.get("vol_50d_avg")

    if not v10 or not v50 or v50 == 0:
        return "MOD", 5, "Volume data unavailable — defaulting to neutral"

    ratio = v10 / v50
    if ratio >= 1.75:
        return "HIGH", 10, f"10d avg volume is {ratio:.1f}x the 50d baseline — strong institutional accumulation signal"
    if ratio >= 1.20:
        return "MOD", 5, f"10d avg volume is {ratio:.1f}x the 50d baseline — elevated but not unusual"
    return "LOW", 0, f"10d avg volume is {ratio:.1f}x the 50d baseline — no unusual volume activity"


def _score_short_interest(data: dict) -> tuple[str, int, str]:
    sf  = data.get("short_float_pct")
    dtc = data.get("days_to_cover")

    if sf is None:
        return "MOD", 5, "Short interest data unavailable — defaulting to neutral"

    if sf >= 15 and (dtc is None or dtc >= 3):
        dtc_str = f", {dtc:.1f} days to cover" if dtc else ""
        return "HIGH", 10, f"{sf:.1f}% short float{dtc_str} — significant short squeeze fuel"
    if sf >= 8:
        dtc_str = f", {dtc:.1f} days to cover" if dtc else ""
        return "MOD", 5, f"{sf:.1f}% short float{dtc_str} — moderate short interest"
    return "LOW", 0, f"{sf:.1f}% short float — low short interest, limited squeeze potential"


def _score_analyst_revisions(data: dict) -> tuple[str, int, str]:
    ups   = data.get("upgrades_90d", 0) or 0
    downs = data.get("downgrades_90d", 0) or 0
    rev   = data.get("est_rev_pct")

    total = ups + downs
    if total == 0 and rev is None:
        return "MOD", 5, "No recent analyst actions or estimate data available"

    upgrade_bias = ups > downs
    strong_rev   = rev is not None and rev > 0.05    # >5% revision up
    any_neg_rev  = rev is not None and rev < -0.05

    if upgrade_bias and strong_rev:
        rev_str = f"+{rev*100:.0f}%" if rev else ""
        return "HIGH", 10, f"{ups} upgrades vs {downs} downgrades (90d); estimates revised {rev_str} — strong positive momentum"
    if upgrade_bias or strong_rev:
        rev_str = f"{rev*100:+.0f}%" if rev else ""
        return "MOD", 5, f"{ups} upgrades vs {downs} downgrades (90d){'; estimates ' + rev_str if rev else ''}"
    if any_neg_rev or downs > ups:
        rev_str = f"{rev*100:+.0f}%" if rev else ""
        return "LOW", 0, f"{ups} upgrades vs {downs} downgrades (90d){'; estimates ' + rev_str if rev else ''} — negative analyst momentum"
    return "MOD", 5, f"{ups} upgrades vs {downs} downgrades (90d) — neutral analyst activity"


def _score_float(data: dict) -> tuple[str, int, str]:
    mktcap      = data.get("market_cap")
    float_sh    = data.get("float_shares")
    shares_out  = data.get("shares_out")
    cur_price   = data.get("current_price")

    float_val = None
    if float_sh and cur_price:
        float_val = float_sh * cur_price

    # Use market cap as primary signal (easier to get consistently)
    if mktcap:
        mktcap_b = mktcap / 1e9
        if mktcap_b < 1.5:
            size_tier = "HIGH"
            size_note = f"${mktcap_b:.1f}B market cap (micro/small cap) — high potential for explosive moves"
        elif mktcap_b < 10:
            size_tier = "MOD"
            size_note = f"${mktcap_b:.1f}B market cap (small/mid cap) — meaningful but not micro-cap explosive"
        else:
            size_tier = "LOW"
            size_note = f"${mktcap_b:.1f}B market cap (large cap) — needs massive catalysts to move significantly"

        if float_val:
            float_b = float_val / 1e9
            size_note += f"; float ~${float_b:.1f}B"
        return size_tier, TIER_PTS[size_tier], size_note

    return "MOD", 5, "Market cap data unavailable — defaulting to neutral"


def _score_insider(data: dict) -> tuple[str, int, str]:
    buys    = data.get("insider_buys", 0) or 0
    buy_usd = data.get("insider_buy_usd", 0) or 0
    sells   = data.get("insider_sells", 0) or 0

    if buys == 0 and sells == 0:
        return "MOD", 5, "No insider transactions in the last 90 days"
    if buys >= 2 and buy_usd >= 200_000 and buys > sells:
        return "HIGH", 10, f"{buys} open-market insider purchase(s), ${buy_usd/1e3:.0f}K total (90d) — strong smart-money signal"
    if buys >= 1 and buys >= sells:
        return "MOD", 5, f"{buys} insider purchase(s), ${buy_usd/1e3:.0f}K total vs {sells} sale(s) — modest positive signal"
    return "LOW", 0, f"Net insider selling ({sells} sales vs {buys} buys, 90d) — negative signal"


def _score_downside_floor(data: dict) -> tuple[str, int, str]:
    net_debt  = data.get("net_debt")     # positive = net debt, negative = net cash
    ebitda    = data.get("ebitda_ttm")
    cash      = data.get("cash", 0) or 0
    tot_debt  = data.get("total_debt", 0) or 0

    if net_debt is None:
        return "MOD", 5, "Balance sheet data unavailable"

    if net_debt <= 0:
        # Net cash position
        net_cash_m = abs(net_debt) / 1e6
        return "HIGH", 10, f"Net cash of ${net_cash_m:.0f}M — downside protected, no debt cliff risk"

    # Net debt — check leverage
    if ebitda and ebitda > 0:
        lev = net_debt / ebitda
        if lev < 2.0:
            return "MOD", 5, f"Net debt ${net_debt/1e6:.0f}M, D/EBITDA {lev:.1f}x — manageable leverage"
        if lev < 4.0:
            return "LOW", 0, f"Net debt ${net_debt/1e6:.0f}M, D/EBITDA {lev:.1f}x — elevated leverage, catalyst must arrive before debt pressure"
        return "LOW", 0, f"Net debt ${net_debt/1e6:.0f}M, D/EBITDA {lev:.1f}x — high leverage, binary risk if catalyst fails"

    if tot_debt > cash * 3:
        return "LOW", 0, f"Debt ${tot_debt/1e6:.0f}M significantly exceeds cash ${cash/1e6:.0f}M — material downside risk"
    return "MOD", 5, f"Cash ${cash/1e6:.0f}M vs debt ${tot_debt/1e6:.0f}M — moderate balance sheet"


def _score_options(data: dict) -> tuple[str, int, str]:
    pc  = data.get("put_call_ratio")
    c_oi = data.get("call_oi")
    p_oi = data.get("put_oi")

    if pc is None:
        err = data.get("options_error", "")
        note = f"Options data unavailable — {err}" if err else "Options data unavailable (no listed options or data error)"
        return "MOD", 5, note

    if pc < 0.45:
        return "HIGH", 10, f"Put/Call ratio {pc:.2f} — heavy call-side positioning, smart money expressing bullish conviction"
    if pc < 0.75:
        return "MOD", 5, f"Put/Call ratio {pc:.2f} — moderately bullish options positioning"
    if pc > 1.20:
        return "LOW", 0, f"Put/Call ratio {pc:.2f} — put-heavy positioning, market hedging against downside"
    return "MOD", 5, f"Put/Call ratio {pc:.2f} — roughly neutral options positioning"


def _score_social_trend(data: dict) -> tuple[str, int, str]:
    """Score social/trend momentum across three sub-signals:

      1. Polygon news sentiment  — bullish vs bearish ratio from Polygon NLP (last 20 articles)
      2. Reddit crowd positioning — INVERTED: crowded/bullish crowd = late; under-the-radar = early
      3. Google Trends           — recent 30d search interest vs prior 60d baseline (rising = bullish)

    HIGH = ≥2 bullish sub-signals · MOD = 1 bullish or all unavailable · LOW = 0 bullish with ≥1 bearish
    """
    bullish_signals = []
    notes = []

    # Sub-signal 1: Polygon news sentiment
    news_bp    = data.get("news_bullish_pct")
    news_bull  = data.get("news_bullish", 0) or 0
    news_bear  = data.get("news_bearish", 0) or 0
    news_count = data.get("news_count") or 0
    news_7d    = data.get("news_7d") or 0

    if news_bp is not None:
        activity = f"{news_7d} articles in last 7d" if news_7d else f"{news_count} total"
        if news_bp >= 0.65:
            bullish_signals.append(True)
            notes.append(f"News: {news_bp*100:.0f}% bullish ({news_bull}+ / {news_bear}−, {activity})")
        elif news_bp >= 0.40:
            notes.append(f"News: {news_bp*100:.0f}% bullish — mixed sentiment ({activity})")
        else:
            bullish_signals.append(False)
            notes.append(f"News: {news_bp*100:.0f}% bullish — bearish headline bias ({activity})")
    elif news_count == 0:
        notes.append("News: no recent articles found")
    else:
        notes.append("News: unavailable")

    # Sub-signal 2: Reddit sentiment (WSB + stocks + investing) — INVERTED SCORING
    # Contrarian logic: extreme crowd bullishness on retail forums means you're late.
    # What we want is UNDER-THE-RADAR positioning, not confirmation that WSB already piled in.
    rd_bp    = data.get("reddit_bullish_pct")
    rd_bull  = data.get("reddit_bullish", 0) or 0
    rd_bear  = data.get("reddit_bearish", 0) or 0
    rd_total = data.get("reddit_mentions") or 0
    rd_24h   = data.get("reddit_24h") or 0

    if rd_bp is not None:
        activity = f"{rd_24h} posts in last 24h" if rd_24h else f"{rd_total} posts this week"
        if rd_bp >= 0.75 and rd_total >= 15:
            # Crowded retail trade — late entry risk
            bullish_signals.append(False)
            notes.append(f"Reddit: {rd_bp*100:.0f}% bullish ({rd_bull}+ / {rd_bear}−, {activity}) — crowded retail trade, asymmetry eroded")
        elif rd_total < 8:
            # Under the radar — early positioning opportunity
            bullish_signals.append(True)
            notes.append(f"Reddit: low visibility ({rd_total} mentions this week) — not yet retail-crowded, potential early setup")
        elif rd_bp < 0.35:
            # Crowd is skeptical — confirms weak setup signal
            bullish_signals.append(False)
            notes.append(f"Reddit: {rd_bp*100:.0f}% bullish — crowd skeptical ({activity})")
        else:
            # Moderate interest, mixed sentiment — neutral; don't penalise or reward
            notes.append(f"Reddit: {rd_bp*100:.0f}% bullish — moderate interest ({activity}), no extreme crowding")
    elif rd_total == 0:
        # No mentions at all — stock under the radar (positive for asymmetry)
        bullish_signals.append(True)
        notes.append("Reddit: no mentions found — not on retail radar, early positioning potential")
    else:
        notes.append("Reddit: credentials not configured")

    # Sub-signal 3: Google Trends — search interest vs 60-day baseline
    # Rising search interest (ratio > 1) suggests growing awareness/discovery.
    # Unlike Reddit, this is NOT inverted — a spike from a low base is bullish.
    gt_ratio  = data.get("best_theme_ratio") or data.get("trend_ratio")
    gt_kw     = data.get("best_theme_kw")    or data.get("trend_kw", "ticker")
    gt_err    = data.get("trend_error")

    if gt_ratio is not None:
        if gt_ratio >= 2.0:
            bullish_signals.append(True)
            notes.append(f"Google Trends: {gt_ratio:.1f}× spike in '{gt_kw}' searches vs 60d baseline — significant discovery surge")
        elif gt_ratio >= 1.3:
            # Elevated but not extreme — neutral (don't reward moderately rising stocks)
            notes.append(f"Google Trends: {gt_ratio:.1f}× for '{gt_kw}' — elevated search interest, not extreme")
        elif gt_ratio <= 0.7:
            bullish_signals.append(False)
            notes.append(f"Google Trends: {gt_ratio:.1f}× for '{gt_kw}' — declining search interest, fading attention")
        else:
            notes.append(f"Google Trends: {gt_ratio:.1f}× for '{gt_kw}' — near-baseline search volume")
    else:
        err_msg = f" ({gt_err})" if gt_err else ""
        notes.append(f"Google Trends: unavailable{err_msg}")

    n_bullish = sum(1 for b in bullish_signals if b is True)
    n_bearish = sum(1 for b in bullish_signals if b is False)

    if n_bullish >= 2:
        tier, pts = "HIGH", 10
    elif n_bullish == 1:
        tier, pts = "MOD", 5
    elif n_bullish == 0 and n_bearish == 0:
        tier, pts = "MOD", 5   # all unavailable — neutral, don't penalise
    else:
        tier, pts = "LOW", 0

    return tier, pts, " · ".join(notes)


def _score_technical_analysis(data: dict) -> tuple[str, int, str]:
    """Score the technical setup using 7 sub-signals that together describe a
    pre-parabolic structure.

    The constellation we are looking for is the classic "coiled-spring + smart-
    money accumulation" pattern that historically precedes outsized moves:

      Trend-confirmation block (3 signals):
        1. MACD crossover           — momentum has turned positive
        2. 52-week high proximity   — price is in a breakout watch zone
        3. EMA stack alignment      — price > EMA21 > EMA50 (uptrend)

      Pre-parabolic precursor block (4 signals):
        4. Bollinger Band squeeze   — volatility compression (coiled spring)
        5. OBV accumulation         — cumulative volume rising (institutional bid)
        6. Up-day volume dominance  — accumulation vs distribution character
        7. ROC acceleration         — rate of change turning up from flat/negative

    Scoring:
        HIGH = 5+ bullish votes (strong setup, multiple precursors firing)
        MOD  = 3–4 bullish votes
        LOW  = 0–2 bullish votes
    """
    votes = []  # list of (key, is_bullish, label)
    notes = []

    def _vote(key, bullish, label):
        if bullish is None:
            return
        votes.append((key, bool(bullish), label))

    # ── 1. MACD ─────────────────────────────────────────────────────────────
    macd      = data.get("macd")
    macd_sig  = data.get("macd_signal")
    macd_hist = data.get("macd_hist")
    if macd is not None and macd_sig is not None and macd_hist is not None:
        if macd_hist > 0 and macd > macd_sig:
            _vote("macd", True, f"MACD+ ({macd_hist:+.2f})")
            notes.append(f"MACD bullish · hist {macd_hist:+.2f}")
        elif macd_hist < 0:
            _vote("macd", False, f"MACD− ({macd_hist:+.2f})")
            notes.append(f"MACD bearish · hist {macd_hist:+.2f}")
        else:
            notes.append("MACD neutral")

    # ── 2. 52-week high proximity ──────────────────────────────────────────
    pct_off = data.get("pct_from_52w_high")
    if pct_off is not None:
        if pct_off >= -0.08:
            _vote("52w_high", True, "near 52w high")
            notes.append(f"{pct_off*100:+.1f}% from 52w high — breakout zone")
        elif pct_off >= -0.25:
            notes.append(f"{pct_off*100:+.1f}% from 52w high — mid-range")
        else:
            _vote("52w_high", False, "deep off 52w high")
            notes.append(f"{pct_off*100:+.1f}% from 52w high — deep")

    # ── 3. EMA stack ───────────────────────────────────────────────────────
    price = data.get("current_price")
    ema21 = data.get("ema_21")
    ema50 = data.get("ema_50")
    if price and ema21 and ema50:
        if price > ema21 > ema50:
            _vote("ema_stack", True, "EMA21>EMA50, price above")
            notes.append("EMA stack bullish")
        elif price < ema21 or ema21 < ema50:
            _vote("ema_stack", False, "EMA stack broken")
            notes.append("EMA stack bearish")
        else:
            notes.append("EMA stack mixed")

    # ── 4. Bollinger Band squeeze / expansion ──────────────────────────────
    bb_sq  = data.get("bb_squeeze")
    bb_exp = data.get("bb_expansion")
    bb_w   = data.get("bb_width_pct")
    if bb_exp:
        _vote("bb", True, "BB expanding from squeeze")
        notes.append(f"BB expansion from squeeze · width {bb_w*100:.1f}%" if bb_w else "BB expansion")
    elif bb_sq:
        # A pure squeeze without expansion isn't yet a directional signal — it's
        # a "watch" condition. Don't vote either way; just narrate.
        notes.append(f"BB squeeze active · coiled-spring setup (width {bb_w*100:.1f}%)" if bb_w else "BB squeeze")
    elif bb_sq is False and bb_w is not None:
        notes.append(f"BB normal-range volatility (width {bb_w*100:.1f}%)")

    # ── 5. OBV accumulation / divergence ───────────────────────────────────
    obv_rising = data.get("obv_rising")
    obv_div    = data.get("obv_divergence")
    obv_slope  = data.get("obv_slope_pct")
    if obv_div:
        _vote("obv", True, "OBV bullish divergence")
        notes.append(f"OBV diverging · {obv_slope*100:+.1f}% with price flat")
    elif obv_rising:
        _vote("obv", True, "OBV rising")
        notes.append(f"OBV rising · {obv_slope*100:+.1f}%")
    elif obv_slope is not None and obv_slope < -0.05:
        _vote("obv", False, "OBV declining")
        notes.append(f"OBV declining · {obv_slope*100:+.1f}%")
    elif obv_slope is not None:
        notes.append(f"OBV flat · {obv_slope*100:+.1f}%")

    # ── 6. Accumulation/Distribution character ─────────────────────────────
    tier   = data.get("acc_dist_tier")
    share  = data.get("up_vol_share")
    if tier == "accumulation":
        _vote("acc_dist", True, f"up-vol {share*100:.0f}%")
        notes.append(f"Accumulation · {share*100:.0f}% of 20d volume on up days")
    elif tier == "distribution":
        _vote("acc_dist", False, f"up-vol {share*100:.0f}%")
        notes.append(f"Distribution · {share*100:.0f}% of 20d volume on up days")
    elif share is not None:
        notes.append(f"Neutral acc/dist · {share*100:.0f}% up-day vol")

    # ── 7. ROC acceleration ────────────────────────────────────────────────
    roc      = data.get("roc_20d")
    roc_acc  = data.get("roc_accelerating")
    if roc_acc:
        _vote("roc", True, f"ROC accelerating ({roc*100:+.1f}%)")
        notes.append(f"20d ROC accelerating · {roc*100:+.1f}%")
    elif roc is not None and roc < -0.10:
        _vote("roc", False, f"ROC weak ({roc*100:+.1f}%)")
        notes.append(f"20d ROC weak · {roc*100:+.1f}%")
    elif roc is not None:
        notes.append(f"20d ROC {roc*100:+.1f}% · no acceleration")

    n_bullish = sum(1 for _, b, _ in votes if b)
    n_bearish = sum(1 for _, b, _ in votes if not b)

    if n_bullish >= 5:
        tier, pts = "HIGH", 10
    elif n_bullish >= 3 and n_bearish <= n_bullish:
        tier, pts = "MOD", 5
    elif n_bullish >= 1 and n_bearish == 0:
        tier, pts = "MOD", 5
    elif n_bearish >= 3:
        tier, pts = "LOW", 0
    else:
        tier, pts = "MOD", 5

    header = f"[{n_bullish}↑ / {n_bearish}↓ of {len(votes)} signals]"
    return tier, pts, header + " · " + " · ".join(notes)


def _score_narrative(narrative_theme: str, narrative_strength: str) -> tuple[str, int, str]:
    """Narrative theme scored from user-supplied dropdown + strength tier."""
    if not narrative_theme or narrative_theme == "None":
        return "LOW", 0, "No narrative theme identified — speculative plays without a story rarely outperform"

    strength_pts = {"HIGH": 10, "MOD": 5, "LOW": 0}.get(narrative_strength.upper() if narrative_strength else "", 5)

    # Hot themes get a boost to HIGH more easily
    hot_themes = {"AI / Machine Learning", "Defence & Aerospace", "GLP-1 / Weight Loss",
                  "Nuclear / SMR", "Crypto-Adjacent", "Biotech Catalyst"}

    if strength_pts == 10:
        tier = "HIGH"
    elif strength_pts == 5:
        tier = "MOD"
    else:
        tier = "LOW"

    # Override tier upward if it's a hot theme with at least MOD strength
    if narrative_theme in hot_themes and strength_pts >= 5:
        tier = "HIGH"
        strength_pts = 10

    note = f"Theme: {narrative_theme} — {narrative_strength or 'MOD'} narrative connection to company"
    return tier, strength_pts, note


def _score_catalyst(catalyst_desc: str, catalyst_timing: str) -> tuple[str, int, str]:
    """Catalyst quality — combines timing dropdown with content-quality of the description.

    Previously scored purely on the timing dropdown ignoring what the user actually wrote
    — meaning "stuff happens" + near-term = HIGH same as a detailed FDA PDUFA citation.
    Now also rewards:
      - concrete date markers (Q3 2026, March, H1 2026, etc.)
      - dollar/scale figures ($500M contract, $2B TAM)
      - known event-type keywords (PDUFA, FDA approval, earnings beat, M&A, etc.)
    """
    import re

    if not catalyst_desc or len(catalyst_desc.strip()) < 5:
        return "LOW", 0, "No specific catalyst identified — without a trigger, timing a move is speculative at best"

    desc       = catalyst_desc.strip()
    desc_lower = desc.lower()

    # ── Content-quality signals ───────────────────────────────────────────────
    has_date = bool(re.search(
        r'\b(q[1-4]\s*20\d{2}|h[12]\s*20\d{2}|'
        r'jan(uary)?|feb(ruary)?|mar(ch)?|apr(il)?|may|jun(e)?|'
        r'jul(y)?|aug(ust)?|sep(t(ember)?)?|oct(ober)?|nov(ember)?|dec(ember)?|'
        r'20\d{2}-\d{1,2}|20\d{2}/\d{1,2}|20[2-9]\d)\b',
        desc_lower,
    ))
    has_dollar = bool(re.search(r'\$\s*\d', desc))
    event_keywords = [
        "pdufa", "fda approval", "fda decision", "phase 3", "phase 2", "phase iii", "phase ii",
        "data readout", "topline data", "interim data", "clinical trial",
        "earnings beat", "guidance raise", "guidance cut", "preannouncement",
        "contract award", "contract win", "design win", "partnership", "licensing deal",
        "acquisition", "merger", "buyout", "takeover", "spin-off", "spinoff",
        "approval", "launch", "product launch", "ipo", "uplisting",
        "court ruling", "patent", "split", "buyback", "dividend",
        "milestone payment", "tender", "rfp",
    ]
    matched_kw   = [kw for kw in event_keywords if kw in desc_lower]
    n_event_kw   = len(matched_kw)

    # Content quality 0-3: dates / $ figures / event-type keywords
    quality_pts = (1 if has_date else 0) + (1 if has_dollar else 0) + (1 if n_event_kw >= 1 else 0)

    timing = (catalyst_timing or "vague").lower()
    short_desc = desc[:120] + ("…" if len(desc) > 120 else "")

    # ── Combined scoring: timing × content quality ────────────────────────────
    if timing == "near":
        if quality_pts >= 2:
            return ("HIGH", 10,
                    f"Near-term catalyst with concrete detail (date={has_date}, $={has_dollar}, "
                    f"event-type={n_event_kw}): {short_desc}")
        elif quality_pts == 1:
            return ("HIGH", 10,
                    f"Near-term catalyst — some specificity: {short_desc}")
        else:
            return ("MOD", 5,
                    f"Near-term claim but vague text — add a date, dollar figure, or event type "
                    f"to qualify as HIGH: {short_desc}")

    if timing == "medium":
        if quality_pts >= 2:
            return ("MOD", 5,
                    f"Medium-term catalyst with detail: {short_desc}")
        elif quality_pts == 1:
            return ("MOD", 5,
                    f"Medium-term catalyst (limited specifics): {short_desc}")
        else:
            return ("LOW", 0,
                    f"Medium-term claim but text is generic — no qualifying detail: {short_desc}")

    # timing == "vague" or other
    if quality_pts >= 2:
        return ("MOD", 5,
                f"Detailed catalyst but no timing anchor: {short_desc}")
    return ("LOW", 0,
            f"Vague catalyst — no timing, no specifics, no event type: {short_desc}")


# ── Main scorecard builder ─────────────────────────────────────────────────────

def build_speculative_scorecard(
    data: dict,
    narrative_theme: str   = "",
    narrative_strength: str = "MOD",
    catalyst_desc: str     = "",
    catalyst_timing: str   = "vague",    # "near" | "medium" | "vague"
) -> dict:
    """Score all 10 dimensions and return the full scorecard metrics dict."""

    scores = {}

    t, p, n = _score_momentum(data)
    scores["momentum"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_volume(data)
    scores["volume"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_short_interest(data)
    scores["short_interest"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_analyst_revisions(data)
    scores["analyst_revisions"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_float(data)
    scores["float_size"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_insider(data)
    scores["insider_buying"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_downside_floor(data)
    scores["downside_floor"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_options(data)
    scores["options_activity"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_technical_analysis(data)
    scores["technical_setup"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_social_trend(data)
    scores["social_trend"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_narrative(narrative_theme, narrative_strength)
    scores["narrative"] = {"tier": t, "pts": p, "note": n}

    t, p, n = _score_catalyst(catalyst_desc, catalyst_timing)
    scores["catalyst"] = {"tier": t, "pts": p, "note": n}

    total = sum(s["pts"] for s in scores.values())
    verdict = _verdict(total)

    # Stamp all auto-scored signals with the data date so readers know signal freshness
    _as_of = datetime.date.today().isoformat()
    _manual_keys = {"narrative", "catalyst"}
    for key, s in scores.items():
        if key not in _manual_keys and s.get("note"):
            s["note"] = s["note"] + f"  [data: {_as_of}]"

    return {
        "scores":            scores,
        "total_score":       total,
        "verdict":           verdict,
        "narrative_theme":   narrative_theme,
        "narrative_strength": narrative_strength,
        "catalyst_desc":     catalyst_desc,
        "catalyst_timing":   catalyst_timing,
    }


# ── Scenario / re-rating model ─────────────────────────────────────────────────

def build_scenario_model(data: dict, scorecard: dict, hold_months: int = 6) -> dict:
    """Build a bear/base/bull scenario model with sector-calibrated EV/Revenue re-rating.

    Bear/bull multiple factors are now sector-specific (biotech catalyst plays have
    wider tails than defence contractors). Score modulates the bull factor on top
    of the sector default — high-conviction scans get more aggressive bull multiples,
    low-conviction scans get more conservative ones.

    Also computes a probability-weighted expected return: E[ret] = p_bull*bull_ret +
    p_base*base_ret + p_bear*bear_ret, with probabilities calibrated to the score.
    """
    price     = data.get("current_price") or 0
    mktcap    = data.get("market_cap")    or 0
    net_debt  = data.get("net_debt")      or 0
    shares    = data.get("shares_out")    or 0
    trail_rev = data.get("trailing_rev")  or 0
    fwd_rev   = data.get("fwd_rev_1yr")   or trail_rev
    cur_ev    = data.get("current_ev")    or mktcap + net_debt

    # Current EV/fwd revenue multiple
    cur_ev_rev_mult = (cur_ev / fwd_rev) if fwd_rev else None

    # Sector calibration — narrative theme drives bear/bull tail widths
    total_score    = scorecard.get("total_score", 50)
    narrative_theme = scorecard.get("narrative_theme", "") or ""
    sector_bucket  = THEME_TO_BUCKET.get(narrative_theme, "default")
    bear_factor, default_bull_factor = SECTOR_SCENARIOS[sector_bucket]

    # Score modulates the bull factor (high-conviction → more aggressive re-rating)
    if total_score >= 80:
        score_mult = 1.15
    elif total_score >= 65:
        score_mult = 1.00
    elif total_score >= 50:
        score_mult = 0.85
    else:
        score_mult = 0.70
    bull_factor = default_bull_factor * score_mult

    if cur_ev_rev_mult and cur_ev_rev_mult > 0:
        bear_mult = cur_ev_rev_mult * bear_factor
        base_mult = cur_ev_rev_mult * 1.00   # base = flat multiple, growth carries return
        bull_mult = cur_ev_rev_mult * bull_factor
    else:
        # No EV/Rev data — use simple price-based scenarios
        cur_ev_rev_mult = None
        bear_mult = base_mult = bull_mult = None

    def _price_from_mult(mult):
        if not mult or not fwd_rev or not shares:
            return None
        target_ev     = mult * fwd_rev
        target_equity = target_ev - net_debt
        if target_equity <= 0:
            return None
        return target_equity / shares

    bear_price = _price_from_mult(bear_mult)
    base_price = _price_from_mult(base_mult)
    bull_price = _price_from_mult(bull_mult)

    # Fallback: simple % moves if EV/Rev unavailable
    if bear_price is None and price:
        bear_price = price * 0.65
        base_price = price * 1.20
        bull_price = price * 1.75
        bear_mult = base_mult = bull_mult = None

    def _ret(tp):
        if tp is None or not price:
            return None
        return (tp / price) - 1

    bear_ret = _ret(bear_price)
    base_ret = _ret(base_price)
    bull_ret = _ret(bull_price)

    # ── Probability-weighted expected return ─────────────────────────────────
    # Score-calibrated scenario probabilities. The probabilities sum to 1.0 and
    # shift toward the bull case as conviction (total_score) rises.
    if total_score >= 80:
        p_bull, p_base, p_bear = 0.45, 0.40, 0.15
    elif total_score >= 65:
        p_bull, p_base, p_bear = 0.35, 0.45, 0.20
    elif total_score >= 50:
        p_bull, p_base, p_bear = 0.25, 0.45, 0.30
    elif total_score >= 36:
        p_bull, p_base, p_bear = 0.20, 0.40, 0.40
    else:
        p_bull, p_base, p_bear = 0.15, 0.35, 0.50

    if bear_ret is not None and base_ret is not None and bull_ret is not None:
        expected_ret = p_bull * bull_ret + p_base * base_ret + p_bear * bear_ret
        expected_price = price * (1 + expected_ret) if price else None
    else:
        expected_ret  = None
        expected_price = None

    # Target price to achieve 1.5x
    target_1_5x_price = (price * 1.5) if price else None

    # What multiple is required for 1.5x?
    req_mult_for_1_5x = None
    if price and shares and fwd_rev:
        target_equity_1_5x = target_1_5x_price * shares
        target_ev_1_5x     = target_equity_1_5x + net_debt
        req_mult_for_1_5x  = target_ev_1_5x / fwd_rev if fwd_rev else None

    return {
        "hold_months":          hold_months,
        "current_price":        price,
        "fwd_rev_b":            fwd_rev / 1e9 if fwd_rev else None,
        "trail_rev_b":          trail_rev / 1e9 if trail_rev else None,
        "current_ev_b":         cur_ev / 1e9 if cur_ev else None,
        "current_ev_rev_mult":  cur_ev_rev_mult,
        # Sector calibration
        "sector_bucket":        sector_bucket,
        "bear_factor":          bear_factor,
        "bull_factor":          bull_factor,
        "default_bull_factor":  default_bull_factor,
        # Scenario multiples
        "bear_mult":            bear_mult,
        "base_mult":            base_mult,
        "bull_mult":            bull_mult,
        # Scenario prices
        "bear_price":           bear_price,
        "base_price":           base_price,
        "bull_price":           bull_price,
        # Scenario returns
        "bear_ret":             bear_ret,
        "base_ret":             base_ret,
        "bull_ret":             bull_ret,
        # Probability-weighted expected return
        "p_bear":               p_bear,
        "p_base":               p_base,
        "p_bull":               p_bull,
        "expected_ret":         expected_ret,
        "expected_price":       expected_price,
        # 1.5x analysis
        "target_1_5x_price":    target_1_5x_price,
        "req_mult_for_1_5x":    req_mult_for_1_5x,
        "bull_reaches_1_5x":    bull_price is not None and price > 0 and bull_price >= price * 1.5,
    }


# ── Track-record CSV logging ──────────────────────────────────────────────────

def append_track_record(ticker: str, data: dict, scorecard: dict, scenario: dict,
                        hold_months: int = 6, path: str | None = None) -> str:
    """Append a single MUFF scan to the track-record CSV.

    Foundation for validating the engine over time — every scan becomes a row that
    can later be cross-referenced against actual realised returns to test whether
    higher scores actually produce better outcomes.

    Schema is append-only. Adding new columns is safe; renaming/removing breaks
    historical analysis.
    """
    import csv
    import os

    if path is None:
        path = os.path.join(os.path.dirname(__file__), "muff_track_record.csv")

    HEADERS = [
        "timestamp", "scan_date", "ticker", "company_name", "sector",
        "score", "verdict", "low_data_confidence",
        "current_price", "market_cap_b", "fwd_rev_b", "current_ev_rev_mult",
        "sector_bucket", "bear_factor", "bull_factor",
        "bear_price", "base_price", "bull_price",
        "bear_ret", "base_ret", "bull_ret",
        "p_bear", "p_base", "p_bull", "expected_ret", "expected_price",
        "narrative_theme", "narrative_strength",
        "catalyst_timing", "catalyst_desc",
        "hold_months",
    ]

    now = datetime.datetime.now(datetime.timezone.utc)
    row = {
        "timestamp":            now.isoformat(timespec="seconds"),
        "scan_date":            now.date().isoformat(),
        "ticker":               ticker,
        "company_name":         data.get("company_name") or "",
        "sector":               data.get("sector") or "",
        "score":                scorecard.get("total_score"),
        "verdict":              scorecard.get("verdict"),
        "low_data_confidence":  scorecard.get("low_data_confidence", False),
        "current_price":        scenario.get("current_price"),
        "market_cap_b":         round((data.get("market_cap") or 0) / 1e9, 3) if data.get("market_cap") else None,
        "fwd_rev_b":            scenario.get("fwd_rev_b"),
        "current_ev_rev_mult":  scenario.get("current_ev_rev_mult"),
        "sector_bucket":        scenario.get("sector_bucket"),
        "bear_factor":          scenario.get("bear_factor"),
        "bull_factor":          scenario.get("bull_factor"),
        "bear_price":           scenario.get("bear_price"),
        "base_price":           scenario.get("base_price"),
        "bull_price":           scenario.get("bull_price"),
        "bear_ret":             scenario.get("bear_ret"),
        "base_ret":             scenario.get("base_ret"),
        "bull_ret":             scenario.get("bull_ret"),
        "p_bear":               scenario.get("p_bear"),
        "p_base":               scenario.get("p_base"),
        "p_bull":               scenario.get("p_bull"),
        "expected_ret":         scenario.get("expected_ret"),
        "expected_price":       scenario.get("expected_price"),
        "narrative_theme":      scorecard.get("narrative_theme") or "",
        "narrative_strength":   scorecard.get("narrative_strength") or "",
        "catalyst_timing":      scorecard.get("catalyst_timing") or "",
        "catalyst_desc":        (scorecard.get("catalyst_desc") or "")[:240],
        "hold_months":          hold_months,
    }

    new_file = not os.path.exists(path)
    try:
        with open(path, "a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=HEADERS, extrasaction="ignore")
            if new_file:
                w.writeheader()
            w.writerow(row)
        return path
    except Exception as e:
        # Logging failure must NEVER block a scan from completing
        return f"track-record write failed: {e}"


# ── Excel workbook builder ─────────────────────────────────────────────────────

def build_speculative_excel(
    ticker: str,
    data: dict,
    scorecard: dict,
    scenario: dict,
) -> bytes:
    """Build a simple 3-tab Excel workbook for the speculative analysis."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    wb = Workbook()

    # ── Colour palette ────────────────────────────────────────────────────────
    BG_DARK    = "FF0F0F0F"
    BG_PANEL   = "FF1A1A2E"
    BG_HEADER  = "FF16213E"
    ORANGE     = "FFFF6B35"
    AMBER      = "FFFFA62B"
    GREEN      = "FF34D399"
    RED        = "FFEF4444"
    GOLD       = "FFE6C168"
    WHITE      = "FFFFFFFF"
    GREY       = "FF9CA3AF"

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color=WHITE, size=11):
        return Font(bold=bold, color=color, name="Calibri", size=size)

    def _border():
        s = Side(style="thin", color="FF2D2D2D")
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr(ws, row, col, val, bold=True, bg=BG_HEADER, fg=WHITE, size=11):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=bold, color=fg, size=size)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        return c

    def _cell(ws, row, col, val, bold=False, bg=BG_PANEL, fg=WHITE, num_fmt=None, align="left"):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _font(bold=bold, color=fg, size=10)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = _border()
        if num_fmt:
            c.number_format = num_fmt
        return c

    tier_color = {"HIGH": GREEN, "MOD": AMBER, "LOW": RED}

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Speculative Scorecard
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Speculative Scorecard"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 10
    ws1.column_dimensions["D"].width = 65

    # Title
    ws1.row_dimensions[1].height = 40
    c = ws1.cell(row=1, column=1,
                 value=f"MAYBE UNSAFE FINANCIAL FREEDOM — {ticker}")
    c.font      = Font(bold=True, color=ORANGE, size=16, name="Calibri")
    c.fill      = _fill(BG_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws1.merge_cells("A1:D1")

    # Sub-title
    ws1.row_dimensions[2].height = 20
    c = ws1.cell(row=2, column=1, value=f"Speculative Signal Scorecard · Generated {datetime.date.today()}")
    c.font      = Font(color=GREY, size=9, name="Calibri")
    c.fill      = _fill(BG_DARK)
    ws1.merge_cells("A2:D2")

    # Column headers
    ws1.row_dimensions[3].height = 22
    _hdr(ws1, 3, 1, "Signal", bg=BG_HEADER, fg=GOLD)
    _hdr(ws1, 3, 2, "Tier",   bg=BG_HEADER, fg=GOLD)
    _hdr(ws1, 3, 3, "Points", bg=BG_HEADER, fg=GOLD)
    _hdr(ws1, 3, 4, "Rationale", bg=BG_HEADER, fg=GOLD)

    LABELS = {
        "momentum":          "Price Momentum",
        "volume":            "Volume Signal",
        "short_interest":    "Short Interest / Squeeze",
        "analyst_revisions": "Analyst Revision Momentum",
        "float_size":        "Float Size / Market Cap",
        "insider_buying":    "Insider Buying (90d)",
        "downside_floor":    "Downside Floor",
        "options_activity":  "Options Activity (P/C)",
        "technical_setup":   "Technical Setup",
        "social_trend":      "Social / Trend Momentum",
        "narrative":         "Narrative Theme",
        "catalyst":          "Catalyst Quality",
    }

    r = 4
    scores = scorecard.get("scores", {})
    for key, label in LABELS.items():
        s = scores.get(key, {})
        tier  = s.get("tier", "MOD")
        pts   = s.get("pts",  5)
        note  = s.get("note", "")
        tc    = tier_color.get(tier, AMBER)
        ws1.row_dimensions[r].height = 32
        _cell(ws1, r, 1, label, bold=True,  fg=WHITE)
        c2 = _cell(ws1, r, 2, tier, bold=True, fg=tc, align="center")
        _cell(ws1, r, 3, pts,  bold=True,  fg=WHITE, num_fmt="0", align="center")
        c4 = _cell(ws1, r, 4, note, fg=GREY)
        c4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        r += 1

    # Totals row
    ws1.row_dimensions[r].height = 28
    total  = scorecard.get("total_score", 0)
    verdict= scorecard.get("verdict", "")
    _hdr(ws1, r, 1, "TOTAL SCORE", bg="FF1F1F1F", fg=GOLD)
    _hdr(ws1, r, 2, verdict,       bg="FF1F1F1F", fg=ORANGE)
    _hdr(ws1, r, 3, total,         bg="FF1F1F1F", fg=WHITE)
    _hdr(ws1, r, 4, "Score out of 120 · Verdict: ≥90 Moonshot · ≥65 Strong Spec · ≥54 Spec Play · ≥36 High Risk · <36 Pass",
         bg="FF1F1F1F", fg=GREY)

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Scenario Model
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Scenario Model")
    ws2.sheet_view.showGridLines = False
    for col, w in [("A", 28), ("B", 18), ("C", 18), ("D", 18)]:
        ws2.column_dimensions[col].width = w

    ws2.row_dimensions[1].height = 40
    c = ws2.cell(row=1, column=1, value=f"RE-RATING SCENARIO MODEL — {ticker}")
    c.font      = Font(bold=True, color=ORANGE, size=16, name="Calibri")
    c.fill      = _fill(BG_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells("A1:D1")

    sc = scenario
    hm = sc.get("hold_months", 6)

    def _kv(ws, row, label, val, val_fmt=None, val_color=WHITE):
        ws.row_dimensions[row].height = 22
        _cell(ws, row, 1, label, bold=True, fg=GOLD)
        c = _cell(ws, row, 2, val, fg=val_color, align="right")
        if val_fmt:
            c.number_format = val_fmt
        ws.merge_cells(f"B{row}:D{row}")

    r2 = 2
    _hdr(ws2, r2, 1, "MODEL INPUTS", bg=BG_HEADER, fg=GOLD)
    ws2.merge_cells(f"A{r2}:D{r2}")
    r2 += 1
    _kv(ws2, r2, "Current Price",         sc.get("current_price"),     '"$"#,##0.00'); r2 += 1
    _kv(ws2, r2, "Current EV ($B)",       sc.get("current_ev_b"),      '#,##0.00'); r2 += 1
    _kv(ws2, r2, "Trailing Revenue ($B)", sc.get("trail_rev_b"),       '#,##0.00'); r2 += 1
    _kv(ws2, r2, "Fwd Revenue Est. ($B)", sc.get("fwd_rev_b"),         '#,##0.00'); r2 += 1
    _kv(ws2, r2, "Current EV/Fwd Rev",    sc.get("current_ev_rev_mult"), '#,##0.0"x"'); r2 += 1
    _kv(ws2, r2, f"Hold Horizon (months)",hm,                          '0 "months"'); r2 += 1

    r2 += 1
    # Scenario table header
    ws2.row_dimensions[r2].height = 22
    _hdr(ws2, r2, 1, "SCENARIO",         bg=BG_HEADER, fg=GOLD)
    _hdr(ws2, r2, 2, "EV/Rev Multiple",  bg=BG_HEADER, fg=GOLD)
    _hdr(ws2, r2, 3, "Target Price",     bg=BG_HEADER, fg=GOLD)
    _hdr(ws2, r2, 4, "Return",           bg=BG_HEADER, fg=GOLD)
    r2 += 1

    scenarios_data = [
        ("BEAR — Narrative Fails",    sc.get("bear_mult"), sc.get("bear_price"), sc.get("bear_ret"), RED),
        ("BASE — Revenue Grows Only", sc.get("base_mult"), sc.get("base_price"), sc.get("base_ret"), AMBER),
        ("BULL — Narrative Plays Out",sc.get("bull_mult"), sc.get("bull_price"), sc.get("bull_ret"), GREEN),
    ]
    for label, mult, tprice, tret, color in scenarios_data:
        ws2.row_dimensions[r2].height = 24
        _cell(ws2, r2, 1, label, bold=True, fg=color)
        _cell(ws2, r2, 2, mult,   fg=WHITE, align="right", num_fmt='#,##0.0"x"')
        _cell(ws2, r2, 3, tprice, fg=WHITE, align="right", num_fmt='"$"#,##0.00')
        c = _cell(ws2, r2, 4, tret,   fg=color, align="right", num_fmt='0.0%')
        r2 += 1

    r2 += 1
    ws2.row_dimensions[r2].height = 24
    _cell(ws2, r2, 1, "1.5× Target Price", bold=True, fg=GOLD)
    _cell(ws2, r2, 2, sc.get("req_mult_for_1_5x"), fg=AMBER, align="right", num_fmt='#,##0.0"x"')
    _cell(ws2, r2, 3, sc.get("target_1_5x_price"),  fg=AMBER, align="right", num_fmt='"$"#,##0.00')
    c = _cell(ws2, r2, 4, "Required for 1.5× return", fg=GREY)
    r2 += 2

    note_txt = ("NOTE: Scenario prices derived from EV/Forward Revenue re-rating. "
                "Bear assumes narrative fails and multiple de-rates. "
                "Base assumes flat multiple with revenue growth only. "
                "Bull assumes narrative takes hold and multiple expands. "
                "All scenarios are illustrative. Adjust multiples as your thesis evolves.")
    ws2.row_dimensions[r2].height = 50
    c = ws2.cell(row=r2, column=1, value=note_txt)
    c.font      = Font(color=GREY, size=9, name="Calibri", italic=True)
    c.fill      = _fill(BG_DARK)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws2.merge_cells(f"A{r2}:D{r2}")

    # ═══════════════════════════════════════════════════════════════════════════
    # Sheet 3 — Market Data
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Market Data")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    ws3.row_dimensions[1].height = 40
    c = ws3.cell(row=1, column=1, value=f"RAW MARKET SIGNALS — {ticker}")
    c.font      = Font(bold=True, color=ORANGE, size=16, name="Calibri")
    c.fill      = _fill(BG_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws3.merge_cells("A1:B1")

    r3 = 2
    raw_fields = [
        ("Company Name",          data.get("company_name"),          None),
        ("Sector",                data.get("sector"),                 None),
        ("Current Price",         data.get("current_price"),          '"$"#,##0.00'),
        ("Market Cap ($B)",       (data.get("market_cap") or 0)/1e9, '#,##0.00'),
        ("Float Shares (M)",      (data.get("float_shares") or 0)/1e6, '#,##0.0'),
        ("RSI (14)",              data.get("rsi_14"),                 '#,##0.0'),
        ("Price vs 50d MA",       data.get("price_vs_50ma"),          '0.0%'),
        ("Price vs 200d MA",      data.get("price_vs_200ma"),         '0.0%'),
        ("10d Avg Volume (M)",    (data.get("vol_10d_avg") or 0)/1e6, '#,##0.00'),
        ("50d Avg Volume (M)",    (data.get("vol_50d_avg") or 0)/1e6, '#,##0.00'),
        ("Short Float %",         data.get("short_float_pct"),        '0.0"%"'),
        ("Days to Cover",         data.get("days_to_cover"),          '#,##0.0'),
        ("Upgrades (90d)",        data.get("upgrades_90d"),           '0'),
        ("Downgrades (90d)",      data.get("downgrades_90d"),         '0'),
        ("Est. Revision %",       data.get("est_rev_pct"),            '0.0%'),
        ("Insider Buys (90d)",    data.get("insider_buys"),           '0'),
        ("Insider Buy Value ($K)",(data.get("insider_buy_usd") or 0)/1e3, '#,##0'),
        ("Insider Sells (90d)",   data.get("insider_sells"),          '0'),
        ("Put/Call Ratio",        data.get("put_call_ratio"),         '#,##0.00'),
        ("Call Open Interest",    data.get("call_oi"),                '#,##0'),
        ("Put Open Interest",     data.get("put_oi"),                 '#,##0'),
        ("Cash ($M)",             (data.get("cash") or 0)/1e6,        '#,##0'),
        ("Total Debt ($M)",       (data.get("total_debt") or 0)/1e6,  '#,##0'),
        ("Net Debt ($M)",         (data.get("net_debt") or 0)/1e6,    '#,##0'),
        ("Trailing Revenue ($M)", (data.get("trailing_rev") or 0)/1e6,'#,##0'),
        ("Fwd Revenue Est. ($M)", (data.get("fwd_rev_1yr") or 0)/1e6, '#,##0'),
        ("EBITDA TTM ($M)",       (data.get("ebitda_ttm") or 0)/1e6,  '#,##0'),
        ("Data Source",           "FMP API + yfinance" + (" [MOCK]" if data.get("mock") else ""), None),
    ]

    for label, val, fmt in raw_fields:
        ws3.row_dimensions[r3].height = 20
        _cell(ws3, r3, 1, label, bold=True, fg=GOLD)
        c = _cell(ws3, r3, 2, val, fg=WHITE, align="right")
        if fmt and val is not None:
            c.number_format = fmt
        r3 += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
