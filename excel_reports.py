"""
excel_reports.py
Generates HTML reports from existing Excel financial model files.
No FMP API calls needed — reads P&L, BS, CF, Scorecard, WACC, DCF tabs directly.

Usage: python excel_reports.py
"""

import os, sys, re
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from report_bridge import build_report_data, render_html_report

OUT_DIR = os.path.join(os.path.dirname(__file__), "static", "reports")
os.makedirs(OUT_DIR, exist_ok=True)

# ── Excel files to process ────────────────────────────────────────────────────
DESKTOP = "C:/Users/justi/OneDrive/Desktop/Investment Automation/"
DLOADS  = "C:/Users/justi/Downloads/"

FILES = {
    "WFC":  DLOADS  + "WFC_FinancialModel_2025.xlsx",
    "INTC": DESKTOP + "INTC_FinancialModel_2025.xlsx",
    "TSLA": DESKTOP + "TSLA_FinancialModel_2025.xlsx",
    "SOFI": DESKTOP + "SOFI_FinancialModel_2025.xlsx",
    "JPM":  DESKTOP + "JPM_FinancialModel_2025.xlsx",
    "C":    DESKTOP + "C_FinancialModel_2025.xlsx",
    "BAC":  DESKTOP + "BAC_FinancialModel_2025.xlsx",
    "UAL":  DLOADS  + "UAL_FinancialModel_2025.xlsx",
}

TIER_VAL = {"HIGH": 10, "MOD-HIGH": 7, "MOD-LOW": 3, "LOW": 0}
CRITERIA_W = {
    "Moat Profile":                10.0, "Management":                7.5,
    "Revenue 3yr CAGR":           10.0, "Cash Quality":             10.0,
    "Capital Returns":              5.0, "ROIC":                      7.5,
    "Credit Risk":                  5.0, "Capital Adequacy":          5.0,
    "Interest Cover":               7.5, "Execution Risk":            5.0,
    "Valuation vs Median  (P/E)": 10.0, "Valuation vs Median  (P/FCF)": 10.0,
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def _fv(v):
    try: return float(v)
    except: return None

def _pct_str(s):
    if not s: return None
    m = re.search(r"([\d.]+)%", str(s))
    return round(float(m.group(1)) / 100, 4) if m else None

def _parse_de(s):
    s = str(s or "")
    if "net cash" in s.lower(): return 0.0
    m = re.search(r"([\d.]+)x", s)
    return float(m.group(1)) if m else None

def _parse_val(s):
    s = str(s or "")
    cur = re.search(r"[Cc]urrent\s+([\d.]+)x", s)
    avg = re.search(r"5yr\s+avg\s+([\d.]+)x", s)
    return (float(cur.group(1)) if cur else None,
            float(avg.group(1)) if avg else None)

def _parse_fcf_ni(s):
    return _pct_str(s)

M = 1_000_000  # Excel stores in $mm — multiply to get raw $


def parse_excel_full(path, ticker):
    """
    Returns (is_data, bs_data, cf_data, years, profile, scorecard_metrics, wacc_val, dcf_prices)
    All financial values are in raw dollars (multiply Excel-mm by 1e6).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames

    # ── Read years from P&L header ────────────────────────────────────────────
    pl_ws = (wb["P&L"] if "P&L" in wb.sheetnames else
             wb["P&L "] if "P&L " in wb.sheetnames else None)
    if not pl_ws:
        raise ValueError("No P&L tab found")

    years = []
    data_col_start = None  # column index (1-based) where year data starts
    for row in pl_ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if str(row[0] or "").strip() == "Fiscal Year Ending":
            for i, v in enumerate(row[1:], start=2):
                if v and str(v).strip().isdigit() and len(str(v).strip()) == 4:
                    years.append(str(v).strip())
                    if data_col_start is None:
                        data_col_start = i
            break

    if not years:
        raise ValueError("Could not find years in P&L tab")

    n = len(years)

    # ── Generic row reader ────────────────────────────────────────────────────
    def _read_rows(ws, labels_map):
        """Scan ws for rows whose col-A label contains any key in labels_map.
        Returns {fmp_field: [val_yr0, val_yr1, ...]} in raw dollars."""
        result = {v: [None] * n for v in labels_map.values()}
        found  = {v: False for v in labels_map.values()}
        for row in ws.iter_rows(min_row=1, values_only=True):
            a = str(row[0] or "").strip()
            for label_sub, fmp_field in labels_map.items():
                if found[fmp_field]:
                    continue
                if label_sub.lower() in a.lower():
                    for i in range(n):
                        col_idx = (data_col_start - 1) + i  # 0-based index into row tuple
                        if col_idx < len(row):
                            v = _fv(row[col_idx])
                            result[fmp_field][i] = (v * M) if v is not None else None
                    found[fmp_field] = True
                    break
        return result

    # ── P&L (income statement) ────────────────────────────────────────────────
    IS_MAP = {
        "Revenue":                       "revenue",
        "Cost of Revenue":               "costOfRevenue",
        "Gross Profit":                  "grossProfit",
        "EBITDA":                        "ebitda",
        "EBIT (Operating Income)":       "operatingIncome",
        "Depreciation & Amortisation":   "depreciationAndAmortization",
        "Interest Expense":              "interestExpense",
        "EBT / Income Before Tax":       "incomeBeforeTax",
        "Income Tax Expense":            "incomeTaxExpense",
        "Net Income from Continuing":    "netIncome",
        "R&D Expenses":                  "researchAndDevelopmentExpenses",
        "SG&A (Combined)":               "sellingGeneralAndAdministrativeExpenses",
    }
    is_raw = _read_rows(pl_ws, IS_MAP)

    # Fallback: "Net Income" if "Net Income from Continuing" not found
    if all(v is None for v in is_raw["netIncome"]):
        is_raw2 = _read_rows(pl_ws, {"Net Income": "netIncome2"})
        is_raw["netIncome"] = is_raw2.get("netIncome2", [None]*n)

    is_data = []
    for i in range(n):
        d = {k: (vals[i] if vals[i] is not None else 0)
             for k, vals in is_raw.items()}
        d["date"] = f"{years[i]}-12-31"
        d["fiscalYear"] = years[i]
        d["calendarYear"] = years[i]
        d["reportedCurrency"] = "USD"
        is_data.append(d)

    # ── Balance Sheet ─────────────────────────────────────────────────────────
    bs_ws = wb["Balance Sheet"] if "Balance Sheet" in wb.sheetnames else None
    BS_MAP = {
        "Cash & Cash Equivalents":    "cashAndCashEquivalents",
        "Short-Term Investments":     "shortTermInvestments",
        "Net Receivables (Total)":    "netReceivables",
        "Total Current Assets":       "totalCurrentAssets",
        "PP&E (Net)":                 "propertyPlantEquipmentNet",
        "Goodwill":                   "goodwill",
        "Intangible Assets":          "intangibleAssets",
        "Total Non-Current Assets":   "totalNonCurrentAssets",
        "TOTAL ASSETS":               "totalAssets",
        "Short-Term Debt":            "shortTermDebt",
        "Total Current Liabilities":  "totalCurrentLiabilities",
        "Long-Term Debt":             "longTermDebt",
        "TOTAL LIABILITIES":          "totalLiabilities",
        "Total Equity (inc. Minority)": "totalStockholdersEquity",
        "Retained Earnings":          "retainedEarnings",
        "Minority Interest":          "minorityInterest",
        "Total Debt (ST + LT)":       "totalDebt",
    }
    bs_raw = _read_rows(bs_ws, BS_MAP) if bs_ws else {k: [None]*n for k in BS_MAP.values()}

    bs_data = []
    for i in range(n):
        d = {k: (vals[i] if vals[i] is not None else 0)
             for k, vals in bs_raw.items()}
        d["date"] = f"{years[i]}-12-31"
        bs_data.append(d)

    # ── Cash Flow ─────────────────────────────────────────────────────────────
    cf_ws = wb["Cash Flow"] if "Cash Flow" in wb.sheetnames else None
    CF_MAP = {
        "Net Cash from Operations (CFO)":      "operatingCashFlow",
        "Capital Expenditures":                "capitalExpenditure",
        "Free Cash Flow (FCF)":                "freeCashFlow",
        "Common Stock Repurchased (Buybacks)": "commonStockRepurchased",
        "Dividends Paid":                      "dividendsPaid",
        "Depreciation & Amortisation":         "depreciationAndAmortization",
        "Stock-Based Compensation":            "stockBasedCompensation",
    }
    cf_raw = _read_rows(cf_ws, CF_MAP) if cf_ws else {k: [None]*n for k in CF_MAP.values()}

    cf_data = []
    for i in range(n):
        d = {k: (vals[i] if vals[i] is not None else 0)
             for k, vals in cf_raw.items()}
        d["date"] = f"{years[i]}-12-31"
        cf_data.append(d)

    # ── Scorecard tab ─────────────────────────────────────────────────────────
    scorecard_metrics = {
        "roic": None, "rev_cagr": None, "fcf_ni": None,
        "d_ebitda": None, "equity_assets": None, "is_bank": False,
        "auto_score": None, "floor_cap": None,
        "pe_current": None, "pe_5yr_avg": None,
        "pfcf_current": None, "pfcf_5yr_avg": None,
    }

    if "Scorecard" in wb.sheetnames:
        sc = wb["Scorecard"]
        label_map = {}
        floor_cap = None
        for row in sc.iter_rows(min_col=1, max_col=5, values_only=True):
            a = str(row[0] or "").strip()
            d = str(row[3] or "").strip()
            e = str(row[4] or "").strip() if row[4] else None
            if a:
                label_map[a] = (d, e)
            if "HARD FLOOR GATE" in a.upper() and "CAPPED AT" in a.upper():
                m = re.search(r"capped at\s+(\d+)", a, re.IGNORECASE)
                if m: floor_cap = int(m.group(1))
            # Bank exemption banner
            if "bank" in a.lower() and "exempt" in a.lower():
                scorecard_metrics["is_bank"] = True
                # Try to extract equity/assets from banner text
                m2 = re.search(r"Equity/Assets\s*=\s*([\d.]+)%", a)
                if m2: scorecard_metrics["equity_assets"] = float(m2.group(1)) / 100

        def _get(sub):
            for k, v in label_map.items():
                if sub.lower() in k.lower():
                    return v
            return ("", None)

        scorecard_metrics["roic"]     = _pct_str(_get("ROIC")[0])
        scorecard_metrics["rev_cagr"] = _pct_str(_get("Revenue 3yr CAGR")[0])
        scorecard_metrics["fcf_ni"]   = _parse_fcf_ni(_get("Cash Quality")[0])
        scorecard_metrics["d_ebitda"] = _parse_de(_get("Credit Risk")[0])
        scorecard_metrics["floor_cap"] = floor_cap

        pe_v, pe5_v     = _parse_val(_get("Valuation vs Median  (P/E)")[0])
        pfcf_v, pfcf5_v = _parse_val(_get("Valuation vs Median  (P/FCF)")[0])
        scorecard_metrics["pe_current"]   = pe_v
        scorecard_metrics["pe_5yr_avg"]   = pe5_v
        scorecard_metrics["pfcf_current"] = pfcf_v
        scorecard_metrics["pfcf_5yr_avg"] = pfcf5_v

        # Compute auto_score from all scored criteria
        scored = []
        for label_sub, weight in CRITERIA_W.items():
            _, tier = _get(label_sub)
            if tier and tier.upper() in TIER_VAL:
                scored.append((TIER_VAL[tier.upper()], weight))
        if scored:
            auto_score = round(sum((s / 10) * w for s, w in scored), 1)
            if floor_cap is not None:
                auto_score = min(auto_score, floor_cap)
            scorecard_metrics["auto_score"] = auto_score

    # ── WACC tab ──────────────────────────────────────────────────────────────
    wacc_val = None
    wacc_components = {}
    if "WACC" in wb.sheetnames:
        wm = {}
        for row in wb["WACC"].iter_rows(min_col=1, max_col=2, values_only=True):
            a = str(row[0] or "").strip()
            if row[1] is not None:
                try: wm[a] = float(row[1])
                except: pass
        rf   = next((wm[k] for k in wm if k.startswith("► Selected Rf")),   None)
        beta = next((wm[k] for k in wm if k.startswith("► Selected beta")), None)
        erp  = next((wm[k] for k in wm if k.startswith("► Selected ERP")),  None)
        rd   = next((wm[k] for k in wm if k.startswith("► Selected pre-tax Rd")), None)
        t_w  = next((wm[k] for k in wm if k.startswith("► Selected tax rate")),   None)
        eq   = next((wm[k] for k in wm if k.startswith("Equity ")), None)
        debt = next((wm[k] for k in wm if k.startswith("Debt ")),   None)
        if all(v is not None for v in [rf, beta, erp, rd, t_w, eq, debt]) and (eq + debt) > 0:
            r_e      = rf + beta * erp
            v_tot    = eq + debt
            wacc_val = (eq / v_tot) * r_e + (debt / v_tot) * rd * (1 - t_w)
            print(f"    WACC = {wacc_val:.2%}")
            wacc_components = {
                "rf": rf, "beta": beta, "erp": erp, "rd": rd, "t_w": t_w,
                "ke": r_e, "ew": eq / v_tot, "dw": debt / v_tot,
            }

    # ── DCF tab — price, shares, net debt, projections ────────────────────────
    dcf_prices = {}
    dcf_ev_output = {}
    price = None
    if "DCF" in wb.sheetnames:
        dcf_ws = wb["DCF"]
        proj_cols = []; hist_cols = []
        for row_t in dcf_ws.iter_rows(min_col=1, max_col=15, values_only=True):
            if str(row_t[0] or "").strip() == "Fiscal Year":
                for i, h in enumerate(row_t[1:]):
                    hs = str(h or "")
                    if hs.endswith("E"): proj_cols.append(i)
                    elif hs.isdigit() and len(hs) == 4: hist_cols.append(i)
                break

        g = shares = net_debt = mi = exit_mult = None
        da_pct = capex_pct = nwc_pct = tax_dcf = None
        proj_rev = []; proj_ebitda = []; hist_rev = []; hist_ebitda = []

        for row_t in dcf_ws.iter_rows(min_col=1, max_col=15, values_only=True):
            a  = str(row_t[0] or "").strip().lower()
            vals = list(row_t[1:])
            def _fvl(v):
                try: return float(v)
                except: return None
            if "current market price" in a:
                price = _fvl(vals[0])
            elif "shares outstanding" in a and "diluted" in a:
                shares = _fvl(vals[0])
            elif "less: net debt" in a:
                net_debt = _fvl(vals[0])
                if net_debt is not None:
                    dcf_ev_output["net_debt_mm"] = net_debt
            elif "less: minority interest" in a:
                mi = _fvl(vals[0])
            elif "terminal growth rate" in a:
                g = _fvl(vals[0])
            elif "terminal ev/ebitda multiple" in a:
                exit_mult = _fvl(vals[0])
            elif "d&a as % of revenue" in a:
                da_pct = _fvl(vals[hist_cols[-1]]) if hist_cols else _fvl(vals[0])
            elif "capex as % of revenue" in a:
                capex_pct = _fvl(vals[hist_cols[-1]]) if hist_cols else _fvl(vals[0])
            elif "change in nwc as % of revenue" in a:
                nwc_pct = _fvl(vals[hist_cols[-1]]) if hist_cols else _fvl(vals[0])
            elif "effective tax rate" in a and "user input" in a:
                tax_dcf = _fvl(vals[hist_cols[-1]]) if hist_cols else _fvl(vals[0])
            elif row_t[0] and str(row_t[0]).strip() == "Revenue" and proj_cols:
                hist_rev   = [v for i in hist_cols if (v := _fvl(vals[i])) is not None]
                proj_rev   = [v for i in proj_cols if (v := _fvl(vals[i])) is not None]
            elif row_t[0] and str(row_t[0]).strip() == "EBITDA" and proj_cols:
                hist_ebitda = [v for i in hist_cols if (v := _fvl(vals[i])) is not None]
                proj_ebitda = [v for i in proj_cols if (v := _fvl(vals[i])) is not None]
            # EV bridge output rows (Excel typically puts these in a results block)
            elif "pv of fcf" in a or "pv of free cash flow" in a:
                dcf_ev_output["pv_fcfs"] = _fvl(vals[0])
            elif "pv of terminal value" in a or "pv of tv" in a:
                dcf_ev_output["pv_tv"]   = _fvl(vals[0])
            elif ("enterprise value" in a and "dcf" in a) or ("implied ev" in a) or (a.strip() == "enterprise value (dcf)"):
                dcf_ev_output["ev"]      = _fvl(vals[0])
            elif "equity value" in a and "implied" in a:
                dcf_ev_output["equity_val"] = _fvl(vals[0])
            elif "tv as % of ev" in a or "terminal value %" in a:
                dcf_ev_output["tv_pct"]  = _fvl(vals[0])

        # Compute GG / EM prices
        try:
            _g   = g         if g         is not None else 0.03
            _tev = exit_mult if exit_mult is not None else 20.0
            _nd  = net_debt  if net_debt  is not None else 0.0
            _mi  = mi        if mi        is not None else 0.0
            _tax = tax_dcf   if tax_dcf   is not None else 0.20
            _da  = da_pct    if da_pct    is not None else 0.08
            _cx  = capex_pct if capex_pct is not None else 0.05
            _nwc = nwc_pct   if nwc_pct   is not None else 0.01
            _lm  = (hist_ebitda[-1] / hist_rev[-1]) if hist_rev and hist_ebitda else 0.20

            if (wacc_val and (wacc_val - _g) > 0.001
                    and shares and shares > 0 and proj_rev and proj_ebitda):
                def _ufcf(r, e):
                    da = r * _da
                    return (e - da) * (1 - _tax) + da - r * _cx - r * _nwc
                n_p = min(len(proj_rev), len(proj_ebitda))
                sum_pv = sum(
                    _ufcf(proj_rev[i], proj_ebitda[i]) / (1 + wacc_val) ** (i + 0.5)
                    for i in range(n_p)
                )
                tv_disc    = (1 + wacc_val) ** n_p
                t_rev      = proj_rev[-1] * (1 + _g)
                t_ebitda   = t_rev * _lm
                t_ufcf     = _ufcf(t_rev, t_ebitda)
                ip_gg = (sum_pv + t_ufcf / (wacc_val - _g) / tv_disc - _nd - _mi) / shares
                ip_em = (sum_pv + t_ebitda * _tev / tv_disc - _nd - _mi) / shares
                if ip_gg > 0 and ip_gg < 10000:
                    dcf_prices["gg_price"] = round(ip_gg, 2)
                if ip_em > 0 and ip_em < 10000:
                    dcf_prices["em_price"] = round(ip_em, 2)
                print(f"    DCF prices: GG=${ip_gg:.2f}  EM=${ip_em:.2f}")
        except Exception as e:
            print(f"    DCF computation skipped: {e}")

    # ── Bank detection and equity/assets override ─────────────────────────────
    # Old Excel files (pre-fix) used D/EBITDA for banks — detect and fix here
    BANK_TICKERS = {"WFC", "JPM", "C", "BAC", "SOFI", "GS", "MS", "USB", "PNC"}
    if ticker.upper() in BANK_TICKERS:
        scorecard_metrics["is_bank"] = True
        # Compute equity/assets from the last available BS year
        if bs_data:
            eq = bs_data[-1].get("totalStockholdersEquity") or 0
            ta = bs_data[-1].get("totalAssets") or 0
            if ta > 0 and eq > 0:
                scorecard_metrics["equity_assets"] = eq / ta
        # Recalculate auto_score: exclude D/EBITDA gate (floor_cap)
        # For old Excel, floor was triggered by D/EBITDA — remove it for banks
        scorecard_metrics["floor_cap"] = None
        # Remove D/EBITDA from auto_score re-calc: rescore without floor gate
        # The Excel scorecard had D/EBITDA scoring LOW for banks (0 pts for 5.0 weight)
        # With equity/assets, banks get MOD-HIGH or HIGH — add those pts back
        ea = scorecard_metrics.get("equity_assets")
        if ea is not None:
            if ea > 0.10:   ea_tier_pts = 10
            elif ea > 0.08: ea_tier_pts = 7
            elif ea > 0.06: ea_tier_pts = 3
            else:            ea_tier_pts = 0
            # Old auto_score treated D/EBITDA as LOW (0 pts × 5.0 weight = 0)
            # New: add bank leverage pts (ea_tier_pts/10 × 5.0)
            old_score = scorecard_metrics.get("auto_score") or 0
            # Remove old gate cap if applied
            # Add equity_assets pts (replacing D/EBITDA 0 pts)
            new_score = round(old_score + (ea_tier_pts / 10) * 5.0, 1)
            scorecard_metrics["auto_score"] = new_score
            print(f"    Bank fix: equity_assets={ea:.1%}  old_score={old_score}  new_score={new_score}")

    # ── Profile stub ──────────────────────────────────────────────────────────
    profile = {
        "symbol":       ticker,
        "companyName":  ticker,
        "price":        price,
        "mktCap":       None,
        "sector":       "Financials" if scorecard_metrics["is_bank"] else "",
        "industry":     "Banking" if scorecard_metrics["is_bank"] else "",
    }

    return is_data, bs_data, cf_data, years, profile, scorecard_metrics, wacc_val, dcf_prices, wacc_components, dcf_ev_output


# ── Main ──────────────────────────────────────────────────────────────────────
results = []

for ticker, path in FILES.items():
    print(f"\n{'='*55}\n  {ticker}\n{'='*55}")
    if not os.path.exists(path):
        print(f"  SKIP — file not found: {path}")
        results.append({"ticker": ticker, "score": None, "error": "file not found"})
        continue
    try:
        is_data, bs_data, cf_data, years, profile, scorecard_metrics, wacc_val, dcf_prices, wacc_components, dcf_ev_output = \
            parse_excel_full(path, ticker)

        auto_score = scorecard_metrics.get("auto_score") or 0
        adj_score  = round(auto_score, 1)
        floor_cap  = scorecard_metrics.get("floor_cap")
        if floor_cap is not None:
            adj_score = min(adj_score, floor_cap)

        report_data = build_report_data(
            ticker            = ticker,
            profile           = profile,
            is_data           = is_data,
            bs_data           = bs_data,
            cf_data           = cf_data,
            years             = years,
            wacc_val          = wacc_val,
            dcf_prices        = dcf_prices,
            scorecard_metrics = scorecard_metrics,
            manual_rating     = None,
            current_price     = profile.get("price"),
            market_cap        = profile.get("mktCap"),
            biz_clarity       = None,
            ltp               = None,
            adj_score         = adj_score,
        )

        # ── Override WACC keys with actual Excel WACC tab values ──────────────
        if wacc_components.get("rf") is not None:
            wc = wacc_components
            rf_v   = wc["rf"]
            beta_v = wc["beta"]
            erp_v  = wc["erp"]
            ke_v   = wc["ke"]
            rd_v   = wc["rd"]
            tw_v   = wc["t_w"]
            ew_v   = wc["ew"]
            dw_v   = wc["dw"]
            wacc_b = wacc_val or 0.09
            report_data.update({
                "WACC_RF":        f"{rf_v*100:.1f}%",
                "WACC_BETA":      f"{beta_v:.2f}",
                "WACC_BETA_SOURCE": "Excel WACC tab (5yr monthly vs S&P500)",
                "WACC_ERP":       f"{erp_v*100:.1f}%",
                "WACC_KE":        f"{ke_v*100:.1f}%",
                "WACC_KD_PRETAX": f"{rd_v*100:.1f}%",
                "WACC_TAX_RATE":  f"{tw_v*100:.1f}%",
                "WACC_EW":        f"{ew_v*100:.0f}%",
                "WACC_DW":        f"{dw_v*100:.0f}%",
                "WACC_NOTE":      (
                    f"WACC = {ew_v*100:.0f}% × {ke_v*100:.1f}% (Ke) "
                    f"+ {dw_v*100:.0f}% × {rd_v*100:.1f}% × (1 − {tw_v*100:.0f}% tax) "
                    f"= {wacc_b*100:.1f}%. Source: Excel WACC tab. "
                    f"Rf: FRED DGS10 {rf_v*100:.1f}%. ERP: Damodaran implied/hist. avg {erp_v*100:.1f}%."
                ),
            })

        # ── Override EV Bridge keys with actual Excel DCF tab values ──────────
        if dcf_ev_output:
            ev_o   = dcf_ev_output.get("ev")       # $mm in Excel
            pv_tv  = dcf_ev_output.get("pv_tv")
            pv_fcs = dcf_ev_output.get("pv_fcfs")
            eq_v   = dcf_ev_output.get("equity_val")
            tv_pct = dcf_ev_output.get("tv_pct")

            def _b_mm(v):
                """Excel mm value → billions string."""
                if v is None: return "—"
                b = v / 1e3  # mm → B
                sign = "-" if b < 0 else ""
                return f"{sign}${abs(b):.1f}B"

            # Net debt from DCF tab (Less: Net Debt row, stored in $mm)
            nd_mm = dcf_ev_output.get("net_debt_mm")

            updates = {}
            if ev_o   is not None: updates["DCF_EV"]       = _b_mm(ev_o)
            if pv_tv  is not None: updates["DCF_PV_TV"]    = _b_mm(pv_tv)
            if pv_fcs is not None: updates["DCF_PV_FCFS"]  = _b_mm(pv_fcs)
            if eq_v   is not None: updates["DCF_EQUITY_VAL"] = _b_mm(eq_v)
            if tv_pct is not None:
                updates["DCF_TV_PCT"] = f"{tv_pct*100:.0f}%" if tv_pct < 2 else f"{tv_pct:.0f}%"
            if updates:
                report_data.update(updates)

        html = render_html_report(report_data)

        out = os.path.join(OUT_DIR, f"{ticker}_report.html")
        with open(out, "w", encoding="utf-8") as f:
            f.write(html)

        gg = dcf_prices.get("gg_price")
        em = dcf_prices.get("em_price")
        p  = profile.get("price")
        gg_up = round((gg - p) / p, 3) if gg and p else None
        em_up = round((em - p) / p, 3) if em and p else None

        print(f"  OK  score={adj_score}  auto={auto_score}  floor={floor_cap}"
              f"  is_bank={scorecard_metrics.get('is_bank',False)}"
              f"  gg={gg_up}  em={em_up}")
        results.append({
            "ticker": ticker, "score": adj_score, "auto": auto_score,
            "floor": floor_cap, "is_bank": scorecard_metrics.get("is_bank", False),
            "roic": scorecard_metrics.get("roic"),
            "cagr": scorecard_metrics.get("rev_cagr"),
            "fcfni": scorecard_metrics.get("fcf_ni"),
            "de": scorecard_metrics.get("d_ebitda"),
            "ea": scorecard_metrics.get("equity_assets"),
            "pe": scorecard_metrics.get("pe_current"),
            "pe5": scorecard_metrics.get("pe_5yr_avg"),
            "pfcf": scorecard_metrics.get("pfcf_current"),
            "pfcf5": scorecard_metrics.get("pfcf_5yr_avg"),
            "gg": gg_up, "em": em_up, "price": p,
        })
    except Exception as e:
        import traceback
        print(f"  FAIL: {e}")
        traceback.print_exc()
        results.append({"ticker": ticker, "score": None, "error": str(e)})

print("\n\n" + "="*60)
print("RESULTS SUMMARY")
print("="*60)
for r in results:
    if r.get("error"):
        print(f"  {r['ticker']:6s}  ERROR: {r['error']}")
    else:
        bank = " [BANK]" if r.get("is_bank") else ""
        cap  = f"  cap={r['floor']}" if r.get("floor") else ""
        print(f"  {r['ticker']:6s}  score={r['score']:<6}{cap}{bank}"
              f"  ROIC={r['roic']:.1%}" if r.get("roic") else
              f"  {r['ticker']:6s}  score={r.get('score')}")
