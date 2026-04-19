"""
backfill_heatmap.py
-------------------
Reads existing local Excel models and writes rows to outputs.csv on GitHub.
No FMP API calls — uses only the already-generated .xlsx files.

Run: python backfill_heatmap.py

Requires:
  - GITHUB_TOKEN env variable (or paste directly below)
  - pip install openpyxl requests
"""

import base64
import datetime
import glob
import os
import re
import sys
import requests
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import csv_schema as _schema

# ── Config ────────────────────────────────────────────────────────────────────
GITHUB_TOKEN  = os.environ.get("GITHUB_TOKEN", "")   # or paste token here
GITHUB_REPO   = "jaysang2908/Investment-Automation"
GITHUB_BRANCH = "main"
FOLDER        = os.path.dirname(os.path.abspath(__file__))

CSV_HEADER = _schema.HEADER

# Scorecard criteria: (label_substring, weight)
CRITERIA_WEIGHTS = {
    "Moat Profile":              10.0,
    "Management":                 7.5,
    "Revenue 3yr CAGR":          10.0,
    "Cash Quality":              10.0,
    "Capital Returns":            5.0,
    "ROIC":                       7.5,
    "Credit Risk":                5.0,
    "Interest Cover":             7.5,
    "Execution Risk":             5.0,
    "Valuation vs Median  (P/E)":10.0,
    "Valuation vs Median  (P/FCF)":10.0,
}
TIER_VAL = {"HIGH": 10, "MOD-HIGH": 7, "MOD-LOW": 3, "LOW": 0}

# ── GitHub helpers ────────────────────────────────────────────────────────────
GH_HEADERS = {
    "Authorization": f"token {GITHUB_TOKEN}",
    "Accept":        "application/vnd.github.v3+json",
}
GH_API = f"https://api.github.com/repos/{GITHUB_REPO}/contents/outputs.csv"


def _read_csv():
    r = requests.get(GH_API, headers=GH_HEADERS, params={"ref": GITHUB_BRANCH}, timeout=8)
    if r.status_code == 200:
        info = r.json()
        return info["sha"], base64.b64decode(info["content"]).decode()
    return None, CSV_HEADER


def _write_csv(sha, content):
    payload = {
        "message": "Backfill heatmap from local Excel models",
        "branch":  GITHUB_BRANCH,
        "content": base64.b64encode(content.encode()).decode(),
    }
    if sha:
        payload["sha"] = sha
    r = requests.put(GH_API, headers=GH_HEADERS, json=payload, timeout=15)
    if r.status_code not in (200, 201):
        print(f"  ERROR: {r.status_code} — {r.json().get('message','')}")
    else:
        print("  Written to GitHub.")


# ── Excel parsing helpers ─────────────────────────────────────────────────────
def _pct(s):
    """Parse '37.5%' → 0.375, or None."""
    if not s:
        return None
    m = re.search(r"([\d.]+)%", str(s))
    return round(float(m.group(1)) / 100, 4) if m else None


def _num(s):
    """Parse first float from a string, or None."""
    if not s:
        return None
    m = re.search(r"([\d.]+)", str(s))
    return float(m.group(1)) if m else None


def _parse_de(s):
    """Parse D/EBITDA note: '2.1x' → 2.1, 'Net cash ...' → 0.0"""
    s = str(s or "")
    if "net cash" in s.lower():
        return 0.0
    m = re.search(r"([\d.]+)x", s)
    return float(m.group(1)) if m else None


def _parse_val(s):
    """Parse 'Current 51.7x  |  5yr avg 44.9x ...' → (51.7, 44.9)"""
    s = str(s or "")
    cur  = re.search(r"[Cc]urrent\s+([\d.]+)x", s)
    avg  = re.search(r"5yr\s+avg\s+([\d.]+)x", s)
    return (
        float(cur.group(1)) if cur else None,
        float(avg.group(1)) if avg else None,
    )


def _parse_fcf_ni(s):
    """Parse FCF/NI note: '97%' or 'FCF/NI 97% ...' → 0.97"""
    return _pct(s)


def parse_excel(path):
    """Extract all heatmap fields from one Excel workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}

    # ── Scorecard tab ─────────────────────────────────────────────────────────
    if "Scorecard" not in wb.sheetnames:
        return None
    sc = wb["Scorecard"]

    # Scan rows: build {stripped_label: (col_D_value, col_E_tier)}
    label_map = {}
    floor_cap = None
    for row in sc.iter_rows(min_col=1, max_col=5, values_only=True):
        a = str(row[0] or "").strip()
        d = str(row[3] or "").strip()
        e = str(row[4] or "").strip() if row[4] else None
        if a:
            label_map[a] = (d, e)
        # Detect floor gate row
        if "HARD FLOOR GATE" in a.upper() and "CAPPED AT" in a.upper():
            m = re.search(r"capped at\s+(\d+)", a, re.IGNORECASE)
            if m:
                floor_cap = int(m.group(1))

    def _get(label_substr):
        for k, v in label_map.items():
            if label_substr.lower() in k.lower():
                return v
        return ("", None)

    # Parse KPIs
    result["roic"]     = _pct(_get("ROIC")[0])
    result["rev_cagr"] = _pct(_get("Revenue 3yr CAGR")[0])
    result["fcf_ni"]   = _parse_fcf_ni(_get("Cash Quality")[0])
    result["d_ebitda"] = _parse_de(_get("Credit Risk")[0])

    pe_note   = _get("Valuation vs Median  (P/E)")[0]
    pfcf_note = _get("Valuation vs Median  (P/FCF)")[0]
    result["pe_current"],   result["pe_5yr_avg"]   = _parse_val(pe_note)
    result["pfcf_current"], result["pfcf_5yr_avg"] = _parse_val(pfcf_note)

    # Compute auto_score
    scored = []
    for label_substr, weight in CRITERIA_WEIGHTS.items():
        _, tier = _get(label_substr)
        if tier and tier.upper() in TIER_VAL:
            scored.append((TIER_VAL[tier.upper()], weight))

    if scored:
        auto_score = round(sum((s / 10) * w for s, w in scored), 1)
        if floor_cap is not None:
            auto_score = min(auto_score, floor_cap)
        result["auto_score"] = auto_score
    else:
        result["auto_score"] = None

    result["floor_cap"] = floor_cap

    # ── DCF tab — read all inputs needed for Python-side GG/EM computation ──────
    price = shares = net_debt = mi = g = exit_mult = None
    da_pct = capex_pct = nwc_pct = tax_dcf = None
    proj_rev = []
    proj_ebitda = []
    hist_rev = []
    hist_ebitda = []

    if "DCF" in wb.sheetnames:
        dcf_ws = wb["DCF"]
        # Pass 1: find fiscal year row and identify projection vs historical columns
        proj_cols = []   # 0-based indices into data values (col B = index 0)
        hist_cols = []
        for row_t in dcf_ws.iter_rows(min_col=1, max_col=15, values_only=True):
            if str(row_t[0] or "").strip() == "Fiscal Year":
                for i, h in enumerate(row_t[1:]):
                    hs = str(h or "")
                    if hs.endswith("E"):
                        proj_cols.append(i)
                    elif hs.isdigit() and len(hs) == 4:
                        hist_cols.append(i)
                break

        # Pass 2: read scalar and vector values
        for row_t in dcf_ws.iter_rows(min_col=1, max_col=15, values_only=True):
            a  = str(row_t[0] or "").strip()
            al = a.lower()
            vals = list(row_t[1:])   # indices match proj_cols / hist_cols

            def _fv(v):
                try: return float(v)
                except: return None

            if "current market price" in al:
                price = _fv(vals[0])
            elif "shares outstanding" in al and "diluted" in al:
                shares = _fv(vals[0])
            elif "less: net debt" in al:
                net_debt = _fv(vals[0])
            elif "less: minority interest" in al:
                mi = _fv(vals[0])
            elif "terminal growth rate" in al:
                g = _fv(vals[0])
            elif "terminal ev/ebitda multiple" in al:
                exit_mult = _fv(vals[0])
            elif "d&a as % of revenue" in al:
                # Use last historical year — that's what the Excel projections default to
                da_pct = _fv(vals[hist_cols[-1]]) if hist_cols else _fv(vals[0])
            elif "capex as % of revenue" in al:
                capex_pct = _fv(vals[hist_cols[-1]]) if hist_cols else _fv(vals[0])
            elif "change in nwc as % of revenue" in al:
                nwc_pct = _fv(vals[hist_cols[-1]]) if hist_cols else _fv(vals[0])
            elif "effective tax rate" in al and "user input" in al:
                tax_dcf = _fv(vals[hist_cols[-1]]) if hist_cols else _fv(vals[0])
            elif a == "Revenue" and proj_cols:
                hist_rev    = [v for i in hist_cols if (v := _fv(vals[i])) is not None]
                proj_rev    = [v for i in proj_cols if (v := _fv(vals[i])) is not None]
            elif a == "EBITDA" and proj_cols:
                hist_ebitda = [v for i in hist_cols if (v := _fv(vals[i])) is not None]
                proj_ebitda = [v for i in proj_cols if (v := _fv(vals[i])) is not None]

    # ── WACC tab — compute WACC from plain-number inputs ─────────────────────
    wacc_val = None
    if "WACC" in wb.sheetnames:
        wm = {}
        for row_t in wb["WACC"].iter_rows(min_col=1, max_col=2, values_only=True):
            a = str(row_t[0] or "").strip()
            b = row_t[1]
            if b is not None:
                try: wm[a] = float(b)
                except: pass
        # Read selected override values (prefixed with ►)
        rf   = next((wm[k] for k in wm if k.startswith("► Selected Rf")),   None)
        beta = next((wm[k] for k in wm if k.startswith("► Selected beta")), None)
        erp  = next((wm[k] for k in wm if k.startswith("► Selected ERP")),  None)
        rd   = next((wm[k] for k in wm if k.startswith("► Selected pre-tax Rd")), None)
        t_w  = next((wm[k] for k in wm if k.startswith("► Selected tax rate")),   None)
        eq   = next((wm[k] for k in wm if k.startswith("Equity ")), None)
        debt = next((wm[k] for k in wm if k.startswith("Debt ")),   None)
        if all(v is not None for v in [rf, beta, erp, rd, t_w, eq, debt]) and (eq + debt) > 0:
            v_tot    = eq + debt
            r_e      = rf + beta * erp
            wacc_val = (eq / v_tot) * r_e + (debt / v_tot) * rd * (1 - t_w)

    # ── Python-side GG/EM computation (mirrors build_dcf logic) ──────────────
    gg_price = em_price = None
    _dbg = {}   # collects intermediate values for validation output
    try:
        _g    = g         if g         is not None else 0.03
        _tev  = exit_mult if exit_mult is not None else 20.0
        _nd   = net_debt  if net_debt  is not None else 0.0
        _mi   = mi        if mi        is not None else 0.0
        _tax  = tax_dcf   if tax_dcf   is not None else 0.20
        _da   = da_pct    if da_pct    is not None else 0.08
        _cx   = capex_pct if capex_pct is not None else 0.05
        _nwc  = nwc_pct   if nwc_pct   is not None else 0.01
        _last_margin = (hist_ebitda[-1] / hist_rev[-1]) if hist_rev and hist_ebitda else 0.20

        _dbg = dict(wacc=wacc_val, g=_g, tev=_tev, nd=_nd, mi=_mi,
                    tax=_tax, da=_da, cx=_cx, nwc=_nwc,
                    ebitda_margin=_last_margin, shares=shares,
                    n_proj=len(proj_rev))

        if wacc_val and (wacc_val - _g) > 0.001 and shares and shares > 0 and proj_rev and proj_ebitda:
            def _ufcf(rev, ebitda):
                da    = rev * _da
                nopat = (ebitda - da) * (1 - _tax)
                return nopat + da - rev * _cx - rev * _nwc

            n = min(len(proj_rev), len(proj_ebitda))
            sum_pv = sum(
                _ufcf(proj_rev[i], proj_ebitda[i]) / (1 + wacc_val) ** (i + 0.5)
                for i in range(n)
            )
            tv_disc     = (1 + wacc_val) ** n
            term_rev    = proj_rev[-1] * (1 + _g)
            term_ebitda = term_rev * _last_margin
            term_ufcf   = _ufcf(term_rev, term_ebitda)

            _wacc_g = wacc_val - _g
            _ip_gg = (sum_pv + term_ufcf / _wacc_g / tv_disc - _nd - _mi) / shares
            _ip_em = (sum_pv + term_ebitda * _tev / tv_disc - _nd - _mi) / shares
            gg_price = round(_ip_gg, 2)
            em_price = round(_ip_em, 2)
            _dbg.update(sum_pv=round(sum_pv), tv_disc=round(tv_disc, 4),
                        term_ufcf=round(term_ufcf), term_ebitda=round(term_ebitda),
                        gg_price=gg_price, em_price=em_price)
    except Exception as _e:
        _dbg["error"] = str(_e)

    result["_dbg"] = _dbg   # stripped before writing to CSV

    result["price"]    = round(price, 2) if price else None
    mkt_cap_b = round(price * shares / 1000, 2) if price and shares else None
    result["mkt_cap_b"] = mkt_cap_b
    result["gg_price"]  = gg_price
    result["em_price"]  = em_price
    result["gg_upside"] = round(gg_price / price - 1, 4) if gg_price and price else None
    result["em_upside"] = round(em_price / price - 1, 4) if em_price and price else None

    return result


def _f(v, dp=4):
    return "" if v is None else f"{v:.{dp}f}"


# ── Main ──────────────────────────────────────────────────────────────────────
def run():
    if not GITHUB_TOKEN:
        sys.exit("ERROR: GITHUB_TOKEN not set.")

    # Find all model files
    files = sorted(glob.glob(os.path.join(FOLDER, "*_FinancialModel_*.xlsx")))
    if not files:
        sys.exit("No *_FinancialModel_*.xlsx files found.")

    print(f"Found {len(files)} Excel files.")
    print("Reading outputs.csv from GitHub...")
    sha, content = _read_csv()

    existing = set()
    for line in content.splitlines()[1:]:
        if line.strip():
            existing.add(line.split(",")[0].strip())
    # Migrate any old schema before appending
    content = _schema.migrate(content)
    sha = None  # sha will be refreshed on write anyway; re-fetch after migrate
    r2 = requests.get(GH_API, headers=GH_HEADERS, params={"ref": GITHUB_BRANCH}, timeout=8)
    if r2.status_code == 200:
        sha = r2.json()["sha"]

    existing = set()
    for line in content.splitlines()[1:]:
        if line.strip():
            existing.add(line.split(",")[0].strip())
    print(f"Already present: {sorted(existing) or 'none'}\n")

    today      = datetime.date.today().isoformat()
    rows_added = 0

    for path in files:
        fname  = os.path.basename(path)
        ticker = fname.split("_")[0]

        if ticker in existing:
            print(f"  Skipping {ticker} — already in CSV")
            continue

        print(f"Processing {ticker} ({fname})...")
        try:
            m = parse_excel(path)
            if m is None:
                print(f"  No Scorecard tab — skipping.")
                continue

            new_row = {
                "Ticker":    ticker,
                "Price":     _f(m.get("price"),        2),
                "MktCap_B":  _f(m.get("mkt_cap_b"),    2),
                "GG_Price":  _f(m.get("gg_price"),  2),
                "GG_Upside": _f(m.get("gg_upside"), 4),
                "EM_Price":  _f(m.get("em_price"),  2),
                "EM_Upside": _f(m.get("em_upside"), 4),
                "PE_Current":    _f(m.get("pe_current"),   1),
                "PE_5yr":        _f(m.get("pe_5yr_avg"),   1),
                "PFCF_Current":  _f(m.get("pfcf_current"), 1),
                "PFCF_5yr":      _f(m.get("pfcf_5yr_avg"), 1),
                "ROIC":          _f(m.get("roic")),
                "Rev_CAGR":      _f(m.get("rev_cagr")),
                "FCF_NI":        _f(m.get("fcf_ni")),
                "D_EBITDA":      _f(m.get("d_ebitda"),     2),
                "Revenue_B":     "",
                "OCF_B":         "",
                "FCF_B":         "",
                "Auto_Score":    "" if m.get("auto_score") is None else str(m["auto_score"]),
                "Floor_Cap":     "" if m.get("floor_cap")  is None else str(m["floor_cap"]),
                "Manual_Clarity": "",
                "Manual_LTP":    "",
                "Date":          today,
            }
            row = ",".join(new_row.get(c, "") for c in _schema.COLUMNS) + "\n"

            content += row
            rows_added += 1
            dbg = m.get("_dbg", {})
            wacc_pct = f"{dbg.get('wacc', 0)*100:.2f}%" if dbg.get("wacc") else "n/a"
            print(f"  score={m.get('auto_score')}  price=${m.get('price')}  "
                  f"gg=${m.get('gg_price')}  em=${m.get('em_price')}  wacc={wacc_pct}")
            print(f"    inputs: da={dbg.get('da'):.3f}  cx={dbg.get('cx'):.3f}  "
                  f"tax={dbg.get('tax'):.3f}  margin={dbg.get('ebitda_margin', 0):.3f}  "
                  f"nd={dbg.get('nd')}  shares={dbg.get('shares')}mm  "
                  f"n_proj={dbg.get('n_proj')}"
                  if dbg.get("da") is not None else "    (no DCF inputs found)")

        except Exception as e:
            import traceback
            print(f"  ERROR: {e}")
            traceback.print_exc()

    if rows_added == 0:
        print("\nNo new rows to write.")
        return

    print(f"\nWriting {rows_added} new row(s) to GitHub...")
    _write_csv(sha, content)
    print("Done.")


if __name__ == "__main__":
    run()
