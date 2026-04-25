"""
read_excel_models.py
Reads all Excel financial models from the Desktop folder and populates
static/data/TICKER_data.json with exact Excel model assumptions.

The DCF calculator uses these as the base case — no FMP calls needed.

Usage:  python read_excel_models.py
        python read_excel_models.py AAPL MSFT    # specific tickers only
        python read_excel_models.py --force      # overwrite existing cache
"""

import os, sys, json, glob, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data_store import DATA_DIR
import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

# ── Location of Excel models ────────────────────────────────────────────────
EXCEL_DIR = os.path.expanduser(
    r"C:\Users\justi\OneDrive\Desktop\Investment Automation"
)

# ── DCF sheet row map (1-indexed) ───────────────────────────────────────────
# Columns: A=label, B=yr1, C=yr2, D=yr3, E=yr4, F=yr5, G=p1..K=p5, L=terminal
DCF_ROWS = {
    "header":         3,   # Fiscal Year | 2021 | 2022 | ...
    "rev_growth":     7,
    "ebitda_margin":  8,
    "da_pct":         9,
    "capex_pct":      10,
    "nwc_pct":        11,
    "tax_rate":       12,
    "revenue":        15,
    "ebitda":         20,
    "ufcf":           32,
    "tgr":            37,   # col B (label in col A)
    "exit_multiple":  38,   # col B
    "net_debt":       55,   # col B
    "shares_mm":      57,   # col B
    "current_price":  61,   # col B
}

WACC_ROWS = {
    "equity_mm":  5,    # col B
    "debt_mm":    6,    # col B
    "rf":         13,   # col B — Selected Rf
    "beta":       21,   # col B — Selected beta
    "erp":        26,   # col B — Selected ERP
    "kd_pretax":  38,   # col B — Selected pre-tax Rd
    "tax_rate":   42,   # col B — Selected tax rate
    # WACC output (row 45) is a formula — we compute it instead
}


def _v(ws, row, col=2):
    """Read cell value, return None if empty/non-numeric string."""
    val = ws.cell(row, col).value
    if val is None:
        return None
    if isinstance(val, str):
        # strip em-dashes, en-dashes, whitespace
        clean = val.strip().replace("\u2212", "").replace("\u2013", "").replace("\u2014", "")
        if not clean or clean in ("—", "-", "N/A", "n/a"):
            return None
        try:
            return float(clean.replace(",", ""))
        except ValueError:
            return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _row_values(ws, row, cols):
    """Return list of values for given (1-indexed) columns."""
    return [_v(ws, row, c) for c in cols]


def find_excel(ticker):
    """Find the Excel file for a ticker in EXCEL_DIR."""
    pattern = os.path.join(EXCEL_DIR, f"{ticker}_FinancialModel*.xlsx")
    matches = sorted(glob.glob(pattern), reverse=True)  # latest version first
    if matches:
        return matches[0]
    # Fallback: case-insensitive search
    for f in os.listdir(EXCEL_DIR):
        if f.upper().startswith(ticker.upper() + "_") and f.endswith(".xlsx"):
            return os.path.join(EXCEL_DIR, f)
    return None


def extract_excel(ticker, path):
    """
    Extract DCF model data from Excel and return a data-store-compatible dict.
    Returns None on failure.
    """
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  {ticker:6s}  Cannot open: {e}")
        return None

    if "DCF" not in wb.sheetnames or "WACC" not in wb.sheetnames:
        print(f"  {ticker:6s}  Missing DCF or WACC sheet")
        return None

    dcf  = wb["DCF"]
    wacc = wb["WACC"]

    # ── Read year headers ──────────────────────────────────────────────────
    # Row 3: Fiscal Year | 2021 | 2022 | 2023 | 2024 | 2025 | 2026E | ... | Terminal
    header_row = [dcf.cell(DCF_ROWS["header"], c).value for c in range(1, 14)]
    years_all  = [str(v).strip() if v else "" for v in header_row[1:13]]

    # Determine split: historicals end where 'E' suffix begins
    hist_cols = []
    proj_cols = []
    term_col  = None
    for i, lbl in enumerate(years_all):
        col = i + 2  # 1-indexed column
        if re.match(r'^\d{4}E?$', lbl.strip()):
            if lbl.strip().endswith("E"):
                proj_cols.append((lbl.strip(), col))
            else:
                hist_cols.append((lbl.strip(), col))
        elif "terminal" in lbl.lower():
            term_col = col
        # anything else (e.g. "Source / Notes") is ignored

    if not hist_cols:
        print(f"  {ticker:6s}  Cannot parse year headers: {years_all[:8]}")
        return None

    n_hist = len(hist_cols)
    n_proj = len(proj_cols)
    all_data_cols = [c for _, c in hist_cols] + [c for _, c in proj_cols]

    # ── Read assumption rows ───────────────────────────────────────────────
    def _row(key):
        return _row_values(dcf, DCF_ROWS[key], all_data_cols)

    rev_growth    = _row("rev_growth")
    ebitda_margin = _row("ebitda_margin")
    da_pct        = _row("da_pct")
    capex_pct     = _row("capex_pct")
    nwc_pct       = _row("nwc_pct")
    tax_rate      = _row("tax_rate")
    revenue_mm    = _row("revenue")    # $mm
    ebitda_mm     = _row("ebitda")     # $mm
    ufcf_mm       = _row("ufcf")       # $mm

    tgr           = _v(dcf, DCF_ROWS["tgr"])
    exit_multiple = _v(dcf, DCF_ROWS["exit_multiple"])
    net_debt_mm   = _v(dcf, DCF_ROWS["net_debt"])
    shares_mm     = _v(dcf, DCF_ROWS["shares_mm"])
    current_price = _v(dcf, DCF_ROWS["current_price"])

    # ── WACC inputs ────────────────────────────────────────────────────────
    equity_mm = _v(wacc, WACC_ROWS["equity_mm"])   # total market cap $mm
    debt_mm   = _v(wacc, WACC_ROWS["debt_mm"])     # $mm
    rf        = _v(wacc, WACC_ROWS["rf"])
    beta      = _v(wacc, WACC_ROWS["beta"])
    erp       = _v(wacc, WACC_ROWS["erp"])
    kd_pre    = _v(wacc, WACC_ROWS["kd_pretax"])
    wacc_tax  = _v(wacc, WACC_ROWS["tax_rate"])

    # Compute WACC (formula cell is None from openpyxl)
    rf        = rf   or 0.043
    beta      = beta or 1.0
    erp       = erp  or 0.045
    kd_pre    = kd_pre or 0.04
    wacc_tax  = wacc_tax or 0.21

    ke = rf + beta * erp
    if equity_mm and debt_mm:
        total_cap = equity_mm + debt_mm
        ew = equity_mm / total_cap
        dw = debt_mm   / total_cap
    else:
        ew, dw = 0.90, 0.10
    wacc_computed = ew * ke + dw * kd_pre * (1 - wacc_tax)

    # ── Build synthetic FMP-format data rows ──────────────────────────────
    # The server's _build_dcf_response reads is_data/bs_data/cf_data.
    # We reconstruct them with exact values from the Excel model.
    years_hist = [lbl for lbl, _ in hist_cols]
    years_proj = [lbl for lbl, _ in proj_cols]

    is_data, bs_data, cf_data = [], [], []
    n_total = n_hist  # only historical rows for the stored FMP-format arrays
    for i in range(n_total):
        rev   = (revenue_mm[i]  or 0) * 1e6   # convert $mm → $
        ebit  = (ebitda_mm[i]   or 0) * 1e6
        ni_est= rev * ((ebitda_margin[i] or 0) - (da_pct[i] or 0)) * (1 - (tax_rate[i] or 0.21))
        da    = rev * (da_pct[i]    or 0.03)
        capex = rev * (capex_pct[i] or 0.03)
        ufcf  = (ufcf_mm[i] or 0) * 1e6
        pti   = ni_est / (1 - (tax_rate[i] or 0.21)) if (tax_rate[i] or 0.21) < 1 else 0
        te    = pti * (tax_rate[i] or 0.21)
        ocf   = ufcf + capex   # approximate: UFCF ≈ OCF - CapEx (ignores NWC for simplicity)

        yr = years_hist[i]
        is_data.append({
            "fiscalYear":                  yr,
            "date":                        f"{yr[:4]}-09-30",
            "revenue":                     rev,
            "ebitda":                      ebit,
            "operatingIncome":             ebit - da,
            "netIncome":                   ni_est,
            "depreciationAndAmortization": da,
            "incomeBeforeTax":             pti,
            "incomeTaxExpense":            te,
            "interestExpense":             0,
            "grossProfit":                 ebit,
        })
        bs_data.append({
            "fiscalYear":              yr,
            "date":                    f"{yr[:4]}-09-30",
            "cashAndCashEquivalents":  0,
            "shortTermDebt":           0,
            "longTermDebt":            (net_debt_mm or 0) * 1e6 if i == n_total - 1 else 0,
            "totalStockholdersEquity": ni_est * 5,
            "totalAssets":             rev * 1.5,
        })
        cf_data.append({
            "fiscalYear":                  yr,
            "date":                        f"{yr[:4]}-09-30",
            "operatingCashFlow":           ocf,
            "capitalExpenditure":          -capex,
            "freeCashFlow":                ufcf,
            "depreciationAndAmortization": da,
        })

    # ── Excel DCF projection assumptions (for DCF calculator defaults) ────
    proj_rows = []
    for i in range(n_proj):
        j = n_hist + i
        proj_rows.append({
            "year":          years_proj[i],
            "rev_growth":    rev_growth[j],
            "ebitda_margin": ebitda_margin[j],
            "da_pct":        da_pct[j],
            "capex_pct":     capex_pct[j],
            "nwc_pct":       nwc_pct[j],
            "tax_rate":      tax_rate[j],
        })

    # ── Profile ────────────────────────────────────────────────────────────
    mkt_cap_mm = equity_mm or 0
    profile = {
        "symbol":            ticker,
        "companyName":       ticker,
        "price":             current_price or 0,
        "mktCap":            mkt_cap_mm * 1e6,
        "sharesOutstanding": (shares_mm or 0) * 1e6,
        "beta":              beta,
    }

    # ── DCF prices (compute GGM from UFCF + WACC if available) ───────────
    gg_price = None
    em_price = None
    if ufcf_mm and any(v for v in ufcf_mm[:n_hist]) and tgr and wacc_computed > tgr:
        # Use last historical UFCF and project 5 years, then terminal
        last_ufcf = None
        for u in reversed(ufcf_mm[:n_hist]):
            if u: last_ufcf = u * 1e6; break
        if last_ufcf:
            # Project 5 years using first projection year's growth
            proj_ufcfs = []
            base_rev = (revenue_mm[n_hist-1] or 0) * 1e6
            for i in range(n_proj or 5):
                j = n_hist + i
                g  = rev_growth[j]    if j < len(rev_growth)    else 0.05
                m  = ebitda_margin[j] if j < len(ebitda_margin) else (ebitda_margin[n_hist-1] or 0.20)
                da = da_pct[j]        if j < len(da_pct)        else (da_pct[n_hist-1] or 0.03)
                cx = capex_pct[j]     if j < len(capex_pct)     else (capex_pct[n_hist-1] or 0.03)
                tx = tax_rate[j]      if j < len(tax_rate)      else (tax_rate[n_hist-1] or 0.21)
                nw = nwc_pct[j]       if j < len(nwc_pct)       else 0.005
                base_rev = base_rev * (1 + (g or 0.05))
                ebitda_f = base_rev * (m or 0.20)
                da_f     = base_rev * (da or 0.03)
                ebit_f   = ebitda_f - da_f
                nopat_f  = ebit_f * (1 - (tx or 0.21))
                capex_f  = base_rev * (cx or 0.03)
                nwc_f    = base_rev * (nw or 0.005)
                ufcf_f   = nopat_f + da_f - capex_f - nwc_f
                proj_ufcfs.append(ufcf_f)

            if proj_ufcfs:
                terminal_ufcf = proj_ufcfs[-1] * (1 + tgr)
                tv_gg = terminal_ufcf / (wacc_computed - tgr)
                # Discount
                pv_fcfs = sum(u / (1 + wacc_computed) ** (i + 0.5)
                              for i, u in enumerate(proj_ufcfs))
                pv_tv   = tv_gg / (1 + wacc_computed) ** (n_proj or 5)
                ev_gg   = pv_fcfs + pv_tv
                net_d   = (net_debt_mm or 0) * 1e6
                eq_val  = ev_gg - net_d
                shs     = (shares_mm or 0) * 1e6
                if shs > 0:
                    gg_price = round(eq_val / shs, 2)

                # Exit multiple
                if exit_multiple:
                    last_ebitda = (ebitda_mm[n_hist-1] or 0) * 1e6 if ebitda_mm else 0
                    # Use terminal year EBITDA
                    term_rev    = base_rev
                    term_ebitda = term_rev * (ebitda_margin[n_hist + n_proj - 1]
                                              if len(ebitda_margin) > n_hist + n_proj - 1
                                              else (ebitda_margin[n_hist-1] or 0.20))
                    tv_em       = term_ebitda * exit_multiple
                    pv_tv_em    = tv_em / (1 + wacc_computed) ** (n_proj or 5)
                    ev_em       = pv_fcfs + pv_tv_em
                    eq_em       = ev_em - net_d
                    if shs > 0:
                        em_price = round(eq_em / shs, 2)

    dcf_prices = {
        "gg_price": gg_price,
        "em_price": em_price,
    }

    # ── Assemble final stored dict ─────────────────────────────────────────
    return {
        "ticker":    ticker,
        "fetched":   datetime.date.today().isoformat(),
        "profile":   profile,
        "years":     years_hist,
        "is_data":   is_data,
        "bs_data":   bs_data,
        "cf_data":   cf_data,
        "wacc_val":  round(wacc_computed, 4),
        "dcf_prices": dcf_prices,
        "scorecard_metrics": {},
        "analyst_ests":      [],
        # Full Excel projection assumptions — used by _build_dcf_response
        "excel_dcf": {
            "years_hist":   years_hist,
            "years_proj":   years_proj,
            "tgr":          tgr   or 0.03,
            "exit_multiple":exit_multiple or 15.0,
            "net_debt_mm":  net_debt_mm or 0,
            "shares_mm":    shares_mm   or 0,
            "current_price":current_price or 0,
            "hist": [
                {
                    "year":          years_hist[i],
                    "rev_mm":        revenue_mm[i],
                    "rev_growth":    rev_growth[i],
                    "ebitda_margin": ebitda_margin[i],
                    "da_pct":        da_pct[i],
                    "capex_pct":     capex_pct[i],
                    "nwc_pct":       nwc_pct[i],
                    "tax_rate":      tax_rate[i],
                    "ufcf_mm":       ufcf_mm[i],
                }
                for i in range(n_hist)
            ],
            "proj": proj_rows,
            "wacc_inputs": {
                "rf":          round(rf, 4),
                "beta":        round(beta, 3),
                "erp":         round(erp, 4),
                "kd_pretax":   round(kd_pre, 4),
                "tax_rate":    round(wacc_tax, 4),
                "equity_mm":   round(equity_mm or 0, 2),
                "debt_mm":     round(debt_mm   or 0, 2),
                "equity_weight": round(ew, 4),
                "wacc":        round(wacc_computed, 4),
            },
        },
    }


def main():
    force = "--force" in sys.argv
    args  = [a for a in sys.argv[1:] if not a.startswith("--")]

    if args:
        tickers = [t.upper() for t in args]
    else:
        # All tickers with Excel models in desktop folder
        tickers = []
        if os.path.isdir(EXCEL_DIR):
            for f in sorted(os.listdir(EXCEL_DIR)):
                m = re.match(r"^([A-Z]+)_FinancialModel.*\.xlsx$", f)
                if m:
                    tickers.append(m.group(1))
        if not tickers:
            print(f"No Excel models found in {EXCEL_DIR}")
            return

    os.makedirs(DATA_DIR, exist_ok=True)
    print(f"\nReading {len(tickers)} Excel model(s) from:")
    print(f"  {EXCEL_DIR}")
    print(f"Output: {DATA_DIR}")
    print("=" * 55)

    ok = 0; skipped = 0; failed = 0
    for ticker in tickers:
        path = find_excel(ticker)
        if not path:
            print(f"  {ticker:6s}  SKIP — no Excel file found")
            skipped += 1
            continue

        # Check existing cache
        cache_path = os.path.join(DATA_DIR, f"{ticker}_data.json")
        if not force and os.path.exists(cache_path):
            with open(cache_path) as f:
                existing = json.load(f)
            if existing.get("excel_dcf"):
                print(f"  {ticker:6s}  already has Excel data — skip (use --force)")
                skipped += 1
                continue

        data = extract_excel(ticker, path)
        if data is None:
            failed += 1
            continue

        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(data, f)

        ex = data["excel_dcf"]
        w  = ex["wacc_inputs"]["wacc"]
        gg = data["dcf_prices"].get("gg_price")
        cp = ex.get("current_price")
        n_proj = len(ex["proj"])
        print(f"  {ticker:6s}  OK  {len(ex['hist'])}yr hist + {n_proj}yr proj"
              f"  WACC={w:.1%}"
              + (f"  price=${cp:.2f}" if cp else "")
              + (f"  GG=${gg:.0f}" if gg else ""))
        ok += 1

    print("=" * 55)
    print(f"Done: {ok} extracted, {skipped} skipped, {failed} failed")
    if ok:
        print(f"\nNext: git add static/data/ && git commit -m 'Excel model data' && git push")


if __name__ == "__main__":
    main()
