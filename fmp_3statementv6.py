import os
import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Output files always save next to this script, regardless of working directory
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION — update API_KEY before running
# ═══════════════════════════════════════════════════════════════════════════════
API_KEY      = "tOPLCq7cEELfef0FA6AKNuVoO549gAS1"
YEARS        = 5    # historical years to fetch
YEARS_PROJ   = 5    # minimum projection years in DCF (auto-extended if FMP has more)

# ── Damodaran tables (Jan 2025 US data — update annually) ─────────────────────
# ICR band → (synthetic rating, default spread)
DAMODARAN_SPREADS = [
    (12.5, 1e9,  "AAA",  0.0054),
    ( 9.5, 12.5, "AA",   0.0072),
    ( 7.5,  9.5, "A+",   0.0096),
    ( 6.0,  7.5, "A",    0.0108),
    ( 4.5,  6.0, "A-",   0.0132),
    ( 4.0,  4.5, "BBB+", 0.0156),
    ( 3.5,  4.0, "BBB",  0.0180),
    ( 3.0,  3.5, "BBB-", 0.0240),
    ( 2.5,  3.0, "BB+",  0.0288),
    ( 2.0,  2.5, "BB",   0.0360),
    ( 1.5,  2.0, "B+",   0.0432),
    ( 1.25, 1.5, "B",    0.0540),
    ( 0.8,  1.25,"B-",   0.0648),
    ( 0.5,  0.8, "CCC",  0.0900),
    ( 0.0,  0.5, "CC",   0.1200),
]
# Moody's → S&P rating equivalents
MOODY_TO_SP = {
    "Aaa": "AAA", "Aa1": "AA+", "Aa2": "AA",  "Aa3": "AA-",
    "A1":  "A+",  "A2":  "A",   "A3":  "A-",
    "Baa1":"BBB+","Baa2":"BBB", "Baa3":"BBB-",
    "Ba1": "BB+", "Ba2": "BB",  "Ba3": "BB-",
    "B1":  "B+",  "B2":  "B",   "B3":  "B-",
    "Caa1":"CCC+","Caa2":"CCC", "Caa3":"CCC-","Ca": "CC",
}
VALID_SP_RATINGS = {
    "AAA","AA+","AA","AA-","A+","A","A-",
    "BBB+","BBB","BBB-","BB+","BB","BB-",
    "B+","B","B-","CCC+","CCC","CCC-","CC","C","D",
}
# Industry → unlevered beta (US, Jan 2025)
DAMODARAN_BETAS = {
    "Semiconductor":         1.15,
    "Software":              1.08,
    "Technology":            1.07,
    "Computer":              0.95,
    "Internet":              1.12,
    "Electronics":           1.05,
    "Telecom":               0.68,
    "Retail":                0.82,
    "Healthcare":            0.76,
    "Pharmaceutical":        0.74,
    "Financial":             0.55,
    "Insurance":             0.60,
    "Oil":                   0.78,
    "Energy":                0.82,
    "Automobile":            0.88,
    "Consumer":              0.80,
    "Industrial":            0.88,
    "Default":               1.00,
}
# Damodaran implied ERP — US market (update annually from pages.stern.nyu.edu)
DAMODARAN_ERP_IMPLIED  = 0.0472   # Jan 2026
DAMODARAN_ERP_HIST_AVG = 0.0420   # arithmetic avg 1928–2025
# Peer tickers by sector (for beta comparison)
SECTOR_PEERS = {
    "Semiconductors":  ["AMD", "INTC", "QCOM", "AVGO", "TSM"],
    "Software":        ["MSFT", "CRM", "ORCL", "ADBE", "NOW"],
    "Technology":      ["AAPL", "MSFT", "GOOGL", "META", "AMZN"],
    "Healthcare":      ["JNJ",  "UNH",  "ABT",  "MDT",  "BMY"],
    "Financials":      ["JPM",  "BAC",  "GS",   "MS",   "WFC"],
    "Consumer":        ["AMZN", "HD",   "MCD",  "NKE",  "SBUX"],
    "Energy":          ["XOM",  "CVX",  "COP",  "SLB",  "PSX"],
}

# CIK lookup for EDGAR bank credit data fetch (SEC EDGAR XBRL API)
_BANK_CIKS = {
    "JPM": "0000019617",
    "BAC": "0000070858",
    "C":   "0000831001",
    "WFC": "0000072971",
    "GS":  "0000886982",
    "MS":  "0000895421",
}

# ── Colours ───────────────────────────────────────────────────────────────────
C_TITLE      = "1F2D3D"
C_SECTION    = "2E4057"
C_SUMMARY_HD = "1A3A5C"
C_SUMMARY_BG = "EAF2FB"
C_DETAIL_HD  = "34495E"
C_ALT        = "F4F8FB"
C_SUBTOTAL   = "D6E4F0"
C_WHITE      = "FFFFFF"
C_BLUE       = "0000FF"   # hardcoded inputs
C_AI_BG      = "FFF9C4"   # amber  — AI recommendation rows
C_AI_RAT     = "FFFDE7"   # pale   — AI rationale rows
C_FLAG_BG    = "FFEBEE"   # red    — warning / flag rows
C_OVR_BG     = "F1F8E9"   # green  — selected / override rows
C_BLACK      = "000000"   # formula outputs
C_GREEN      = "006400"   # cross-sheet links
# DCF-specific colours
C_SUB        = "D6E4F0"   # subtotal / header rows  (same as C_SUBTOTAL)
C_ASSM       = "EBF5FB"   # assumption / input rows
C_HIST       = "F8FBFD"   # historical data rows
C_CONS       = "D4E6F1"   # FMP consensus-driven projection rows
C_BG         = "D4E6F1"   # alias: consensus/projection background in DCF
C_SECT       = "2E4057"   # DCF section header background (same as C_SECTION)
C_HD         = "1A3A5C"   # DCF sub-section header (slightly lighter)

# ── Fonts / fills / borders ───────────────────────────────────────────────────
def fnt(bold=False, color=C_BLACK, size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def fll(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def brd(color="B0B8C1"):
    t = Side(style="thin", color=color)
    return Border(left=t, right=t, top=t, bottom=t)

def pct_fmt(cell):   cell.number_format = '0.0%;(0.0%);"-"'
def num_fmt(cell):   cell.number_format = '#,##0.0;(#,##0.0);"-"'
def ratio_fmt(cell): cell.number_format = '0.0x;(0.0x);"-"'
def days_fmt(cell):  cell.number_format = '#,##0.0;(#,##0.0);"-"'

def cl(col): return get_column_letter(col)

# ═══════════════════════════════════════════════════════════════════════════════
# API FETCH
# ═══════════════════════════════════════════════════════════════════════════════
def fetch(endpoint, ticker, extra_params=""):
    url = (f"https://financialmodelingprep.com/stable/{endpoint}"
           f"?symbol={ticker}&limit={YEARS}{extra_params}&apikey={API_KEY}")
    print(f"  GET {endpoint}...")
    r = requests.get(url)
    print(f"  -> {r.status_code}")
    if r.status_code != 200:
        raise ValueError(f"HTTP {r.status_code} on {endpoint}")
    if not r.text.strip():
        raise ValueError(f"Empty response: {endpoint}/{ticker}")
    try:
        data = r.json()
    except Exception as e:
        raise ValueError(f"JSON parse failed: {e}\nRaw: {r.text[:200]}")
    if isinstance(data, dict):
        msg = data.get("Error Message") or data.get("message", "")
        if msg:
            raise ValueError(f"API error: {msg}")
    if not isinstance(data, list) or len(data) == 0:
        raise ValueError(f"No data for '{ticker}' on {endpoint}.")
    return data

# ── Ratios cache (per-process) ────────────────────────────────────────────────
# build_dcf() and build_scorecard() both need the ratios endpoint.
# Cache the response so the second caller gets a free hit — saves 1 FMP call
# per ticker per run (15 tickers × 1 = 15 calls saved on a typical batch).
_RATIOS_CACHE: dict = {}

def _fetch_ratios(ticker: str, limit: int = 5) -> list:
    """Return FMP /stable/ratios for ticker, cached for this process lifetime."""
    key = f"{ticker.upper()}:{limit}"
    if key not in _RATIOS_CACHE:
        try:
            url = (f"https://financialmodelingprep.com/stable/ratios"
                   f"?symbol={ticker}&limit={limit}&apikey={API_KEY}")
            resp = requests.get(url, timeout=10).json()
            _RATIOS_CACHE[key] = resp if isinstance(resp, list) else []
        except Exception as _e_rat_cache:
            print(f"  ratios fetch failed for {ticker}: {_e_rat_cache}")
            _RATIOS_CACHE[key] = []
    return _RATIOS_CACHE[key]

# ── TTM Net Income cache (quarterly IS, 4 most recent quarters) ──────────────
_TTM_NI_CACHE: dict = {}

def _fetch_ttm_ni(ticker: str) -> float | None:
    """Sum last 4 quarterly net incomes → TTM NI in raw USD.
    Needed for live TTM P/E: current_mktcap / ttm_ni (avoids FMP annual
    snapshot which uses price-at-FY-end ÷ FY-EPS, not current price).
    Returns None on any failure — callers degrade gracefully to stale ratio.
    One FMP call per ticker per process (cached)."""
    key = ticker.upper()
    if key not in _TTM_NI_CACHE:
        try:
            url = (f"https://financialmodelingprep.com/stable/income-statement"
                   f"?symbol={ticker}&period=quarter&limit=4&apikey={API_KEY}")
            r = requests.get(url, timeout=10)
            if r.status_code != 200:
                print(f"  [TTM-NI] HTTP {r.status_code} for {ticker}")
                _TTM_NI_CACHE[key] = None
            else:
                data = r.json()
                if isinstance(data, list) and len(data) >= 2:
                    qs = data[:4]   # newest-first from FMP
                    _ttm = sum(q.get("netIncome") or 0 for q in qs)
                    print(f"  [TTM-NI] {len(qs)} qtrs  ttm_ni=${_ttm/1e9:.2f}B")
                    _TTM_NI_CACHE[key] = _ttm
                else:
                    _TTM_NI_CACHE[key] = None
        except Exception as _e_ttm:
            print(f"  [TTM-NI] failed for {ticker}: {_e_ttm}")
            _TTM_NI_CACHE[key] = None
    return _TTM_NI_CACHE[key]

def fetch_segment(endpoint, ticker):
    """Fetch segmentation — returns None gracefully if not on plan."""
    try:
        url = (f"https://financialmodelingprep.com/stable/{endpoint}"
               f"?symbol={ticker}&apikey={API_KEY}")
        r = requests.get(url)
        if r.status_code != 200:
            return None
        data = r.json()
        if not data or (isinstance(data, dict) and ("Error" in str(data) or "message" in data)):
            return None
        return data if isinstance(data, list) else None
    except:
        return None

# ═══════════════════════════════════════════════════════════════════════════════
# WACC HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_fred(series_id):
    """Return (value_as_decimal, date_string) for latest FRED observation.
    Uses public CSV endpoint — no API key required."""
    try:
        csv = requests.get(
            f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}",
            timeout=10
        ).text.strip().split("\n")
        last = next(r for r in reversed(csv)
                    if r and r.split(",")[1] not in (".", ""))
        date, val = last.split(",")
        return float(val) / 100, date
    except Exception:
        return None, None

def fetch_analyst_estimates(ticker, last_hist_year):
    """Fetch FMP annual analyst estimates, return only forward years (sorted oldest→newest).
    Each record: {year, rev_avg, rev_low, rev_high, ebitda_avg, ebitda_low, ebitda_high,
                  ni_avg, eps_avg, n_analysts_rev, n_analysts_eps}
    Returns [] gracefully on any failure."""
    try:
        url = (f"https://financialmodelingprep.com/stable/analyst-estimates"
               f"?symbol={ticker}&period=annual&limit=10&apikey={API_KEY}")
        r = requests.get(url, timeout=10)
        if r.status_code != 200:
            print(f"  [Estimates] HTTP {r.status_code}")
            return []
        raw = r.json()
        if not isinstance(raw, list):
            return []
        out = []
        for rec in raw:
            yr = str(rec.get("date", ""))[:4]
            if yr <= str(last_hist_year):
                continue          # skip historical estimate years
            out.append({
                "year":          yr,
                "rev_avg":       (rec.get("revenueAvg")   or 0) / 1e6,
                "rev_low":       (rec.get("revenueLow")   or 0) / 1e6,
                "rev_high":      (rec.get("revenueHigh")  or 0) / 1e6,
                "ebitda_avg":    (rec.get("ebitdaAvg")    or 0) / 1e6,
                "ebitda_low":    (rec.get("ebitdaLow")    or 0) / 1e6,
                "ebitda_high":   (rec.get("ebitdaHigh")   or 0) / 1e6,
                "ni_avg":        (rec.get("netIncomeAvg") or 0) / 1e6,
                "eps_avg":       rec.get("epsAvg"),
                "n_rev":         rec.get("numAnalystsRevenue") or 0,
                "n_eps":         rec.get("numAnalystsEps")     or 0,
            })
        # Sort oldest → newest
        out.sort(key=lambda x: x["year"])
        print(f"  [Estimates] {len(out)} forward years: "
              f"{[e['year'] for e in out]}")
        return out
    except Exception as e:
        print(f"  [Estimates] Failed: {e}")
        return []

def get_synthetic_rating(icr):
    """Map Interest Coverage Ratio to Damodaran synthetic rating + spread."""
    for lo, hi, rating, spread in DAMODARAN_SPREADS:
        if lo <= icr < hi:
            return rating, spread
    return "CC", 0.12


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def g(rec, key):
    v = rec.get(key)
    try:   return float(v) if v is not None else None
    except: return None

def gm(rec, key):
    v = g(rec, key)
    return round(v / 1e6, 2) if v is not None else None

def g_any(rec, *keys):
    """Try multiple field names, return first non-None value found."""
    for k in keys:
        v = g(rec, k)
        if v is not None:
            return v
    return None

def gm_any(rec, *keys):
    """g_any but scaled to $mm."""
    v = g_any(rec, *keys)
    return round(v / 1e6, 2) if v is not None else None

def setup_ws(ws, years, col_a_width=42):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = col_a_width
    for i in range(len(years)):
        ws.column_dimensions[cl(i+2)].width = 16
    ws.freeze_panes = "B3"

def write_tab_title(ws, row, text, ncols, subtitle=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = fnt(bold=True, color=C_WHITE, size=13)
    c.fill  = fll(C_TITLE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    if subtitle:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        s = ws.cell(row=row, column=1, value=subtitle)
        s.font = fnt(size=9, italic=True, color="888888")
        ws.row_dimensions[row].height = 14
    return row + 1

def write_section_hdr(ws, row, text, ncols, color=None):
    bg = color or C_SECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font  = fnt(bold=True, color=C_WHITE, size=10)
    c.fill  = fll(bg)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 18
    return row + 1

def write_year_hdr(ws, row, years, ncols, label="Fiscal Year Ending"):
    c = ws.cell(row=row, column=1, value=label)
    c.font  = fnt(bold=True, size=10)
    c.fill  = fll(C_SUBTOTAL)
    c.border = brd()
    c.alignment = Alignment(horizontal="left", indent=1)
    for i, yr in enumerate(years):
        cell = ws.cell(row=row, column=i+2, value=yr)
        cell.font  = fnt(bold=True, size=10)
        cell.fill  = fll(C_SUBTOTAL)
        cell.alignment = Alignment(horizontal="right")
        cell.border = brd()
    ws.row_dimensions[row].height = 18
    return row + 1

def write_data_row(ws, row, label, values, years,
                   bold=False, bg=None, indent=0,
                   is_pct=False, is_ratio=False, is_days=False, color=None):
    bg = bg or C_WHITE
    tc = color or (C_BLUE if not is_pct and not is_ratio and not is_days else C_BLACK)
    c = ws.cell(row=row, column=1, value=label)
    c.font  = fnt(bold=bold, size=10)
    c.fill  = fll(bg)
    c.border = brd()
    c.alignment = Alignment(horizontal="left", indent=1+indent)
    for i in range(len(years)):
        cell = ws.cell(row=row, column=i+2)
        cell.value = values[i] if i < len(values) else None
        cell.font  = fnt(bold=bold, color=tc, size=10)
        cell.fill  = fll(bg)
        cell.border = brd()
        cell.alignment = Alignment(horizontal="right")
        if is_pct:   pct_fmt(cell)
        elif is_ratio: ratio_fmt(cell)
        elif is_days:  days_fmt(cell)
        else:          num_fmt(cell)
    return row + 1

def write_formula_row(ws, row, label, formula_fn, n_years,
                      bold=False, bg=None, indent=0,
                      is_pct=False, is_ratio=False, is_days=False):
    bg = bg or C_WHITE
    c = ws.cell(row=row, column=1, value=label)
    c.font  = fnt(bold=bold, size=10)
    c.fill  = fll(bg)
    c.border = brd()
    c.alignment = Alignment(horizontal="left", indent=1+indent)
    for i in range(n_years):
        col = i+2
        cell = ws.cell(row=row, column=col)
        cell.value = formula_fn(row, col)
        cell.font  = fnt(bold=bold, color=C_BLACK, size=10)
        cell.fill  = fll(bg)
        cell.border = brd()
        cell.alignment = Alignment(horizontal="right")
        if is_pct:    pct_fmt(cell)
        elif is_ratio: ratio_fmt(cell)
        elif is_days:  days_fmt(cell)
        else:          num_fmt(cell)
    return row + 1

def patch_formula_cells(ws, target_row, n_years, formula_fn,
                        bold=False, bg=None,
                        is_pct=False, is_ratio=False, is_days=False):
    """
    Overwrite data cells in target_row with formulas (fix-after pattern).
    Also corrects font color to black so cells look like formula outputs.
    """
    bg = bg or C_WHITE
    for i in range(n_years):
        col = i + 2
        cell = ws.cell(row=target_row, column=col)
        cell.value = formula_fn(target_row, col)
        cell.font  = fnt(bold=bold, color=C_BLACK, size=10)
        cell.fill  = fll(bg)
        cell.border = brd()
        cell.alignment = Alignment(horizontal="right")
        if is_pct:    pct_fmt(cell)
        elif is_ratio: ratio_fmt(cell)
        elif is_days:  days_fmt(cell)
        else:          num_fmt(cell)

def blank_row(ws, row, ncols):
    ws.row_dimensions[row].height = 6
    return row + 1

# ═══════════════════════════════════════════════════════════════════════════════
# P&L TAB
# v4 changes:
#   1. Net Interest formula fixed: Interest Income − Interest Expense
#      (positive = net earner, e.g. NVDA cash-rich)
#   2. Added "Other Non-Operating Income / (Expenses)" line in summary
#      between Net Interest and EBT, formula = EBT − EBIT − Net Interest
# ═══════════════════════════════════════════════════════════════════════════════
def build_pl(wb, data, years, ticker):
    ws = wb.create_sheet("P&L")
    n  = len(years)
    nc = n + 1
    setup_ws(ws, years)

    row = write_tab_title(ws, 1,
        f"{ticker} — Income Statement ($mm)",
        nc, subtitle="All figures in USD millions. Blue = source data, Black = formula.")
    row = write_year_hdr(ws, row, years, nc)

    def v(key): return [gm(d, key) for d in data]

    # ── SUMMARY ───────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "SUMMARY — INCOME STATEMENT", nc, C_SUMMARY_HD)

    rev_row = row
    row = write_data_row(ws, row, "(1)  Revenue", v("revenue"), years, bold=True)

    cogs_row = row
    row = write_data_row(ws, row, "(2)  Cost of Revenue (COGS)", v("costOfRevenue"), years)

    gp_row = row
    row = write_formula_row(ws, row, "(3)  Gross Profit", bold=True, bg=C_SUMMARY_BG,
        formula_fn=lambda r,c: f"={cl(c)}{rev_row}-{cl(c)}{cogs_row}", n_years=n)
    row = write_formula_row(ws, row, "     Gross Margin %", indent=1, is_pct=True,
        formula_fn=lambda r,c: f"=IFERROR({cl(c)}{gp_row}/{cl(c)}{rev_row},\"\")", n_years=n)

    sga_row = row
    row = write_data_row(ws, row, "(4)  SG&A", v("sellingGeneralAndAdministrativeExpenses"), years)

    # Other OPEX placeholder — fixed after opex_row is known
    _other_opex_r = row
    row = write_formula_row(ws, row, "(5)  Other OPEX (ex-SG&A)",
        lambda r,c: '=""', n_years=n)

    opex_row = row
    row = write_data_row(ws, row, "     Total Operating Expenses", v("operatingExpenses"), years, indent=1)

    # Fix Other OPEX formula
    patch_formula_cells(ws, _other_opex_r, n,
        lambda r,c: f"=IFERROR({cl(c)}{opex_row}-{cl(c)}{sga_row},\"\")")

    # v4 FIX: EBITDA as formula = EBIT + D&A (pure GAAP EBITDA).
    # FMP's ebitda field adds back SBC and other non-cash items (Adjusted EBITDA).
    # Using formula ensures internal consistency. FMP's adjusted figure is in the Detail section.
    ebitda_row = row
    row = write_formula_row(ws, row, "(6)  EBITDA", bold=True, bg=C_SUMMARY_BG,
        formula_fn=lambda r,c: '=""', n_years=n)   # placeholder — fixed after ebit_row
    row = write_formula_row(ws, row, "     EBITDA Margin %", indent=1, is_pct=True,
        formula_fn=lambda r,c: f"=IFERROR({cl(c)}{ebitda_row}/{cl(c)}{rev_row},\"\")", n_years=n)

    da_row = row
    row = write_data_row(ws, row, "(7)  Depreciation & Amortisation", v("depreciationAndAmortization"), years)

    ebit_row = row
    row = write_data_row(ws, row, "(8)  EBIT (Operating Income)", v("operatingIncome"), years, bold=True, bg=C_SUMMARY_BG)
    row = write_formula_row(ws, row, "     EBIT Margin %", indent=1, is_pct=True,
        formula_fn=lambda r,c: f"=IFERROR({cl(c)}{ebit_row}/{cl(c)}{rev_row},\"\")", n_years=n)

    # Fix EBITDA = EBIT + D&A
    patch_formula_cells(ws, ebitda_row, n,
        lambda r,c: f"={cl(c)}{ebit_row}+{cl(c)}{da_row}",
        bold=True, bg=C_SUMMARY_BG)

    int_exp_row = row
    row = write_data_row(ws, row, "     Interest Expense", v("interestExpense"), years, indent=1)
    int_inc_row = row
    row = write_data_row(ws, row, "     Interest Income", v("interestIncome"), years, indent=1)

    # v4 FIX: Interest Income (+) minus Interest Expense
    # Positive = net earner (e.g. cash-rich companies like NVDA)
    net_int_row = row
    row = write_formula_row(ws, row, "(9)  Net Interest Income / (Expense)",
        formula_fn=lambda r,c: f"=IFERROR({cl(c)}{int_inc_row}-{cl(c)}{int_exp_row},\"\")", n_years=n)

    # v4 NEW: Other Non-Operating Income / (Expenses) placeholder — fixed after ebt_row
    _other_no_r = row
    row = write_formula_row(ws, row, "     Other Non-Operating Income / (Expenses)",
        lambda r,c: '=""', n_years=n, indent=1)

    ebt_row = row
    row = write_data_row(ws, row, "(10) EBT / NPBT", v("incomeBeforeTax"), years, bold=True, bg=C_SUMMARY_BG)

    # Fix Other Non-Operating: EBT − EBIT − Net Interest
    patch_formula_cells(ws, _other_no_r, n,
        lambda r,c: f"=IFERROR({cl(c)}{ebt_row}-{cl(c)}{ebit_row}-{cl(c)}{net_int_row},\"\")")

    tax_row = row
    row = write_data_row(ws, row, "(11) Income Tax Expense", v("incomeTaxExpense"), years)
    row = write_formula_row(ws, row, "     Effective Tax Rate %", indent=1, is_pct=True,
        formula_fn=lambda r,c: f"=IFERROR({cl(c)}{tax_row}/{cl(c)}{ebt_row},\"\")", n_years=n)

    ni_row = row
    row = write_data_row(ws, row, "(12) Net Income / NPAT", v("netIncome"), years, bold=True, bg=C_SUMMARY_BG)
    row = write_formula_row(ws, row, "     Net Margin %", indent=1, is_pct=True,
        formula_fn=lambda r,c: f"=IFERROR({cl(c)}{ni_row}/{cl(c)}{rev_row},\"\")", n_years=n)

    row = blank_row(ws, row, nc)

    # ── DETAIL ────────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "DETAIL — ALL LINE ITEMS FROM FMP", nc, C_DETAIL_HD)
    row = write_year_hdr(ws, row, years, nc)

    row = write_section_hdr(ws, row, "Revenue & Cost", nc)
    row = write_data_row(ws, row, "Revenue",                            v("revenue"),                                    years, bold=True)
    row = write_data_row(ws, row, "Cost of Revenue",                    v("costOfRevenue"),                              years)
    row = write_data_row(ws, row, "Gross Profit",                       v("grossProfit"),                                years, bold=True, bg=C_ALT)

    row = write_section_hdr(ws, row, "Operating Expenses", nc)
    row = write_data_row(ws, row, "R&D Expenses",                       v("researchAndDevelopmentExpenses"),             years)
    row = write_data_row(ws, row, "General & Administrative",           v("generalAndAdministrativeExpenses"),           years)
    row = write_data_row(ws, row, "Selling & Marketing",                v("sellingAndMarketingExpenses"),                years)
    row = write_data_row(ws, row, "SG&A (Combined)",                    v("sellingGeneralAndAdministrativeExpenses"),    years)
    row = write_data_row(ws, row, "Other Expenses",                     v("otherExpenses"),                              years)
    row = write_data_row(ws, row, "Total Operating Expenses",           v("operatingExpenses"),                          years, bold=True, bg=C_ALT)
    row = write_data_row(ws, row, "Cost & Expenses (COGS + Opex)",      v("costAndExpenses"),                            years)

    row = write_section_hdr(ws, row, "Operating & EBITDA", nc)
    row = write_data_row(ws, row, "EBIT (Operating Income)",            v("operatingIncome"),                            years, bold=True)
    row = write_data_row(ws, row, "EBIT (FMP field)",                   v("ebit"),                                       years)
    row = write_data_row(ws, row, "EBITDA",                             v("ebitda"),                                     years, bold=True, bg=C_ALT)
    row = write_data_row(ws, row, "Depreciation & Amortisation",        v("depreciationAndAmortization"),                years)

    row = write_section_hdr(ws, row, "Below the Line", nc)
    row = write_data_row(ws, row, "Interest Income",                    v("interestIncome"),                             years)
    row = write_data_row(ws, row, "Interest Expense",                   v("interestExpense"),                            years)
    row = write_data_row(ws, row, "Net Interest Income",                v("netInterestIncome"),                          years)
    row = write_data_row(ws, row, "Non-Operating Income (ex-interest)", v("nonOperatingIncomeExcludingInterest"),        years)
    total_other_row = row
    row = write_data_row(ws, row, "Total Other Income / (Expenses)",    v("totalOtherIncomeExpensesNet"),                years)
    row = write_data_row(ws, row, "EBT / Income Before Tax",            v("incomeBeforeTax"),                            years, bold=True, bg=C_ALT)
    row = write_data_row(ws, row, "Income Tax Expense",                 v("incomeTaxExpense"),                           years)

    row = write_section_hdr(ws, row, "Net Income", nc)
    row = write_data_row(ws, row, "Net Income from Continuing Ops",     v("netIncomeFromContinuingOperations"),          years)
    row = write_data_row(ws, row, "Net Income from Discontinued Ops",   v("netIncomeFromDiscontinuedOperations"),        years)
    row = write_data_row(ws, row, "Other Adjustments to Net Income",    v("otherAdjustmentsToNetIncome"),                years)
    row = write_data_row(ws, row, "Net Income Deductions",              v("netIncomeDeductions"),                        years)
    row = write_data_row(ws, row, "Net Income",                         v("netIncome"),                                  years, bold=True, bg=C_ALT)
    row = write_data_row(ws, row, "Bottom Line Net Income",             v("bottomLineNetIncome"),                        years, bold=True)

    row = write_section_hdr(ws, row, "Per Share", nc)
    row = write_data_row(ws, row, "EPS (Basic)",                        [g(d,"eps") for d in data],                      years)
    row = write_data_row(ws, row, "EPS (Diluted)",                      [g(d,"epsdiluted") for d in data],               years)
    row = write_data_row(ws, row, "Shares Outstanding — Basic (mm)",    [gm(d,"weightedAverageShsOut") for d in data],   years)
    row = write_data_row(ws, row, "Shares Outstanding — Diluted (mm)",  [gm(d,"weightedAverageShsOutDil") for d in data],years)

    row = write_section_hdr(ws, row, "Metadata", nc)
    row = write_data_row(ws, row, "Reported Currency",                  [data[i].get("reportedCurrency","") for i in range(min(n,len(data)))], years)

    return {"rev": rev_row, "cogs": cogs_row, "gp": gp_row,
            "sga": sga_row, "opex": opex_row,
            "ebitda": ebitda_row, "da": da_row,
            "ebit": ebit_row, "int_exp": int_exp_row, "int_inc": int_inc_row,
            "net_int": net_int_row,
            "ebt": ebt_row, "tax": tax_row, "ni": ni_row}

# ═══════════════════════════════════════════════════════════════════════════════
# BALANCE SHEET TAB
# v4 changes (summary):
#   3.  (4)  Other Current Assets  → formula plug: TCA − Cash − Rec − Inv
#   4.  (9)  Other LT Assets       → formula plug: TLTA − PPE − Goodwill − DTA
#   5.  (14) Short-Term Leases     → uses correct FMP current-lease field
#   6.  (15) Other Current Liabs   → formula plug: TCL − AP − STDebt − STLeases
#   7.  (18) Long-Term Leases      → uses correct FMP LT-lease field
#   8.  (19) Other LT Liabilities  → formula plug: TL − TCL − LTDebt − LTLeases
#   9.  (23) Other Equity          → formula plug: TE − CommonStock − RetainedEarnings
#   10. (25) Total L&E             → formula: TL + TE
# v4 changes (detail):
#   11. Added "Accrued & Other Current Liabilities" to current liabilities section
#   12. Moved "Minority Interest" from Non-Current Liabilities → Shareholders' Equity
#   13. Added "Long-Term Operating Lease Liabilities" in non-current liabilities
#   14. Total Non-Current Liabilities → formula sum of components
#   15. Total Stockholders' Equity (detail) → formula sum of components
# ═══════════════════════════════════════════════════════════════════════════════
def build_bs(wb, data, years, ticker):
    ws = wb.create_sheet("Balance Sheet")
    n  = len(years)
    nc = n + 1
    setup_ws(ws, years)

    row = write_tab_title(ws, 1, f"{ticker} — Balance Sheet ($mm)", nc,
        subtitle="All figures in USD millions. Blue = source data, Black = formula.")
    row = write_year_hdr(ws, row, years, nc)

    def v(key): return [gm(d, key) for d in data]

    # Lease field helpers: try the most specific FMP fields first,
    # then fall back to broader fields.  Returns $mm list.
    def v_st_leases():
        return [gm_any(d,
            "shortTermCapitalLeaseObligation",   # FMP newer naming
            "currentPortionLeaseLiabilities",
            "shortTermLeaseLiabilities",
            "operatingLeaseLiabilityCurrentPortion",
        ) for d in data]

    def v_lt_leases():
        return [gm_any(d,
            "longTermCapitalLeaseObligation",    # FMP newer naming
            "longTermLeaseLiabilities",
            "operatingLeaseLiabilityNoncurrentPortion",
            "operatingLeaseLiabilityNonCurrent",
        ) for d in data]

    # ── SUMMARY ───────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "SUMMARY — BALANCE SHEET", nc, C_SUMMARY_HD)

    cash_row = row
    row = write_data_row(ws, row, "(1)  Cash & Cash Equivalents",      v("cashAndCashEquivalents"),    years)
    rec_row  = row
    row = write_data_row(ws, row, "(2)  Receivables",                  v("netReceivables"),            years)
    inv_row  = row
    row = write_data_row(ws, row, "(3)  Inventory",                    v("inventory"),                 years)

    # v4: (4) Other Current Assets = TCA − Cash − Rec − Inv (plug)
    _oca_r = row
    row = write_formula_row(ws, row, "(4)  Other Current Assets",
        lambda r,c: '=""', n_years=n)

    tca_row  = row
    row = write_data_row(ws, row, "(5)  Total Current Assets",         v("totalCurrentAssets"),        years, bold=True, bg=C_SUMMARY_BG)

    # Fix OCA plug now that tca_row is known
    patch_formula_cells(ws, _oca_r, n,
        lambda r,c: f"=IFERROR({cl(c)}{tca_row}-{cl(c)}{cash_row}-{cl(c)}{rec_row}-{cl(c)}{inv_row},\"\")")

    ppe_row  = row
    row = write_data_row(ws, row, "(6)  PP&E (Net)",                   v("propertyPlantEquipmentNet"), years)
    gw_row   = row
    row = write_data_row(ws, row, "(7)  Goodwill",                     v("goodwill"),                  years)
    dta_row  = row
    row = write_data_row(ws, row, "(8)  Deferred Tax Assets",          v("taxAssets"),                 years)

    # v4: (9) Other LT Assets = TLTA − PPE − Goodwill − DTA (plug)
    _olta_r = row
    row = write_formula_row(ws, row, "(9)  Other LT Assets",
        lambda r,c: '=""', n_years=n)

    tlta_row = row
    row = write_data_row(ws, row, "(10) Total LT Assets",              v("totalNonCurrentAssets"),     years, bold=True, bg=C_SUMMARY_BG)

    # Fix Other LT Assets plug
    patch_formula_cells(ws, _olta_r, n,
        lambda r,c: f"=IFERROR({cl(c)}{tlta_row}-{cl(c)}{ppe_row}-{cl(c)}{gw_row}-{cl(c)}{dta_row},\"\")")

    tot_assets_row = row
    row = write_data_row(ws, row, "(11) Total Assets",                 v("totalAssets"),               years, bold=True, bg=C_SUMMARY_BG)

    row = blank_row(ws, row, nc)

    ap_row   = row
    row = write_data_row(ws, row, "(12) Accounts Payable",             v("accountPayables"),           years)
    std_row  = row
    row = write_data_row(ws, row, "(13) Short-Term Borrowings",        v("shortTermDebt"),             years)

    # v4: Short-Term Leases uses proper current-portion field
    stl_row  = row
    row = write_data_row(ws, row, "(14) Short-Term Leases",            v_st_leases(),                  years)

    # v4: (15) Other Current Liabilities = TCL − AP − STDebt − STLeases (plug)
    _ocl_r = row
    row = write_formula_row(ws, row, "(15) Other Current Liabilities",
        lambda r,c: '=""', n_years=n)

    tcl_row  = row
    row = write_data_row(ws, row, "(16) Total Current Liabilities",    v("totalCurrentLiabilities"),   years, bold=True, bg=C_SUMMARY_BG)

    # Fix Other CL plug
    patch_formula_cells(ws, _ocl_r, n,
        lambda r,c: f"=IFERROR({cl(c)}{tcl_row}-{cl(c)}{ap_row}-{cl(c)}{std_row}-{cl(c)}{stl_row},\"\")")

    ltd_row  = row
    row = write_data_row(ws, row, "(17) Long-Term Debt",               v("longTermDebt"),              years)

    # v4: Long-Term Leases uses proper LT-lease field
    ltl_row  = row
    row = write_data_row(ws, row, "(18) Long-Term Leases",             v_lt_leases(),                  years)

    # v4: (19) Other LT Liabilities = TL − TCL − LTDebt − LTLeases (plug)
    _oltl_r = row
    row = write_formula_row(ws, row, "(19) Other LT Liabilities",
        lambda r,c: '=""', n_years=n)

    tl_row   = row
    row = write_data_row(ws, row, "(20) Total Liabilities",            v("totalLiabilities"),          years, bold=True, bg=C_SUMMARY_BG)

    # Fix Other LT Liabilities plug
    patch_formula_cells(ws, _oltl_r, n,
        lambda r,c: f"=IFERROR({cl(c)}{tl_row}-{cl(c)}{tcl_row}-{cl(c)}{ltd_row}-{cl(c)}{ltl_row},\"\")")

    row = blank_row(ws, row, nc)

    cs_row   = row
    row = write_data_row(ws, row, "(21) Common Stock & APIC",          v("commonStock"),               years)
    re_row   = row
    row = write_data_row(ws, row, "(22) Retained Earnings",            v("retainedEarnings"),          years)

    # v4: (23) Other Equity = TE − Common Stock − Retained Earnings (plug)
    _oe_r = row
    row = write_formula_row(ws, row, "(23) Other Equity",
        lambda r,c: '=""', n_years=n)

    te_row   = row
    row = write_data_row(ws, row, "(24) Total Equity",                 v("totalStockholdersEquity"),   years, bold=True, bg=C_SUMMARY_BG)

    # Fix Other Equity plug
    patch_formula_cells(ws, _oe_r, n,
        lambda r,c: f"=IFERROR({cl(c)}{te_row}-{cl(c)}{cs_row}-{cl(c)}{re_row},\"\")")

    # v4: (25) Total Liabilities & Equity = TL + TE (formula, not raw data)
    tle_row  = row
    row = write_formula_row(ws, row, "(25) Total Liabilities & Equity",
        formula_fn=lambda r,c: f"={cl(c)}{tl_row}+{cl(c)}{te_row}",
        n_years=n, bold=True, bg=C_SUMMARY_BG)

    row = blank_row(ws, row, nc)

    # ── DETAIL ────────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "DETAIL — ALL LINE ITEMS FROM FMP", nc, C_DETAIL_HD)
    row = write_year_hdr(ws, row, years, nc)

    row = write_section_hdr(ws, row, "Current Assets", nc)
    row = write_data_row(ws, row, "Cash & Cash Equivalents",               v("cashAndCashEquivalents"),           years)
    row = write_data_row(ws, row, "Short-Term Investments",                v("shortTermInvestments"),             years)
    row = write_data_row(ws, row, "Cash & Short-Term Investments",         v("cashAndShortTermInvestments"),      years, bg=C_ALT)
    row = write_data_row(ws, row, "Accounts Receivable (Trade)",           v("accountsReceivables"),              years)
    row = write_data_row(ws, row, "Other Receivables",                     v("otherReceivables"),                 years)
    row = write_data_row(ws, row, "Net Receivables (Total)",               v("netReceivables"),                   years, bg=C_ALT)
    row = write_data_row(ws, row, "Inventory",                             v("inventory"),                        years)
    row = write_data_row(ws, row, "Prepaids",                              v("prepaids"),                         years)
    row = write_data_row(ws, row, "Other Current Assets",                  v("otherCurrentAssets"),               years)
    row = write_data_row(ws, row, "Total Current Assets",                  v("totalCurrentAssets"),               years, bold=True, bg=C_ALT)

    row = write_section_hdr(ws, row, "Non-Current Assets", nc)
    row = write_data_row(ws, row, "PP&E (Net)",                            v("propertyPlantEquipmentNet"),        years)
    row = write_data_row(ws, row, "Goodwill",                              v("goodwill"),                         years)
    row = write_data_row(ws, row, "Intangible Assets",                     v("intangibleAssets"),                 years)
    row = write_data_row(ws, row, "Long-Term Investments",                 v("longTermInvestments"),              years)
    row = write_data_row(ws, row, "Tax Assets (Deferred)",                 v("taxAssets"),                        years)
    row = write_data_row(ws, row, "Total Investments",                     v("totalInvestments"),                 years)
    row = write_data_row(ws, row, "Other Non-Current Assets",              v("otherNonCurrentAssets"),            years)
    row = write_data_row(ws, row, "Total Non-Current Assets",              v("totalNonCurrentAssets"),            years, bold=True, bg=C_ALT)
    row = write_data_row(ws, row, "TOTAL ASSETS",                          v("totalAssets"),                      years, bold=True, bg=C_SUBTOTAL)

    # v4 DETAIL — Current Liabilities
    row = write_section_hdr(ws, row, "Current Liabilities", nc)
    d_ap_r = row
    row = write_data_row(ws, row, "Accounts Payable",                      v("accountPayables"),                  years)
    d_std_r = row
    row = write_data_row(ws, row, "Short-Term Debt",                       v("shortTermDebt"),                    years)
    d_stl_r = row
    row = write_data_row(ws, row, "Short-Term Lease Liabilities",          v_st_leases(),                         years)
    d_drev_cur_r = row
    row = write_data_row(ws, row, "Deferred Revenue (Current)",            v("deferredRevenue"),                  years)
    d_accrued_r = row
    row = write_data_row(ws, row, "Accrued & Other Current Liabilities",
        [gm_any(d,
            "accruedLiabilities",
            "accruedAndOtherCurrentLiabilities",
            "otherCurrentLiabilities",
        ) for d in data], years)
    # Other Current Liabilities = plug: Total CL − AP − ST Debt − ST Leases − Def Rev − Accrued
    _d_ocl_r = row
    row = write_formula_row(ws, row, "Other Current Liabilities",
        lambda r,c: '=""', n_years=n)
    d_tcl_detail_r = row
    row = write_data_row(ws, row, "Total Current Liabilities",             v("totalCurrentLiabilities"),          years, bold=True, bg=C_ALT)
    # Fix Other CL plug
    patch_formula_cells(ws, _d_ocl_r, n,
        lambda r,c: (
            f"=IFERROR({cl(c)}{d_tcl_detail_r}"
            f"-{cl(c)}{d_ap_r}-{cl(c)}{d_std_r}-{cl(c)}{d_stl_r}"
            f"-{cl(c)}{d_drev_cur_r}-{cl(c)}{d_accrued_r},\"\")"
        ))

    # v4 DETAIL — Non-Current Liabilities
    row = write_section_hdr(ws, row, "Non-Current Liabilities", nc)
    d_ltd_row = row
    row = write_data_row(ws, row, "Long-Term Debt",                        v("longTermDebt"),                     years)
    d_ltl_row = row
    row = write_data_row(ws, row, "Long-Term Operating Lease Liabilities", v_lt_leases(),                         years)
    d_drev_nc_r = row
    row = write_data_row(ws, row, "Deferred Revenue (Non-Current)",        v("deferredRevenueNonCurrent"),        years)
    d_dtl_row  = row
    row = write_data_row(ws, row, "Deferred Tax Liabilities",              v("deferredTaxLiabilitiesNonCurrent"), years)
    d_oltl_row = row
    row = write_data_row(ws, row, "Other Non-Current Liabilities",         v("otherNonCurrentLiabilities"),       years)
    d_tncl_row = row
    row = write_formula_row(ws, row, "Total Non-Current Liabilities",
        formula_fn=lambda r,c: (
            f"=IFERROR({cl(c)}{d_ltd_row}+{cl(c)}{d_ltl_row}"
            f"+{cl(c)}{d_drev_nc_r}+{cl(c)}{d_dtl_row}+{cl(c)}{d_oltl_row},\"\")"
        ), n_years=n, bold=True, bg=C_ALT)
    row = write_data_row(ws, row, "TOTAL LIABILITIES",                     v("totalLiabilities"),                 years, bold=True, bg=C_SUBTOTAL)

    # v4 DETAIL — Shareholders' Equity
    row = write_section_hdr(ws, row, "Shareholders' Equity", nc)
    d_cs_row = row
    row = write_data_row(ws, row, "Common Stock & APIC",                   v("commonStock"),                      years)
    d_re_row = row
    row = write_data_row(ws, row, "Retained Earnings",                     v("retainedEarnings"),                 years)
    # Other Total SE = plug (fix-after d_te_inc_min is written)
    _d_ose_r = row
    row = write_formula_row(ws, row, "Other Total Stockholders Equity",
        lambda r,c: '=""', n_years=n)
    # Total Stockholders' Equity = formula sum of components
    d_tse_row = row
    row = write_formula_row(ws, row, "Total Stockholders Equity",
        formula_fn=lambda r,c: (
            f"=IFERROR({cl(c)}{d_cs_row}+{cl(c)}{d_re_row}+{cl(c)}{_d_ose_r},\"\")"
        ), n_years=n, bold=True, bg=C_ALT)
    d_mi_row = row
    row = write_data_row(ws, row, "Minority Interest",                     v("minorityInterest"),                 years)
    d_te_inc_min_row = row
    row = write_data_row(ws, row, "Total Equity (inc. Minority)",          v("totalEquity"),                      years, bold=True, bg=C_ALT)
    # Fix Other Total SE plug = Total Equity (inc. minority) − Common − Retained − Minority
    patch_formula_cells(ws, _d_ose_r, n,
        lambda r,c: (
            f"=IFERROR({cl(c)}{d_te_inc_min_row}"
            f"-{cl(c)}{d_cs_row}-{cl(c)}{d_re_row}-{cl(c)}{d_mi_row},\"\")"
        ))

    row = write_section_hdr(ws, row, "Key Derived Balances", nc)
    td_kd_row = row
    row = write_formula_row(ws, row, "Total Debt (ST + LT)",
        formula_fn=lambda r, c: f"={cl(c)}{std_row}+{cl(c)}{ltd_row}",
        n_years=n, bold=True)
    nd_kd_row = row
    row = write_formula_row(ws, row, "Net Debt",
        formula_fn=lambda r, c: f"={cl(c)}{td_kd_row}-{cl(c)}{cash_row}",
        n_years=n, bold=True)
    # Working Capital = formula: Total CA − Total CL (from summary rows, same sheet)
    row = write_formula_row(ws, row, "Working Capital",
        formula_fn=lambda r,c: f"=IFERROR({cl(c)}{tca_row}-{cl(c)}{tcl_row},\"\")",
        n_years=n, bold=True)
    # Total L&E = reference the summary Total L&E formula row
    row = write_formula_row(ws, row, "TOTAL LIABILITIES & EQUITY",
        formula_fn=lambda r,c: f"={cl(c)}{tle_row}",
        n_years=n, bold=True, bg=C_SUBTOTAL)

    return {"cash": cash_row, "rec": rec_row, "inv": inv_row, "tca": tca_row,
            "ppe": ppe_row, "dta": dta_row, "tlta": tlta_row, "tot_assets": tot_assets_row,
            "ap": ap_row, "tcl": tcl_row, "ltd": ltd_row,
            "tl": tl_row, "te": te_row, "nd": nd_kd_row}

# ═══════════════════════════════════════════════════════════════════════════════
# CASH FLOW TAB
# v4 changes:
#   16. Added "Other Investing Activities" in detail investing section
#       so the sub-items reconcile to total CFI
#   17. Added "Other Financing Activities" in detail financing section
#       so the sub-items reconcile to total CFF
# ═══════════════════════════════════════════════════════════════════════════════
def build_cf(wb, data, years, ticker):
    ws = wb.create_sheet("Cash Flow")
    n  = len(years)
    nc = n + 1
    setup_ws(ws, years)

    row = write_tab_title(ws, 1, f"{ticker} — Cash Flow Statement ($mm)", nc,
        subtitle="All figures in USD millions. Blue = source data, Black = formula.")
    row = write_year_hdr(ws, row, years, nc)

    def v(key): return [gm(d, key) for d in data]

    # ── SUMMARY ───────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "SUMMARY — CASH FLOW STATEMENT", nc, C_SUMMARY_HD)

    cfo_row   = row
    row = write_data_row(ws, row, "(1)  Net Cash from Operations (CFO)", v("netCashProvidedByOperatingActivities"), years, bold=True, bg=C_SUMMARY_BG)

    capex_row = row
    row = write_data_row(ws, row, "(2)  Capital Expenditures",            v("capitalExpenditure"),                  years)

    # Summary CFI and CFF: placeholders — fixed after detail totals are computed as formula sums
    _cfi_summary = row
    cfi_row = row
    row = write_formula_row(ws, row, "(3)  Net Cash from Investing (CFI)",
        lambda r,c: '=""', n_years=n, bold=True, bg=C_SUMMARY_BG)

    row = blank_row(ws, row, nc)

    draw_row  = row
    row = write_data_row(ws, row, "(4)  Debt Drawdowns",                  v("netDebtIssuance"),                     years)
    rep_row   = row
    row = write_data_row(ws, row, "(5)  Debt Repayments",
        [gm_any(d, "debtRepayment", "repaymentOfDebt", "longTermDebtRepayment") for d in data], years)
    iss_row   = row
    row = write_data_row(ws, row, "(6)  Issuance of Common Stock",        v("commonStockIssuance"),                 years)
    div_row   = row
    row = write_data_row(ws, row, "(7)  Dividends Paid",
        [gm_any(d, "dividendsPaid", "commonDividendsPaid", "paymentOfDividends") for d in data], years)
    _cff_summary = row
    cff_row = row
    row = write_formula_row(ws, row, "(8)  Net Cash from Financing (CFF)",
        lambda r,c: '=""', n_years=n, bold=True, bg=C_SUMMARY_BG)

    row = blank_row(ws, row, nc)
    fcf_row   = row
    row = write_data_row(ws, row, "     Free Cash Flow (FCF)",            v("freeCashFlow"),                        years, bold=True, bg=C_SUMMARY_BG)
    row = write_data_row(ws, row, "     Net Change in Cash",              v("netChangeInCash"),                     years, bold=True)

    row = blank_row(ws, row, nc)

    # ── DETAIL ────────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "DETAIL — ALL LINE ITEMS FROM FMP", nc, C_DETAIL_HD)
    row = write_year_hdr(ws, row, years, nc)

    row = write_section_hdr(ws, row, "Operating Activities", nc)
    row = write_data_row(ws, row, "Net Income",                           v("netIncome"),                           years)
    row = write_data_row(ws, row, "Depreciation & Amortisation",          v("depreciationAndAmortization"),         years)
    row = write_data_row(ws, row, "Deferred Income Tax",                  v("deferredIncomeTax"),                   years)
    row = write_data_row(ws, row, "Stock-Based Compensation",             v("stockBasedCompensation"),              years)
    row = write_data_row(ws, row, "Change in Working Capital",            v("changeInWorkingCapital"),              years)
    row = write_data_row(ws, row, "  — Accounts Receivable",              v("accountsReceivables"),                 years, indent=1)
    row = write_data_row(ws, row, "  — Inventory",                        v("inventory"),                           years, indent=1)
    row = write_data_row(ws, row, "  — Accounts Payable",                 v("accountsPayables"),                    years, indent=1)
    row = write_data_row(ws, row, "  — Other Working Capital",            v("otherWorkingCapital"),                 years, indent=1)
    row = write_data_row(ws, row, "Other Non-Cash Items",                 v("otherNonCashItems"),                   years)
    row = write_data_row(ws, row, "Net Cash from Operations (CFO)",       v("netCashProvidedByOperatingActivities"),years, bold=True, bg=C_ALT)

    # Investing detail — track rows for formula sum
    row = write_section_hdr(ws, row, "Investing Activities", nc)
    d_capex_r = row
    row = write_data_row(ws, row, "Capital Expenditures",                 v("capitalExpenditure"),                  years)
    row = write_data_row(ws, row, "  (Alt: Invest. in PP&E)",             v("investmentsInPropertyPlantAndEquipment"), years, indent=1)
    d_acq_r = row
    row = write_data_row(ws, row, "Acquisitions (Net)",                   v("acquisitionsNet"),                     years)
    d_purch_r = row
    row = write_data_row(ws, row, "Purchases of Investments",             v("purchasesOfInvestments"),              years)
    d_sales_r = row
    row = write_data_row(ws, row, "Sales / Maturities of Investments",    v("salesMaturitiesOfInvestments"),        years)
    d_other_inv_r = row
    row = write_data_row(ws, row, "Other Investing Activities",
        [gm_any(d, "otherInvestingActivities", "otherInvestingActivitiesNet") for d in data], years)
    # CFI total = formula sum (Alt PP&E row excluded to avoid double-count with Capex)
    d_cfi_r = row
    row = write_formula_row(ws, row, "Net Cash from Investing (CFI)",
        formula_fn=lambda r,c: (
            f"=IFERROR({cl(c)}{d_capex_r}+{cl(c)}{d_acq_r}"
            f"+{cl(c)}{d_purch_r}+{cl(c)}{d_sales_r}+{cl(c)}{d_other_inv_r},\"\")"
        ), n_years=n, bold=True, bg=C_ALT)

    # Financing detail — track rows for formula sum
    row = write_section_hdr(ws, row, "Financing Activities", nc)
    d_debt_iss_r = row
    row = write_data_row(ws, row, "Debt Issuance (Net)",                  v("netDebtIssuance"),                     years)
    d_debt_rep_r = row
    row = write_data_row(ws, row, "Debt Repayment",
        [gm_any(d, "debtRepayment", "repaymentOfDebt", "longTermDebtRepayment") for d in data], years)
    d_stk_iss_r = row
    row = write_data_row(ws, row, "Common Stock Issuance",                v("commonStockIssuance"),                 years)
    d_buyback_r = row
    row = write_data_row(ws, row, "Common Stock Repurchased (Buybacks)",  v("commonStockRepurchased"),              years)
    d_div_r = row
    row = write_data_row(ws, row, "Dividends Paid",
        [gm_any(d, "dividendsPaid", "commonDividendsPaid", "paymentOfDividends") for d in data], years)
    d_other_fin_r = row
    row = write_data_row(ws, row, "Other Financing Activities",
        [gm_any(d, "otherFinancingActivities", "otherFinancingActivitiesNet") for d in data], years)
    # CFF total = formula sum of all financing line items
    d_cff_r = row
    row = write_formula_row(ws, row, "Net Cash from Financing (CFF)",
        formula_fn=lambda r,c: (
            f"=IFERROR({cl(c)}{d_debt_iss_r}+{cl(c)}{d_debt_rep_r}"
            f"+{cl(c)}{d_stk_iss_r}+{cl(c)}{d_buyback_r}"
            f"+{cl(c)}{d_div_r}+{cl(c)}{d_other_fin_r},\"\")"
        ), n_years=n, bold=True, bg=C_ALT)

    # Patch summary CFI and CFF to reference detail formula totals
    patch_formula_cells(ws, _cfi_summary, n,
        lambda r,c: f"={cl(c)}{d_cfi_r}", bold=True, bg=C_SUMMARY_BG)
    patch_formula_cells(ws, _cff_summary, n,
        lambda r,c: f"={cl(c)}{d_cff_r}", bold=True, bg=C_SUMMARY_BG)

    row = write_section_hdr(ws, row, "Cash Summary", nc)
    row = write_data_row(ws, row, "Effect of Forex on Cash",              v("effectOfForexChangesOnCash"),          years)
    net_change_row = row
    row = write_data_row(ws, row, "Net Change in Cash",                   v("netChangeInCash"),                     years, bold=True)
    row = write_data_row(ws, row, "Cash at Beginning of Period",          v("cashAtBeginningOfPeriod"),             years)
    row = write_data_row(ws, row, "Cash at End of Period",                v("cashAtEndOfPeriod"),                   years, bold=True, bg=C_ALT)
    row = write_data_row(ws, row, "Free Cash Flow",                       v("freeCashFlow"),                        years, bold=True, bg=C_ALT)
    row = write_data_row(ws, row, "Operating Cash Flow per Share",        [g(d,"operatingCashFlowPerShare") for d in data], years)
    row = write_data_row(ws, row, "Free Cash Flow per Share",             [g(d,"freeCashFlowPerShare") for d in data],      years)

    return {"cfo": cfo_row, "capex": capex_row, "cfi": cfi_row,
            "cff": cff_row, "fcf": fcf_row, "div": div_row,
            "net_change": net_change_row}

# ═══════════════════════════════════════════════════════════════════════════════
# EDGAR BANK CREDIT DATA FETCH
# ═══════════════════════════════════════════════════════════════════════════════
def fetch_bank_credit_data(ticker):
    """Fetch bank credit quality metrics from SEC EDGAR XBRL API.
    Returns dict with keys: nco_rates (dict yr->float), nco_latest (float),
    nco_2yr_delta (float), chargeoffs (dict yr->$), gross_loans (dict yr->$),
    allowance (dict yr->$), provision (dict yr->$).
    Returns empty dict on any failure — callers must handle missing keys gracefully.
    Only meaningful for bank tickers; non-bank tickers return {}.
    """
    cik = _BANK_CIKS.get(ticker.upper())
    if not cik:
        return {}

    _HDR = {"User-Agent": "justin.song91@gmail.com"}

    def _fetch(tag, min_year="2020"):
        try:
            url = (f"https://data.sec.gov/api/xbrl/companyconcept/"
                   f"CIK{cik}/us-gaap/{tag}.json")
            r = requests.get(url, headers=_HDR, timeout=10)
            if r.status_code != 200:
                return {}
            units = r.json().get("units", {}).get("USD", [])
            annual = [x for x in units
                      if x.get("form") in ("10-K", "10-K/A")
                      and x.get("end", "") >= f"{min_year}-01-01"]
            by_year = {}
            for row in annual:
                yr = row["end"][:4]
                if yr not in by_year or row.get("filed","") > by_year[yr].get("filed",""):
                    by_year[yr] = row
            return {yr: row["val"] for yr, row in by_year.items()}
        except Exception:
            return {}

    # Net charge-offs — try two tags, use whichever has recent data
    chargeoffs = _fetch("FinancingReceivableExcludingAccruedInterestAllowanceForCreditLossWriteoffAfterRecovery")
    if not chargeoffs or max(chargeoffs.keys(), default="0") < "2022":
        chargeoffs = _fetch("FinancingReceivableAllowanceForCreditLossesWriteOffs")

    # Gross loans
    gross_loans = _fetch("FinancingReceivableExcludingAccruedInterestBeforeAllowanceForCreditLoss")

    # Allowance for credit losses
    allowance = _fetch("FinancingReceivableAllowanceForCreditLossExcludingAccruedInterest")
    if not allowance or max(allowance.keys(), default="0") < "2022":
        allowance = _fetch("FinancingReceivableAllowanceForCreditLosses")

    # Provision for credit losses
    provision = _fetch("FinancingReceivableExcludingAccruedInterestCreditLossExpenseReversal")
    if not provision or max(provision.keys(), default="0") < "2022":
        provision = _fetch("ProvisionForLoanLeaseAndOtherLosses")
    if not provision or max(provision.keys(), default="0") < "2022":
        provision = _fetch("ProvisionForLoanLossesExpensed")

    # Compute NCO rates: net_chargeoffs / avg(gross_loans[yr], gross_loans[yr-1])
    nco_rates = {}
    common_years = sorted(set(chargeoffs.keys()) & set(gross_loans.keys()), reverse=True)
    for yr in common_years:
        loan_curr = gross_loans[yr]
        loan_prev = gross_loans.get(str(int(yr) - 1), loan_curr)
        avg_loans = (loan_curr + loan_prev) / 2
        if avg_loans > 0:
            nco_rates[yr] = chargeoffs[yr] / avg_loans

    years_sorted = sorted(nco_rates.keys(), reverse=True)
    nco_latest  = nco_rates[years_sorted[0]] if years_sorted else None

    # 3yr average — more representative of structural credit quality than a single year.
    # Smooths one-off spikes (e.g. COVID 2020) and reflects the bank's underwriting
    # standard across a full rate cycle rather than the most recent 12 months.
    _last3 = [nco_rates[y] for y in years_sorted[:3] if nco_rates.get(y) is not None]
    nco_3yr_avg = sum(_last3) / len(_last3) if _last3 else None

    # Trend: latest vs 2 years ago (directional signal for management proxy)
    nco_2yr_ago   = nco_rates[years_sorted[2]] if len(years_sorted) >= 3 else None
    nco_2yr_delta = (nco_latest - nco_2yr_ago) if (nco_latest is not None and nco_2yr_ago is not None) else None

    print(f"  EDGAR bank credit ({ticker}): NCO 3yr avg={nco_3yr_avg:.2%}  latest={nco_latest:.2%}" if nco_3yr_avg else
          f"  EDGAR bank credit ({ticker}): no NCO data")

    return {
        "nco_rates":    nco_rates,
        "nco_latest":   nco_latest,
        "nco_3yr_avg":  nco_3yr_avg,
        "nco_2yr_delta": nco_2yr_delta,
        "chargeoffs":   chargeoffs,
        "gross_loans":  gross_loans,
        "allowance":    allowance,
        "provision":    provision,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# RATIOS & FCF BRIDGE TAB  (unchanged from v3)
# ═══════════════════════════════════════════════════════════════════════════════
def build_ratios(wb, is_data, bs_data, cf_data, years, ticker, pl_refs, bs_refs, cf_refs, bank_credit=None):
    ws = wb.create_sheet("Ratios & FCF")
    n  = len(years)
    nc = n + 1
    setup_ws(ws, years)

    row = write_tab_title(ws, 1, f"{ticker} — Key Ratios & Free Cash Flow Bridge", nc,
        subtitle="All formulas. Black = calculated. Cross-sheet references pull from P&L, Balance Sheet, Cash Flow tabs.")
    row = write_year_hdr(ws, row, years, nc)

    def pl(r, col):  return f"'P&L'!{cl(col)}{r}"
    def bs(r, col):  return f"'Balance Sheet'!{cl(col)}{r}"
    def cf(r, col):  return f"'Cash Flow'!{cl(col)}{r}"

    rev   = pl_refs["rev"];   cogs  = pl_refs["cogs"]; gp    = pl_refs["gp"]
    ebitda= pl_refs["ebitda"];da    = pl_refs["da"];   ebit  = pl_refs["ebit"]
    ebt   = pl_refs["ebt"];   tax   = pl_refs["tax"];  ni    = pl_refs["ni"]

    tca   = bs_refs["tca"];   tcl   = bs_refs["tcl"];  tot_a = bs_refs["tot_assets"]
    te    = bs_refs["te"];    ltd   = bs_refs["ltd"];  cash  = bs_refs["cash"]
    rec   = bs_refs["rec"];   inv   = bs_refs["inv"];  ap    = bs_refs["ap"]
    nd    = bs_refs["nd"]

    cfo   = cf_refs["cfo"];   capex = cf_refs["capex"];fcf   = cf_refs["fcf"]

    # ── UNLEVERED FREE CASH FLOW BRIDGE ───────────────────────────────────────
    row = write_section_hdr(ws, row, "UNLEVERED FREE CASH FLOW (UFCF) — STEP-BY-STEP BRIDGE", nc, C_SUMMARY_HD)
    row = write_section_hdr(ws, row, "Note: UFCF = NOPAT + D&A − ΔNWC − Capex  |  Used as input to DCF (unlevered / WACC-based)", nc, "555555")

    ebit_r = row
    row = write_formula_row(ws, row, "EBIT (Operating Income)",
        lambda r,c: f"={pl(ebit, c)}", n, bold=True)

    nopat_r = row
    row = write_formula_row(ws, row, "  (−) Taxes on EBIT  [EBIT × Eff. Tax Rate]",
        lambda r,c: f"=IFERROR(-{cl(c)}{ebit_r}*MAX(0,MIN(0.5,{pl(tax,c)}/{pl(ebt,c)})),0)",
        n, indent=1)

    nopat_total = row
    row = write_formula_row(ws, row, "NOPAT  (Net Operating Profit After Tax)",
        lambda r,c: f"={cl(c)}{ebit_r}+{cl(c)}{nopat_r}",
        n, bold=True, bg=C_SUMMARY_BG)

    da_r = row
    row = write_formula_row(ws, row, "  (+) Depreciation & Amortisation",
        lambda r,c: f"={pl(da, c)}", n, indent=1)

    ebitda_r = row
    row = write_formula_row(ws, row, "  = EBITDA (cross-check)",
        lambda r,c: f"={cl(c)}{nopat_total}+{cl(c)}{da_r}",
        n, indent=1, bg=C_ALT)

    nwc_r = row
    row = write_formula_row(ws, row, "  (−) Increase in Net Working Capital  [ΔRec + ΔInv − ΔAP]",
        lambda r,c: (
            f"=IFERROR(('Balance Sheet'!{cl(c)}{rec}-'Balance Sheet'!{cl(c-1)}{rec})"
            f"+('Balance Sheet'!{cl(c)}{inv}-'Balance Sheet'!{cl(c-1)}{inv})"
            f"-('Balance Sheet'!{cl(c)}{ap}-'Balance Sheet'!{cl(c-1)}{ap}),0)"
            if c > 2 else "=0"
        ), n, indent=1)

    capex_r = row
    row = write_formula_row(ws, row, "  (−) Capital Expenditures",
        lambda r,c: f"={cf(capex, c)}", n, indent=1)

    ufcf_row = row
    row = write_formula_row(ws, row, "UNLEVERED FREE CASH FLOW (UFCF)",
        lambda r,c: f"={cl(c)}{nopat_total}+{cl(c)}{da_r}-{cl(c)}{nwc_r}+{cl(c)}{capex_r}",
        n, bold=True, bg=C_SUMMARY_BG)
    row = write_formula_row(ws, row, "  UFCF Margin %",
        lambda r,c: f"=IFERROR({cl(c)}{ufcf_row}/{pl(rev,c)},\"\")",
        n, is_pct=True, indent=1)

    row = blank_row(ws, row, nc)

    # ── LEVERED FREE CASH FLOW BRIDGE ─────────────────────────────────────────
    row = write_section_hdr(ws, row, "LEVERED FREE CASH FLOW (LFCF) — STEP-BY-STEP BRIDGE", nc, C_SUMMARY_HD)
    row = write_section_hdr(ws, row, "Note: LFCF = Net Income + D&A − ΔNWC − Capex  |  Represents cash available to equity holders", nc, "555555")

    ni_r = row
    row = write_formula_row(ws, row, "Net Income",
        lambda r,c: f"={pl(ni, c)}", n, bold=True)

    da_lev = row
    row = write_formula_row(ws, row, "  (+) Depreciation & Amortisation",
        lambda r,c: f"={pl(da, c)}", n, indent=1)

    nwc_lev = row
    row = write_formula_row(ws, row, "  (−) Increase in Net Working Capital",
        lambda r,c: (
            f"=IFERROR(('Balance Sheet'!{cl(c)}{rec}-'Balance Sheet'!{cl(c-1)}{rec})"
            f"+('Balance Sheet'!{cl(c)}{inv}-'Balance Sheet'!{cl(c-1)}{inv})"
            f"-('Balance Sheet'!{cl(c)}{ap}-'Balance Sheet'!{cl(c-1)}{ap}),0)"
            if c > 2 else "=0"
        ), n, indent=1)

    capex_lev = row
    row = write_formula_row(ws, row, "  (−) Capital Expenditures",
        lambda r,c: f"={cf(capex, c)}", n, indent=1)

    lfcf_row = row
    row = write_formula_row(ws, row, "LEVERED FREE CASH FLOW (LFCF)",
        lambda r,c: f"={cl(c)}{ni_r}+{cl(c)}{da_lev}-{cl(c)}{nwc_lev}+{cl(c)}{capex_lev}",
        n, bold=True, bg=C_SUMMARY_BG)
    row = write_formula_row(ws, row, "  LFCF Margin %",
        lambda r,c: f"=IFERROR({cl(c)}{lfcf_row}/{pl(rev,c)},\"\")",
        n, is_pct=True, indent=1)
    row = write_formula_row(ws, row, "  FCF Conversion (LFCF / Net Income)",
        lambda r,c: f"=IFERROR({cl(c)}{lfcf_row}/{cl(c)}{ni_r},\"\")",
        n, is_pct=True, indent=1)

    row = blank_row(ws, row, nc)

    # ── PROFITABILITY ─────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "PROFITABILITY RATIOS", nc)
    row = write_formula_row(ws, row, "Gross Margin %",
        lambda r,c: f"=IFERROR({pl(gp,c)}/{pl(rev,c)},\"\")", n, is_pct=True)
    row = write_formula_row(ws, row, "EBITDA Margin %",
        lambda r,c: f"=IFERROR({pl(ebitda,c)}/{pl(rev,c)},\"\")", n, is_pct=True)
    row = write_formula_row(ws, row, "EBIT Margin %",
        lambda r,c: f"=IFERROR({pl(ebit,c)}/{pl(rev,c)},\"\")", n, is_pct=True)
    row = write_formula_row(ws, row, "Net Margin %",
        lambda r,c: f"=IFERROR({pl(ni,c)}/{pl(rev,c)},\"\")", n, is_pct=True)
    row = write_formula_row(ws, row, "Return on Equity (ROE)",
        lambda r,c: f"=IFERROR({pl(ni,c)}/{bs(te,c)},\"\")", n, is_pct=True)
    row = write_formula_row(ws, row, "Return on Assets (ROA)",
        lambda r,c: f"=IFERROR({pl(ni,c)}/{bs(tot_a,c)},\"\")", n, is_pct=True)
    row = write_formula_row(ws, row, "ROIC  [NOPAT / (Equity + Net Debt)]",
        lambda r,c: f"=IFERROR(({pl(ebit,c)}*MAX(0,MIN(0.5,1-{pl(tax,c)}/{pl(ebt,c)})))/({bs(te,c)}+'Balance Sheet'!{cl(c)}{nd}),\"\")",
        n, is_pct=True)

    row = blank_row(ws, row, nc)

    # ── LEVERAGE ──────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "LEVERAGE & CREDIT RATIOS", nc)
    row = write_formula_row(ws, row, "Net Debt / EBITDA",
        lambda r,c: f"=IFERROR(('Balance Sheet'!{cl(c)}{ltd}-'Balance Sheet'!{cl(c)}{cash})/{pl(ebitda,c)},\"\")",
        n, is_ratio=True)
    row = write_formula_row(ws, row, "Total Debt / EBITDA",
        lambda r,c: f"=IFERROR('Balance Sheet'!{cl(c)}{ltd}/{pl(ebitda,c)},\"\")",
        n, is_ratio=True)
    row = write_formula_row(ws, row, "Interest Coverage  (EBIT / Interest Expense)",
        lambda r,c: f"=IFERROR({pl(ebit,c)}/ABS({pl(ebt,c)}-{pl(ebit,c)}),\"\")",
        n, is_ratio=True)
    row = write_formula_row(ws, row, "Debt / Equity",
        lambda r,c: f"=IFERROR('Balance Sheet'!{cl(c)}{ltd}/{bs(te,c)},\"\")",
        n, is_ratio=True)
    row = write_formula_row(ws, row, "Total Debt / Total Assets",
        lambda r,c: f"=IFERROR('Balance Sheet'!{cl(c)}{ltd}/{bs(tot_a,c)},\"\")",
        n, is_pct=True)

    row = blank_row(ws, row, nc)

    # ── LIQUIDITY ─────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "LIQUIDITY RATIOS", nc)
    row = write_formula_row(ws, row, "Current Ratio  (CA / CL)",
        lambda r,c: f"=IFERROR({bs(tca,c)}/{bs(tcl,c)},\"\")", n, is_ratio=True)
    row = write_formula_row(ws, row, "Quick Ratio  (CA − Inventory) / CL",
        lambda r,c: f"=IFERROR(({bs(tca,c)}-'Balance Sheet'!{cl(c)}{inv})/{bs(tcl,c)},\"\")",
        n, is_ratio=True)
    row = write_formula_row(ws, row, "Cash Ratio  (Cash / CL)",
        lambda r,c: f"=IFERROR({bs(cash,c)}/{bs(tcl,c)},\"\")", n, is_ratio=True)

    row = blank_row(ws, row, nc)

    # ── EFFICIENCY ────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "EFFICIENCY & WORKING CAPITAL", nc)
    row = write_formula_row(ws, row, "Asset Turnover  (Revenue / Assets)",
        lambda r,c: f"=IFERROR({pl(rev,c)}/{bs(tot_a,c)},\"\")", n, is_ratio=True)
    # Average balance helpers:
    # For cols 3..n+1 use AVERAGE(EOP, BOP) where BOP = prior column (older year).
    # For col 2 (earliest year) use EOP only — no prior year available.
    def avg_bs(row_ref, c):
        if c > 2:
            return f"AVERAGE('Balance Sheet'!{cl(c)}{row_ref},'Balance Sheet'!{cl(c-1)}{row_ref})"
        return f"'Balance Sheet'!{cl(c)}{row_ref}"

    rec_days_row = row
    row = write_formula_row(ws, row, "Receivables Days  (Avg Rec / Rev × 365)",
        lambda r,c: f"=IFERROR({avg_bs(rec,c)}/{pl(rev,c)}*365,\"\")", n, is_days=True)
    inv_days_row = row
    row = write_formula_row(ws, row, "Inventory Days  (Avg Inv / COGS × 365)",
        lambda r,c: f"=IFERROR({avg_bs(inv,c)}/{pl(cogs,c)}*365,\"\")", n, is_days=True)
    ap_days_row = row
    row = write_formula_row(ws, row, "Payables Days  (Avg AP / COGS × 365)",
        lambda r,c: f"=IFERROR({avg_bs(ap,c)}/{pl(cogs,c)}*365,\"\")", n, is_days=True)
    row = write_formula_row(ws, row, "Cash Conversion Cycle  (Rec Days + Inv Days − AP Days)",
        lambda r,c: f"=IFERROR({cl(c)}{rec_days_row}+{cl(c)}{inv_days_row}-{cl(c)}{ap_days_row},\"\")",
        n, is_days=True, bold=True, bg=C_ALT)

    row = blank_row(ws, row, nc)

    # ── CASH FLOW QUALITY ─────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "CASH FLOW QUALITY", nc)
    row = write_formula_row(ws, row, "FCF Margin  (FCF / Revenue)",
        lambda r,c: f"=IFERROR({cf(fcf,c)}/{pl(rev,c)},\"\")", n, is_pct=True)
    row = write_formula_row(ws, row, "FCF Conversion  (FCF / Net Income)",
        lambda r,c: f"=IFERROR({cf(fcf,c)}/{pl(ni,c)},\"\")", n, is_pct=True)
    row = write_formula_row(ws, row, "Capex as % of Revenue",
        lambda r,c: f"=IFERROR(ABS({cf(capex,c)})/{pl(rev,c)},\"\")", n, is_pct=True)
    row = write_formula_row(ws, row, "CFO / Net Income  (Cash Quality)",
        lambda r,c: f"=IFERROR({cf(cfo,c)}/{pl(ni,c)},\"\")", n, is_pct=True)

    row = blank_row(ws, row, nc)

    # ── PER SHARE ─────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "PER SHARE METRICS", nc)
    dilsh = [gm(d,"weightedAverageShsOutDil") for d in is_data]
    shares_row = row
    row = write_data_row(ws, row, "Diluted Shares (mm)", dilsh, years, color=C_BLUE)
    row = write_formula_row(ws, row, "FCF per Share",
        lambda r,c: f"=IFERROR({cf(fcf,c)}/'Ratios & FCF'!{cl(c)}{shares_row},\"\")", n)
    row = write_formula_row(ws, row, "Book Value per Share",
        lambda r,c: f"=IFERROR({bs(te,c)}/'Ratios & FCF'!{cl(c)}{shares_row},\"\")", n)

    row = blank_row(ws, row, nc)

    # ── MODEL CONTROLS ────────────────────────────────────────────────────────
    # Each check shows the difference between two values that should reconcile.
    # A zero (or blank) difference = PASS.  Any non-zero = investigate.
    row = write_section_hdr(ws, row, "MODEL CONTROLS — KEY RECONCILIATION CHECKS", nc, "8B0000")
    row = write_section_hdr(ws, row,
        "Zero = OK  |  Non-zero = investigate  |  Blank = FMP returned no data for that field",
        nc, "555555")

    # ── Master check placeholder — fixed after all individual checks are written ──
    _master_row = row
    row = write_formula_row(ws, row, "MASTER CHECK", lambda r,c: '=""', n_years=n,
        bold=True, bg=C_SUBTOTAL)

    row = blank_row(ws, row, nc)

    check_rows = []   # accumulate row numbers of numeric check cells

    def add_check(label, fml_fn, note=False):
        nonlocal row
        r = row
        row = write_formula_row(ws, row, label, fml_fn, n)
        if not note:
            check_rows.append(r)

    # ── P&L checks ────────────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "P&L Checks", nc, C_DETAIL_HD)

    add_check("Gross Profit: (Revenue - COGS) vs Reported  [= 0]",
        lambda r,c: f"=IFERROR({pl(rev,c)}-{pl(cogs,c)}-{pl(gp,c)},\"\")")

    # EBITDA is now formula = EBIT + DA so this will always be 0 — kept as integrity check
    add_check("EBITDA: (EBIT + D&A) vs Reported  [= 0  — formula-driven, always passes]",
        lambda r,c: f"=IFERROR({pl(ebit,c)}+{pl(da,c)}-{pl(ebitda,c)},\"\")")

    _int_inc = pl_refs["int_inc"]
    _int_exp = pl_refs["int_exp"]
    add_check("Below-the-Line Residual: (EBT - EBIT) - Net Interest  [non-zero = Other Non-Op items, informational]",
        lambda r,c: (
            f"=IFERROR(({pl(ebt,c)}-{pl(ebit,c)})"
            f"-({pl(_int_inc,c)}-{pl(_int_exp,c)}),\"\")"
        ), note=True)   # informational only — excluded from master check

    add_check("Net Income: (EBT - Tax) vs Reported  [= 0]",
        lambda r,c: f"=IFERROR({pl(ebt,c)}-{pl(tax,c)}-{pl(ni,c)},\"\")")

    # ── Balance Sheet checks ──────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "Balance Sheet Checks", nc, C_DETAIL_HD)

    bs_tot_a  = bs_refs["tot_assets"]
    bs_tl     = bs_refs["tl"]
    bs_te_r   = bs_refs["te"]
    bs_tca_r  = bs_refs["tca"]

    add_check("BS Balanced: Total Assets vs (Total Liabilities + Total Equity)  [= 0]",
        lambda r,c: f"=IFERROR({bs(bs_tot_a,c)}-{bs(bs_tl,c)}-{bs(bs_te_r,c)},\"\")")

    add_check("Assets: (Total CA + Total LT Assets) vs Total Assets  [= 0]",
        lambda r,c: (
            f"=IFERROR({bs(bs_tca_r,c)}"
            f"+('Balance Sheet'!{cl(c)}{bs_refs['tlta']})"
            f"-{bs(bs_tot_a,c)},\"\")"
        ))

    # ── Cash Flow checks ──────────────────────────────────────────────────────
    row = write_section_hdr(ws, row, "Cash Flow Checks", nc, C_DETAIL_HD)

    cf_cfo_r  = cf_refs["cfo"]
    cf_cfi_r  = cf_refs["cfi"]
    cf_cff_r  = cf_refs["cff"]
    cf_capex_r = cf_refs["capex"]
    cf_fcf_r   = cf_refs["fcf"]

    add_check("Cash Change: CFO + CFI + CFF vs Reported Net Change  [= 0]",
        lambda r,c: (
            f"=IFERROR({cf(cf_cfo_r,c)}+{cf(cf_cfi_r,c)}+{cf(cf_cff_r,c)}"
            f"-'Cash Flow'!{cl(c)}{cf_refs['net_change']},\"\")"
        ))

    add_check("FCF: (CFO + Capex) vs Reported FCF  [= 0]",
        lambda r,c: f"=IFERROR({cf(cf_cfo_r,c)}+{cf(cf_capex_r,c)}-{cf(cf_fcf_r,c)},\"\")")

    # ── UFCF directional sense check (informational) ──────────────────────────
    row = write_section_hdr(ws, row, "UFCF Bridge Check  [informational — not exact]", nc, C_DETAIL_HD)
    add_check("UFCF vs (CFO + Capex)  [difference = NWC and tax adjustments]",
        lambda r,c: f"=IFERROR({cl(c)}{ufcf_row}-({cf(cf_cfo_r,c)}+{cf(cf_capex_r,c)}),\"\")",
        note=True)

    # ── Master check: counts how many check cells are non-zero ────────────────
    def master_fml(r, c):
        fail_parts = "+".join(
            f"IFERROR((ABS({cl(c)}{cr})>0.01)*1,0)" for cr in check_rows
        )
        return (
            f'=IF(({fail_parts})=0,'
            f'"ALL PASS","FAIL: "&({fail_parts})&" check(s) non-zero")'
        )

    # Write master check label and formula (overwrite placeholder)
    ws.cell(row=_master_row, column=1).value = "MASTER CHECK — ALL CONTROLS"
    ws.cell(row=_master_row, column=1).font  = fnt(bold=True, color=C_WHITE, size=10)
    ws.cell(row=_master_row, column=1).fill  = fll("8B0000")
    ws.cell(row=_master_row, column=1).border = brd()
    ws.cell(row=_master_row, column=1).alignment = Alignment(horizontal="left", indent=1)
    for i in range(n):
        col = i + 2
        cell = ws.cell(row=_master_row, column=col)
        cell.value = master_fml(_master_row, col)
        cell.font  = fnt(bold=True, color=C_WHITE, size=10)
        cell.fill  = fll("8B0000")
        cell.border = brd()
        cell.alignment = Alignment(horizontal="center")

    # ── BANK CREDIT QUALITY (EDGAR XBRL) ──────────────────────────────────────
    if bank_credit and bank_credit.get("nco_rates"):
        row = write_section_hdr(ws, row, "BANK CREDIT QUALITY  (source: SEC EDGAR 10-K XBRL)", nc, C_SUMMARY_HD)
        row = write_section_hdr(ws, row,
            "NCO Rate = Net Charge-offs / Avg Gross Loans  |  Coverage = Allowance / Net Charge-offs  |  Prov/NII = Provision / Net Interest Income",
            nc, "555555")

        def _edgar_vals(src_dict, scalar=1.0):
            """Return list of values aligned to years list, None if year missing."""
            return [src_dict.get(y) and src_dict[y] * scalar if src_dict.get(y) is not None else None for y in years]

        def _nco_rate_vals():
            return [bank_credit["nco_rates"].get(y) for y in years]

        def _coverage_vals():
            out = []
            for y in years:
                co  = bank_credit["chargeoffs"].get(y)
                al  = bank_credit["allowance"].get(y)
                out.append(al / co if (co and al and co > 0) else None)
            return out

        def _prov_nii_vals():
            nii_map = {}
            for is_, yr in zip(is_data, years):
                nii = is_.get("netInterestIncome") or 0
                if nii:
                    nii_map[yr] = nii
            out = []
            for y in years:
                prov = bank_credit["provision"].get(y)
                nii  = nii_map.get(y)
                out.append(prov / nii if (prov and nii and nii > 0) else None)
            return out

        nco_r  = _nco_rate_vals()
        co_r   = _edgar_vals(bank_credit["chargeoffs"], 1/1e9)
        gl_r   = _edgar_vals(bank_credit["gross_loans"], 1/1e9)
        al_r   = _edgar_vals(bank_credit["allowance"],  1/1e9)
        pv_r   = _edgar_vals(bank_credit["provision"],  1/1e9)
        cov_r  = _coverage_vals()
        pni_r  = _prov_nii_vals()

        row = write_data_row(ws, row, "Net Charge-offs ($B)",       co_r,  years, color=C_BLUE)
        row = write_data_row(ws, row, "Gross Loans ($B)",           gl_r,  years, color=C_BLUE)
        row = write_data_row(ws, row, "NCO Rate %",                 nco_r, years, color=C_BLUE, is_pct=True)
        row = write_data_row(ws, row, "Allowance for Credit Losses ($B)", al_r, years, color=C_BLUE)
        row = write_data_row(ws, row, "Coverage Ratio (Allowance / NCO)", cov_r, years, color=C_BLUE, is_ratio=True)
        row = write_data_row(ws, row, "Provision for Credit Losses ($B)", pv_r, years, color=C_BLUE)
        row = write_data_row(ws, row, "Provision / Net Interest Income %", pni_r, years, color=C_BLUE, is_pct=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SEGMENTATION TAB  (unchanged from v3)
# ═══════════════════════════════════════════════════════════════════════════════
def build_segments(wb, ticker, years):
    ws = wb.create_sheet("Segments")
    n  = len(years)
    nc = max(n + 1, 4)
    setup_ws(ws, years)

    row = write_tab_title(ws, 1, f"{ticker} — Revenue Segmentation", nc,
        subtitle="Requires FMP Starter plan or above. Data sourced from company filings.")

    prod_data = fetch_segment("revenue-product-segmentation", ticker)
    geo_data  = fetch_segment("revenue-geographic-segments",  ticker)

    def render_segment(data, title, start_row):
        r = write_section_hdr(ws, start_row, title, nc, C_SUMMARY_HD)
        if not data:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc)
            msg = ws.cell(row=r, column=1,
                value="Data not available on your current FMP plan (requires Starter or above). "
                      "Upgrade at financialmodelingprep.com to unlock segment data.")
            msg.font = fnt(italic=True, color="AA0000", size=10)
            ws.row_dimensions[r].height = 24
            return r + 2

        recent = data[:n]
        meta = {"date","symbol","reportedCurrency","cik","filingDate",
                "acceptedDate","fiscalYear","period"}

        def extract_val(d, seg):
            v = d.get(seg)
            if v is None:
                return None
            if isinstance(v, dict):
                v = list(v.values())[0] if v else None
            try:
                return round(float(v) / 1e6, 2) if v is not None else None
            except:
                return None

        segments = []
        for d in recent:
            for k in d:
                if k not in meta and k not in segments:
                    segments.append(k)

        seg_years = [d.get("date","")[:4] for d in recent]
        r = write_year_hdr(ws, r, seg_years, nc)

        for seg in segments:
            vals = [extract_val(d, seg) for d in recent]
            r = write_data_row(ws, r, seg, vals, seg_years)
        return r + 1

    row = render_segment(prod_data, "PRODUCT / BUSINESS SEGMENT REVENUE ($mm)", row)
    row = blank_row(ws, row, nc)
    row = render_segment(geo_data, "GEOGRAPHIC SEGMENT REVENUE ($mm)", row)

# ═══════════════════════════════════════════════════════════════════════════════
# COVER TAB  (unchanged from v3)
# ═══════════════════════════════════════════════════════════════════════════════
def build_cover(wb, ticker, years, is_data):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22
    nc = 2

    _ccy = is_data[0].get("reportedCurrency", "USD") if is_data else "USD"
    _ccy_note = f"All figures {_ccy} millions" if _ccy == "USD" else f"All figures {_ccy} millions  |  ⚠ Non-USD currency — DCF implied prices converted to USD"
    row = write_tab_title(ws, 1, f"{ticker.upper()} — Financial Model", nc,
        subtitle=f"Source: Financial Modeling Prep API  |  {_ccy_note}  |  FY {years[0]}–{years[-1]}")

    row += 1
    d = is_data[-1]

    # Write reporting currency so backfill can read it back without an API call
    _rc = ws.cell(row=row, column=1, value="Reporting Currency")
    _rc.font = fnt(size=10); _rc.border = brd()
    _rc.alignment = Alignment(horizontal="left", indent=1)
    _rv = ws.cell(row=row, column=2, value=_ccy)
    _rv.font = fnt(color=C_BLUE, size=10)
    _rv.alignment = Alignment(horizontal="right"); _rv.border = brd()
    row += 1

    row = write_section_hdr(ws, row, "KEY METRICS — MOST RECENT YEAR", nc, C_SUMMARY_HD)

    def cov_row(ws, r, label, val, fmt):
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = fnt(size=10); c1.border = brd()
        c1.alignment = Alignment(horizontal="left", indent=1)
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = fnt(color=C_BLUE, size=10)
        c2.alignment = Alignment(horizontal="right")
        c2.border = brd()
        c2.number_format = fmt
        return r + 1

    row = cov_row(ws, row, f"Fiscal Year",          years[-1],                  "@")
    row = cov_row(ws, row, "Revenue ($mm)",          gm(d,"revenue"),           "#,##0.0")
    row = cov_row(ws, row, "Gross Profit ($mm)",     gm(d,"grossProfit"),       "#,##0.0")
    row = cov_row(ws, row, "EBITDA ($mm)",           gm(d,"ebitda"),            "#,##0.0")
    row = cov_row(ws, row, "EBIT ($mm)",             gm(d,"operatingIncome"),   "#,##0.0")
    row = cov_row(ws, row, "Net Income ($mm)",       gm(d,"netIncome"),         "#,##0.0")
    row = cov_row(ws, row, "Free Cash Flow ($mm)",   gm(d,"freeCashFlow") if d.get("freeCashFlow") else None, "#,##0.0")
    row = cov_row(ws, row, "EPS Diluted",            g(d,"epsdiluted"),         "$#,##0.00")
    row = cov_row(ws, row, "Gross Margin %",         g(d,"grossProfitRatio"),   "0.0%")
    row = cov_row(ws, row, "EBITDA Margin %",        g(d,"ebitdaratio"),        "0.0%")
    row = cov_row(ws, row, "Net Margin %",           g(d,"netIncomeRatio"),     "0.0%")

    row += 1
    row = write_section_hdr(ws, row, "WORKBOOK STRUCTURE", nc, C_DETAIL_HD)
    tabs = [
        ("Cover",         "This page — key metrics snapshot"),
        ("P&L",           "Income statement: summary + all FMP line items"),
        ("Balance Sheet", "Balance sheet: summary + all FMP line items"),
        ("Cash Flow",     "Cash flow: summary + all FMP line items"),
        ("Ratios & FCF",  "UFCF bridge, LFCF bridge, and full ratio suite"),
        ("Segments",      "Product & geographic revenue segments (plan-dependent)"),
    ]
    for tab, desc in tabs:
        c1 = ws.cell(row=row, column=1, value=tab)
        c1.font = fnt(bold=True, size=10); c1.border = brd()
        c1.alignment = Alignment(horizontal="left", indent=1)
        c2 = ws.cell(row=row, column=2, value=desc)
        c2.font = fnt(size=10); c2.border = brd()
        c2.alignment = Alignment(horizontal="left", indent=1)
        row += 1

    row += 1
    note = ws.cell(row=row, column=1,
        value="Colour convention: Blue = raw API data input | Black = formula/calculated | Green = cross-sheet link")
    note.font = fnt(size=9, italic=True, color="666666")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

# ═══════════════════════════════════════════════════════════════════════════════
# WACC TAB
# ═══════════════════════════════════════════════════════════════════════════════
def build_wacc(wb, ticker, is_data, bs_data, manual_rating=None, profile=None):
    """Build WACC & Cost of Capital sheet.

    profile (optional): pre-fetched FMP profile dict for this ticker.  When
    provided, skips the redundant internal /stable/profile call (the same one
    the caller already made).  Saves 1 FMP call per report.
    """

    NC = 3   # columns: label | value | source/note

    ws = wb.create_sheet("WACC")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 42
    ws.freeze_panes = "A3"

    # ── Sheet-local helpers ───────────────────────────────────────────────────
    def wrow(r, label, val, note="", bold=False, bg=C_WHITE,
             val_color=C_BLACK, is_pct=False):
        ca = ws.cell(row=r, column=1, value=label)
        ca.font = fnt(bold=bold, size=10)
        ca.fill = fll(bg); ca.border = brd()
        ca.alignment = Alignment(horizontal="left", indent=1)
        cb = ws.cell(row=r, column=2, value=val)
        cb.font = fnt(bold=bold, color=val_color, size=10)
        cb.fill = fll(bg); cb.border = brd()
        cb.alignment = Alignment(horizontal="right")
        if is_pct:
            cb.number_format = '0.00%;(0.00%);"-"'
        elif isinstance(val, (int, float)):
            cb.number_format = '#,##0.00;(#,##0.00);"-"'
        elif isinstance(val, str) and val.startswith("="):
            cb.number_format = ('#,##0.00;(#,##0.00);"-"'
                                if not is_pct else '0.00%;(0.00%);"-"')
        cc = ws.cell(row=r, column=3, value=note)
        cc.font = fnt(size=9, italic=True, color="777777")
        cc.fill = fll(bg); cc.border = brd()
        cc.alignment = Alignment(horizontal="left", indent=1)
        return r + 1

    def prow(r, label, val, note="", is_pct=False):
        """Percentage formula row — thin wrapper around wrow."""
        return wrow(r, label, val, note, is_pct=True, val_color=C_BLACK)

    def input_row(r, label, val, note="", is_pct=False):
        """Blue override cell pre-filled with suggested value."""
        ca = ws.cell(row=r, column=1, value=label)
        ca.font = fnt(bold=True, size=10)
        ca.fill = fll(C_OVR_BG); ca.border = brd()
        ca.alignment = Alignment(horizontal="left", indent=1)
        cb = ws.cell(row=r, column=2, value=val)
        cb.font = fnt(bold=True, color=C_BLUE, size=10)
        cb.fill = fll(C_OVR_BG); cb.border = brd()
        cb.alignment = Alignment(horizontal="right")
        cb.number_format = ('0.00%;(0.00%);"-"' if is_pct
                            else '0.000;(0.000);"-"')
        cc = ws.cell(row=r, column=3, value=note)
        cc.font = fnt(size=9, italic=True, color="2E7D32")
        cc.fill = fll(C_OVR_BG); cc.border = brd()
        cc.alignment = Alignment(horizontal="left", indent=1)
        return r + 1

    def rat_row(r, text):
        """Merged rationale row (AI output)."""
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=NC)
        c = ws.cell(row=r, column=1, value=text)
        c.font = fnt(size=9, italic=True, color="5D4037")
        c.fill = fll(C_AI_RAT); c.border = brd()
        c.alignment = Alignment(horizontal="left", indent=2,
                                wrap_text=True)
        ws.row_dimensions[r].height = 48
        return r + 1

    def flag_row(r, text):
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=NC)
        c = ws.cell(row=r, column=1, value=text)
        c.font = fnt(size=9, bold=True, color="B71C1C")
        c.fill = fll(C_FLAG_BG); c.border = brd()
        c.alignment = Alignment(horizontal="left", indent=2)
        return r + 1

    def blank(r):
        for col in range(1, NC + 1):
            c = ws.cell(row=r, column=col)
            c.fill = fll(C_WHITE); c.border = brd()
        return r + 1

    def shdr(r, text, color=None):
        return write_section_hdr(ws, r, text, NC, color or C_SECTION)

    # ── Fetch data ────────────────────────────────────────────────────────────
    print("  Fetching WACC inputs...")

    # FMP profile — reuse caller's profile if provided (saves 1 FMP call)
    prof = profile or {}
    if not prof:
        try:
            p = requests.get(
                f"https://financialmodelingprep.com/stable/profile"
                f"?symbol={ticker}&apikey={API_KEY}", timeout=10
            ).json()
            prof = (p[0] if isinstance(p, list) and p
                    else p if isinstance(p, dict) else {})
        except Exception:
            pass

    raw_beta = float(prof.get("beta") or 0) or None
    mktcap   = float(prof.get("marketCap") or 0) or None
    sector   = prof.get("industry") or prof.get("sector") or ""
    price    = prof.get("price", "")
    print(f"    Beta={raw_beta}  MktCap={mktcap}  Sector={sector}")

    # Credit rating — manual input takes priority, then FMP, then synthetic fallback
    if manual_rating:
        fmp_rating    = manual_rating
        rating_source = "User input"
        print(f"    Rating={fmp_rating}  (manual)")
    else:
        fmp_rating = None
        rating_source = "FMP /ratings endpoint"
        try:
            rat = requests.get(
                f"https://financialmodelingprep.com/stable/ratings"
                f"?symbol={ticker}&apikey={API_KEY}", timeout=10
            ).json()
            if isinstance(rat, list) and rat:
                fmp_rating = rat[0].get("rating") or rat[0].get("ratingScore")
        except Exception:
            pass
        print(f"    Rating={fmp_rating}")

    # Balance sheet / income statement — most recent year
    is0 = is_data[-1]
    bs0 = bs_data[-1]
    bs1 = bs_data[-2] if len(bs_data) > 1 else bs_data[-1]

    debt0    = ((bs0.get("shortTermDebt") or 0) +
                (bs0.get("longTermDebt")  or 0))
    debt1    = ((bs1.get("shortTermDebt") or 0) +
                (bs1.get("longTermDebt")  or 0))
    avg_debt = (debt0 + debt1) / 2

    ebit     = abs(is0.get("operatingIncome")   or 0)
    int_exp  = abs(is0.get("interestExpense")    or 0)
    int_inc  = abs(is0.get("interestIncome")     or 0)
    tax_exp  = abs(is0.get("incomeTaxExpense")   or 0)
    pretax   = abs(is0.get("incomeBeforeTax")    or 0)
    eff_tax  = tax_exp / pretax if pretax else 0
    icr      = ebit / int_exp if int_exp else 999

    # Capital structure
    E   = (mktcap or 0) / 1e6
    D   = debt0 / 1e6
    V   = E + D
    w_e = E / V if V else 1.0
    w_d = D / V if V else 0.0

    # FRED rates
    rf,     rf_date = fetch_fred("DGS10")
    rd_aaa, _       = fetch_fred("BAMLC0A1CAAAEY")
    rd_aa,  _       = fetch_fred("BAMLC0A2CAAEY")
    rd_a,   _       = fetch_fred("BAMLC0A3CAEY")
    rd_bbb, _       = fetch_fred("BAMLC0A4CBBBEY")
    rd_hy,  _       = fetch_fred("BAMLH0A0HYM2EY")
    rf = rf or 0.043

    RATING_FRED = {
        "AAA":  (rd_aaa, "FRED BAMLC0A1CAAAEY — AAA"),
        "AA+":  (rd_aa,  "FRED BAMLC0A2CAAEY — AA"),
        "AA":   (rd_aa,  "FRED BAMLC0A2CAAEY — AA"),
        "AA-":  (rd_aa,  "FRED BAMLC0A2CAAEY — AA"),
        "A+":   (rd_a,   "FRED BAMLC0A3CAEY — A"),
        "A":    (rd_a,   "FRED BAMLC0A3CAEY — A"),
        "A-":   (rd_a,   "FRED BAMLC0A3CAEY — A"),
        "BBB+": (rd_bbb, "FRED BAMLC0A4CBBBEY — BBB"),
        "BBB":  (rd_bbb, "FRED BAMLC0A4CBBBEY — BBB"),
        "BBB-": (rd_bbb, "FRED BAMLC0A4CBBBEY — BBB"),
    }
    fred_rd, fred_rd_src = RATING_FRED.get(fmp_rating, (None, "No matched rating"))
    if not fred_rd and fmp_rating and fmp_rating[:2] in ("BB", "B-", "B+",
                                                          "B ", "CC"):
        fred_rd, fred_rd_src = rd_hy, "FRED BAMLH0A0HYM2EY — High Yield"

    # Synthetic rating
    synth_rating, synth_spread = get_synthetic_rating(icr)
    rd_synthetic = rf + synth_spread
    rd_acctg     = int_exp / avg_debt if avg_debt else None

    # Peer betas
    peer_list = []
    for key, peers in SECTOR_PEERS.items():
        if key.lower() in sector.lower() or sector.lower() in key.lower():
            peer_list = [p for p in peers if p != ticker]
            break
    peer_betas = []
    for p in peer_list[:5]:
        try:
            pp = requests.get(
                f"https://financialmodelingprep.com/stable/profile"
                f"?symbol={p}&apikey={API_KEY}", timeout=8
            ).json()
            pp = pp[0] if isinstance(pp, list) and pp else pp
            b  = float(pp.get("beta") or 0)
            if b:
                peer_betas.append((p, round(b, 3)))
        except Exception:
            pass
    peer_vals   = sorted([b for _, b in peer_betas])
    peer_median = peer_vals[len(peer_vals) // 2] if peer_vals else None
    print(f"    Peers: {peer_betas}")

    # Damodaran industry beta → re-lever
    dama_unlevered = DAMODARAN_BETAS["Default"]
    for key, val in DAMODARAN_BETAS.items():
        if (key.lower() in sector.lower() or
                sector.lower() in key.lower()):
            dama_unlevered = val
            break
    de_ratio      = D / E if E else 0
    dama_relevered = round(dama_unlevered * (1 + (1 - eff_tax) * de_ratio), 3)
    blume         = round(0.67 * raw_beta + 0.33, 3) if raw_beta else None

    # ── Average-based defaults (no AI) ───────────────────────────────────────
    net_int = int_inc - int_exp

    # Beta: average of all non-None data points
    beta_candidates = [v for v in [raw_beta, blume, dama_relevered, peer_median]
                       if v is not None]
    sel_beta = round(sum(beta_candidates) / len(beta_candidates), 3) \
               if beta_candidates else 1.0

    # ERP: average of Damodaran implied + historical
    sel_erp = round((DAMODARAN_ERP_IMPLIED + DAMODARAN_ERP_HIST_AVG) / 2, 4)

    # Rd: average of all non-None data points
    rd_candidates = [v for v in [fred_rd, rd_synthetic, rd_acctg]
                     if v is not None]
    sel_rd = round(sum(rd_candidates) / len(rd_candidates), 4) \
             if rd_candidates else 0.05

    # ── Write sheet ───────────────────────────────────────────────────────────
    row = 1
    row = write_tab_title(
        ws, row, f"{ticker.upper()} — WACC & COST OF CAPITAL", NC,
        subtitle=("Blue = user override  |  Green = selected input  |  "
                  "Default pre-filled with average of all sources  |  All figures USD millions"))
    row = blank(row)

    # ── Capital structure ─────────────────────────────────────────────────────
    row = shdr(row, "CAPITAL STRUCTURE")
    eq_row = row
    row = wrow(row, "Equity  (market capitalisation, $mm)", E or None,
               f"FMP marketCap  |  price ${price}", val_color=C_BLUE)
    dbt_row = row
    row = wrow(row, "Debt  (gross book value, $mm)", D or None,
               "FMP: shortTermDebt + longTermDebt, most recent yr-end",
               val_color=C_BLUE)
    tot_row = row
    row = wrow(row, "Total capital  (V = E + D)", f"=B{eq_row}+B{dbt_row}",
               "", bold=True, bg=C_SUMMARY_BG)
    ws.cell(row=tot_row, column=2).number_format = '#,##0.0;(#,##0.0);"-"'
    ew_row = row
    row = prow(row, "Equity weight  (E / V)",
               f"=IFERROR(B{eq_row}/B{tot_row},1)",
               "Weight applied to Re in WACC")
    dw_row = row
    row = prow(row, "Debt weight  (D / V)",
               f"=IFERROR(B{dbt_row}/B{tot_row},0)",
               "Weight applied to after-tax Rd in WACC")
    row = blank(row)

    # ── Risk-free rate ────────────────────────────────────────────────────────
    row = shdr(row, "STEP 1 — RISK-FREE RATE")
    row = wrow(row, "10-yr US Treasury yield  (FRED DGS10)",
               rf, f"FRED series DGS10  |  latest: {rf_date}",
               val_color=C_BLACK, is_pct=True)
    rf_row = row
    row = input_row(row, "► Selected Rf  (override if needed)",
                    round(rf, 4),
                    "Pre-filled with live FRED DGS10 rate", is_pct=True)
    row = blank(row)

    # ── Beta ──────────────────────────────────────────────────────────────────
    row = shdr(row, "STEP 2 — BETA")
    row = wrow(row, "Raw beta  (FMP — 5yr monthly vs S&P 500)",
               raw_beta, "FMP company profile endpoint")
    row = wrow(row, "Blume-adjusted  (0.67 × raw + 0.33)",
               blume, "Mean-reversion toward market beta of 1.0")
    row = wrow(row, f"Damodaran industry unlevered  ({sector or 'sector'})",
               dama_unlevered,
               "Damodaran.com — betas by industry, Jan 2025 US")
    row = wrow(row, f"Damodaran re-levered  "
               f"(D/E = {de_ratio:.2f}x,  t = {eff_tax*100:.1f}%)",
               dama_relevered,
               "= unlevered × (1 + (1 − t) × D/E)")
    if peer_betas:
        peers_str = "  |  ".join(f"{p}: {b}" for p, b in peer_betas)
        row = wrow(row, "Peer median beta",
                   peer_median, peers_str[:52])
    if (raw_beta and dama_unlevered and
            raw_beta > dama_unlevered * 1.8):
        row = flag_row(row,
            f"FLAG: Raw beta ({raw_beta:.2f}) is "
            f"{raw_beta/dama_unlevered:.1f}x Damodaran industry avg "
            f"({dama_unlevered:.2f}) — historical window may include "
            f"structural break or regime change.")
    beta_row = row
    row = input_row(row, "► Selected beta  (user override)",
                    round(sel_beta, 3),
                    f"Default = average of {len(beta_candidates)} source(s) above — override freely")
    row = blank(row)

    # ── Equity risk premium ───────────────────────────────────────────────────
    row = shdr(row, "STEP 3 — EQUITY RISK PREMIUM  (ERP)")
    row = wrow(row, "Damodaran implied ERP — US market (current)",
               DAMODARAN_ERP_IMPLIED,
               "Damodaran.com implied ERP — Jan 2026", is_pct=True)
    row = wrow(row, "Damodaran historical avg — US 1928–2025",
               DAMODARAN_ERP_HIST_AVG,
               "Arithmetic average excess return over T-bill", is_pct=True)
    erp_row = row
    row = input_row(row, "► Selected ERP  (user override)",
                    round(sel_erp, 4),
                    "Default = average of Damodaran implied & historical avg — override freely",
                    is_pct=True)
    row = blank(row)

    # ── Cost of equity ────────────────────────────────────────────────────────
    row = shdr(row, "STEP 4 — COST OF EQUITY  (CAPM:  Re = Rf + β × ERP)")
    re_row = row
    re_cell = ws.cell(row=re_row, column=2,
                      value=f"=B{rf_row}+B{beta_row}*B{erp_row}")
    re_label = ws.cell(row=re_row, column=1,
                       value="Cost of equity  (Re)")
    re_note  = ws.cell(row=re_row, column=3,
                       value="CAPM formula — references selected inputs above")
    for c_ in (re_label, re_cell, re_note):
        c_.fill = fll(C_SUMMARY_BG); c_.border = brd()
    re_label.font  = fnt(bold=True, size=10)
    re_label.alignment = Alignment(horizontal="left", indent=1)
    re_cell.font   = fnt(bold=True, color=C_BLACK, size=10)
    re_cell.alignment = Alignment(horizontal="right")
    re_cell.number_format = '0.00%;(0.00%);"-"'
    re_note.font   = fnt(size=9, italic=True, color="777777")
    re_note.alignment = Alignment(horizontal="left", indent=1)
    row += 1
    row = blank(row)

    # ── Cost of debt ──────────────────────────────────────────────────────────
    row = shdr(row, "STEP 5 — COST OF DEBT  (Pre-tax Rd)")
    row = wrow(row, "Credit rating  (S&P / Moody's)",
               fmp_rating or "Not available",
               rating_source)
    if fred_rd:
        row = wrow(row, "FRED matched yield  (rating-tier index)",
                   fred_rd, fred_rd_src, is_pct=True)
    else:
        row = wrow(row, "FRED matched yield",
                   None, "No rating returned — FRED method not applicable")
    row = wrow(row,
               f"Interest Coverage Ratio  (EBIT / Gross Int Exp)",
               round(icr, 1) if icr < 900 else None,
               f"EBIT ${ebit/1e6:.0f}mm  /  Int Exp ${int_exp/1e6:.0f}mm")
    row = wrow(row, "Synthetic rating  (Damodaran ICR table)",
               synth_rating,
               f"ICR {icr:.1f}x  →  {synth_rating}")
    row = wrow(row, "Synthetic Rd  (Rf + Damodaran default spread)",
               rd_synthetic,
               f"Rf {rf*100:.2f}% + spread {synth_spread*100:.2f}%",
               is_pct=True)
    row = wrow(row, "Accounting Rd  (Gross int exp / Avg gross debt)",
               rd_acctg, "Cross-check only — backward-looking", is_pct=True)
    if net_int > 0:
        row = wrow(row,
                   f"  Note: net interest INCOME of ${net_int/1e6:.0f}mm  "
                   f"(int inc ${int_inc/1e6:.0f}mm > int exp ${int_exp/1e6:.0f}mm)",
                   None,
                   "Gross Rd is the correct basis here — net figure is distorted")
    if D < 1:
        row = flag_row(row,
                       "FLAG: Zero / negligible debt detected — "
                       "Rd is immaterial.  WACC ≈ Re.")
    rd_row = row
    row = input_row(row, "► Selected pre-tax Rd  (user override)",
                    round(sel_rd, 4),
                    f"Default = average of {len(rd_candidates)} source(s) above — override freely",
                    is_pct=True)
    row = blank(row)

    # ── Tax rate ──────────────────────────────────────────────────────────────
    row = shdr(row, "STEP 6 — TAX RATE")
    row = wrow(row, "Effective tax rate  (tax expense / pre-tax income)",
               round(eff_tax, 4),
               f"FMP: ${tax_exp/1e6:.0f}mm  /  ${pretax/1e6:.0f}mm",
               is_pct=True)
    tax_row = row
    row = input_row(row, "► Selected tax rate  (user override)",
                    round(eff_tax, 4),
                    "Adjust to normalised / marginal rate if preferred",
                    is_pct=True)
    row = blank(row)

    # ── WACC output ───────────────────────────────────────────────────────────
    row = shdr(row, "WACC OUTPUT", C_SUMMARY_HD)
    wacc_row = row
    # D-001: floor WACC at 8.5% (matches Python engine). Override in /dcf calculator.
    wacc_formula = (f"=MAX(0.085, B{ew_row}*B{re_row}"
                    f"+B{dw_row}*B{rd_row}*(1-B{tax_row}))")
    wl = ws.cell(row=wacc_row, column=1,
                 value="WACC  =  MAX(8.5%, (E/V × Re) + (D/V × Rd × (1 − t)))")
    wv = ws.cell(row=wacc_row, column=2, value=wacc_formula)
    wn = ws.cell(row=wacc_row, column=3,
                 value="Floored 8.5% (D-001); raw formula below")
    for c_ in (wl, wv, wn):
        c_.fill = fll(C_SUMMARY_BG); c_.border = brd()
    wl.font = fnt(bold=True, size=11)
    wl.alignment = Alignment(horizontal="left", indent=1)
    wv.font = fnt(bold=True, color=C_BLACK, size=11)
    wv.alignment = Alignment(horizontal="right")
    wv.number_format = '0.00%;(0.00%);"-"'
    wn.font = fnt(size=10, italic=True, color="555555")
    wn.alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[wacc_row].height = 22
    row += 1
    row = blank(row)

    # ── Sensitivity table ─────────────────────────────────────────────────────
    row = shdr(row, "SENSITIVITY — WACC (%)  by Beta offset × ERP offset",
               C_DETAIL_HD)
    beta_deltas = [-1.0, -0.5, 0.0, +0.5, +1.0]
    erp_deltas  = [-0.01, -0.005, 0.0, +0.005, +0.01]
    erp_labels  = ["-1.0%", "-0.5%", "Base ERP", "+0.5%", "+1.0%"]
    beta_labels = ["-1.0",  "-0.5",  "Base β",   "+0.5",  "+1.0"]

    # Header row
    hdr_cell = ws.cell(row=row, column=1,
                       value="Beta offset  \\  ERP offset →")
    hdr_cell.font = fnt(bold=True, size=9)
    hdr_cell.fill = fll(C_SUBTOTAL); hdr_cell.border = brd()
    hdr_cell.alignment = Alignment(horizontal="center")
    for ci, lbl in enumerate(erp_labels):
        c_ = ws.cell(row=row, column=ci + 2, value=lbl)
        c_.font = fnt(bold=True, size=9)
        c_.fill = fll(C_SUBTOTAL); c_.border = brd()
        c_.alignment = Alignment(horizontal="center")
    row += 1

    for bd, bl in zip(beta_deltas, beta_labels):
        is_base_row = (bd == 0.0)
        bg_ = C_SUMMARY_BG if is_base_row else C_WHITE
        lc = ws.cell(row=row, column=1, value=bl)
        lc.font = fnt(bold=is_base_row, size=9)
        lc.fill = fll(bg_); lc.border = brd()
        lc.alignment = Alignment(horizontal="center")
        for ci, ed in enumerate(erp_deltas):
            is_base_cell = is_base_row and ed == 0.0
            f = (f"=B{ew_row}*(B{rf_row}+(B{beta_row}+({bd}))"
                 f"*(B{erp_row}+({ed})))"
                 f"+B{dw_row}*B{rd_row}*(1-B{tax_row})")
            vc = ws.cell(row=row, column=ci + 2, value=f)
            vc.font = fnt(bold=is_base_cell, size=9)
            vc.fill = fll(C_SUMMARY_BG if is_base_cell else bg_)
            vc.border = brd()
            vc.alignment = Alignment(horizontal="right")
            vc.number_format = '0.00%;(0.00%);"-"'
        row += 1

    sel_re   = rf + sel_beta * sel_erp
    wacc_raw = w_e * sel_re + w_d * sel_rd * (1 - eff_tax)

    # D-001: Global WACC floor at 8.5%.
    # No equity investment should be discounted below the lowest reasonable equity return.
    # 8.5% ≈ Damodaran composite Ke for a market-beta US name (Rf 4.3% + 1.0 × ERP 4.5%
    # with rounding buffer). FMP 5yr regression betas systematically understate equity
    # risk for stable compounders (e.g. PEP β=0.41 → 6%; AAPL β=1.07 was producing 2.1%
    # due to an upstream Kd zero-out). Users can override interactively at /dcf.
    _WACC_FLOOR  = 0.085
    wacc_floored = wacc_raw < _WACC_FLOOR
    wacc_val     = max(wacc_raw, _WACC_FLOOR)

    return {
        "wacc_row":     wacc_row, "re_row":   re_row,
        "rf_row":       rf_row,   "beta_row": beta_row,
        "erp_row":      erp_row,  "rd_row":   rd_row,
        "tax_row":      tax_row,
        "wacc_val":     round(wacc_val, 4),
        "wacc_raw":     round(wacc_raw, 4),
        "wacc_floored": wacc_floored,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# DCF TAB
# ═══════════════════════════════════════════════════════════════════════════════
def build_dcf(wb, ticker, is_data, bs_data, cf_data, years, pl_refs, bs_refs, wacc_refs, current_price=None, cf_refs=None, profile=None):
    """Build DCF sheet — consensus years auto-populated from FMP, remainder user input."""

    last_hist_year = years[-1]
    estimates      = fetch_analyst_estimates(ticker, last_hist_year)

    # Projection year list: all FMP forward years, extended to at least YEARS_PROJ
    est_years = [e["year"] for e in estimates]
    last_yr   = int(last_hist_year)
    n_proj    = max(YEARS_PROJ, len(est_years))
    proj_years = []
    for i in range(1, n_proj + 1):
        proj_years.append(str(last_yr + i))

    # Lookup dict: year → estimate record
    est_map = {e["year"]: e for e in estimates}
    n_hist  = len(years)
    n_term  = 1

    # ── Growth tier: based on revenue growth rate ────────────────────────────
    # Drives base-case TGR, exit multiple, and bear/bull ranges throughout.
    # F-E: Cyclical companies use 5-year MEDIAN YoY growth instead of 3-year
    # average. A 3-year window at cycle peak overstates the structural trend;
    # at trough it understates it. Median of all available years is stable.
    _gt_revs = [d.get("revenue") or 0 for d in is_data]
    _gt_yoys = []
    for _k in range(max(1, len(_gt_revs) - 3), len(_gt_revs)):
        if _gt_revs[_k-1] > 0 and _gt_revs[_k] > 0:
            _gt_yoys.append(_gt_revs[_k] / _gt_revs[_k-1] - 1)
    _rev_3yr_avg_dcf = sum(_gt_yoys) / len(_gt_yoys) if _gt_yoys else 0.05

    # F-E / F-C: Sector bucket detection for both cyclical smoothing and quality premium
    _sector_str_dcf = ((profile or {}).get("industry") or
                       (profile or {}).get("sector") or "")
    _sector_bucket_dcf     = _sector_bucket(_sector_str_dcf, ticker)
    _is_cyclical_dcf       = _sector_bucket_dcf == "cyclical"
    _is_stable_compounder_dcf = _sector_bucket_dcf == "stable_compounder"
    if _is_cyclical_dcf:
        _all_yoys = []
        for _k in range(1, len(_gt_revs)):
            if _gt_revs[_k-1] > 0 and _gt_revs[_k] > 0:
                _all_yoys.append(_gt_revs[_k] / _gt_revs[_k-1] - 1)
        if len(_all_yoys) >= 3:
            _sorted = sorted(_all_yoys)
            _n = len(_sorted)
            _median = (_sorted[(_n - 1) // 2] + _sorted[_n // 2]) / 2.0
            _rev_3yr_avg_dcf = _median
            print(f"  F-E: cyclical tier smoothing — 5yr yoys={[round(y,3) for y in _all_yoys]}"
                  f"  median={_median:.3f}  (was 3yr avg={sum(_gt_yoys)/len(_gt_yoys):.3f})"
                  if _gt_yoys else f"  F-E: cyclical median={_median:.3f}")

    if _rev_3yr_avg_dcf < 0.05:
        _TIER          = "low"
        _DCF_TGR_BASE  = 0.025
        _DCF_TGR_BEAR  = round(0.025 * 0.80, 4)   # 2.0%
        _DCF_TGR_BULL  = round(0.025 * 1.20, 4)   # 3.0%
        _DCF_TEV_BASE  = 10.0
        _DCF_TEV_BEAR  = round(_DCF_TEV_BASE * 0.80)  # 8x
        _DCF_TEV_BULL  = round(_DCF_TEV_BASE * 1.20)  # 12x
    elif _rev_3yr_avg_dcf < 0.12:
        _TIER          = "medium"
        _DCF_TGR_BASE  = 0.030
        _DCF_TGR_BEAR  = round(0.030 * 0.75, 4)   # 2.25%
        _DCF_TGR_BULL  = round(0.030 * 1.25, 4)   # 3.75%
        _DCF_TEV_BASE  = 15.0
        _DCF_TEV_BEAR  = round(_DCF_TEV_BASE * 0.75)  # 11x
        _DCF_TEV_BULL  = round(_DCF_TEV_BASE * 1.25)  # 19x
    else:
        _TIER          = "high"
        _DCF_TGR_BASE  = 0.040
        _DCF_TGR_BEAR  = round(0.040 * 0.75, 4)   # 3.0%
        _DCF_TGR_BULL  = round(0.040 * 1.25, 4)   # 5.0%
        _DCF_TEV_BASE  = 18.0
        _DCF_TEV_BEAR  = round(_DCF_TEV_BASE * 0.75)  # 14x
        _DCF_TEV_BULL  = round(_DCF_TEV_BASE * 1.25)  # 23x

    # ── F-C: ROIC quality premium on EM multiple ─────────────────────────────
    # Standard tier multiples (10x/15x/18x) reflect sector-average EV/EBITDA for
    # medium-quality names. Stable compounders with ROIC > 25% sustain structural
    # advantages (membership moats, network effects, brand pricing power) that
    # command a premium vs commodity peers — historically 20-30x vs 10-15x average.
    # +5x is calibrated as the minimum quality increment between sector-average and
    # quality-tier multiples observed across COST/V/NKE class names.
    # Only applies to stable_compounder bucket (not banks, not EVS, not cyclicals).
    _trailing_ebit_fc_mm    = (is_data[-1].get("ebit") or is_data[-1].get("operatingIncome") or 0) / 1e6
    _trailing_tax_fc        = min(0.50, max(0.0,
        abs(is_data[-1].get("incomeTaxExpense") or 0) /
        max(abs(is_data[-1].get("incomeBeforeTax") or 1), 1)
    ))
    _trailing_nopat_fc_mm   = _trailing_ebit_fc_mm * (1.0 - _trailing_tax_fc)
    _trailing_equity_fc_mm  = (bs_data[-1].get("totalStockholdersEquity") or 0) / 1e6
    _trailing_td_fc_mm      = (bs_data[-1].get("totalDebt") or 0) / 1e6
    _trailing_cash_fc_mm    = (bs_data[-1].get("cashAndShortTermInvestments") or 0) / 1e6
    _trailing_ic_fc_mm      = _trailing_equity_fc_mm + (_trailing_td_fc_mm - _trailing_cash_fc_mm)
    _trailing_roic_fc       = (_trailing_nopat_fc_mm / max(_trailing_ic_fc_mm, 100.0)
                               if _trailing_ic_fc_mm > 100.0 else 0.0)

    # FCF margin — stored in dcf_prices for thin-margin EM-primary override in report_bridge
    _fcf_trailing_fc_mm  = (cf_data[-1].get("freeCashFlow") or 0) / 1e6
    _rev_trailing_fc_mm  = (is_data[-1].get("revenue") or 1) / 1e6
    _fcf_margin_trailing = round(_fcf_trailing_fc_mm / max(_rev_trailing_fc_mm, 1.0), 4)

    # F-D: Bank detection — needed here for quality-premium and EM-anchoring guards
    # (full detection block is repeated later in the DCF write-out section)
    _BANK_DCF_EXCLUDE_EARLY = {"V", "MA", "PYPL", "FIS", "FISV", "GPN", "WU", "DFS", "TRMK"}
    _BANK_DCF_KW_EARLY = {"bank", "banking", "financial services", "savings",
                           "thrift", "mortgage", "credit union", "investment bank",
                           "diversified financial"}
    _prof_industry_early = (profile or {}).get("industry") or (profile or {}).get("sector") or ""
    _is_bank_dcf = (
        any(kw in _prof_industry_early.lower() for kw in _BANK_DCF_KW_EARLY)
        and ticker.upper() not in _BANK_DCF_EXCLUDE_EARLY
    )

    _quality_em_premium  = False
    _FC_ROIC_THRESHOLD   = 0.25   # 25% ROIC = structural moat / quality compounder
    _FC_PREMIUM_X        = 5.0    # +5x added to all EM scenario multiples
    if (_is_stable_compounder_dcf and not _is_bank_dcf
            and _trailing_roic_fc > _FC_ROIC_THRESHOLD):
        _DCF_TEV_BASE  += _FC_PREMIUM_X
        _DCF_TEV_BEAR  += round(_FC_PREMIUM_X * 0.8)   # +4x bear
        _DCF_TEV_BULL  += round(_FC_PREMIUM_X * 1.0)   # +5x bull
        _quality_em_premium = True
        print(f"  F-C quality premium: ROIC {_trailing_roic_fc*100:.1f}% > 25% "
              f"→ EM {_DCF_TEV_BASE - _FC_PREMIUM_X:.0f}x → {_DCF_TEV_BASE:.0f}x base  "
              f"(bear {_DCF_TEV_BEAR:.0f}x / bull {_DCF_TEV_BULL:.0f}x)")

    # ── F-C Phase 2: Historical EV/EBITDA anchoring for exit multiple ─────────
    # Stock-specific anchor: 5yr avg EV/EBITDA × 80% mean-reversion discount.
    # Floor = current tier+quality base (preserves quality premium gains).
    # Cap = 28x (prevents bubble-era averages perpetuating into terminal value).
    # Cyclicals excluded — tier smoothing via F-E is a better anchor for those.
    # Banks excluded — EM disabled anyway.
    # Note: _evs_regime is not yet computed here; anchoring is neutralised later
    #       in dcf_prices if EVS fires (EM prices become None regardless of mult).
    _HIST_EM_DISCOUNT      = 0.80
    _HIST_EM_CAP           = 28.0
    _em_anchored           = False
    _em_hist_anchor_raw    = None
    _em_hist_anchor_capped = False

    if not _is_cyclical_dcf and not _is_bank_dcf:
        _rat_ev_vals = []
        try:
            _rat_resp_h = _fetch_ratios(ticker, limit=5)   # cached — no extra FMP call
            if _rat_resp_h:
                _rat_ev_vals = [
                    r["enterpriseValueMultiple"] for r in _rat_resp_h
                    if r.get("enterpriseValueMultiple")
                    and 2 < r["enterpriseValueMultiple"] < 200
                ]
        except Exception as _e_h:
            print(f"  F-C hist anchor: ratios parse failed — {_e_h}")

        if len(_rat_ev_vals) >= 2:
            _hist_ev_avg        = round(sum(_rat_ev_vals) / len(_rat_ev_vals), 1)
            _em_hist_anchor_raw = _hist_ev_avg
            _anchored_raw       = _hist_ev_avg * _HIST_EM_DISCOUNT
            _hist_floor         = _DCF_TEV_BASE   # post-quality-premium floor
            _em_hist_anchor_capped = _anchored_raw > _HIST_EM_CAP
            _anchored_final     = max(_hist_floor, min(_HIST_EM_CAP, _anchored_raw))

            if abs(_anchored_final - _DCF_TEV_BASE) > 0.5:
                _h_scale      = _anchored_final / _DCF_TEV_BASE
                _DCF_TEV_BEAR = round(_DCF_TEV_BEAR * _h_scale, 1)
                _DCF_TEV_BULL = min(round(_DCF_TEV_BULL * _h_scale, 1), _HIST_EM_CAP + 4.0)
                _DCF_TEV_BASE = round(_anchored_final, 1)
                _em_anchored  = True
                _h_reason = ("capped" if _em_hist_anchor_capped
                             else "floored" if _anchored_raw < _hist_floor
                             else "anchored")
                print(f"  F-C hist anchor: {len(_rat_ev_vals)}-yr avg {_hist_ev_avg:.1f}x "
                      f"x {_HIST_EM_DISCOUNT} = {_anchored_raw:.1f}x -> {_h_reason} "
                      f"{_DCF_TEV_BASE:.1f}x base "
                      f"(bear {_DCF_TEV_BEAR:.1f}x / bull {_DCF_TEV_BULL:.1f}x)")
            else:
                print(f"  F-C hist anchor: {len(_rat_ev_vals)}-yr avg {_hist_ev_avg:.1f}x "
                      f"-> anchored {_anchored_raw:.1f}x within 0.5x of base -- no change")

    # Column layout: A=labels | hist cols | proj cols | terminal | notes
    NC          = 1 + n_hist + n_proj + n_term + 1
    HIST_COLS   = list(range(2, 2 + n_hist))
    PROJ_COLS   = list(range(2 + n_hist, 2 + n_hist + n_proj))
    TERM_COL    = 2 + n_hist + n_proj
    NOTE_COL    = TERM_COL + 1

    ws = wb.create_sheet("DCF")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 42
    for c in range(2, TERM_COL + 1):
        ws.column_dimensions[cl(c)].width = 13
    ws.column_dimensions[cl(NOTE_COL)].width = 40
    ws.freeze_panes = f"B5"

    # ── Local helpers ─────────────────────────────────────────────────────────
    def wcell(r, c, val, bold=False, bg=C_WHITE, color=C_BLACK,
              italic=False, halign="right", fmt=None, indent=0):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font      = fnt(bold=bold, color=color, size=10, italic=italic)
        cell.fill      = fll(bg)
        cell.border    = brd()
        cell.alignment = Alignment(horizontal=halign, vertical="center",
                                   indent=indent, wrap_text=False)
        if fmt:
            cell.number_format = fmt
        return cell

    def note(r, text, bg=C_WHITE):
        c = ws.cell(row=r, column=NOTE_COL, value=text)
        c.font      = fnt(size=9, italic=True, color="555555")
        c.fill      = fll(bg)
        c.border    = brd()
        c.alignment = Alignment(horizontal="left", indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 20
        return c

    def blank(r):
        for c in range(1, NC + 1):
            ws.cell(row=r, column=c).fill   = fll(C_WHITE)
            ws.cell(row=r, column=c).border = brd()
        ws.row_dimensions[r].height = 6
        return r + 1

    def shdr(r, text, bg=None):
        return write_section_hdr(ws, r, text, NC, bg or C_SECT)

    NUM  = '#,##0.0;(#,##0.0);"-"'
    PCT  = '0.0%;(0.0%);"-"'
    PCT2 = '0.00%;(0.00%);"-"'
    DOLS = '$#,##0.00'

    # ── Title & year header ───────────────────────────────────────────────────
    row = 1
    row = write_tab_title(ws, row,
        f"{ticker.upper()} — DCF VALUATION", NC,
        subtitle=("Grey = historical actual  |  Blue (darker) = FMP analyst consensus  |  "
                  "Blue (lighter) = user input  |  Amber = key assumption"))

    wcell(row, 1, "Fiscal Year", bold=True, bg=C_SUB, halign="left", indent=1)
    for i, yr in enumerate(years):
        wcell(row, HIST_COLS[i], yr, bold=True, bg=C_SUB)
    for i, yr in enumerate(proj_years):
        has_est = yr in est_map
        bg_ = C_BG if has_est else C_ALT
        wcell(row, PROJ_COLS[i], f"{yr}E", bold=True, bg=bg_)
    wcell(row, TERM_COL, "Terminal", bold=True, bg=C_ASSM)
    wcell(row, NOTE_COL, "Source / Notes", bold=True, bg=C_SUB,
          halign="left", indent=1)
    ws.row_dimensions[row].height = 16
    row += 1

    # ── Consensus boundary flag ───────────────────────────────────────────────
    if estimates:
        last_est_yr = est_years[-1]
        n_consensus = len([y for y in proj_years if y in est_map])
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=NC)
        flag = ws.cell(row=row, column=1,
            value=(f"FMP analyst consensus: {est_years[0]}E – {est_years[-1]}E  "
                   f"({n_consensus} year{'s' if n_consensus != 1 else ''})   |   "
                   f"User estimates required from: "
                   f"{next((y for y in proj_years if y not in est_map), 'N/A')}E onwards"))
        flag.font      = fnt(size=9, bold=True, color="1A5276")
        flag.fill      = fll(C_BG)
        flag.border    = brd()
        flag.alignment = Alignment(horizontal="left", indent=2)
        ws.row_dimensions[row].height = 16
        row += 1
    row = blank(row)

    # ── SECTION 1: PROJECTION ASSUMPTIONS ────────────────────────────────────
    row = shdr(row,
        "SECTION 1 — PROJECTION ASSUMPTIONS  "
        "(darker blue = FMP consensus-derived  |  lighter blue = user input)", C_HD)

    # Helper: write one assumption row
    def assm_row(r, label, hist_vals, proj_fn, term_val, is_pct=True,
                 note_text="", term_color=C_BLUE):
        wcell(r, 1, label, bold=True, bg=C_ASSM, halign="left", indent=1)
        fmt = PCT if is_pct else NUM
        # Historical (greyed out — actuals, not assumptions)
        for i, c in enumerate(HIST_COLS):
            v = hist_vals[i] if hist_vals and i < len(hist_vals) else None
            cell = wcell(r, c, v, bg=C_HIST, color="999999", fmt=fmt)
            if v is None:
                cell.value = "—"; cell.number_format = "@"
        # Projection
        for i, c in enumerate(PROJ_COLS):
            yr = proj_years[i]
            val, color_, bg_ = proj_fn(i, yr, c)
            wcell(r, c, val, bg=bg_, color=color_, fmt=fmt)
        # Terminal
        wcell(r, TERM_COL, term_val, bg=C_ASSM, color=term_color, fmt=fmt)
        note(r, note_text, bg=C_ASSM)
        return r + 1

    # Extract historical actuals for back-reference
    hist_rev    = [(g(d, "revenue")        or 0) / 1e6 for d in is_data]
    hist_ebitda = [(g(d, "ebitda")         or 0) / 1e6 for d in is_data]
    hist_da     = [(g(d, "depreciationAndAmortization") or 0) / 1e6 for d in is_data]
    hist_capex  = [(abs(g(d, "capitalExpenditure") or 0)) / 1e6 for d in cf_data]
    hist_tax    = [(abs(g(d, "incomeTaxExpense") or 0) /
                    max(abs(g(d, "incomeBeforeTax") or 1), 1))
                   for d in is_data]

    # Prior-year revenue for growth calc (last historical as base)
    def prior_rev(i):
        """Return $mm revenue for the year before projection index i."""
        if i == 0:
            return hist_rev[-1] if hist_rev[-1] else 1
        # Use the consensus/projected revenue for the prior projection year
        prev_yr = proj_years[i - 1]
        if prev_yr in est_map:
            return est_map[prev_yr]["rev_avg"]
        return None   # can't compute; formula will handle

    # Shared P&L row refs used throughout assumption rows
    _pl_rev    = pl_refs["rev"]
    _pl_ebitda = pl_refs["ebitda"]
    _pl_da     = pl_refs["da"]
    _pl_tax    = pl_refs["tax"]
    _pl_ebt    = pl_refs["ebt"]

    # Revenue growth % — Yrs 1-3: analyst consensus; Yrs 4-5: = Yr 3 formula
    _rg_row = row  # capture row before assm_row writes it (closure will ref this)
    def rev_growth_fn(i, yr, c):
        if i < 3:  # Years 1-3: analyst consensus where available
            if yr in est_map:
                e    = est_map[yr]
                prev = prior_rev(i)
                val  = round(e["rev_avg"] / prev - 1, 4) if prev else None
                return val, "1A5276", C_BG
            return 0.08, C_BLUE, C_ALT
        # Years 4-5: formula = Year 3 growth rate (PROJ_COLS[2])
        return f"={cl(PROJ_COLS[2])}{_rg_row}", C_BLUE, C_ALT

    rev_growth_defaults = [rev_growth_fn(i, yr, c)[0]
                           for i, (yr, c) in enumerate(zip(proj_years, PROJ_COLS))]

    # Historical Revenue Growth: formula from P&L (first year = None, no prior)
    hist_rev_growth = [None] + [
        f"='P&L'!{cl(HIST_COLS[i])}{_pl_rev}/'P&L'!{cl(HIST_COLS[i-1])}{_pl_rev}-1"
        for i in range(1, n_hist)
    ]
    row = assm_row(row, "Revenue Growth %",
        hist_rev_growth[:n_hist], rev_growth_fn, 0.03,
        note_text=("Yrs 1-3: FMP analyst consensus implied growth (back-calc from rev estimates).  "
                   "Yrs 4-5: formula = Yr 3 growth.  Terminal = long-run growth (2-4%)."),
        term_color=C_BLUE)
    rev_growth_row = row - 1

    # EBITDA Margin % — Yrs 1-3: analyst consensus; Yrs 4-5: = Yr 3 formula
    _em_row = row
    def ebitda_margin_fn(i, yr, c):
        if i < 3:
            if yr in est_map:
                e   = est_map[yr]
                rev = e["rev_avg"] or 1
                val = round(e["ebitda_avg"] / rev, 4) if rev else None
                return val, "1A5276", C_BG
            last_known = (est_map[est_years[-1]]["ebitda_avg"] /
                          max(est_map[est_years[-1]]["rev_avg"], 1)
                          if estimates else
                          (hist_ebitda[-1] / max(hist_rev[-1], 1) if hist_rev[-1] else 0.50))
            return round(last_known, 4), C_BLUE, C_ALT
        return f"={cl(PROJ_COLS[2])}{_em_row}", C_BLUE, C_ALT  # Yrs 4-5: = Yr 3

    # Historical EBITDA margin: formula from P&L
    hist_ebitda_margins_f = [
        f"='P&L'!{cl(HIST_COLS[i])}{_pl_ebitda}/'P&L'!{cl(HIST_COLS[i])}{_pl_rev}"
        for i in range(n_hist)
    ]
    hist_ebitda_margins = [  # numeric fallback for terminal value default
        round(hist_ebitda[i] / max(hist_rev[i], 1), 4) if hist_rev[i] else None
        for i in range(n_hist)
    ]
    row = assm_row(row, "EBITDA Margin %",
        hist_ebitda_margins_f, ebitda_margin_fn, None,
        note_text=("Yrs 1-3: FMP analyst consensus margin.  Yrs 4-5: = Yr 3 margin.  "
                   "Terminal: long-run normalised EBITDA margin (user input)."),
        term_color=C_BLUE)
    ebitda_margin_row = row - 1
    # Terminal EBITDA margin — manual blue input
    wcell(ebitda_margin_row, TERM_COL,
          hist_ebitda_margins[-1] if hist_ebitda_margins else 0.50,
          bg=C_ASSM, color=C_BLUE, fmt=PCT)

    # D&A % revenue — historical from P&L; Year 1 = avg historicals; Years 2+ = Year 1
    hist_da_pct = [round(hist_da[i] / max(hist_rev[i], 1), 4)
                   if hist_rev[i] else None for i in range(n_hist)]
    last_da_pct = next((v for v in reversed(hist_da_pct) if v), 0.02)
    hist_da_pct_f = [
        f"='P&L'!{cl(HIST_COLS[i])}{_pl_da}/'P&L'!{cl(HIST_COLS[i])}{_pl_rev}"
        for i in range(n_hist)
    ]
    _da_pct_row = row
    def da_pct_fn(i, yr, c):
        if i == 0:
            return (f"=AVERAGE({cl(HIST_COLS[0])}{_da_pct_row}:{cl(HIST_COLS[-1])}{_da_pct_row})",
                    C_BLUE, C_ALT)
        return f"={cl(PROJ_COLS[0])}{_da_pct_row}", C_BLUE, C_ALT

    row = assm_row(row, "D&A as % of Revenue",
        hist_da_pct_f, da_pct_fn, f"={cl(PROJ_COLS[0])}{_da_pct_row}",
        note_text="Historical linked from P&L.  Yr 1 = avg of historicals; Yrs 2-5 & terminal = Yr 1.")
    da_pct_row = row - 1

    # CapEx % revenue — historical from Cash Flow; Year 1 = avg; Years 2+ = Year 1
    hist_capex_pct = [round(hist_capex[i] / max(hist_rev[i], 1), 4)
                      if hist_rev[i] else None for i in range(n_hist)]
    last_capex_pct = next((v for v in reversed(hist_capex_pct) if v), 0.02)
    _cx_pct_row = row
    if cf_refs:
        _cf_capex = cf_refs["capex"]
        hist_capex_pct_f = [
            f"=-'Cash Flow'!{cl(HIST_COLS[i])}{_cf_capex}/'P&L'!{cl(HIST_COLS[i])}{_pl_rev}"
            for i in range(n_hist)
        ]
    else:
        hist_capex_pct_f = hist_capex_pct
    def capex_pct_fn(i, yr, c):
        if i == 0:
            return (f"=AVERAGE({cl(HIST_COLS[0])}{_cx_pct_row}:{cl(HIST_COLS[-1])}{_cx_pct_row})",
                    C_BLUE, C_ALT)
        return f"={cl(PROJ_COLS[0])}{_cx_pct_row}", C_BLUE, C_ALT

    row = assm_row(row, "CapEx as % of Revenue",
        hist_capex_pct_f, capex_pct_fn, f"={cl(PROJ_COLS[0])}{_cx_pct_row}",
        note_text="Historical linked from Cash Flow.  Yr 1 = avg of historicals; Yrs 2-5 & terminal = Yr 1.")
    capex_pct_row = row - 1

    # NWC change % revenue — historical from Balance Sheet (ΔNWC/Revenue); Year 1 = avg; Years 2+ = Year 1
    _bs_tca = bs_refs["tca"]
    _bs_tcl = bs_refs["tcl"]
    hist_nwc_pct_f = [None]  # first year: no prior period for delta
    for _i in range(1, n_hist):
        _c = HIST_COLS[_i]; _p = HIST_COLS[_i - 1]
        hist_nwc_pct_f.append(
            f"=(('Balance Sheet'!{cl(_c)}{_bs_tca}-'Balance Sheet'!{cl(_c)}{_bs_tcl})"
            f"-('Balance Sheet'!{cl(_p)}{_bs_tca}-'Balance Sheet'!{cl(_p)}{_bs_tcl}))"
            f"/'P&L'!{cl(_c)}{_pl_rev}"
        )
    _nwc_pct_row = row
    def nwc_pct_fn(i, yr, c):
        if i == 0:
            # Year 1: clamped historical average [-2%, +5%] — prevents outlier years (e.g. inventory
            # destocks) from permanently inflating or depressing UFCF.
            _avg_start = cl(HIST_COLS[1]) if n_hist > 1 else cl(HIST_COLS[0])
            return (f"=MAX(-0.02,MIN(0.05,AVERAGE({_avg_start}{_nwc_pct_row}:{cl(HIST_COLS[-1])}{_nwc_pct_row})))",
                    C_BLUE, C_ALT)
        # Years 2–n_proj: linear fade to 0 by terminal year.
        # Factor = (n_proj − i) / n_proj  →  Year 2 = Y1*(n−1)/n … Year n = Y1*1/n
        _factor_num = n_proj - i
        return (f"={cl(PROJ_COLS[0])}{_nwc_pct_row}*{_factor_num}/{n_proj}",
                C_BLUE, C_ALT)

    row = assm_row(row, "Change in NWC as % of Revenue",
        hist_nwc_pct_f, nwc_pct_fn, 0,
        note_text="Yr 1 = clamped avg of historicals [-2%, +5%]. Yrs 2-5 fade linearly to 0. Terminal = 0 (steady-state: NWC tracks revenue).")
    nwc_pct_row = row - 1

    # Tax rate — historical from P&L; Year 1 = avg; Years 2+ = Year 1
    last_tax = round(hist_tax[-1], 4) if hist_tax else 0.15
    hist_tax_f = [
        f"=IFERROR(MAX(0,MIN(0.5,'P&L'!{cl(HIST_COLS[i])}{_pl_tax}/'P&L'!{cl(HIST_COLS[i])}{_pl_ebt})),0)"
        for i in range(n_hist)
    ]
    _tax_row = row
    def tax_fn(i, yr, c):
        if i == 0:
            return (f"=AVERAGE({cl(HIST_COLS[0])}{_tax_row}:{cl(HIST_COLS[-1])}{_tax_row})",
                    C_BLUE, C_ALT)
        return f"={cl(PROJ_COLS[0])}{_tax_row}", C_BLUE, C_ALT

    row = assm_row(row, "Effective Tax Rate",
        hist_tax_f, tax_fn, f"={cl(PROJ_COLS[0])}{_tax_row}",
        note_text="Historical linked from P&L (tax / pretax income).  Forecast: avg of historicals, constant thereafter.")
    tax_row_dcf = row - 1

    row = blank(row)

    # ── SECTION 2: REVENUE & EBITDA ───────────────────────────────────────────
    row = shdr(row, "SECTION 2 — REVENUE & EBITDA  ($mm)", C_SECT)

    # Revenue
    wcell(row, 1, "Revenue", bold=True, bg=C_BG, halign="left", indent=1)
    for i, c in enumerate(HIST_COLS):
        wcell(row, c, f"='P&L'!{cl(c)}{_pl_rev}",
              bg=C_HIST, color=C_BLUE, fmt=NUM)
    for i, c in enumerate(PROJ_COLS):
        yr = proj_years[i]
        if yr in est_map and i < 3:
            e = est_map[yr]
            wcell(row, c, round(e["rev_avg"], 1), bg=C_BG, color="1A5276", fmt=NUM)
        else:
            prior_c = HIST_COLS[-1] if i == 0 else PROJ_COLS[i - 1]
            wcell(row, c, f"={cl(prior_c)}{row}*(1+{cl(c)}{rev_growth_row})",
                  bg=C_ALT, fmt=NUM)
    prior_proj = PROJ_COLS[-1]
    wcell(row, TERM_COL,
          f"={cl(prior_proj)}{row}*(1+{cl(TERM_COL)}{rev_growth_row})",
          bg=C_ASSM, fmt=NUM)
    note(row, ("FMP consensus years: analyst revenue average ($mm).  "
               "Range: Low–High shown below.  User years: prior × (1+growth)."))
    rev_row = row; row += 1

    # Revenue low / high (consensus years only)
    wcell(row, 1, "  Analyst Range:  Low — High  ($mm)",
          italic=True, halign="left", indent=2)
    for i, c in enumerate(PROJ_COLS):
        yr = proj_years[i]
        if yr in est_map:
            e = est_map[yr]
            wcell(row, c,
                  f"{e['rev_low']:,.0f} — {e['rev_high']:,.0f}",
                  italic=True, color="555555", bg=C_BG)
        else:
            wcell(row, c, "—", italic=True, color="999999")
    wcell(row, TERM_COL, "—", italic=True, color="999999", bg=C_ASSM)
    note(row, "Sell-side analyst low / high revenue estimates for consensus years")
    row += 1

    # Analyst count
    wcell(row, 1, "  Number of Analysts", italic=True, halign="left", indent=2)
    for i, c in enumerate(PROJ_COLS):
        yr = proj_years[i]
        if yr in est_map:
            n_ = est_map[yr]["n_rev"]
            color_ = ("B71C1C" if n_ < 5 else
                      "E65100" if n_ < 10 else C_BLACK)
            wcell(row, c, n_, italic=True, color=color_,
                  fmt='#,##0', bg=C_BG)
        else:
            wcell(row, c, "—", italic=True, color="999999")
    wcell(row, TERM_COL, "—", italic=True, color="999999", bg=C_ASSM)
    note(row, "Red < 5 analysts — treat estimate with caution.  Orange < 10.")
    row += 1

    # Revenue growth % (display row)
    wcell(row, 1, "  YoY Revenue Growth %", italic=True, halign="left", indent=2)
    for i, c in enumerate(HIST_COLS):
        if i > 0:
            f = f"=IFERROR({cl(c)}{rev_row}/{cl(HIST_COLS[i-1])}{rev_row}-1,\"\")"
            cell = wcell(row, c, f, italic=True, bg=C_HIST, fmt=PCT)
            cell.font = fnt(italic=True, color=C_BLACK)
        else:
            wcell(row, c, "—", italic=True, color="999999", bg=C_HIST)
    for i, c in enumerate(PROJ_COLS):
        prior_c = HIST_COLS[-1] if i == 0 else PROJ_COLS[i - 1]
        cell = wcell(row, c,
                     f"=IFERROR({cl(c)}{rev_row}/{cl(prior_c)}{rev_row}-1,\"\")",
                     italic=True, bg=C_BG if proj_years[i] in est_map else C_ALT, fmt=PCT)
        cell.font = fnt(italic=True, color=C_BLACK)
    wcell(row, TERM_COL,
          f"={cl(TERM_COL)}{rev_growth_row}",
          italic=True, bg=C_ASSM, fmt=PCT)
    note(row, "= this year / prior year − 1  (formula cross-check on assumptions)")
    row += 1
    row = blank(row)

    # EBITDA
    wcell(row, 1, "EBITDA", bold=True, bg=C_BG, halign="left", indent=1)
    for i, c in enumerate(HIST_COLS):
        wcell(row, c, f"='P&L'!{cl(c)}{_pl_ebitda}",
              bg=C_HIST, color=C_BLUE, fmt=NUM)
    for i, c in enumerate(PROJ_COLS):
        yr = proj_years[i]
        if yr in est_map and i < 3:
            e = est_map[yr]
            wcell(row, c, round(e["ebitda_avg"], 1), bg=C_BG, color="1A5276", fmt=NUM)
        else:
            wcell(row, c, f"={cl(c)}{rev_row}*{cl(c)}{ebitda_margin_row}",
                  bg=C_ALT, fmt=NUM)
    wcell(row, TERM_COL,
          f"={cl(TERM_COL)}{rev_row}*{cl(TERM_COL)}{ebitda_margin_row}",
          bg=C_ASSM, fmt=NUM)
    note(row, ("FMP consensus years: analyst EBITDA average.  "
               "User years: Revenue × EBITDA margin assumption."))
    ebitda_row = row; row += 1

    # EBITDA range
    wcell(row, 1, "  Analyst Range:  Low — High  ($mm)",
          italic=True, halign="left", indent=2)
    for i, c in enumerate(PROJ_COLS):
        yr = proj_years[i]
        if yr in est_map:
            e = est_map[yr]
            wcell(row, c,
                  f"{e['ebitda_low']:,.0f} — {e['ebitda_high']:,.0f}",
                  italic=True, color="555555", bg=C_BG)
        else:
            wcell(row, c, "—", italic=True, color="999999")
    wcell(row, TERM_COL, "—", italic=True, color="999999", bg=C_ASSM)
    note(row, "Sell-side analyst low / high EBITDA estimates for consensus years")
    row += 1

    # EBITDA margin display
    wcell(row, 1, "  EBITDA Margin %", italic=True, halign="left", indent=2)
    for i, c in enumerate(HIST_COLS):
        cell = wcell(row, c,
                     f"=IFERROR({cl(c)}{ebitda_row}/{cl(c)}{rev_row},\"\")",
                     italic=True, bg=C_HIST, fmt=PCT)
        cell.font = fnt(italic=True, color=C_BLACK)
    for i, c in enumerate(PROJ_COLS):
        cell = wcell(row, c,
                     f"=IFERROR({cl(c)}{ebitda_row}/{cl(c)}{rev_row},\"\")",
                     italic=True,
                     bg=C_BG if proj_years[i] in est_map else C_ALT, fmt=PCT)
        cell.font = fnt(italic=True, color=C_BLACK)
    cell = wcell(row, TERM_COL,
                 f"=IFERROR({cl(TERM_COL)}{ebitda_row}/{cl(TERM_COL)}{rev_row},\"\")",
                 italic=True, bg=C_ASSM, fmt=PCT)
    cell.font = fnt(italic=True, color=C_BLACK)
    note(row, "= EBITDA / Revenue  (formula — cross-check on margin assumption)")
    row += 1
    row = blank(row)

    # ── SECTION 3: FCF BUILD ──────────────────────────────────────────────────
    row = shdr(row, "SECTION 3 — UNLEVERED FREE CASH FLOW BUILD  ($mm)", C_SECT)

    # D&A
    wcell(row, 1, "  Less: D&A", halign="left", indent=2)
    for i, c in enumerate(HIST_COLS):
        wcell(row, c, f"=-'P&L'!{cl(c)}{_pl_da}",
              bg=C_HIST, color=C_BLUE, fmt=NUM)
    for c in PROJ_COLS + [TERM_COL]:
        bg_ = C_BG if (proj_years[PROJ_COLS.index(c)] in est_map
                       if c in PROJ_COLS else False) else C_ALT
        if c == TERM_COL: bg_ = C_ASSM
        wcell(row, c, f"=-{cl(c)}{rev_row}*{cl(c)}{da_pct_row}",
              bg=bg_, fmt=NUM)
    note(row, "= Revenue × D&A % assumption  (negative = P&L charge)")
    da_row = row; row += 1

    # EBIT
    wcell(row, 1, "EBIT  (Operating Profit)", bold=True, bg=C_SUB, halign="left", indent=1)
    for i, c in enumerate(HIST_COLS):
        wcell(row, c, f"={cl(c)}{ebitda_row}+{cl(c)}{da_row}", bold=True, bg=C_HIST, fmt=NUM)
        ws.cell(row=row, column=c).font = fnt(bold=True, color=C_BLACK)
    for c in PROJ_COLS + [TERM_COL]:
        wcell(row, c, f"={cl(c)}{ebitda_row}+{cl(c)}{da_row}",
              bold=True, bg=C_SUB, fmt=NUM)
    note(row, "= EBITDA + D&A  (GAAP operating income, EBIT)")
    ebit_row = row; row += 1

    # Tax on EBIT
    wcell(row, 1, "  Less: Tax on EBIT  (unlevered)", halign="left", indent=2)
    for i, c in enumerate(HIST_COLS):
        wcell(row, c, f"=-{cl(c)}{ebit_row}*{cl(c)}{tax_row_dcf}",
              bg=C_HIST, fmt=NUM)
        ws.cell(row=row, column=c).font = fnt(color=C_BLACK)
    for c in PROJ_COLS + [TERM_COL]:
        wcell(row, c, f"=-{cl(c)}{ebit_row}*{cl(c)}{tax_row_dcf}",
              fmt=NUM)
    note(row, "= EBIT × tax rate  (no interest tax shield — UFCF is pre-debt)")
    tax_ebit_row = row; row += 1

    # NOPAT
    wcell(row, 1, "NOPAT", bold=True, bg=C_BG, halign="left", indent=1)
    for i, c in enumerate(HIST_COLS):
        wcell(row, c, f"={cl(c)}{ebit_row}+{cl(c)}{tax_ebit_row}",
              bold=True, bg=C_HIST, fmt=NUM)
        ws.cell(row=row, column=c).font = fnt(bold=True, color=C_BLACK)
    for c in PROJ_COLS + [TERM_COL]:
        wcell(row, c, f"={cl(c)}{ebit_row}+{cl(c)}{tax_ebit_row}",
              bold=True, bg=C_BG, fmt=NUM)
    note(row, "= EBIT × (1 − tax rate)")
    nopat_row = row; row += 1

    # D&A add-back
    wcell(row, 1, "  (+) D&A add-back  (non-cash)", halign="left", indent=2)
    for i, c in enumerate(HIST_COLS):
        wcell(row, c, f"=-{cl(c)}{da_row}", bg=C_HIST, fmt=NUM)
        ws.cell(row=row, column=c).font = fnt(color=C_BLACK)
    for c in PROJ_COLS + [TERM_COL]:
        wcell(row, c, f"=-{cl(c)}{da_row}", fmt=NUM)
    note(row, "= D&A added back (non-cash charge — converts NOPAT to cash basis)")
    da_back_row = row; row += 1

    # CapEx
    wcell(row, 1, "  (−) Capital Expenditures", halign="left", indent=2)
    for i, c in enumerate(HIST_COLS):
        if cf_refs:
            wcell(row, c, f"='Cash Flow'!{cl(c)}{cf_refs['capex']}",
                  bg=C_HIST, color=C_BLUE, fmt=NUM)
        else:
            wcell(row, c, -round(hist_capex[i], 1) if hist_capex[i] else None,
                  bg=C_HIST, color=C_BLUE, fmt=NUM)
    for c in PROJ_COLS + [TERM_COL]:
        wcell(row, c, f"=-{cl(c)}{rev_row}*{cl(c)}{capex_pct_row}", fmt=NUM)
    note(row, "= Revenue × CapEx % assumption  (negative = cash outflow)")
    capex_row = row; row += 1

    # NWC
    wcell(row, 1, "  (−) Increase in Net Working Capital", halign="left", indent=2)
    for i, c in enumerate(HIST_COLS):
        if i == 0:
            wcell(row, c, None, bg=C_HIST, fmt=NUM)
        else:
            _pc = HIST_COLS[i - 1]
            wcell(row, c,
                  f"=-(('Balance Sheet'!{cl(c)}{_bs_tca}-'Balance Sheet'!{cl(c)}{_bs_tcl})"
                  f"-('Balance Sheet'!{cl(_pc)}{_bs_tca}-'Balance Sheet'!{cl(_pc)}{_bs_tcl}))",
                  bg=C_HIST, fmt=NUM)
    for c in PROJ_COLS + [TERM_COL]:
        wcell(row, c, f"=-{cl(c)}{rev_row}*{cl(c)}{nwc_pct_row}", fmt=NUM)
    note(row, "= −Revenue × NWC%  (positive NWC% → cash outflow, reduces UFCF)")
    nwc_row = row; row += 1

    # UFCF
    wcell(row, 1, "UNLEVERED FREE CASH FLOW  (UFCF)", bold=True,
          bg=C_BG, halign="left", indent=1)
    for i, c in enumerate(HIST_COLS):
        wcell(row, c,
              f"={cl(c)}{nopat_row}+{cl(c)}{da_back_row}+{cl(c)}{capex_row}+{cl(c)}{nwc_row}",
              bold=True, bg=C_HIST, fmt=NUM)
        ws.cell(row=row, column=c).font = fnt(bold=True, color=C_BLACK)
    for c in PROJ_COLS + [TERM_COL]:
        wcell(row, c,
              f"={cl(c)}{nopat_row}+{cl(c)}{da_back_row}+{cl(c)}{capex_row}+{cl(c)}{nwc_row}",
              bold=True, bg=C_BG, fmt=NUM)
    note(row, "= NOPAT + D&A − CapEx − ΔNWC  (pre-debt, pre-interest free cash flow)")
    ufcf_row = row; row += 1

    wcell(row, 1, "  UFCF Margin %", italic=True, halign="left", indent=2)
    for i, c in enumerate(HIST_COLS):
        cell = wcell(row, c, f"={cl(c)}{ufcf_row}/{cl(c)}{rev_row}",
                     italic=True, bg=C_HIST, fmt=PCT)
        cell.font = fnt(italic=True, color=C_BLACK)
    for c in PROJ_COLS + [TERM_COL]:
        cell = wcell(row, c,
                     f"=IFERROR({cl(c)}{ufcf_row}/{cl(c)}{rev_row},\"\")",
                     italic=True, fmt=PCT)
        cell.font = fnt(italic=True, color=C_BLACK)
    note(row, "Sense check — should approximate EBITDA margin less capex intensity")
    row += 1
    row = blank(row)

    # ── SECTION 4: TERMINAL VALUE ─────────────────────────────────────────────
    row = shdr(row, "SECTION 4 — TERMINAL VALUE  ($mm)", C_HD)

    # WACC ref
    wacc_ref = (f"=WACC!B{wacc_refs['wacc_row']}"
                if wacc_refs else None)
    wcell(row, 1, "WACC  (from WACC tab)", bold=True, bg=C_ASSM, halign="left", indent=1)
    wv = wcell(row, 2, wacc_ref or 0.12, bold=True, bg=C_ASSM,
               color=C_GREEN if wacc_ref else C_BLUE, fmt=PCT2)
    for c in range(3, NC + 1): wcell(row, c, None, bg=C_ASSM)
    note(row, ("Auto-linked from WACC tab selected output.  "
               "Override manually if needed."), bg=C_ASSM)
    wacc_dcf_row = row; row += 1

    wcell(row, 1, "Terminal Growth Rate  (g)", bold=True, bg=C_ASSM, halign="left", indent=1)
    wcell(row, 2, _DCF_TGR_BASE, bold=True, bg=C_ASSM, color=C_BLUE, fmt=PCT2)
    for c in range(3, NC + 1): wcell(row, c, None, bg=C_ASSM)
    note(row, (f"Growth tier: {_TIER.upper()} ({_rev_3yr_avg_dcf*100:.1f}% 3yr avg rev growth).  "
               f"Bear {_DCF_TGR_BEAR*100:.2f}% / Base {_DCF_TGR_BASE*100:.1f}% / Bull {_DCF_TGR_BULL*100:.2f}%.  "
               f"Keep below WACC."),
         bg=C_ASSM)
    tg_row = row; row += 1

    wcell(row, 1, "Terminal EV/EBITDA Multiple  (exit multiple method)",
          bold=True, bg=C_ASSM, halign="left", indent=1)
    wcell(row, 2, _DCF_TEV_BASE, bold=True, bg=C_ASSM, color=C_BLUE, fmt='0.0x')
    for c in range(3, NC + 1): wcell(row, c, None, bg=C_ASSM)
    note(row, (f"Growth tier: {_TIER.upper()} ({_rev_3yr_avg_dcf*100:.1f}% 3yr avg rev growth).  "
               f"Bear {_DCF_TEV_BEAR:.0f}x / Base {_DCF_TEV_BASE:.0f}x / Bull {_DCF_TEV_BULL:.0f}x.  "
               f"Override manually if needed."),
         bg=C_ASSM)
    tev_row = row; row += 1
    row = blank(row)

    wcell(row, 1, "Terminal Year UFCF  (grown by g)", halign="left", indent=1)
    wcell(row, 2, f"={cl(TERM_COL)}{ufcf_row}",
          color=C_GREEN, fmt=NUM)
    ws.cell(row=row, column=2).fill = fll(C_WHITE)
    ws.cell(row=row, column=2).border = brd()
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    for c in range(3, NC + 1): wcell(row, c, None)
    note(row, "Cross-ref from Section 3 — terminal year UFCF (already grown by g in assumptions)")
    tv_ufcf_row = row; row += 1

    wcell(row, 1, "Terminal Value  [Gordon Growth:  UFCF / (WACC − g)]",
          bold=True, bg=C_BG, halign="left", indent=1)
    tv_gg = wcell(row, 2, f"=B{tv_ufcf_row}/(B{wacc_dcf_row}-B{tg_row})",
                  bold=True, bg=C_BG, fmt=NUM)
    tv_gg.font = fnt(bold=True, color=C_BLACK)
    for c in range(3, NC + 1): wcell(row, c, None, bg=C_BG)
    note(row, "Sensitive to g — always cross-check vs Exit Multiple below")
    tv_gg_row = row; row += 1

    wcell(row, 1, "Terminal Value  [Exit Multiple:  Terminal EBITDA × Multiple]",
          bold=True, bg=C_BG, halign="left", indent=1)
    tv_em = wcell(row, 2, f"={cl(TERM_COL)}{ebitda_row}*B{tev_row}",
                  bold=True, bg=C_BG, fmt=NUM)
    tv_em.font = fnt(bold=True, color=C_BLACK)
    for c in range(3, NC + 1): wcell(row, c, None, bg=C_BG)
    note(row, "Anchored to observable market multiples — less model-sensitive than Gordon Growth")
    tv_em_row = row; row += 1
    row = blank(row)

    # ── SECTION 5: DISCOUNTING ────────────────────────────────────────────────
    row = shdr(row, "SECTION 5 — DISCOUNTING  &  ENTERPRISE VALUE  ($mm)", C_SECT)

    wcell(row, 1, "Discount Period  (mid-year convention)", halign="left", indent=1)
    for i, c in enumerate(PROJ_COLS):
        wcell(row, c, i + 0.5, fmt='0.0')
    wcell(row, TERM_COL, len(proj_years), fmt='0.0', bg=C_ASSM)
    note(row, "Mid-year: 0.5, 1.5, 2.5...  assumes FCF received evenly through each year")
    disc_period_row = row; row += 1

    wcell(row, 1, "Discount Factor  =  1 / (1 + WACC) ^ period", halign="left", indent=1)
    for c in PROJ_COLS + [TERM_COL]:
        bg_ = C_ASSM if c == TERM_COL else C_WHITE
        wcell(row, c, f"=1/(1+B{wacc_dcf_row})^{cl(c)}{disc_period_row}",
              bg=bg_, fmt='0.000')
    note(row, "= 1 / (1 + WACC) ^ discount period")
    disc_factor_row = row; row += 1

    wcell(row, 1, "PV of UFCF", bold=True, bg=C_BG, halign="left", indent=1)
    for c in PROJ_COLS:
        wcell(row, c, f"={cl(c)}{ufcf_row}*{cl(c)}{disc_factor_row}",
              bold=True, bg=C_BG, fmt=NUM)
    wcell(row, TERM_COL, None, bg=C_BG)
    note(row, "= UFCF × discount factor")
    pv_ufcf_row = row; row += 1

    sum_f = "+".join(f"{cl(c)}{pv_ufcf_row}" for c in PROJ_COLS)
    wcell(row, 1, "Sum of PV of FCFs  (explicit period)", bold=True,
          bg=C_SUB, halign="left", indent=1)
    wcell(row, 2, f"={sum_f}", bold=True, bg=C_SUB, fmt=NUM)
    for c in range(3, NC + 1): wcell(row, c, None, bg=C_SUB)
    note(row, f"Sum of {len(proj_years)} discounted annual FCFs")
    sum_pv_row = row; row += 1

    for label, tv_r, bg_ in [("PV of Terminal Value  [Gordon Growth]",  tv_gg_row, C_SUB),
                              ("PV of Terminal Value  [Exit Multiple]",  tv_em_row, C_SUB)]:
        wcell(row, 1, label, bold=True, bg=bg_, halign="left", indent=1)
        wcell(row, 2, f"=B{tv_r}*{cl(TERM_COL)}{disc_factor_row}",
              bold=True, bg=bg_, fmt=NUM)
        for c in range(3, NC + 1): wcell(row, c, None, bg=bg_)
        note(row, "Terminal value × terminal-year discount factor")
        if "Gordon" in label: pvtv_gg_row = row
        else:                 pvtv_em_row = row
        row += 1
    row = blank(row)

    # ── SECTION 6: EQUITY BRIDGE ──────────────────────────────────────────────
    row = shdr(row, "SECTION 6 — EQUITY VALUE BRIDGE  &  IMPLIED SHARE PRICE", C_HD)

    ev_rows = {}
    for label, pv_tv in [("Gordon Growth", pvtv_gg_row),
                          ("Exit Multiple",  pvtv_em_row)]:
        wcell(row, 1, f"Enterprise Value  [{label}]",
              bold=True, bg=C_BG, halign="left", indent=1)
        wcell(row, 2, f"=B{sum_pv_row}+B{pv_tv}",
              bold=True, bg=C_BG, fmt=NUM)
        for c in range(3, NC + 1): wcell(row, c, None, bg=C_BG)
        ev_rows[label] = row; row += 1

    bs0 = bs_data[-1]  # used here and below for minority interest / shares
    # net_debt in $mm — always computed for the Python-side DCF price calculation below
    net_debt = (((bs0.get("shortTermDebt") or 0) + (bs0.get("longTermDebt") or 0))
                - (bs0.get("cashAndCashEquivalents") or 0)) / 1e6
    wcell(row, 1, "  Less: Net Debt  (Debt − Cash)", halign="left", indent=2)
    _nd_col = cl(HIST_COLS[-1])
    if "nd" in bs_refs:
        _nd_val = f"='Balance Sheet'!{_nd_col}{bs_refs['nd']}"
    else:
        _nd_val = round(net_debt, 1)
    wcell(row, 2, _nd_val, color=C_GREEN, fmt=NUM)
    ws.cell(row=row, column=2).fill = fll(C_WHITE); ws.cell(row=row, column=2).border = brd()
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    for c in range(3, NC + 1): wcell(row, c, None)
    note(row, "Linked from Balance Sheet: Total Debt (ST+LT) − Cash & Equivalents  (negative = net cash)")
    nd_row = row; row += 1

    wcell(row, 1, "  Less: Minority Interest", halign="left", indent=2)
    mi = (bs0.get("minorityInterest") or 0) / 1e6
    wcell(row, 2, round(mi, 1), color=C_GREEN, fmt=NUM)
    ws.cell(row=row, column=2).fill = fll(C_WHITE); ws.cell(row=row, column=2).border = brd()
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    for c in range(3, NC + 1): wcell(row, c, None)
    note(row, "Auto-linked from Balance Sheet: minorityInterest")
    mi_row = row; row += 1

    shares = (bs0.get("commonStockSharesOutstanding") or
              is_data[-1].get("weightedAverageShsOutDil") or 0) / 1e6

    # ── S-001: YoY share-count split detector ────────────────────────────────
    # Catches stock splits/reverse-splits before they silently corrupt per-share
    # multiples (P/E, P/FCF, EPS).  >25% YoY move is far outside normal buyback
    # or issuance ranges — almost always signals a split adjustment issue.
    _s001_prev = 0.0
    if len(bs_data) >= 2:
        _s001_prev = (bs_data[-2].get("commonStockSharesOutstanding") or 0) / 1e6
    if _s001_prev == 0 and len(is_data) >= 2:
        _s001_prev = (is_data[-2].get("weightedAverageShsOutDil") or 0) / 1e6
    if shares > 0 and _s001_prev > 0:
        _s001_chg = (shares - _s001_prev) / _s001_prev
        if abs(_s001_chg) > 0.25:
            _s001_dir = "increase" if _s001_chg > 0 else "decrease"
            print(
                f"[S-001 WARNING] {ticker}: shares changed {_s001_chg:+.1%} YoY "
                f"({_s001_prev:.1f}mm → {shares:.1f}mm). "
                f"Large {_s001_dir} — likely stock split/reverse-split. "
                f"Verify FMP price & EPS split-adjustment before trusting any "
                f"per-share multiples (P/E, P/FCF, EPS)."
            )
    # ─────────────────────────────────────────────────────────────────────────

    wcell(row, 1, "  Shares Outstanding — Diluted  (mm)", halign="left", indent=2)
    wcell(row, 2, round(shares, 1), color=C_GREEN, fmt=NUM)
    ws.cell(row=row, column=2).fill = fll(C_WHITE); ws.cell(row=row, column=2).border = brd()
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    for c in range(3, NC + 1): wcell(row, c, None)
    note(row, "Auto-linked: weightedAverageShsOutDil from income statement (diluted)")
    sh_row = row; row += 1
    row = blank(row)

    if current_price:
        price = float(current_price)
    elif profile and profile.get("price"):
        # Reuse caller's profile (no extra FMP call)
        price = float(profile.get("price") or 0)
    else:
        price = float(is_data[-1].get("price") or 0)
        try:
            prof = requests.get(
                f"https://financialmodelingprep.com/stable/profile"
                f"?symbol={ticker}&apikey={API_KEY}", timeout=8
            ).json()
            price = float((prof[0] if isinstance(prof, list) else prof).get("price") or price)
        except Exception:
            pass

    for label in ["Gordon Growth", "Exit Multiple"]:
        ev_r = ev_rows[label]
        wcell(row, 1, f"Implied Share Price  [{label}]  ($)",
              bold=True, bg=C_BG, halign="left", indent=1)
        ip = wcell(row, 2,
                   f"=IFERROR((B{ev_r}-B{nd_row}-B{mi_row})/B{sh_row},\"\")",
                   bold=True, bg=C_BG, fmt=DOLS)
        ip.font = fnt(bold=True, color=C_BLACK)
        for c in range(3, NC + 1): wcell(row, c, None, bg=C_BG)
        if label == "Gordon Growth": ip_gg_row = row
        else:                        ip_em_row = row
        row += 1

    wcell(row, 1, "Current Market Price  ($)", halign="left", indent=1)
    cp = wcell(row, 2, round(price, 2) if price else None, color=C_GREEN, fmt=DOLS)
    cp.fill = fll(C_WHITE); cp.border = brd()
    cp.alignment = Alignment(horizontal="right")
    for c in range(3, NC + 1): wcell(row, c, None)
    note(row, "Auto-linked from FMP company profile — price")
    cp_row = row; row += 1

    for label, ip_r in [("Gordon Growth", ip_gg_row),
                         ("Exit Multiple",  ip_em_row)]:
        wcell(row, 1, f"Upside / (Downside)  [{label}]",
              bold=True, bg=C_SUB, halign="left", indent=1)
        cell = wcell(row, 2,
                     f"=IFERROR(B{ip_r}/B{cp_row}-1,\"\")",
                     bold=True, bg=C_SUB, fmt=PCT)
        cell.font = fnt(bold=True, color=C_BLACK)
        for c in range(3, NC + 1): wcell(row, c, None, bg=C_SUB)
        row += 1

    # ── Python-side implied prices — exact mirror of Excel assumption rows ──────
    # All rates use AVERAGES matching Excel (Year 1 = avg historicals, Yrs 2-5 = Yr 1).
    # Revenue / EBITDA years 4-5 use Year 3 growth + Year 3 margin (matches Excel formula).
    dcf_prices = {"gg_price": None, "em_price": None,
                  "gg_upside": None, "em_upside": None}
    try:
        _g    = _DCF_TGR_BASE   # tier-calibrated terminal growth rate
        _tev  = _DCF_TEV_BASE  # tier-calibrated exit EV/EBITDA multiple
        _wacc = (wacc_refs or {}).get("wacc_val")

        # FX: financials in reportedCurrency; implied price must be in USD.
        # Reads latest historical (is_data[-1]) — previous version read [0] which is
        # the OLDEST year, occasionally misreporting currency for tickers that switched.
        _ccy_dcf   = is_data[-1].get("reportedCurrency", "USD") if is_data else "USD"
        _fx_to_usd = 1.0
        _fx_fetched = False
        if _ccy_dcf != "USD":
            try:
                import requests as _rx
                _fxr = _rx.get(
                    f"https://financialmodelingprep.com/api/v3/fx/{_ccy_dcf}USD"
                    f"?apikey={API_KEY}", timeout=5
                ).json()
                if isinstance(_fxr, list) and _fxr:
                    _fx_to_usd = float(_fxr[0]["ask"])
                    _fx_fetched = True
            except Exception:
                pass

        # F-P: Foreign-reporter guard. If financials are non-USD AND the FX fetch
        # failed (or returned the 1.0 fallback), refuse to compute USD prices.
        # Silent FX failure was producing nonsense (TSM in TWD → +5,550% upside).
        _foreign_reporter_unsupported = (_ccy_dcf != "USD" and not _fx_fetched)

        # F-D / F-Q: Bank-charter detection for DCF.
        # Payment networks (V, MA, PYPL, FIS, FISV, GPN, etc.) share "Financial"
        # industry / sector tags with deposit-funded banks but are NOT balance-
        # sheet lenders. Exclude them so F-D does not wrongly disable their DCF.
        _BANK_DCF_EXCLUDE = {"V", "MA", "PYPL", "FIS", "FISV", "GPN", "WU",
                              "DFS", "TRMK"}
        _BANK_DCF_KW = {"bank", "banking", "financial services", "savings",
                        "thrift", "mortgage", "credit union", "investment bank",
                        "diversified financial"}
        _prof_industry_dcf = (
            (profile or {}).get("industry") or (profile or {}).get("sector") or ""
        )
        _is_bank_dcf = (
            any(kw in _prof_industry_dcf.lower() for kw in _BANK_DCF_KW)
            and ticker.upper() not in _BANK_DCF_EXCLUDE
        )

        # Assumption averages — match Excel "Year 1 = AVERAGE(historicals)" rows
        def _avg(vals): return sum(v for v in vals if v) / max(sum(1 for v in vals if v), 1)
        avg_da_pct    = _avg(hist_da_pct)
        avg_capex_pct = _avg(hist_capex_pct)
        avg_tax       = _avg(hist_tax) if hist_tax else last_tax

        # NWC%: historical average ΔNWC/Revenue, clamped to [-2%, +5%], then faded to 0
        # by terminal year — mirrors the Excel cap-and-fade assumption row.
        _nwc_hist = []
        for _ni in range(1, n_hist):
            _pa = (bs_data[_ni-1].get("totalCurrentAssets")     or 0) / 1e6
            _pl = (bs_data[_ni-1].get("totalCurrentLiabilities") or 0) / 1e6
            _ca = (bs_data[_ni].get("totalCurrentAssets")        or 0) / 1e6
            _cl = (bs_data[_ni].get("totalCurrentLiabilities")   or 0) / 1e6
            _dnwc = (_ca - _cl) - (_pa - _pl)
            if hist_rev[_ni] > 0:
                _nwc_hist.append(_dnwc / hist_rev[_ni])
        _raw_nwc_pct   = _avg(_nwc_hist) if _nwc_hist else 0.01
        _clamped_nwc   = max(-0.02, min(0.05, _raw_nwc_pct))  # cap: [-2%, +5%]

        if _wacc and (_wacc - _g) > 0.001 and shares > 0:
            # Project revenues and EBITDA (years 1-3 from analyst estimates, 4-5 from yr 3)
            _proj_revs, _proj_ebitda = [], []
            for _i, _yr in enumerate(proj_years):
                if _yr in est_map and _i < 3 and est_map[_yr].get("rev_avg"):
                    _r = est_map[_yr]["rev_avg"]
                    _e = est_map[_yr].get("ebitda_avg") or (
                        _r * (est_map[est_years[-1]]["ebitda_avg"] /
                              max(est_map[est_years[-1]]["rev_avg"], 1))
                        if estimates else _r * (hist_ebitda[-1] / max(hist_rev[-1], 1))
                    )
                else:
                    _prior_r = _proj_revs[-1] if _proj_revs else hist_rev[-1]
                    if _i >= 3 and len(_proj_revs) >= 3:
                        # Years 4-5: use Year 3 growth rate and Year 3 EBITDA margin
                        _yr3_r = _proj_revs[2]; _yr2_r = _proj_revs[1]
                        _yr3_g = (_yr3_r / _yr2_r - 1) if _yr2_r > 0 else _g
                        _yr3_m = _proj_ebitda[2] / _yr3_r if _yr3_r > 0 else (hist_ebitda[-1] / max(hist_rev[-1], 1))
                        _r = _prior_r * (1 + _yr3_g)
                        _e = _r * _yr3_m
                    else:
                        _last_known_m = (hist_ebitda[-1] / max(hist_rev[-1], 1) if hist_rev[-1] else 0.20)
                        _r = _prior_r * (1 + _g)
                        _e = _r * _last_known_m
                _proj_revs.append(_r); _proj_ebitda.append(_e)

            # Terminal year — uses Year 5 margin (trailing EBITDA margin in Excel)
            _trailing_margin = hist_ebitda[-1] / max(hist_rev[-1], 1) if hist_rev[-1] else 0.20
            _term_rev    = _proj_revs[-1] * (1 + _g)
            _term_ebitda = _term_rev * _trailing_margin

            def _py_ufcf(rev, ebitda, nwc_pct_=0.0):
                da    = rev * avg_da_pct
                nopat = (ebitda - da) * (1 - avg_tax)
                return nopat + da - rev * avg_capex_pct - rev * nwc_pct_

            # Each projection year gets a faded NWC% (Year 1 = clamped, linear → 0 by terminal).
            _n_py = len(proj_years)
            _sum_pv  = sum(
                _py_ufcf(_proj_revs[i], _proj_ebitda[i],
                         nwc_pct_=_clamped_nwc * (_n_py - i) / _n_py)
                / (1 + _wacc) ** (i + 0.5)
                for i in range(_n_py)
            )
            _tv_disc = (1 + _wacc) ** len(proj_years)

            # Terminal UFCF uses nwc_pct_=0 (steady state: NWC growth tracks revenue exactly)
            _tv_gg   = _py_ufcf(_term_rev, _term_ebitda, nwc_pct_=0.0) / (_wacc - _g)
            _ip_gg   = (_sum_pv + _tv_gg / _tv_disc - net_debt - mi) / shares

            _tv_em   = _term_ebitda * _tev
            _ip_em   = (_sum_pv + _tv_em / _tv_disc - net_debt - mi) / shares

            _ip_gg_usd = _ip_gg * _fx_to_usd
            _ip_em_usd = _ip_em * _fx_to_usd

            # F-I: Block negative price targets. A negative equity value per share is
            # mathematically possible (positive EBITDA × multiple minus enormous net
            # debt) but unpublishable. Affects banks (deposit liabilities in net-debt),
            # captive-finance autos (Ford), and over-levered names. Rule 5 violation
            # to display these. Clamp to None → downstream "N/A — Insufficient inputs".
            if _ip_gg_usd is not None and _ip_gg_usd <= 0:
                _ip_gg_usd = None
            if _ip_em_usd is not None and _ip_em_usd <= 0:
                _ip_em_usd = None

            # F-P: If foreign-reporter and FX fetch failed, void the prices entirely.
            if _foreign_reporter_unsupported:
                _ip_gg_usd = None
                _ip_em_usd = None

            # GG sensitivity: vary both TGR and WACC (±0.5pp) — bear/bull are
            # more conservative/optimistic on both levers simultaneously.
            _WACC_SHIFT = 0.005   # 0.5 percentage points
            _wacc_bear  = round(_wacc + _WACC_SHIFT, 4)  # higher WACC → lower price
            _wacc_bull  = round(_wacc - _WACC_SHIFT, 4)  # lower WACC  → higher price

            def _gg_px_at(tgr_s, wacc_s=None):
                w = wacc_s if wacc_s is not None else _wacc
                if (w - tgr_s) <= 0.001:
                    return None
                _trv = _proj_revs[-1] * (1 + tgr_s)
                _teb = _trv * _trailing_margin
                _ip  = (_sum_pv + _py_ufcf(_trv, _teb) / (w - tgr_s) / _tv_disc
                        - net_debt - mi) / shares
                return round(_ip * _fx_to_usd, 2)

            # EM sensitivity: tier-calibrated bear/bull multiples
            _tev_bear = _DCF_TEV_BEAR
            _tev_bull = _DCF_TEV_BULL
            def _em_px_at(mult):
                _ip = (_sum_pv + _term_ebitda * mult / _tv_disc - net_debt - mi) / shares
                return round(_ip * _fx_to_usd, 2)

            # ── Negative-earnings regime detection ─────────────────────────────
            # Gordon Growth requires stable positive UFCF growing forever. If trailing
            # FCF or EBIT is negative, the perpetuity formula produces nonsense
            # (negative terminal value → negative implied price). Disable GG entirely
            # in that case and fall back to EV/EBITDA Exit Multiple as primary.
            _trailing_ebit_raw  = (is_data[-1].get("operatingIncome")  or 0) / 1e6
            _trailing_ocf_raw   = (cf_data[-1].get("operatingCashFlow") or 0) / 1e6
            _trailing_capex_raw = (cf_data[-1].get("capitalExpenditure") or 0) / 1e6
            _trailing_fcf_raw   = _trailing_ocf_raw + _trailing_capex_raw  # capex is negative

            _neg_earnings_regime = (_trailing_fcf_raw < 0) or (_trailing_ebit_raw < 0)
            _gg_disabled_reason = None
            if _neg_earnings_regime:
                _gg_disabled_reason = (
                    f"Gordon Growth disabled — trailing FCF "
                    f"${_trailing_fcf_raw/1e3:+,.1f}B, trailing EBIT "
                    f"${_trailing_ebit_raw/1e3:+,.1f}B. "
                    f"GG perpetuity formula requires stable positive UFCF; "
                    f"EV/EBITDA Exit Multiple used as sole primary method."
                )

            # ── EV/Sales regime ────────────────────────────────────────────────
            # Triggered when GG is disabled AND trailing EBITDA is also negative.
            # In that case even the EM formula is unreliable (negative terminal EBITDA
            # produces a negative or nonsense EM price). Use forward EV/Revenue
            # with a sector-appropriate mature multiple, discounted back at WACC.
            #
            # F-L: Extend EVS trigger for marginal/turnaround situations:
            # (a) EBITDA < 5% of revenue (de minimis — EM exit will produce tiny EV)
            # (b) FCF worse than −10% of revenue (heavy cash consumption)
            # Both thresholds are designed to catch INTC-style situations where
            # EBITDA is technically positive but the multiple is distorted by
            # the severity of the trough. Cyclicals are excluded — their EBITDA
            # compression is transient and EM at 10x trough still makes sense.
            _trailing_ebitda_mm = hist_ebitda[-1]  # $mm, already computed above
            _trailing_rev_mm_fl  = hist_rev[-1] if hist_rev else 1.0  # $mm, for F-L ratio
            _ebitda_near_zero = (
                _trailing_ebitda_mm < 0.05 * max(_trailing_rev_mm_fl, 1)
                and not _is_cyclical_dcf   # cyclicals' trough EBITDA is transient
            )
            _fcf_deeply_neg = (
                _trailing_fcf_raw < -0.10 * max(_trailing_rev_mm_fl / 1e3, 1)
                and not _is_cyclical_dcf
            )
            _evs_regime = _neg_earnings_regime and (
                _trailing_ebitda_mm < 0
                or _ebitda_near_zero
                or _fcf_deeply_neg
            )

            _evs_price         = None
            _evs_implied_cagr  = None
            _evs_required_rev  = None   # $B
            _evs_mature_mult   = None
            _evs_subtype       = None
            _evs_yr5_rev_b     = None

            if _evs_regime:
                _evs_subtype     = _secular_growth_subtype(ticker)
                _evs_mature_mult = EVS_MATURE_MULTS.get(_evs_subtype, 4.0)

                _trailing_rev_mm = hist_rev[-1]
                # Year-5 revenue from DCF projections (already computed above)
                _yr5_rev_mm = _proj_revs[-1] if _proj_revs else (
                    _trailing_rev_mm * (1 + _rev_3yr_avg_dcf) ** 5
                    if _trailing_rev_mm > 0 else None
                )
                _evs_yr5_rev_b = round(_yr5_rev_mm / 1e3, 2) if _yr5_rev_mm else None

                if _yr5_rev_mm and _yr5_rev_mm > 0:
                    _yr5_ev_mm  = _yr5_rev_mm * _evs_mature_mult
                    _yr5_eq_mm  = _yr5_ev_mm - net_debt - mi
                    _evs_pv_mm  = _yr5_eq_mm / (1 + _wacc) ** 5
                    _evs_px_loc = _evs_pv_mm / shares if shares > 0 else None
                    _evs_price  = (round(_evs_px_loc * _fx_to_usd, 2)
                                   if _evs_px_loc and _evs_px_loc > 0 else None)

                # Reverse check: what 5yr revenue CAGR does today's EV imply?
                if price and price > 0 and shares > 0 and _trailing_rev_mm > 0:
                    _curr_ev_mm = price / _fx_to_usd * shares + net_debt + mi
                    _req_rev_mm = _curr_ev_mm / _evs_mature_mult
                    _evs_required_rev = round(_req_rev_mm / 1e3, 2)   # $B
                    if _req_rev_mm > 0:
                        _evs_implied_cagr = round(
                            (_req_rev_mm / _trailing_rev_mm) ** (1 / 5) - 1, 4
                        )

            # F-I post-clamp: _ip_gg_usd / _ip_em_usd may now be None (negative or
            # foreign-reporter-unsupported). Helpers below stay None-safe.
            _gg_final = None if (_neg_earnings_regime or _ip_gg_usd is None) else round(_ip_gg_usd, 2)
            _em_final = None if _ip_em_usd is None else round(_ip_em_usd, 2)
            _gg_upside_final = (round(_ip_gg_usd / price - 1, 4)
                                if (_ip_gg_usd is not None and price and not _neg_earnings_regime) else None)
            _em_upside_final = (round(_ip_em_usd / price - 1, 4)
                                if (_ip_em_usd is not None and price) else None)

            # F-P: foreign-reporter flag surfaced to bridge so the hero can explain.
            if _foreign_reporter_unsupported:
                _gg_disabled_reason = (
                    f"Financials reported in {_ccy_dcf}; FX conversion to USD unavailable. "
                    f"Pricing skipped — see /dcf calculator to enter manually."
                )

            # F-D Phase 1 + F-M: Force-disable GG and EM for bank-charter institutions.
            # Neither Gordon Growth (perpetuity on FCF) nor EV/EBITDA Exit Multiple applies
            # to deposit-funded balance sheets — DDM / Justified P/B is the correct
            # methodology (pending Phase 2 implementation).
            # F-M: Also resets neg_earnings_regime / evs_regime to False so that bank
            # "negative FCF" (loan origination and deposit accounting noise) does not
            # mistakenly trigger the EVS fallback path — that path is for pre-profit
            # secular-growth companies, not solvent deposit-funded institutions.
            _bank_disabled_reason = None
            if _is_bank_dcf:
                _bank_disabled_reason = (
                    "Bank-charter institution — Gordon Growth and EV/EBITDA Exit Multiple "
                    "do not apply to deposit-funded balance sheets. "
                    "DDM / Justified P/B methodology pending."
                )
                _gg_final           = None
                _em_final           = None
                _gg_upside_final    = None
                _em_upside_final    = None
                _gg_disabled_reason = _bank_disabled_reason
                # F-M: bank FCF/EBIT accounting noise must not trigger EVS regime
                _neg_earnings_regime = False
                _evs_regime          = False
                _evs_price           = None
                _evs_implied_cagr    = None
                _evs_required_rev    = None
                _evs_mature_mult     = None
                _evs_subtype         = None
                _evs_yr5_rev_b       = None

            dcf_prices = {
                "gg_price":      _gg_final,
                "gg_bear_price": None if (_neg_earnings_regime or _foreign_reporter_unsupported) else _gg_px_at(_DCF_TGR_BEAR, wacc_s=_wacc_bear),
                "gg_bull_price": None if (_neg_earnings_regime or _foreign_reporter_unsupported) else _gg_px_at(_DCF_TGR_BULL, wacc_s=_wacc_bull),
                "gg_disabled_reason": _gg_disabled_reason,
                "wacc_bear":     _wacc_bear,
                "wacc_bull":     _wacc_bull,
                "em_price":      _em_final,
                "em_bear_price": None if _foreign_reporter_unsupported else _em_px_at(_tev_bear),
                "em_bull_price": None if _foreign_reporter_unsupported else _em_px_at(_tev_bull),
                "em_base_mult":  _tev,
                "em_bear_mult":  _tev_bear,
                "em_bull_mult":  _tev_bull,
                "tgr_base":      _DCF_TGR_BASE,
                "tgr_bear":      _DCF_TGR_BEAR,
                "tgr_bull":      _DCF_TGR_BULL,
                "growth_tier":   _TIER,
                "rev_3yr_avg":   round(_rev_3yr_avg_dcf, 4),
                "trailing_fcf_b":    round(_trailing_fcf_raw / 1e3, 2),
                "trailing_ebit_b":   round(_trailing_ebit_raw / 1e3, 2),
                "trailing_ebitda_b": round(_trailing_ebitda_mm / 1e3, 2),
                "neg_earnings_regime": _neg_earnings_regime,
                "evs_regime":        _evs_regime,
                "evs_price":         _evs_price,
                "evs_implied_cagr":  _evs_implied_cagr,
                "evs_required_rev":  _evs_required_rev,
                "evs_mature_mult":   _evs_mature_mult,
                "evs_subtype":       _evs_subtype,
                "evs_yr5_rev_b":     _evs_yr5_rev_b,
                "gg_upside":  _gg_upside_final,
                "em_upside":  _em_upside_final,
                "evs_upside": round(_evs_price / price - 1, 4) if (_evs_price and price) else None,
                # D-001 + F-P + F-N transparency fields
                "wacc_raw":         (wacc_refs or {}).get("wacc_raw"),
                "wacc_floored":     (wacc_refs or {}).get("wacc_floored", False),
                "foreign_reporter": _foreign_reporter_unsupported,
                "reported_currency": _ccy_dcf,
                # F-D Phase 1: bank-disabled flag — bridges reads this to show
                # "N/A — Bank methodology pending" instead of "N/A — Insufficient inputs"
                "bank_disabled":        _is_bank_dcf,
                "bank_disabled_reason": _bank_disabled_reason,
                # F-C: quality premium and thin-margin flag for report_bridge
                "quality_em_premium":   _quality_em_premium,
                "fcf_margin_trailing":  _fcf_margin_trailing,
                # F-C Phase 2: historical EV/EBITDA anchor metadata for cap banner
                "em_anchored":          _em_anchored and not _evs_regime,
                "em_hist_anchor_raw":   _em_hist_anchor_raw if (_em_anchored and not _evs_regime) else None,
                "em_hist_anchor_capped": _em_hist_anchor_capped if (_em_anchored and not _evs_regime) else None,
            }
    except Exception:
        pass  # never break model generation

    return {
        "ufcf_row": ufcf_row, "rev_row": rev_row,
        "ebitda_row": ebitda_row, "wacc_dcf_row": wacc_dcf_row,
        "dcf_prices": dcf_prices,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# SECTOR-AWARE SCORING — bucket classification + threshold tables
# ═══════════════════════════════════════════════════════════════════════════════

def _sector_bucket(sector_str, ticker):
    """Map company to one of 4 scoring buckets based on sector/industry string and ticker."""
    s = (sector_str or "").lower()
    t = (ticker or "").upper()

    # Banks/financials get special treatment (already have is_bank flag)
    if any(x in s for x in ["bank", "financial service", "insurance", "capital market"]):
        return "bank"
    if t in {"JPM","BAC","WFC","C","GS","MS","SOFI","BLK","SCHW","AXP","COF","USB","PNC"}:
        return "bank"

    # Tech/growth: high ROIC acceptable at higher threshold, high rev growth expected
    if any(x in s for x in ["software", "semiconductor", "technology", "internet", "media"]):
        return "tech_growth"
    if t in {"NVDA","MSFT","AAPL","ADBE","AMD","META","NFLX","TSLA","SOFI","GOOGL","AMZN",
             "ORCL","CRM","NOW","SNOW","PANW","CRWD","INTU","AMAT","KLAC","LRCX","MU","AVGO",
             "QCOM","TSM","DDOG","TEAM","NET","MDB"}:
        return "tech_growth"

    # Cyclical: lower thresholds, higher leverage tolerance
    if any(x in s for x in ["auto", "airline", "aerospace", "industrial", "steel", "mining",
                              "oil", "energy", "chemical"]):
        return "cyclical"
    if t in {"F","UAL","INTC","BA","GE","CAT","DE","XOM","CVX","COP","SLB","OXY","RCL","CCL"}:
        return "cyclical"

    # Default: stable compounder (consumer staples, healthcare, retail, utilities, etc.)
    return "stable_compounder"


def _secular_growth_subtype(ticker):
    """EV/Sales subtype for pre-profit secular-growth companies.

    Used when both GG and EM are unreliable (negative EBITDA).
    Returns one of: secular_growth_deeptech, secular_growth_software,
    secular_growth_resources, tech_growth, stable_compounder, cyclical.
    """
    t = (ticker or "").upper()
    # Space, deep hardware, quantum, autonomous, biotech/genomics
    if t in {
        "RKLB","SPCE","ASTR","RDW","ASTS","LUNR","ACHR","JOBY","LILM",
        "IONQ","RGTI","QUBT","ARQQ","RXRX","SANA","NTLA","BEAM","EDIT","CRSP",
        "ABCL","PACB","NNOX","LAZR","LIDR","VLDR","OUST","MVIS",
    }:
        return "secular_growth_deeptech"
    # High-growth SaaS / software / data platforms with negative earnings
    if t in {
        "AI","PATH","GTLB","TASK","DOCN","CFLT","APPN","SUMO","BASE","BRZE",
        "AMPL","SEMR","ZETA","RELY","OKTA","ZS","S","ESTC","MSTR",
    }:
        return "secular_growth_software"
    # Clean energy, hydrogen, EV, critical materials
    if t in {
        "PLUG","FCEL","BLOOM","MP","CLNE","NKLA","HYLN","HYZN","FFIE","GOEV",
        "FSR","PTRA","WKHS","RUN","NOVA",
    }:
        return "secular_growth_resources"
    # F-O: Social media / ad-tech platforms — revenue-generating but pre-profitability.
    # SNAP, PINS, RDDT are ad-driven networks, not deep-tech; tech_growth (4.5×) is
    # more conservative than secular_growth_software (6×) for this stage.
    if t in {"SNAP", "PINS", "RDDT", "BMBL", "MTCH"}:
        return "tech_growth"
    # Default for other pre-profit companies
    return "secular_growth_deeptech"


# EV/Sales mature multiples — used when trailing EBITDA < 0 makes EM unreliable.
# Represents a normalised enterprise-value-to-revenue multiple at scale/maturity.
EVS_MATURE_MULTS = {
    "secular_growth_deeptech":  4.0,
    "secular_growth_software":  6.0,
    "secular_growth_resources": 2.5,
    "tech_growth":              4.5,
    "stable_compounder":        3.5,
    "cyclical":                 1.5,
}


# Per-metric thresholds: (HIGH, MOD-HIGH, MOD-LOW).  Values below MOD-LOW → LOW.
# d_ebitda is inverted (lower = better).
SECTOR_THRESHOLDS = {
    "tech_growth": {
        "rev_cagr":    (0.15, 0.10, 0.06),   # HIGH, MOD-HIGH, MOD-LOW
        "roic":        (0.25, 0.15, 0.08),
        "fcf_ni":      (0.80, 0.60, 0.40),
        "d_ebitda":    (1.0,  2.0,  3.5),     # lower = better (inverted)
    },
    "stable_compounder": {
        "rev_cagr":    (0.08, 0.05, 0.03),
        "roic":        (0.20, 0.12, 0.06),
        "fcf_ni":      (0.90, 0.70, 0.50),
        "d_ebitda":    (1.5,  3.0,  4.5),
    },
    "cyclical": {
        "rev_cagr":    (0.10, 0.05, 0.02),
        "roic":        (0.15, 0.10, 0.05),
        "fcf_ni":      (0.70, 0.50, 0.30),
        "d_ebitda":    (2.0,  3.5,  5.0),
    },
    "bank": {
        "rev_cagr":    (0.06, 0.03, 0.01),
        "roic":        (0.12, 0.08, 0.04),    # ROE for banks
        "fcf_ni":      (1.20, 0.90, 0.60),    # less meaningful for banks
        "d_ebitda":    (5.0,  8.0,  12.0),    # banks have high leverage by nature
    },
}


def _tier(value, thresholds, inverted=False):
    """Map a metric value to HIGH/MOD-HIGH/MOD-LOW/LOW based on thresholds."""
    h, mh, ml = thresholds
    if inverted:  # lower is better (e.g. D/EBITDA)
        if value <= h:  return "HIGH"
        if value <= mh: return "MOD-HIGH"
        if value <= ml: return "MOD-LOW"
        return "LOW"
    else:
        if value >= h:  return "HIGH"
        if value >= mh: return "MOD-HIGH"
        if value >= ml: return "MOD-LOW"
        return "LOW"


# Continuous scoring — eliminates threshold cliffs of the discrete tier system.
# Anchors: LOW=0, MOD-LOW=3 (at ml), MOD-HIGH=7 (at mh), HIGH=10 (at h).
# HIGH is the ceiling — no bonus above the top tier so each criterion's maximum
# contribution equals its stated weight, keeping the scorecard self-consistent.
SCORE_CAP = 10.0

def _score(value, thresholds, inverted=False):
    """Continuous piecewise-linear score [0, SCORE_CAP] from a metric value.

    Higher score = better. Used to drive scorecard totals so a 0.1pp move near
    a tier boundary no longer produces a 3-4pt swing in weighted score.
    Returns 0.0 if value is None.
    """
    if value is None:
        return 0.0
    h, mh, ml = thresholds

    if not inverted:
        # Higher is better (ROIC, FCF/NI, rev CAGR, etc.)
        if value >= h:
            extra = (value - h) / max(h - mh, 1e-9) * 3.0
            return min(SCORE_CAP, 10.0 + max(0.0, extra))
        if value >= mh:
            return 7.0 + (value - mh) / max(h - mh, 1e-9) * 3.0
        if value >= ml:
            return 3.0 + (value - ml) / max(mh - ml, 1e-9) * 4.0
        if value > 0 and ml > 0:
            return max(0.0, value / ml * 3.0)
        return 0.0

    # Inverted: lower is better (D/EBITDA — h is tightest, ml is loosest acceptable)
    if value <= h:
        extra = (h - value) / max(mh - h, 1e-9) * 3.0
        return min(SCORE_CAP, 10.0 + max(0.0, extra))
    if value <= mh:
        return 7.0 + (mh - value) / max(mh - h, 1e-9) * 3.0
    if value <= ml:
        return 3.0 + (ml - value) / max(ml - mh, 1e-9) * 4.0
    # Beyond MOD-LOW: linearly decay to 0 over the next 50% of the ml range
    if value <= 1.5 * ml:
        return max(0.0, (1.5 * ml - value) / max(0.5 * ml, 1e-9) * 3.0)
    return 0.0


def _proxy_score(n_pos, n_total):
    """Continuous score for proxy criteria (moat, mgmt, exec) based on indicator pass count.

    4/4 indicators → 10, 3/4 → 7.5, 2/4 → 5.0, 1/4 → 2.5, 0/4 → 0.
    Smoother than the discrete HIGH/MOD-HIGH/MOD-LOW/LOW mapping (10/7/3/0)
    that previously rewarded crossing arbitrary indicator-count thresholds.
    """
    if not n_total:
        return 0.0
    return round(n_pos / n_total * 10.0, 2)


def _exec_quality(series, expansion_threshold=0.05):
    """Score a metric series 0-3 on execution quality (direction + stability).

    Pure σ conflates two opposite cases: a company whose op margin expanded
    25% → 65% (exceptional execution) has the same σ profile as one whose
    margin oscillated unpredictably (poor execution). This function rewards
    directional improvement while still penalising erratic behaviour.

    Returns:
      3 (HIGH)     — expanding ≥ expansion_threshold AND recent trajectory
                     stronger than earlier years; OR rock-steady (|Δ| ≤ 2pp,
                     σ < 4%) which captures mature stable compounders.
      2 (MOD-HIGH) — broadly stable (no material decline) with σ < 4%.
      1 (MOD-LOW)  — modest decline (≤ 5pp) OR moderate volatility (σ < 8%).
      0 (LOW)      — significant decline OR high volatility.
    """
    vals = [v for v in series if v is not None]
    if len(vals) < 3:
        return 0

    earliest, latest = vals[0], vals[-1]
    direction = latest - earliest
    mu        = sum(vals) / len(vals)
    sigma     = (sum((v - mu) ** 2 for v in vals) / len(vals)) ** 0.5

    if len(vals) >= 4:
        recent_avg = sum(vals[-2:]) / 2
        prior_avg  = sum(vals[:-2]) / (len(vals) - 2)
        recent_improving = recent_avg >= prior_avg
    else:
        recent_improving = latest >= earliest

    # HIGH — meaningful expansion with sustained recent strength
    if direction >= expansion_threshold and recent_improving:
        return 3
    # HIGH — rock-steady mature business
    if abs(direction) <= 0.02 and sigma < 0.04:
        return 3
    # MOD-HIGH — broadly stable with minor noise, no material decline
    if direction >= -0.02 and sigma < 0.04:
        return 2
    # MOD-LOW — modest decline or moderate volatility
    if direction >= -0.05 and sigma < 0.08:
        return 1
    return 0


def _val_score(delta, premium_ok=False):
    """Continuous valuation score from delta vs benchmark.

    delta = (current_multiple - benchmark) / benchmark.
    Negative delta (cheap) → high score; positive (expensive) → low score.
    `premium_ok` (high ROIC + high growth) softens the penalty on premium names.
    """
    if delta is None:
        return None
    # Anchor points: -20% delta → 10, 0% → 7, +25% → 3, +50% → 0
    if delta <= -0.20:
        # Reward deep discounts up to -40% → SCORE_CAP
        extra = (-0.20 - delta) / 0.20 * 2.0
        return min(SCORE_CAP, 10.0 + max(0.0, extra))
    if delta <= 0.0:
        return 7.0 + (0.0 - delta) / 0.20 * 3.0
    if delta <= 0.25:
        return 3.0 + (0.25 - delta) / 0.25 * 4.0
    if delta <= 0.50:
        base = max(0.0, (0.50 - delta) / 0.25 * 3.0)
        # Premium partly justified by quality+growth → softer floor
        return base + (1.5 if premium_ok else 0.0)
    return 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# SCORECARD
# ═══════════════════════════════════════════════════════════════════════════════
def build_scorecard(wb, ticker, is_data, bs_data, cf_data, years,
                    biz_clarity=None, ltp=None, dcf_gg_price=None,
                    evs_regime=False, bank_credit=None, analyst_ests=None,
                    profile=None):
    """
    JS Scorecard tab — auto-scores 11 of 13 criteria.
    Quantitative: Revenue CAGR, FCF/NI, Capital Returns, ROIC, D/EBITDA, EBIT/Int
    Proxy-based:  Moat Profile, Management, Execution Risk, P/E vs Median, P/FCF vs Median
    Manual only:  Business Clarity (needs segment data), Long-Term Potential
    Scoring engine follows Master Prompt v2 thresholds.

    biz_clarity / ltp (optional): user-supplied tier values from the web form.
    When provided, they pre-fill the corresponding tier cells in the Excel
    scorecard so the workbook reflects what the HTML report shows.

    dcf_gg_price (optional): Gordon Growth fair value from build_dcf. When
    provided, used to compute DCF-implied P/E and P/FCF as a third valuation
    anchor (alongside 5yr historical avg and sector peer median). The lowest
    of the three becomes the benchmark — prevents post-2020 multiple expansion
    from making everything look "in line with history".
    """
    # is_bank is set later via profile fetch; initialise here so equity_assets
    # block (which runs before the fetch) can reference it safely.
    is_bank = False
    ws = wb.create_sheet("Scorecard")
    NC = 8   # columns A–H

    # ── Column widths ─────────────────────────────────────────────────────────
    for col, w in zip("ABCDEFGH", [44, 7, 9, 26, 13, 8, 12, 52]):
        ws.column_dimensions[col].width = w

    # ── Pre-calculate quantitative metrics ───────────────────────────────────

    # 1. Revenue 3yr CAGR
    rev_cagr = None
    if len(is_data) >= 4:
        r_now = is_data[-1].get("revenue") or 0
        r_3ya = is_data[-4].get("revenue") or 0
        if r_now and r_3ya > 0:
            rev_cagr = (r_now / r_3ya) ** (1 / 3) - 1

    # 2. FCF/NI series
    def _fcf(cf):
        v = cf.get("freeCashFlow")
        if v:
            return v
        ocf = cf.get("operatingCashFlow") or 0
        cap = abs(cf.get("capitalExpenditure") or 0)
        return ocf - cap

    fcf_ni_series = []
    for i in range(min(len(is_data), len(cf_data))):
        ni = is_data[i].get("netIncome") or 0
        fcf_ni_series.append(_fcf(cf_data[i]) / ni if ni else None)

    fcf_ni_latest = fcf_ni_series[-1] if fcf_ni_series else None
    fcf_ni_3ya    = fcf_ni_series[-4] if len(fcf_ni_series) >= 4 else None
    fcf_ni_trend  = (fcf_ni_latest is not None and fcf_ni_3ya is not None
                     and (fcf_ni_3ya - fcf_ni_latest) > 0.15)

    # SBC adjustment (Option B: display metric, no scoring impact)
    _sbc_raw     = (cf_data[-1].get("stockBasedCompensation") or 0) if cf_data else 0
    _fcf_raw_sbc = _fcf(cf_data[-1]) if cf_data else 0
    _fcf_ex_sbc  = _fcf_raw_sbc - _sbc_raw
    _sbc_pct_fcf = (_sbc_raw / _fcf_raw_sbc) if _fcf_raw_sbc > 0 else None

    # 3. ROIC series
    def _roic(is_, bs_):
        ebit    = abs(is_.get("operatingIncome") or 0)
        tax_e   = abs(is_.get("incomeTaxExpense") or 0)
        pretax  = abs(is_.get("incomeBeforeTax") or 1e-9)
        nopat   = ebit * (1 - min(tax_e / pretax, 0.50))
        equity  = bs_.get("totalStockholdersEquity") or 0
        debt    = (bs_.get("shortTermDebt") or 0) + (bs_.get("longTermDebt") or 0)
        cash    = bs_.get("cashAndCashEquivalents") or 0
        ic      = equity + debt - cash
        return (nopat / ic) if ic > 1 else None

    roic_series = [_roic(is_data[i], bs_data[i])
                   for i in range(min(len(is_data), len(bs_data)))]
    roic_latest = roic_series[-1] if roic_series else None
    roic_3ya    = roic_series[-4] if len(roic_series) >= 4 else None
    roic_trend  = (roic_latest is not None and roic_3ya is not None
                   and (roic_3ya - roic_latest) > 0.05)

    # 4. D/EBITDA and EBIT/Interest
    bs0 = bs_data[-1]; is0 = is_data[-1]; cf0 = cf_data[-1]
    total_debt  = (bs0.get("shortTermDebt") or 0) + (bs0.get("longTermDebt") or 0)
    cash0       = bs0.get("cashAndCashEquivalents") or 0
    net_cash_v  = cash0 - total_debt

    ebitda0 = is0.get("ebitda") or 0
    if not ebitda0:
        da = abs(is0.get("depreciationAndAmortization") or
                 cf0.get("depreciationAndAmortization") or 0)
        ebitda0 = (is0.get("operatingIncome") or 0) + da
    d_ebitda = total_debt / ebitda0 if ebitda0 > 0 else None

    ebit0   = abs(is0.get("operatingIncome") or 0)
    int_exp = abs(is0.get("interestExpense") or 0)
    ebit_int = ebit0 / int_exp if int_exp > 0 else None

    # 4b. Bank capital adequacy — computed after profile fetch (is_bank set there)
    equity_assets = None   # populated below once is_bank is confirmed

    # 5. Capital Returns
    def _ret(cf):
        return (abs(cf.get("commonStockRepurchased") or
                    cf.get("stockRepurchase") or 0) +
                abs(cf.get("dividendsPaid") or 0))

    tot_ret       = _ret(cf0)
    ret_yrs_cnt   = sum(1 for cf_ in cf_data if _ret(cf_) > 0)
    debt_prior    = ((bs_data[-2].get("shortTermDebt") or 0) +
                     (bs_data[-2].get("longTermDebt") or 0)) if len(bs_data) >= 2 else total_debt
    debt_funded   = total_debt > debt_prior * 1.05 and tot_ret > 0

    # 6. Gross / operating margin series (used for moat + management proxies)
    rev_series = [is_.get("revenue") or 0 for is_ in is_data]
    gm_series  = [((is_.get("grossProfit") or 0) / rev if rev else None)
                  for is_, rev in zip(is_data, rev_series)]
    om_series  = [(abs(is_.get("operatingIncome") or 0) / rev if rev else None)
                  for is_, rev in zip(is_data, rev_series)]

    gm_latest    = gm_series[-1]
    gm_3yr_delta = ((gm_series[-1] - gm_series[-4])
                    if len(gm_series) >= 4 and gm_series[-1] and gm_series[-4]
                    else None)
    om_latest    = om_series[-1]
    om_3yr_delta = ((om_series[-1] - om_series[-4])
                    if len(om_series) >= 4 and om_series[-1] and om_series[-4]
                    else None)

    # Revenue growth std dev (for moat consistency + execution risk)
    rev_growths = [(rev_series[i] / rev_series[i - 1] - 1)
                   for i in range(1, len(rev_series))
                   if rev_series[i - 1] and rev_series[i]]
    if len(rev_growths) > 1:
        mu_rg  = sum(rev_growths) / len(rev_growths)
        rev_std = (sum((g - mu_rg) ** 2 for g in rev_growths) / len(rev_growths)) ** 0.5
    else:
        rev_std = 0.0

    # Op margin std dev (for execution risk)
    om_valid = [o for o in om_series if o is not None]
    if len(om_valid) > 1:
        mu_om  = sum(om_valid) / len(om_valid)
        om_std = (sum((o - mu_om) ** 2 for o in om_valid) / len(om_valid)) ** 0.5
    else:
        om_std = 0.0

    # ── FMP ratios: P/E, P/FCF — current + 5yr historical average ───────────
    yf_info          = {}   # kept for beta fallback in moat section
    trailing_pe      = None
    forward_pe       = None
    trailing_pfcf    = None
    pe_5yr_avg       = None
    _ttm_ni_live     = None   # set in ratios block if quarterly IS fetch succeeds
    pfcf_5yr_avg     = None
    ev_ebitda_5yr_avg = None
    sector_pe_med    = None
    sector_pfcf_med  = None

    try:
        rat_data = _fetch_ratios(ticker, limit=5)   # cached — shared with build_dcf()
        if isinstance(rat_data, list) and rat_data:
            pe_vals   = [r["priceToEarningsRatio"]    for r in rat_data
                         if r.get("priceToEarningsRatio")    and r["priceToEarningsRatio"]    > 0]
            pfcf_vals = [r["priceToFreeCashFlowRatio"] for r in rat_data
                         if r.get("priceToFreeCashFlowRatio") and r["priceToFreeCashFlowRatio"] > 0]
            ev_vals   = [r["enterpriseValueMultiple"]  for r in rat_data
                         if r.get("enterpriseValueMultiple")  and 2 < r["enterpriseValueMultiple"] < 200]
            _fmp_trailing_pe = round(rat_data[0].get("priceToEarningsRatio")    or 0, 1) or None
            trailing_pfcf = round(rat_data[0].get("priceToFreeCashFlowRatio") or 0, 1) or None
            pe_5yr_avg    = round(sum(pe_vals)   / len(pe_vals),   1) if len(pe_vals)   > 1 else None
            pfcf_5yr_avg  = round(sum(pfcf_vals) / len(pfcf_vals), 1) if len(pfcf_vals) > 1 else None
            ev_ebitda_5yr_avg = round(sum(ev_vals) / len(ev_vals), 1) if len(ev_vals) > 1 else None

            # ── TTM P/E: live quarterly income statement  (1 FMP call, cached) ──
            # FMP's annual ratio snapshot uses price-at-FY-end ÷ FY-EPS, which can
            # diverge materially from current P/E after stock splits or large price
            # moves. True TTM = current_mktcap / sum(last 4 quarterly net incomes).
            _ttm_ni_live = _fetch_ttm_ni(ticker)
            if _ttm_ni_live and _ttm_ni_live > 0:
                _sc_price_early = float(prof_sc.get("price") or 0) or None
                _shs_early = (is_data[-1].get("weightedAverageShsOut") or
                              is_data[-1].get("weightedAverageShsOutDil") or 0)
                if _sc_price_early and _shs_early > 0:
                    _mktcap_live = _sc_price_early * _shs_early
                    _ttm_pe_live = round(_mktcap_live / _ttm_ni_live, 1)
                    trailing_pe  = _ttm_pe_live
                    print(f"  TTM P/E (live): price={_sc_price_early:.2f}  "
                          f"ttm_ni=${_ttm_ni_live/1e9:.2f}B  pe={_ttm_pe_live:.1f}x  "
                          f"(FMP snapshot was {_fmp_trailing_pe}x)")
                else:
                    trailing_pe = _fmp_trailing_pe
            else:
                trailing_pe = _fmp_trailing_pe

            print(f"  FMP ratios: P/E={trailing_pe}  5yr avg={pe_5yr_avg}  "
                  f"P/FCF={trailing_pfcf}  5yr avg={pfcf_5yr_avg}  "
                  f"EV/EBITDA 5yr avg={ev_ebitda_5yr_avg}")
    except Exception as e_rat:
        print(f"  FMP ratios fetch failed: {e_rat}")

    # Sector peer P/E and P/FCF medians (FMP ratios, latest year only)
    prof_sc    = {}
    sector_str = ""
    is_bank    = False
    try:
        # Sector for is_bank detection — reuse caller's profile when present
        # (saves 1 FMP call per report).  Only fetch fresh if caller didn't
        # provide one (legacy call sites or tests).
        prof_sc = profile or {}
        if not prof_sc:
            try:
                p_sc = requests.get(
                    f"https://financialmodelingprep.com/stable/profile"
                    f"?symbol={ticker}&apikey={API_KEY}", timeout=8
                ).json()
                prof_sc = (p_sc[0] if isinstance(p_sc, list) and p_sc
                           else p_sc if isinstance(p_sc, dict) else {})
            except Exception:
                pass
        sector_str = prof_sc.get("industry") or prof_sc.get("sector") or ""
        # Bank/financial sector detection — D/EBITDA is meaningless for deposit-funded institutions
        _BANK_KW = {"bank", "banking", "financial services", "savings", "thrift",
                    "mortgage", "credit union", "investment bank", "diversified financial"}
        is_bank = any(kw in sector_str.lower() for kw in _BANK_KW)
        print(f"  Sector: {sector_str!r}  is_bank={is_bank}")
        # Equity/Assets (CET1 proxy) — computed here so is_bank is already confirmed
        if is_bank:
            total_equity = (bs0.get("totalStockholdersEquity") or
                            bs0.get("totalEquity") or 0)
            total_assets = bs0.get("totalAssets") or 0
            if total_assets > 0 and total_equity > 0:
                equity_assets = total_equity / total_assets
        peer_list_sc = []
        for key, peers in SECTOR_PEERS.items():
            if key.lower() in sector_str.lower() or sector_str.lower() in key.lower():
                peer_list_sc = [p for p in peers if p != ticker]
                break
        peer_pes = []; peer_pfcfs = []
        for peer in peer_list_sc[:4]:
            try:
                pr = requests.get(
                    f"https://financialmodelingprep.com/stable/ratios"
                    f"?symbol={peer}&limit=1&apikey={API_KEY}", timeout=8
                ).json()
                if isinstance(pr, list) and pr:
                    pe = pr[0].get("priceToEarningsRatio")
                    pf = pr[0].get("priceToFreeCashFlowRatio")
                    if pe and 0 < pe < 300: peer_pes.append(pe)
                    if pf and 0 < pf < 300: peer_pfcfs.append(pf)
            except Exception:
                pass
        if peer_pes:
            sector_pe_med   = round(sorted(peer_pes)[len(peer_pes) // 2], 1)
        if peer_pfcfs:
            sector_pfcf_med = round(sorted(peer_pfcfs)[len(peer_pfcfs) // 2], 1)
        print(f"  Sector peer P/E median={sector_pe_med}  P/FCF median={sector_pfcf_med}")
    except Exception as e_peers:
        print(f"  Sector peer fetch failed: {e_peers}")

    # Bank-specific: ROE and NIM series.
    # ROIC is meaningless for deposit-funded institutions — deposit liabilities
    # inflate invested capital by trillions, producing near-zero ROIC even for
    # highly profitable banks. Substitute ROE (netIncome / avg equity) throughout
    # all scoring that references roic_latest / roic_3ya / roic_series.
    _roe_series = []; _nim_series = []
    if is_bank:
        for _i in range(min(len(is_data), len(bs_data))):
            _ni_b  = is_data[_i].get("netIncome") or 0
            _eq_b  = (bs_data[_i].get("totalStockholdersEquity") or
                      bs_data[_i].get("totalEquity") or 0)
            _ta_b  = bs_data[_i].get("totalAssets") or 0
            _nii_b = is_data[_i].get("netInterestIncome") or 0
            _roe_series.append(_ni_b / _eq_b if _eq_b > 0 else None)
            _nim_series.append(_nii_b / _ta_b if _ta_b > 0 else None)
        roic_series = _roe_series
        roic_latest = roic_series[-1] if roic_series else None
        roic_3ya    = roic_series[-4] if len(roic_series) >= 4 else None
        roic_trend  = (roic_latest is not None and roic_3ya is not None
                       and (roic_3ya - roic_latest) > 0.05)
        if roic_latest is not None:
            print(f"  Bank ROE (replaces ROIC): latest={roic_latest:.1%}")
    _nim_latest    = _nim_series[-1] if _nim_series else None
    _nim_3yr_ago   = _nim_series[-4] if len(_nim_series) >= 4 else None
    _nim_3yr_delta = ((_nim_latest - _nim_3yr_ago)
                      if (_nim_latest is not None and _nim_3yr_ago is not None) else None)

    # yfinance: beta only (for rough WACC in moat proxy) — graceful fallback
    try:
        import yfinance as yf
        yf_info = yf.Ticker(ticker).info or {}
    except Exception:
        yf_info = {}

    # ── Scoring helpers ───────────────────────────────────────────────────────
    TIER_ORDER = ["LOW", "MOD-LOW", "MOD-HIGH", "HIGH"]
    TIER_SCORE = {"HIGH": 10, "MOD-HIGH": 7, "MOD-LOW": 3, "LOW": 0}

    def down_tier(t):
        i = TIER_ORDER.index(t)
        return TIER_ORDER[max(i - 1, 0)]

    # Determine sector bucket for threshold lookups
    _bucket = _sector_bucket(sector_str, ticker)
    _thresholds = SECTOR_THRESHOLDS[_bucket]
    print(f"  Sector bucket: {_bucket!r}  (thresholds: {_thresholds})")

    # Continuous scoring penalty for trend-deterioration flags.
    # ≈ 1 tier drop in the discrete system (4-pt average gap), preserves the
    # original behavioural intent without re-introducing cliff edges.
    _TREND_PENALTY = 3.5

    def _t_rev(v):
        if v is None:
            return None, 0.0, "N/A — insufficient data"
        t = _tier(v, _thresholds["rev_cagr"])
        s = _score(v, _thresholds["rev_cagr"])
        return t, s, f"{v:.1%}"

    def _t_fcf(v, pen):
        if v is None:
            return None, 0.0, "N/A — insufficient data"
        v2 = abs(v)
        t  = _tier(v2, _thresholds["fcf_ni"])
        s  = _score(v2, _thresholds["fcf_ni"])
        note = f"{v:.0%}"
        if pen:
            t = down_tier(t)
            s = max(0.0, s - _TREND_PENALTY)
            note += "  [trend penalty: declined >15pp vs 3yr ago]"
        return t, s, note

    # Capital returns: discrete logic (no underlying continuous metric to interpolate).
    # Map tier → score using the same 4-anchor scale as continuous helpers.
    _CAP_RET_SCORE = {"HIGH": 10.0, "MOD-HIGH": 7.0, "MOD-LOW": 3.0, "LOW": 0.0}

    def _t_ret(tot, yrs, df):
        if tot == 0:
            return "LOW", 0.0, "No capital returns in latest year"
        s = f"${tot / 1e6:,.0f}mm latest FY"
        if yrs < 3 or df:
            r = "debt-funded" if df else f"only {yrs}/{len(cf_data)}yr history"
            tier_v, score_v, note_v = "MOD-LOW", 3.0, f"{s} — {r}"
        elif yrs < 5:
            tier_v, score_v, note_v = "MOD-HIGH", 7.0, f"{s} — {yrs}yr equity-funded"
        else:
            tier_v, score_v, note_v = "HIGH", 10.0, f"{s} — {yrs}yr+ consistent equity-funded"

        # FCF payout ratio (#5) — discipline check on top of consistency.
        # Sustained payouts above 100% of FCF are funded by debt or cash burn,
        # not by recurring cash generation; healthy range is roughly 40–80%.
        # Buyback ROI vs current price would also belong here, but is deferred
        # because it requires fetching historical-price-full from FMP (extra
        # daily-quota cost). Existing FMP fields are sufficient for payout ratio.
        fcf0 = cf0.get("freeCashFlow")
        if not fcf0:
            ocf0 = cf0.get("operatingCashFlow") or 0
            cap0 = abs(cf0.get("capitalExpenditure") or 0)
            fcf0 = ocf0 - cap0
        if fcf0 and fcf0 > 0:
            payout = tot / fcf0
            if payout > 1.0:
                score_v = max(0.0, score_v - 2.5)
                note_v += f"  [payout {payout:.0%} of FCF — > 100%, unsustainable]"
            elif payout < 0.20 and tot > 0:
                note_v += f"  [payout {payout:.0%} of FCF — low, building cash]"
            else:
                note_v += f"  [payout {payout:.0%} of FCF]"
        elif fcf0 is not None and fcf0 <= 0:
            score_v = max(0.0, score_v - 2.5)
            note_v += "  [FCF negative — capital returns funded externally]"
        return tier_v, score_v, note_v

    def _t_roic(v, pen):
        if v is None:
            return None, 0.0, "N/A — insufficient data"
        t = _tier(v, _thresholds["roic"])
        s = _score(v, _thresholds["roic"])
        note = f"{v:.1%}"
        if pen:
            t = down_tier(t)
            s = max(0.0, s - _TREND_PENALTY)
            note += "  [trend penalty: declined >5pp vs 3yr ago]"
        return t, s, note

    def _t_de(de, nc):
        if nc > 0:
            # Net cash → cap-out score (treated as "better than HIGH" — outlier reward)
            return "HIGH", SCORE_CAP, f"Net cash ${nc / 1e6:,.0f}mm — no net leverage"
        if de is None:
            return None, 0.0, "N/A"
        t = _tier(de, _thresholds["d_ebitda"], inverted=True)
        s = _score(de, _thresholds["d_ebitda"], inverted=True)
        return t, s, f"{de:.1f}x"

    # EBIT/Interest thresholds (h, mh, ml) — higher = better (not inverted).
    _EBIT_INT_THR = (10.0, 4.0, 2.0)

    def _t_ei(v):
        if v is None:
            return "HIGH", SCORE_CAP, "No interest expense — debt-free"
        t = _tier(v, _EBIT_INT_THR)
        s = _score(v, _EBIT_INT_THR)
        return t, s, f"{v:.1f}x"

    # Equity/Assets thresholds for banks (h, mh, ml).
    _EQUITY_ASSETS_THR = (0.10, 0.08, 0.06)

    def _t_equity_assets(v):
        """Capital adequacy (Equity/Assets) for banks — proxy for CET1 ratio.
        Regulators target >8% (minimum) to >10-12% (well-capitalised)."""
        if v is None:
            return None, 0.0, "N/A — insufficient data"
        t = _tier(v, _EQUITY_ASSETS_THR)
        s = _score(v, _EQUITY_ASSETS_THR)
        return t, s, f"{v:.1%}  [Equity/Assets — CET1 proxy; well-capitalised >10%]"

    # ── Quality-of-earnings checks (#4) ───────────────────────────────────────
    # Two independent red flags that catch aggressive-accounting names without
    # any extra API call. Penalties subtract from the continuous score (not a
    # tier down-shift) so they compose cleanly with the trend penalties.
    #   • Sloan accruals: 3yr avg (NI − CFO) / Total Assets > 10% → high
    #     accruals = earnings driven by non-cash items. Long-documented anomaly
    #     (Sloan 1996); persistently high accruals = future earnings reversal.
    #   • CFO/NI 3yr avg < 80% → cash conversion lagging reported earnings;
    #     either revenue rec is aggressive or working capital is bloating.
    sloan_series = []
    cfo_ni_series_q = []
    for i in range(min(len(is_data), len(bs_data), len(cf_data))):
        ni_i  = is_data[i].get("netIncome") or 0
        cfo_i = cf_data[i].get("operatingCashFlow") or 0
        ta_i  = bs_data[i].get("totalAssets") or 0
        if ta_i > 0 and ni_i:
            sloan_series.append((ni_i - cfo_i) / ta_i)
        if ni_i:
            cfo_ni_series_q.append(cfo_i / ni_i)
    sloan_3yr_avg  = (sum(sloan_series[-3:]) / len(sloan_series[-3:])
                      if len(sloan_series) >= 1 else None)
    cfo_ni_3yr_avg = (sum(cfo_ni_series_q[-3:]) / len(cfo_ni_series_q[-3:])
                      if len(cfo_ni_series_q) >= 1 else None)

    qoe_fcf_ni_penalty = 0.0
    qoe_fcf_ni_note    = ""
    if sloan_3yr_avg is not None and sloan_3yr_avg > 0.10:
        qoe_fcf_ni_penalty = 2.0
        qoe_fcf_ni_note = f"  [QoE: Sloan accruals {sloan_3yr_avg:+.1%} 3yr avg — high]"
    qoe_roic_penalty = 0.0
    qoe_roic_note    = ""
    if cfo_ni_3yr_avg is not None and cfo_ni_3yr_avg < 0.80:
        qoe_roic_penalty = 1.5
        qoe_roic_note = f"  [QoE: CFO/NI {cfo_ni_3yr_avg:.0%} 3yr avg — cash conv weak]"

    # ── Through-cycle normalization (#2) ──────────────────────────────────────
    # Latest-year ROIC / FCF-NI / D-EBITDA can flatter cyclicals at the peak
    # (XOM 2022, F 2021) and panic on them at the trough. Smoothing brings the
    # scoring closer to mid-cycle reality.
    #   • cyclical bucket → 5yr median (full through-cycle view)
    #   • all other buckets → 70/30 blend of latest + 5yr median (mostly latest,
    #     with a modest pull toward trend so a single great/bad year doesn't
    #     dominate the score). Trend penalties run on top of the smoothed value.
    def _median(xs):
        ys = sorted(x for x in xs if x is not None)
        if not ys: return None
        n = len(ys)
        return ys[n // 2] if n % 2 else (ys[n // 2 - 1] + ys[n // 2]) / 2

    def _smooth(latest, series):
        med = _median(series)
        if latest is None: return med
        if med is None:    return latest
        if _bucket == "cyclical":
            return med
        return latest * 0.7 + med * 0.3

    # Build d_ebitda series (only latest was computed above)
    d_ebitda_series = []
    for i in range(min(len(is_data), len(bs_data), len(cf_data))):
        _bsi = bs_data[i]; _isi = is_data[i]; _cfi = cf_data[i]
        _td = (_bsi.get("shortTermDebt") or 0) + (_bsi.get("longTermDebt") or 0)
        _ed = _isi.get("ebitda") or 0
        if not _ed:
            _da = abs(_isi.get("depreciationAndAmortization") or
                      _cfi.get("depreciationAndAmortization") or 0)
            _ed = (_isi.get("operatingIncome") or 0) + _da
        d_ebitda_series.append(_td / _ed if _ed > 0 else None)

    rev_cagr_n      = rev_cagr  # rev CAGR is already a multi-year metric
    fcf_ni_for_t    = _smooth(fcf_ni_latest, fcf_ni_series)
    roic_for_t      = _smooth(roic_latest,   roic_series)
    d_ebitda_for_t  = _smooth(d_ebitda,      d_ebitda_series)
    if _bucket == "cyclical" and (fcf_ni_for_t != fcf_ni_latest or
                                  roic_for_t != roic_latest or
                                  d_ebitda_for_t != d_ebitda):
        print(f"  Through-cycle smoothing (cyclical):  ROIC {roic_latest} → {roic_for_t}  |  "
              f"FCF/NI {fcf_ni_latest} → {fcf_ni_for_t}  |  D/EBITDA {d_ebitda} → {d_ebitda_for_t}")

    tier_rev_cagr,  score_rev_cagr,  note_rev_cagr  = _t_rev(rev_cagr_n)
    tier_fcf_ni,    score_fcf_ni,    note_fcf_ni    = _t_fcf(fcf_ni_for_t, fcf_ni_trend)
    tier_cap_ret,   score_cap_ret,   note_cap_ret   = _t_ret(tot_ret, ret_yrs_cnt, debt_funded)
    tier_roic,      score_roic,      note_roic      = _t_roic(roic_for_t, roic_trend)
    tier_d_ebitda,  score_d_ebitda,  note_d_ebitda  = _t_de(d_ebitda_for_t, net_cash_v)
    tier_ebit_int,  score_ebit_int,  note_ebit_int  = _t_ei(ebit_int)

    # Annotate through-cycle smoothing in notes when it materially shifted the input
    def _smooth_note(latest, smoothed):
        if latest is None or smoothed is None: return ""
        if abs(smoothed - latest) < 0.005:     return ""
        tag = "5yr median" if _bucket == "cyclical" else "70/30 latest+5yr median"
        return f"  [through-cycle: {tag}, latest {latest:.1%} → smoothed {smoothed:.1%}]"
    if isinstance(roic_latest, float):
        note_roic   += _smooth_note(roic_latest,    roic_for_t)
    if isinstance(fcf_ni_latest, float):
        note_fcf_ni += _smooth_note(fcf_ni_latest,  fcf_ni_for_t)
    if isinstance(d_ebitda, float) and isinstance(d_ebitda_for_t, float):
        if abs(d_ebitda - d_ebitda_for_t) >= 0.05:
            tag = "5yr median" if _bucket == "cyclical" else "70/30 latest+5yr median"
            note_d_ebitda += f"  [through-cycle: {tag}, latest {d_ebitda:.1f}x → smoothed {d_ebitda_for_t:.1f}x]"

    # Apply QoE penalties (#4) on top of the tier-derived scores.
    if qoe_fcf_ni_penalty and score_fcf_ni:
        score_fcf_ni = max(0.0, score_fcf_ni - qoe_fcf_ni_penalty)
        note_fcf_ni += qoe_fcf_ni_note
    if qoe_roic_penalty and score_roic:
        score_roic = max(0.0, score_roic - qoe_roic_penalty)
        note_roic  += qoe_roic_note

    # Dynamic leverage criterion — bank-aware
    if is_bank:
        _ea_t, _ea_s, _ea_n = _t_equity_assets(equity_assets)
        tier_leverage  = _ea_t
        score_leverage = _ea_s
        note_leverage  = _ea_n
        leverage_label = "Capital Adequacy  (Equity / Assets)"
    else:
        tier_leverage  = tier_d_ebitda
        score_leverage = score_d_ebitda
        note_leverage  = note_d_ebitda
        leverage_label = "Credit Risk  (D / EBITDA)"

    # ── Moat proxy (4 indicators → tier) ─────────────────────────────────────
    # Rough WACC for ROIC spread — use FMP profile beta (already fetched above)
    beta_yf   = float(prof_sc.get("beta") or 1.0) or 1.0
    avg_erp   = (DAMODARAN_ERP_IMPLIED + DAMODARAN_ERP_HIST_AVG) / 2
    rough_re  = 0.043 + beta_yf * avg_erp
    avg_debt  = (total_debt + debt_prior) / 2 if debt_prior else total_debt
    tax_r_sc  = min(abs(is0.get("incomeTaxExpense") or 0) /
                    abs(is0.get("incomeBeforeTax") or 1), 0.50)
    rough_rd  = int_exp / avg_debt if avg_debt > 0 else 0.05
    mktcap_sc = yf_info.get("marketCap") or 0
    E_sc = mktcap_sc / 1e6; D_sc = total_debt / 1e6; V_sc = E_sc + D_sc
    w_e_sc = E_sc / V_sc if V_sc > 0 else 0.8
    w_d_sc = D_sc / V_sc if V_sc > 0 else 0.2
    rough_wacc = w_e_sc * rough_re + w_d_sc * rough_rd * (1 - tax_r_sc)

    moat_ind = []; moat_parts = []; moat_total = 0
    if is_bank:
        # Bank moat: NIM viability, NIM stability, ROE vs cost of equity, revenue consistency.
        # Gross margin is structurally near-zero for deposit-funded institutions and is
        # not a meaningful moat indicator — NIM and ROE>CoE are the bank equivalents.
        if _nim_latest is not None:
            ok = _nim_latest > 0.025
            if ok: moat_ind.append(True)
            moat_total += 1
            moat_parts.append(f"NIM {_nim_latest:.2%} {'✓' if ok else '✗'} (>2.5%)")
        if _nim_3yr_delta is not None:
            ok = _nim_3yr_delta >= -0.002
            if ok: moat_ind.append(True)
            moat_total += 1
            moat_parts.append(f"NIM trend {_nim_3yr_delta*100:+.1f}bps {'✓' if ok else '✗'} (>=-20bps)")
        if roic_latest is not None:  # roic_latest = ROE for banks
            spread = roic_latest - rough_re  # ROE vs cost of equity
            ok = spread > 0.02
            if ok: moat_ind.append(True)
            moat_total += 1
            moat_parts.append(f"ROE {roic_latest:.1%} vs CoE {rough_re:.1%} spread {spread:+.1%} {'✓' if ok else '✗'} (>+2pp)")
        ok_std = rev_std < 0.08
        if ok_std: moat_ind.append(True)
        moat_total += 1
        moat_parts.append(f"Rev consistency σ={rev_std:.1%} {'✓' if ok_std else '✗'} (<8%)")
        # NCO rate level check — 3yr average for structural credit quality of loan book.
        # 3yr avg smooths one-off spikes and reflects underwriting standard across a rate cycle.
        _nco_rate = (bank_credit or {}).get("nco_3yr_avg")
        if _nco_rate is not None:
            ok = _nco_rate < 0.010  # 3yr avg NCO rate < 1.0%
            if ok: moat_ind.append(True)
            moat_total += 1
            moat_parts.append(f"NCO 3yr avg {_nco_rate:.2%} {'pass' if ok else 'fail'} (<1.0%)")
    else:
        if gm_latest is not None:
            ok = gm_latest > 0.40
            if ok: moat_ind.append(True)
            moat_total += 1
            moat_parts.append(f"GM {gm_latest:.1%} {'✓' if ok else '✗'} (>40%)")
        if gm_3yr_delta is not None:
            ok = gm_3yr_delta > 0.01
            if ok: moat_ind.append(True)
            moat_total += 1
            moat_parts.append(f"GM trend {gm_3yr_delta:+.1%} {'✓' if ok else '✗'} (>+1pp)")
        if roic_latest is not None:
            spread = roic_latest - rough_wacc
            ok = spread > 0.05
            if ok: moat_ind.append(True)
            moat_total += 1
            moat_parts.append(f"ROIC-WACC {spread:+.1%} {'✓' if ok else '✗'} (>+5pp)")
        ok_std = rev_std < 0.08
        if ok_std: moat_ind.append(True)
        moat_total += 1
        moat_parts.append(f"Rev consistency σ={rev_std:.1%} {'✓' if ok_std else '✗'} (<8%)")
    n_moat = len(moat_ind)
    tier_moat = ("HIGH" if n_moat >= 4 else "MOD-HIGH" if n_moat == 3
                 else "MOD-LOW" if n_moat == 2 else "LOW")
    score_moat = _proxy_score(n_moat, moat_total)
    note_moat = "  |  ".join(moat_parts) + f"  [{n_moat}/{moat_total} indicators positive — proxy score]"

    # ── Management proxy (4 indicators → tier) ───────────────────────────────
    mgmt_ind = []; mgmt_parts = []; mgmt_total = 0
    if roic_latest is not None and roic_3ya is not None:
        chg = roic_latest - roic_3ya
        ok  = chg >= -0.02
        if ok: mgmt_ind.append(True)
        mgmt_total += 1
        _trend_label = "ROE trend" if is_bank else "ROIC trend"
        mgmt_parts.append(f"{_trend_label} {chg:+.1%} {'✓' if ok else '✗'} (>=-2pp)")
    elif roic_latest is not None:
        _metric_label = "ROE" if is_bank else "ROIC"
        mgmt_parts.append(f"{_metric_label} {roic_latest:.1%} (no trend data)")
    if is_bank:
        # Replace gross-margin check with NIM stability — the bank equivalent of
        # "are margins being defended?" GM is structurally near-zero for banks.
        if _nim_3yr_delta is not None:
            ok = _nim_3yr_delta >= -0.002
            if ok: mgmt_ind.append(True)
            mgmt_total += 1
            mgmt_parts.append(f"NIM maintained {_nim_3yr_delta*100:+.1f}bps {'✓' if ok else '✗'} (>=-20bps)")
        # NCO trend check — rising charge-offs signal deteriorating underwriting discipline
        _nco_trend = (bank_credit or {}).get("nco_2yr_delta")
        if _nco_trend is not None:
            ok = _nco_trend < 0.003  # rising < 30bps over 2yr
            if ok: mgmt_ind.append(True)
            mgmt_total += 1
            trend_pp = _nco_trend * 100
            mgmt_parts.append(f"NCO trend {trend_pp:+.2f}pp {'pass' if ok else 'fail'} (<+30bps/2yr)")
    else:
        if gm_3yr_delta is not None:
            ok = gm_3yr_delta >= -0.01
            if ok: mgmt_ind.append(True)
            mgmt_total += 1
            mgmt_parts.append(f"GM maintained {gm_3yr_delta:+.1%} {'✓' if ok else '✗'}")
    if om_3yr_delta is not None:
        ok = om_3yr_delta >= -0.02
        if ok: mgmt_ind.append(True)
        mgmt_total += 1
        mgmt_parts.append(f"Op margin {om_3yr_delta:+.1%} {'✓' if ok else '✗'} (≥-2pp)")
    ok_ret = tier_cap_ret in ("HIGH", "MOD-HIGH")
    if ok_ret: mgmt_ind.append(True)
    mgmt_total += 1
    mgmt_parts.append(f"Capital returns {tier_cap_ret or 'N/A'} {'✓' if ok_ret else '✗'}")

    # Share count change over 5yr (#6) — dilution flag.
    # Net dilution >2%/yr suggests SBC + acquisitions overwhelming buybacks.
    # Net buyback (negative dilution) is a positive capital allocation signal.
    sh_old = is_data[0].get("weightedAverageShsOut") or is_data[0].get("weightedAverageShsOutDil")
    sh_new = is_data[-1].get("weightedAverageShsOut") or is_data[-1].get("weightedAverageShsOutDil")
    if sh_old and sh_new and sh_old > 0:
        n_yrs = max(len(is_data) - 1, 1)
        sh_cagr = (sh_new / sh_old) ** (1 / n_yrs) - 1
        ok_sh = sh_cagr <= 0.02
        if ok_sh: mgmt_ind.append(True)
        mgmt_total += 1
        if sh_cagr < 0:
            mgmt_parts.append(f"Share count {sh_cagr*100:+.1f}%/yr ✓ (net buyback)")
        else:
            mgmt_parts.append(f"Share count {sh_cagr*100:+.1f}%/yr {'✓' if ok_sh else '✗'} (≤+2%/yr)")

    # Goodwill / equity (#6) — M&A discipline check.
    # >40% goodwill on equity = balance sheet built by acquisition; pair with
    # flat/declining ROIC = capital allocation red flag.
    gw_latest = bs_data[-1].get("goodwill") or 0
    eq_latest = bs_data[-1].get("totalStockholdersEquity") or bs_data[-1].get("totalEquity") or 0
    if gw_latest and eq_latest and eq_latest > 0:
        gw_eq = gw_latest / eq_latest
        ok_gw = gw_eq < 0.40
        if ok_gw: mgmt_ind.append(True)
        mgmt_total += 1
        mgmt_parts.append(f"Goodwill/Equity {gw_eq:.0%} {'✓' if ok_gw else '✗'} (<40%)")
    elif eq_latest > 0:
        # Zero / no goodwill is a positive (organic balance sheet)
        mgmt_ind.append(True); mgmt_total += 1
        mgmt_parts.append("Goodwill/Equity 0% ✓ (organic)")

    n_mgmt = len(mgmt_ind)
    # Scale tier thresholds to mgmt_total (now 4-6 with #6 additions): HIGH ≥85%
    # of indicators, MOD-HIGH ≥60%, MOD-LOW ≥35%, else LOW.
    if mgmt_total > 0:
        _frac = n_mgmt / mgmt_total
        tier_mgmt = ("HIGH" if _frac >= 0.85 else "MOD-HIGH" if _frac >= 0.60
                     else "MOD-LOW" if _frac >= 0.35 else "LOW")
    else:
        tier_mgmt = "LOW"
    score_mgmt = _proxy_score(n_mgmt, mgmt_total)
    note_mgmt = "  |  ".join(mgmt_parts) + f"  [{n_mgmt}/{mgmt_total} indicators positive — proxy score]"

    # ── Execution Risk proxy (revenue stability + margin trajectory → tier) ──
    # Revenue: σ-based (lumpiness is the right signal for revenue execution —
    # mature businesses should have predictable demand year-over-year).
    rev_risk_idx = (3 if rev_std < 0.05 else 2 if rev_std < 0.10
                    else 1 if rev_std < 0.18 else 0)
    # Op margin: direction-aware quality (not σ alone). A company expanding
    # margins 25% → 65% has high σ but is *exceptional* execution, not poor.
    # _exec_quality() rewards directional improvement and recent strength;
    # σ-only would mis-flag the inflection as instability.
    om_quality   = _exec_quality(om_valid)
    # Legacy σ-bucket kept only for the note string (transparency).
    om_risk_idx  = (3 if om_std < 0.02 else 2 if om_std < 0.04
                    else 1 if om_std < 0.08 else 0)

    _om_first  = om_valid[0]  if om_valid else None
    _om_latest = om_valid[-1] if om_valid else None
    _om_delta  = ((_om_latest - _om_first)
                  if (_om_first is not None and _om_latest is not None) else None)
    _delta_str = f"{_om_delta:+.1%}" if _om_delta is not None else "N/A"

    # Hypergrowth (3yr CAGR > 40%): rev σ is structurally high from the ramp
    # itself, not from operational issues. Score on margin trajectory only.
    _high_growth_exec = rev_cagr is not None and rev_cagr > 0.40
    if _high_growth_exec:
        exec_idx  = om_quality
        note_exec = (f"Rev CAGR={rev_cagr:.0%} >40% — margin trajectory only  |  "
                     f"OM trend {_delta_str} ({_om_first:.1%}→{_om_latest:.1%}), "
                     f"σ={om_std:.1%}  [quality = direction + stability]")
    else:
        exec_idx  = (rev_risk_idx + om_quality) // 2
        note_exec = (f"Rev growth σ={rev_std:.1%}  |  "
                     f"OM trend {_delta_str}, σ={om_std:.1%}  "
                     f"[quality = direction + stability]")
    score_exec = round(exec_idx / 3.0 * 10.0, 2)
    tier_exec  = TIER_ORDER[exec_idx]

    # ── Valuation: P/E and P/FCF vs 5yr historical average ───────────────────
    # ── DCF-implied valuation anchor (#8) ─────────────────────────────────────
    # Compute what P/E and P/FCF the GG fair value implies given current TTM
    # earnings and FCF per share. Using this as a third benchmark protects
    # against multiple-expansion regimes (e.g. 2020-2024 mega-cap tech) where
    # the 5yr historical average is itself elevated. _t_val will take the
    # minimum of (5yr avg, sector peer median, DCF-implied) as the benchmark.
    dcf_implied_pe   = None
    dcf_implied_pfcf = None
    if dcf_gg_price and dcf_gg_price > 0 and is_data and cf_data:
        _ttm_ni  = is_data[-1].get("netIncome") or 0
        _ttm_cfo = cf_data[-1].get("operatingCashFlow") or 0
        _ttm_cap = abs(cf_data[-1].get("capitalExpenditure") or 0)
        _ttm_fcf = cf_data[-1].get("freeCashFlow") or (_ttm_cfo - _ttm_cap)
        _shares  = (is_data[-1].get("weightedAverageShsOut") or
                    is_data[-1].get("weightedAverageShsOutDil") or 0)
        if _shares > 0:
            _eps   = _ttm_ni  / _shares
            _fcfps = _ttm_fcf / _shares
            if _eps > 0:
                dcf_implied_pe = round(dcf_gg_price / _eps, 1)
            if _fcfps > 0:
                dcf_implied_pfcf = round(dcf_gg_price / _fcfps, 1)
        print(f"  DCF-implied multiples (GG ${dcf_gg_price:.2f}): "
              f"P/E={dcf_implied_pe}  P/FCF={dcf_implied_pfcf}")

    def _t_val(current, hist_avg, sect_med, label, roic_v, cagr_v, dcf_imp=None):
        if not current:
            return None, 0.0, f"N/A — {label} not available from yfinance"
        # Negative multiples mean the denominator (earnings or FCF) is negative.
        # The ratio is mathematically defined but economically meaningless — a
        # company losing money does not get cheaper as losses widen. Score LOW
        # so the scorecard does not falsely flag distressed names as bargains.
        if current <= 0:
            return "LOW", 0.0, (f"Current {current:.1f}x — {label} negative "
                                f"(earnings/FCF below zero, multiple meaningless)")
        premium_ok = (roic_v is not None and roic_v > 0.25 and
                      cagr_v is not None and cagr_v > 0.15)
        # Benchmark = min of historical 5yr average and sector peer median.
        # DCF-implied multiple is shown for context only — it is NOT used as a
        # benchmark here. Including the DCF in the benchmark created a circular
        # penalty: when GG says a stock is overvalued, the DCF-implied multiple
        # becomes the binding floor, which *further* penalises the valuation score
        # and double-counts the DCF's scepticism. P/E and P/FCF scoring must be
        # independent of the DCF to avoid this self-reinforcing feedback loop.
        _bench_candidates = [b for b in (hist_avg, sect_med) if b and b > 0]
        benchmark = min(_bench_candidates) if _bench_candidates else None
        parts_v   = [f"Current {current:.1f}x"]
        if hist_avg:    parts_v.append(f"5yr avg {hist_avg:.1f}x")
        if sect_med:    parts_v.append(f"Sector median {sect_med:.1f}x")
        if dcf_imp:     parts_v.append(f"DCF-implied {dcf_imp:.1f}x [ref only — not used as benchmark]")
        note_v = "  |  ".join(parts_v)
        if not benchmark:
            return None, 0.0, note_v + "  [no benchmark — review manually]"
        if benchmark <= 0:
            # 5yr avg distorted by loss years; cannot derive meaningful spread.
            return None, 0.0, note_v + "  [historical avg distorted by loss years — review manually]"
        delta = (current - benchmark) / benchmark
        note_v += f"  [{delta:+.0%} vs benchmark"
        # Tier label preserved for HTML/Excel display
        if delta > 0.25:
            tier_v = "MOD-LOW" if premium_ok else "LOW"
            if premium_ok:
                note_v += " — premium partly justified (ROIC>25% & fwd CAGR>15%)"
        elif delta > 0.10:
            tier_v = "MOD-LOW"
        elif delta >= -0.10:
            tier_v = "MOD-HIGH"
        else:
            tier_v = "HIGH"
        note_v += "]"
        score_v = _val_score(delta, premium_ok=premium_ok)
        return tier_v, score_v, note_v

    # ── Forward P/E and forward P/FCF derivation ──────────────────────────────
    # Forward P/E  : analyst FY+1 consensus EPS estimate (already fetched by caller).
    # Forward P/FCF: trailing FCF margin × analyst FY+1 revenue estimate.
    #   Only used when trailing FCF margin is positive — negative-FCF businesses
    #   cannot extrapolate a reliable forward FCF from the margin anchor.
    _sc_price        = float(prof_sc.get("price") or 0) or None
    forward_pe_val   = None
    forward_pfcf_val = None

    if analyst_ests and _sc_price:
        # Forward P/E
        try:
            _fwd_eps = float((analyst_ests[0] or {}).get("epsAvg") or 0) or None
            if _fwd_eps and _fwd_eps > 0:
                forward_pe_val = round(_sc_price / _fwd_eps, 1)
                print(f"  Fwd P/E: price={_sc_price:.2f}  eps_est={_fwd_eps:.2f}"
                      f"  fwd_pe={forward_pe_val:.1f}x")
        except Exception:
            pass

        # Forward P/FCF via FCF-margin anchor
        try:
            _fwd_rev = float((analyst_ests[0] or {}).get("revenueAvg") or 0) or None
            if _fwd_rev and cf_data and is_data:
                _ttm_rev   = is_data[-1].get("revenue") or 0
                _ttm_fcf   = (cf_data[-1].get("freeCashFlow") or
                              (cf_data[-1].get("operatingCashFlow", 0) +
                               cf_data[-1].get("capitalExpenditure", 0)))
                _fcf_margin = _ttm_fcf / _ttm_rev if _ttm_rev > 0 else None
                if _fcf_margin and _fcf_margin > 0:
                    _shares_sc  = (is_data[-1].get("weightedAverageShsOut") or
                                   is_data[-1].get("weightedAverageShsOutDil") or 0)
                    _fwd_fcf    = _fwd_rev * _fcf_margin
                    _fwd_fcfps  = _fwd_fcf / _shares_sc if _shares_sc > 0 else None
                    if _fwd_fcfps and _fwd_fcfps > 0:
                        forward_pfcf_val = round(_sc_price / _fwd_fcfps, 1)
                        print(f"  Fwd P/FCF: fcf_margin={_fcf_margin:.1%}"
                              f"  fwd_rev=${_fwd_rev/1e9:.1f}B"
                              f"  fwd_pfcf={forward_pfcf_val:.1f}x")
        except Exception:
            pass

    # pe_current: best available P/E for tier label + absolute score component.
    # Priority: (1) analyst FY+1 consensus forward P/E — most actionable for investors
    #           (2) live TTM P/E (from quarterly IS) — accurate trailing snapshot
    #           (3) forward_pe from legacy yfinance path (always None, kept for safety)
    # Note: the 40% forward blend component is computed separately from forward_pe_val
    # so using forward_pe_val here does NOT double-count — it simply anchors the tier
    # label and absolute-score lookup to the consensus view rather than a stale snapshot.
    pe_current    = forward_pe_val or trailing_pe or forward_pe
    # For banks the 5yr P/E average often embeds zero-rate / crisis-era compression
    # (e.g. 2020-2021 when bank P/Es were unusually depressed). Using a compressed
    # benchmark penalises banks that simply re-rated to normal post-rate-normalisation.
    # Floor at 10.0x — below which we consider the historical average distorted.
    _pe_5yr_bench = pe_5yr_avg
    if is_bank and _pe_5yr_bench is not None and _pe_5yr_bench < 10.0:
        print(f"  Bank P/E floor: 5yr avg {_pe_5yr_bench:.1f}x -> 10.0x (zero-rate compression)")
        _pe_5yr_bench = 10.0
    tier_pe,   score_pe,   note_pe   = _t_val(pe_current,    _pe_5yr_bench,  sector_pe_med,
                                               "P/E",   roic_latest, rev_cagr,
                                               dcf_imp=dcf_implied_pe)
    tier_pfcf, score_pfcf, note_pfcf = _t_val(trailing_pfcf, pfcf_5yr_avg, sector_pfcf_med,
                                               "P/FCF", roic_latest, rev_cagr,
                                               dcf_imp=dcf_implied_pfcf)

    # ── Valuation blend: 40% trailing / 40% forward / 20% absolute ───────────
    # Three-way blend anchors the score against:
    #   (1) how cheap the stock is vs its own history / peers  (trailing relative)
    #   (2) how cheap it is on next-year estimates             (forward relative)
    #   (3) whether the absolute multiple is expensive at all  (absolute lookup)
    # If no forward data is available, the forward component falls back to the
    # trailing score, making the effective split 80/20 relative/absolute.
    #
    # Absolute P/E  (0–12): ≤15x=12 | 15–20x=10 | 20–25x=7 | 25–35x=4 | 35–50x=1.5 | >50x=0
    # Absolute PFCF (0–12): ≤15x=12 | 15–20x=10 | 20–30x=7.5 | 30–40x=4 | 40–55x=1.5 | >55x=0
    #
    # Bank cap: banks carry PE weight=20 (2×). Cap the absolute PE score at 6.0
    # for banks to prevent the weight amplification from dominating the blend.
    # Banks have PFCF weight=0 so PFCF blending is skipped for them entirely.
    _TRAIL_W      = 0.40
    _FWD_W        = 0.40
    _ABS_W        = 0.20
    _ABS_BANK_CAP = 6.0

    def _abs_pe_score(pe, peg=None):
        """
        Absolute P/E lookup (0–12 scale), with optional PEG-ratio boost.

        When next-year EPS growth justifies the multiple (PEG < 1.5), the absolute
        penalty is partially relieved — a 37x P/E for a 50%-growth company is very
        different from a 37x P/E for a 5%-growth company.  The boost is capped at
        3.0 pts and the result is clipped to 12.0 so the absolute component never
        dominates the 40/40/20 blend.

        PEG < 1.0  → boost +3.0 (growth more than justifies the multiple)
        PEG 1.0–1.5 → boost +1.5 (growth largely justifies premium)
        PEG 1.5–2.5 → boost +0.5 (partial justification)
        PEG ≥ 2.5  → no boost   (growth does not justify premium)
        """
        if pe is None or pe <= 0: return 0.0
        if pe <= 15: raw = 12.0
        elif pe <= 20: raw = 10.0
        elif pe <= 25: raw =  7.0
        elif pe <= 35: raw =  4.0
        elif pe <= 50: raw =  1.5
        else:          raw =  0.0
        if peg is not None and 0 < peg < 2.5:
            boost = 3.0 if peg < 1.0 else 1.5 if peg < 1.5 else 0.5
            raw   = min(12.0, raw + boost)
        return raw

    def _abs_pfcf_score(pfcf):
        if pfcf is None or pfcf <= 0: return 0.0
        if pfcf <= 15: return 12.0
        if pfcf <= 20: return 10.0
        if pfcf <= 30: return  7.5
        if pfcf <= 40: return  4.0
        if pfcf <= 55: return  1.5
        return 0.0

    # Forward relative scores — use same _t_val vs same benchmarks as trailing
    _fwd_pe_score   = None
    _fwd_pfcf_score = None
    if forward_pe_val and forward_pe_val > 0:
        _, _fwd_pe_score, _ = _t_val(forward_pe_val, _pe_5yr_bench, sector_pe_med,
                                      "Fwd P/E", roic_latest, rev_cagr)
        print(f"  Fwd P/E relative score: {_fwd_pe_score:.2f}")
    if forward_pfcf_val and forward_pfcf_val > 0 and not is_bank:
        _, _fwd_pfcf_score, _ = _t_val(forward_pfcf_val, pfcf_5yr_avg, sector_pfcf_med,
                                         "Fwd P/FCF", roic_latest, rev_cagr)
        print(f"  Fwd P/FCF relative score: {_fwd_pfcf_score:.2f}")

    # ── PEG ratio — modifier for absolute P/E penalty ─────────────────────────
    # Forward PEG = forward P/E ÷ forward EPS growth (vs trailing TTM EPS).
    # Falls back to trailing P/E ÷ revenue CAGR when EPS estimates are absent.
    # A PEG < 1.5 indicates growth meaningfully justifies the absolute multiple,
    # and earns a score boost inside _abs_pe_score (see function docstring).
    _peg_val  = None
    _peg_note = ""
    try:
        _pe_for_peg = forward_pe_val or pe_current
        if _pe_for_peg and _pe_for_peg > 0 and analyst_ests and is_data:
            _fwd_eps_est = float((analyst_ests[0] or {}).get("epsAvg") or 0) or None
            _shs_peg     = (is_data[-1].get("weightedAverageShsOut") or
                            is_data[-1].get("weightedAverageShsOutDil") or 0)
            _ttm_ni_peg  = is_data[-1].get("netIncome") or 0
            _ttm_eps_peg = (_ttm_ni_peg / _shs_peg) if _shs_peg > 0 else None
            if _fwd_eps_est and _ttm_eps_peg and _ttm_eps_peg > 0 and _fwd_eps_est > 0:
                _eps_growth_peg = (_fwd_eps_est / _ttm_eps_peg) - 1
                if _eps_growth_peg > 0:
                    _peg_val  = round(_pe_for_peg / _eps_growth_peg, 2)
                    _peg_note = (f"PEG={_peg_val:.2f} (fwd_pe={_pe_for_peg:.1f}x"
                                 f" / eps_growth={_eps_growth_peg:.0%})")
        # Fallback: use revenue CAGR as growth proxy
        if _peg_val is None and pe_current and pe_current > 0 and rev_cagr and rev_cagr > 0.02:
            _peg_val  = round(pe_current / rev_cagr, 2)
            _peg_note = (f"PEG≈{_peg_val:.2f} (trailing_pe={pe_current:.1f}x"
                         f" / rev_cagr={rev_cagr:.0%}, revenue proxy)")
    except Exception:
        pass
    if _peg_note:
        print(f"  {_peg_note}")

    # ── Earnings revision momentum ────────────────────────────────────────────
    # Proxy: FY+1 consensus EPS vs trailing TTM EPS. Captures whether analysts
    # expect EPS to accelerate or contract — a de-risking or warning signal.
    # Applied as a small modifier (±0.25–0.5 pts) on the blended P/E score.
    _n_analysts_eps        = 0
    _eps_revision_pct      = None
    _eps_revision_dir      = "N/A"
    _rev_momentum_modifier = 0.0
    if analyst_ests and is_data:
        try:
            _n_analysts_eps = int(analyst_ests[0].get("numAnalystsEps") or 0)
            _fwd_eps_rev    = float(analyst_ests[0].get("epsAvg") or 0) or None
            _shs_rev        = (is_data[-1].get("weightedAverageShsOut") or
                               is_data[-1].get("weightedAverageShsOutDil") or 0)
            _ttm_ni_rev     = is_data[-1].get("netIncome") or 0
            _ttm_eps_rev    = (_ttm_ni_rev / _shs_rev) if _shs_rev > 0 else None
            if _fwd_eps_rev and _ttm_eps_rev and _ttm_eps_rev > 0 and _fwd_eps_rev > 0:
                _eps_revision_pct = (_fwd_eps_rev / _ttm_eps_rev) - 1
                if   _eps_revision_pct >  0.20: _eps_revision_dir = "Strong upgrade cycle"
                elif _eps_revision_pct >  0.05: _eps_revision_dir = "Modest upgrade expected"
                elif _eps_revision_pct > -0.05: _eps_revision_dir = "Flat consensus"
                else:                           _eps_revision_dir = "Earnings headwind"
            elif _fwd_eps_rev and _ttm_eps_rev and _ttm_eps_rev < 0 and _fwd_eps_rev > 0:
                _eps_revision_dir = "Turning profitable"
            if _eps_revision_pct is not None:
                if   _eps_revision_pct >  0.20: _rev_momentum_modifier =  0.5
                elif _eps_revision_pct >  0.05: _rev_momentum_modifier =  0.25
                elif _eps_revision_pct < -0.15: _rev_momentum_modifier = -0.5
                elif _eps_revision_pct < -0.05: _rev_momentum_modifier = -0.25
        except Exception:
            pass
        print(f"  EPS revision: dir={_eps_revision_dir}"
              f"  pct={f'{_eps_revision_pct:.0%}' if _eps_revision_pct is not None else 'N/A'}"
              f"  modifier={_rev_momentum_modifier:+.2f}  n_analysts={_n_analysts_eps}")

    # P/E blend
    if score_pe is not None and pe_current and pe_current > 0:
        _abs_pe  = _abs_pe_score(pe_current, peg=_peg_val)
        if is_bank:
            _abs_pe = min(_abs_pe, _ABS_BANK_CAP)
        _rel_pe  = score_pe
        _fpe     = _fwd_pe_score if _fwd_pe_score is not None else _rel_pe
        _blended_pe = round(_TRAIL_W * _rel_pe + _FWD_W * _fpe + _ABS_W * _abs_pe, 4)
        _fwd_note  = f"fwd_pe={forward_pe_val:.1f}x→{_fpe:.2f}" if _fwd_pe_score is not None else "no_fwd_data(fallback)"
        _bank_note = f" [bank cap {_ABS_BANK_CAP}]" if is_bank and _abs_pe == _ABS_BANK_CAP else ""
        _peg_str   = f"  {_peg_note}" if _peg_note else ""
        note_pe   += (f"  [trail={_rel_pe:.2f} | {_fwd_note} | abs={_abs_pe:.1f}{_bank_note}"
                      f" → blended 40/40/20={_blended_pe:.2f}]{_peg_str}")
        print(f"  PE blend 40/40/20: trail={_rel_pe:.2f}  fwd={_fpe:.2f}"
              f"  abs={_abs_pe:.1f}  → {_blended_pe:.2f}"
              f"  {_peg_note}")
        score_pe = _blended_pe

    # Apply earnings revision momentum modifier to P/E score
    if score_pe is not None and _rev_momentum_modifier != 0.0:
        score_pe = round(max(0.0, min(12.0, score_pe + _rev_momentum_modifier)), 4)
        note_pe += f"  [revision:{_eps_revision_dir}→{_rev_momentum_modifier:+.2f}]"
        print(f"  PE revision modifier: {_rev_momentum_modifier:+.2f} → adjusted={score_pe:.2f}")

    # P/FCF blend (non-banks only)
    if score_pfcf is not None and trailing_pfcf and trailing_pfcf > 0 and not is_bank:
        _abs_pfcf  = _abs_pfcf_score(trailing_pfcf)
        _rel_pfcf  = score_pfcf
        _fpfcf     = _fwd_pfcf_score if _fwd_pfcf_score is not None else _rel_pfcf
        _blended_pfcf = round(_TRAIL_W * _rel_pfcf + _FWD_W * _fpfcf + _ABS_W * _abs_pfcf, 4)
        _fwd_pfcf_note = (f"fwd_pfcf={forward_pfcf_val:.1f}x→{_fpfcf:.2f}"
                          if _fwd_pfcf_score is not None else "no_fwd_data(fallback)")
        note_pfcf += (f"  [trail={_rel_pfcf:.2f} | {_fwd_pfcf_note} | abs={_abs_pfcf:.1f}"
                      f" → blended 40/40/20={_blended_pfcf:.2f}]")
        print(f"  PFCF blend 40/40/20: trail={_rel_pfcf:.2f}  fwd={_fpfcf:.2f}"
              f"  abs={_abs_pfcf:.1f}  → {_blended_pfcf:.2f}")
        score_pfcf = _blended_pfcf

    # ── FCF yield vs 10-year treasury spread ─────────────────────────────────
    # FCF yield = 1 / P/FCF. A positive spread above the 10Y treasury means the
    # stock offers a real return premium over risk-free; negative spread warns that
    # treasuries beat the company's FCF yield.
    _fcf_yield        = (1.0 / trailing_pfcf) if trailing_pfcf and trailing_pfcf > 0 else None
    _rf_rate_sc, _    = fetch_fred("DGS10")
    _rf_rate_sc       = _rf_rate_sc or 0.043
    _fcf_yield_spread = (_fcf_yield - _rf_rate_sc) if _fcf_yield is not None else None
    _fcf_spread_modifier = 0.0
    if _fcf_yield_spread is not None and score_pfcf is not None:
        if   _fcf_yield_spread >  0.05: _fcf_spread_modifier =  0.5
        elif _fcf_yield_spread >  0.02: _fcf_spread_modifier =  0.25
        elif _fcf_yield_spread < -0.02: _fcf_spread_modifier = -0.5
        elif _fcf_yield_spread <  0.0:  _fcf_spread_modifier = -0.25
        if _fcf_spread_modifier != 0.0:
            score_pfcf = round(max(0.0, min(12.0, score_pfcf + _fcf_spread_modifier)), 4)
            note_pfcf += (f"  [fcf_yield={_fcf_yield:.1%} vs rf={_rf_rate_sc:.1%}"
                          f"  spread={_fcf_yield_spread:+.1%}→modifier={_fcf_spread_modifier:+.2f}]")
            print(f"  FCF yield spread: {_fcf_yield:.1%} - {_rf_rate_sc:.1%}"
                  f" = {_fcf_yield_spread:+.1%}  modifier={_fcf_spread_modifier:+.2f}"
                  f"  → adjusted score_pfcf={score_pfcf:.2f}")

    # ── Soft credit floor cap (#7) ────────────────────────────────────────────
    # Replaces the binary floor gates (D/EBITDA >4 → cap 64; EBIT/Int <2 → cap 64;
    # both → cap 59). Two problems with cliffs: (1) D/EBITDA = 4.01 vs 3.99 made a
    # 5-pt swing, (2) the cap was the same magnitude regardless of how badly the
    # threshold was breached. Continuous formula:
    #
    #   cap = 100 − 5·max(0, D/EBITDA − 3.0) − 4·max(0, 3.0 − EBIT/Int)
    #
    # Bounded to [40, 100]. Activates from D/EBITDA > 3.0 and EBIT/Int < 3.0,
    # so genuine distress is penalised before crossing the old hard thresholds.
    # Banks remain exempt — the same structural reasons apply (deposit funding).
    if is_bank:
        soft_cap = 100.0
        cap_d_pen = 0.0
        cap_e_pen = 0.0
    else:
        cap_d_pen = (5.0 * max(0.0, (d_ebitda - 3.0))
                     if d_ebitda is not None and net_cash_v <= 0 else 0.0)
        cap_e_pen = (4.0 * max(0.0, (3.0 - ebit_int))
                     if ebit_int is not None else 0.0)
        soft_cap = max(40.0, min(100.0, 100.0 - cap_d_pen - cap_e_pen))
    # Only treat as a "cap" when it actually constrains (i.e. < 100).
    floor_cap = round(soft_cap, 1) if soft_cap < 99.5 else None
    # Legacy gate booleans retained for the warning banner only.
    gate1 = (d_ebitda is not None and d_ebitda > 4.0 and net_cash_v <= 0)
    gate2 = (ebit_int is not None and ebit_int < 2.0)

    # ── Normalise user-supplied qualitative tiers (HIGH / MOD / LOW) ──────────
    def _norm_qual(v):
        v = (v or "").strip().upper()
        # Map condensed "MOD" to MOD-HIGH (centre of the moderate band) so it
        # scores consistently with the Excel 4-tier scale (HIGH / MOD-HIGH /
        # MOD-LOW / LOW). User dropdowns only expose HIGH/MOD/LOW.
        if v == "HIGH": return "HIGH"
        if v == "MOD":  return "MOD-HIGH"
        if v in ("MOD-HIGH", "MOD-LOW"): return v
        if v == "LOW":  return "LOW"
        return None
    _bc_tier  = _norm_qual(biz_clarity)
    _ltp_tier = _norm_qual(ltp)

    # ── Regime-aware weight schema (#9) ───────────────────────────────────────
    # Default weights (sum = 100). Regime modifiers redistribute weight when
    # specific criteria become structurally meaningless:
    #   • banks:  P/FCF is meaningless (deposit-funded, no capex/FCF rhythm) →
    #             zero P/FCF weight, double P/E weight to 20 (single equity-
    #             multiple anchor, since P/TBV is not yet computed).
    #   • evs_regime (pre-profit secular growth): ROIC, FCF/NI, Capital Returns,
    #             Interest Cover, P/E, P/FCF are all distorted or undefined when
    #             trailing earnings/FCF are negative. Zero those (50 weight) and
    #             redistribute to growth + qualitative + moat — the only signals
    #             that meaningfully differentiate pre-profit names.
    W = {
        "BC": 2.5, "Moat": 10.0, "LTP": 10.0, "Mgmt": 7.5,
        "RevCAGR": 10.0, "FCFNI": 10.0, "CapRet": 5.0, "ROIC": 7.5,
        "Lev": 5.0, "EBITInt": 7.5, "Exec": 5.0,
        "PE": 10.0, "PFCF": 10.0,
    }
    if is_bank:
        W["PE"]      = 20.0   # single equity-multiple anchor (P/TBV not yet implemented)
        W["PFCF"]    = 0.0    # meaningless for deposit-funded institutions
        W["EBITInt"] = 0.0    # traditional interest cover nonsensical for banks (interest IS their raw-material cost)
        W["Lev"]     = 12.5   # absorbs the freed 7.5 — capital adequacy is the bank risk metric
    if evs_regime:
        # zero distorted criteria
        W["ROIC"]    = 0.0
        W["FCFNI"]   = 0.0
        W["CapRet"]  = 0.0
        W["EBITInt"] = 0.0
        W["PE"]      = 0.0
        W["PFCF"]    = 0.0
        # redistribute the freed 50.0 weight to growth + qualitative + moat
        W["RevCAGR"] = 25.0   # +15 — primary signal for pre-profit
        W["Moat"]    = 20.0   # +10 — gross-margin trajectory included
        W["LTP"]     = 25.0   # +15 — TAM judgment dominates
        W["BC"]      = 7.5    # +5  — segment economics matter more
        W["Mgmt"]    = 12.5   # +5  — capital stewardship critical pre-profit
        # Lev (5) + Exec (5) unchanged → total = 25+20+25+7.5+12.5+5+5 = 100 ✓
    _wsum = sum(W.values())
    if _wsum and abs(_wsum - 100.0) > 0.01:
        print(f"  ⚠ Regime weights sum to {_wsum} (expected 100)")
    print(f"  Weight regime: is_bank={is_bank}  evs_regime={evs_regime}  weights={W}")

    # ── Criteria table definition ─────────────────────────────────────────────
    # (part, label, weight, auto_tier, auto_score, note, is_auto)
    # auto_score is the continuous score (0..SCORE_CAP). For auto rows the
    # literal value is written to Excel col F so totals reflect the smooth
    # underlying metric, not the discrete tier mapping. Qualitative rows pass
    # None for auto_score and the tier→score formula computes from col E.
    # Business Clarity and Long-Term Potential are qualitative — pre-filled from
    # the web form when provided; otherwise left blank for manual input.
    CRITERIA_RAW = [
        ("P1", "Business Clarity",                  W["BC"],   _bc_tier,      None,
         ("User-supplied via web form" if _bc_tier else
          "Segment data not on current FMP plan — assign manually after reviewing 10-K"),
         _bc_tier is not None),
        ("P1", "Moat Profile",                       W["Moat"], tier_moat,     score_moat,     note_moat,      True),
        ("P1", "Long-Term Potential",                W["LTP"],  _ltp_tier,     None,
         ("User-supplied via web form" if _ltp_tier else
          "Structural/TAM outlook — assign manually (genuinely qualitative)"),
         _ltp_tier is not None),
        ("P1", "Management",                          W["Mgmt"], tier_mgmt,     score_mgmt,     note_mgmt,      True),
        ("P2", "Revenue 3yr CAGR",                  W["RevCAGR"], tier_rev_cagr, score_rev_cagr, note_rev_cagr,  True),
        ("P2", "Cash Quality  (FCF / Net Income)",  W["FCFNI"], tier_fcf_ni,   score_fcf_ni,   note_fcf_ni,    True),
        ("P2", "Capital Returns",                    W["CapRet"], tier_cap_ret,  score_cap_ret,  note_cap_ret,   True),
        ("P2", "ROIC",                               W["ROIC"], tier_roic,     score_roic,     note_roic,      True),
        ("P3", leverage_label,                        W["Lev"],  tier_leverage, score_leverage, note_leverage,  True),
        ("P3", "Interest Cover  (EBIT / Interest)",  W["EBITInt"], tier_ebit_int, score_ebit_int, note_ebit_int,  True),
        ("P3", "Execution Risk",                     W["Exec"], tier_exec,     score_exec,     note_exec,      True),
        ("P4", "Valuation vs Median  (P/E)",        W["PE"],   tier_pe,       score_pe,       note_pe,        tier_pe   is not None),
        ("P4", "Valuation vs Median  (P/FCF)",      W["PFCF"], tier_pfcf,     score_pfcf,     note_pfcf,      tier_pfcf is not None),
    ]
    # Drop zero-weight criteria entirely so the Excel sheet doesn't waste rows
    # on N/A lines for the regime that has zeroed them.
    CRITERIA = [c for c in CRITERIA_RAW if c[2] > 0]

    # ── Cell writing helpers ──────────────────────────────────────────────────
    def wcell(r, col, val="", bold=False, color=C_BLACK, bg=C_WHITE,
              halign="left", indent=0, italic=False, fmt=None, wrap=False):
        c = ws.cell(row=r, column=col, value=val)
        c.font      = fnt(bold=bold, color=color, italic=italic)
        c.fill      = fll(bg)
        c.border    = brd()
        c.alignment = Alignment(horizontal=halign, vertical="center",
                                indent=indent, wrap_text=wrap)
        if fmt:
            c.number_format = fmt
        return c

    def merge_row(r, val, bold=False, color=C_WHITE, bg=C_SECTION,
                  halign="left", size=10, indent=1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(row=r, column=1, value=val)
        c.font      = Font(name="Arial", bold=bold, color=color, size=size)
        c.fill      = fll(bg)
        c.alignment = Alignment(horizontal=halign, vertical="center",
                                indent=indent, wrap_text=True)
        ws.row_dimensions[r].height = 18
        return r + 1

    def blank_row(r):
        for col in range(1, NC + 1):
            ws.cell(row=r, column=col).fill = fll(C_WHITE)
            ws.cell(row=r, column=col).border = brd()
        return r + 1

    def tier_bg(t):
        return {"HIGH": "C8E6C9", "MOD-HIGH": "BBDEFB",
                "MOD-LOW": "FFE0B2", "LOW": "FFCDD2"}.get(t, C_WHITE)

    def tier_fg(t):
        return {"HIGH": "1B5E20", "MOD-HIGH": "1565C0",
                "MOD-LOW": "E65100", "LOW": "B71C1C"}.get(t, C_BLACK)

    SCORE_FORMULA = (
        '=IF({e}="HIGH",10,'
        'IF({e}="MOD-HIGH",7,'
        'IF({e}="MOD-LOW",3,'
        'IF({e}="LOW",0,""))))'
    )

    # ── Dropdown validation for column E ─────────────────────────────────────
    dv = DataValidation(type="list",
                        formula1='"HIGH,MOD-HIGH,MOD-LOW,LOW"',
                        allow_blank=True,
                        showDropDown=False)
    dv.sqref = "E9:E50"
    ws.add_data_validation(dv)

    # ════════════════════════════════════════════════════════════════════════
    # WRITE SHEET
    # ════════════════════════════════════════════════════════════════════════
    row = 1

    # Title
    ws.row_dimensions[row].height = 24
    row = merge_row(row,
                    f"JS SCORECARD — {ticker}  |  Master Prompt v2  |  "
                    f"{datetime.date.today():%d %b %Y}  |  "
                    f"Sector bucket: {_bucket.replace('_', ' ').title()}",
                    bold=True, size=12, bg=C_TITLE)

    # Subtitle / instructions
    ws.row_dimensions[row].height = 28
    row = merge_row(
        row,
        "Blue rows = auto-scored (FMP data + yfinance proxies).  "
        "Yellow rows = manual — select tier from dropdown in column E.  "
        "Only 2 criteria require manual input: Business Clarity + Long-Term Potential.  "
        "Score: HIGH=10 | MOD-HIGH=7 | MOD-LOW=3 | LOW=0",
        bold=False, color=C_BLACK, bg="EAF2FB", size=9
    )

    row = blank_row(row)

    # Gate status
    if is_bank:
        ws.row_dimensions[row].height = 16
        row = merge_row(
            row,
            "✓  Bank/Financial sector — D/EBITDA and EBIT/Interest gates exempt  "
            f"| Capital Adequacy (Equity/Assets) used as P3 leverage criterion"
            + (f"  |  Equity/Assets = {equity_assets:.1%}" if equity_assets else ""),
            bold=False, color="1B5E20", bg="C8E6C9", size=9
        )
    elif floor_cap is not None:
        msgs = []
        if cap_d_pen > 0:
            msgs.append(f"D/EBITDA {d_ebitda:.1f}x → −{cap_d_pen:.1f} pts")
        if cap_e_pen > 0:
            msgs.append(f"EBIT/Interest {ebit_int:.1f}x → −{cap_e_pen:.1f} pts")
        # Severity of cap drives banner color: tighter cap = redder
        if floor_cap < 60:
            bg, fg, ico = "B71C1C", C_WHITE, "⚠"
        elif floor_cap < 75:
            bg, fg, ico = "F57C00", C_WHITE, "⚠"
        else:
            bg, fg, ico = "FFE0B2", "5D4037", "•"
        ws.row_dimensions[row].height = 20
        row = merge_row(
            row,
            f"{ico}  SOFT FLOOR CAP {floor_cap:.0f}/100 applied — "
            + "  |  ".join(msgs)
            + f"  (kicks in at D/EBITDA > 3.0x or EBIT/Int < 3.0x; smooth, no cliffs)",
            bold=True, color=fg, bg=bg, size=10
        )
    else:
        ws.row_dimensions[row].height = 16
        row = merge_row(
            row,
            "✓  No floor cap (D/EBITDA and EBIT/Interest within safe thresholds)",
            bold=False, color="1B5E20", bg="C8E6C9", size=9
        )

    row = blank_row(row)

    # Column headers
    ws.row_dimensions[row].height = 20
    for col, (txt, halign) in enumerate([
        ("CRITERION",        "left"),
        ("PART",             "center"),
        ("WEIGHT %",         "center"),
        ("CALCULATED VALUE", "center"),
        ("TIER  ▼",          "center"),
        ("SCORE",            "center"),
        ("WTD SCORE",        "center"),
        ("NOTES / COMMENTARY (editable)", "left"),
    ], start=1):
        c = ws.cell(row=row, column=col, value=txt)
        c.font      = fnt(bold=True, color=C_WHITE, size=9)
        c.fill      = fll(C_DETAIL_HD)
        c.border    = brd()
        c.alignment = Alignment(horizontal=halign, vertical="center", indent=1)
    row += 1

    hdr_row = row  # first criteria row index
    crit_rows = []  # track for SUM formula

    current_part = None
    for part, label, weight, auto_tier, auto_score, note, is_auto in CRITERIA:
        # Part separator header
        if part != current_part:
            current_part = part
            part_labels = {
                "P1": "PART 1 — BUSINESS QUALITY  (qualitative)",
                "P2": "PART 2 — FINANCIAL PERFORMANCE  (quantitative)",
                "P3": "PART 3 — RISK PROFILE  (quantitative / qualitative)",
                "P4": "PART 4 — VALUATION  (qualitative / user-supplied market data)",
            }
            ws.row_dimensions[row].height = 16
            row = write_section_hdr(ws, row, part_labels[part], NC, C_SECTION)

        # Row background
        row_bg = C_ASSM if is_auto else C_AI_BG   # blue=auto, yellow=qualitative

        ws.row_dimensions[row].height = 18

        # A: Criterion name
        wcell(row, 1, f"  {label}", bold=is_auto, bg=row_bg, halign="left")

        # B: Part
        wcell(row, 2, part, bold=False, bg=row_bg, halign="center", color="555555")

        # C: Weight
        c_wt = wcell(row, 3, weight / 100, bold=False, bg=row_bg, halign="center")
        c_wt.number_format = "0.0%"

        # D: Calculated value
        if is_auto and note:
            wcell(row, 4, note, bold=False, bg=row_bg, halign="left", italic=True,
                  color="1A3A5C")
        else:
            wcell(row, 4, "— user input required", italic=True,
                  color="999999", bg=row_bg)

        # E: Tier (pre-filled for auto; blank for qualitative)
        e_addr = f"E{row}"
        if auto_tier:
            c_tier = ws.cell(row=row, column=5, value=auto_tier)
            c_tier.font      = Font(name="Arial", bold=True, size=10,
                                    color=tier_fg(auto_tier))
            c_tier.fill      = fll(tier_bg(auto_tier))
        else:
            c_tier = ws.cell(row=row, column=5, value=None)
            c_tier.fill = fll(C_WHITE)
        c_tier.border    = brd()
        c_tier.alignment = Alignment(horizontal="center", vertical="center")

        # F: Score
        # For auto-scored rows: write the continuous engine score as a literal
        # value (smooth, no threshold cliffs). For qualitative rows: keep the
        # tier→score formula so user dropdown changes recompute live in Excel.
        if is_auto and auto_score is not None:
            c_score = wcell(row, 6, round(float(auto_score), 2), bold=True, bg=row_bg,
                            halign="center", fmt='0.00;(0.00);"-"')
        else:
            score_f = SCORE_FORMULA.replace("{e}", e_addr)
            c_score = wcell(row, 6, score_f, bold=True, bg=row_bg,
                            halign="center", fmt='0.00;(0.00);"-"')
        c_score.font = fnt(bold=True, color=C_BLACK)

        # G: Weighted score (formula)
        # Weight in col C is stored as a decimal (e.g. 0.10 for 10%).
        # Score in col F is 0–10. Multiply by 10 so the total is on a 0–100 scale.
        wt_col   = get_column_letter(3)
        scr_col  = get_column_letter(6)
        c_wscore = wcell(row, 7,
                         f'=IF({scr_col}{row}="",0,{wt_col}{row}*{scr_col}{row}*10)',
                         bold=False, bg=row_bg, halign="center", fmt='0.00;(0.00);"-"')

        # H: Notes
        wcell(row, 8, note if is_auto else "", italic=True, color="555555",
              bg=row_bg, halign="left", wrap=True)

        crit_rows.append(row)
        row += 1

    # ── Total section ─────────────────────────────────────────────────────────
    row = blank_row(row)

    ws.row_dimensions[row].height = 20
    # Total row
    for col in range(1, NC + 1):
        ws.cell(row=row, column=col).fill = fll(C_SUBTOTAL)
        ws.cell(row=row, column=col).border = brd()
    c_tot_lbl = ws.cell(row=row, column=1, value="TOTAL SCORE  (max = 100.0)")
    c_tot_lbl.font      = fnt(bold=True, color=C_BLACK, size=11)
    c_tot_lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Weighted total formula — sum of all G cells in crit_rows
    g_refs = "+".join(f"G{r}" for r in crit_rows)
    c_tot = ws.cell(row=row, column=7, value=f"={g_refs}")
    c_tot.font         = fnt(bold=True, size=11)
    c_tot.number_format = "0.00"
    c_tot.alignment    = Alignment(horizontal="center", vertical="center")
    c_tot.border        = brd()

    # Note in H
    cap_txt = f"Floor cap {floor_cap} applies — see gate warning above." if floor_cap else "No floor cap."
    c_cap   = ws.cell(row=row, column=8, value=cap_txt)
    c_cap.font      = fnt(bold=(floor_cap is not None), color=("B71C1C" if floor_cap else "1B5E20"))
    c_cap.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c_cap.border    = brd()
    tot_row = row; row += 1

    # Floor-adjusted row (only shown when cap applies)
    if floor_cap is not None:
        ws.row_dimensions[row].height = 18
        for col in range(1, NC + 1):
            ws.cell(row=row, column=col).fill = fll(C_FLAG_BG)
            ws.cell(row=row, column=col).border = brd()
        c_fl = ws.cell(row=row, column=1,
                       value=f"FLOOR-ADJUSTED SCORE  (capped at {floor_cap})")
        c_fl.font      = fnt(bold=True, color="B71C1C", size=11)
        c_fl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_adj = ws.cell(row=row, column=7,
                        value=f"=MIN(G{tot_row},{floor_cap})")
        c_adj.font         = fnt(bold=True, color="B71C1C", size=11)
        c_adj.number_format = "0.00"
        c_adj.alignment    = Alignment(horizontal="center", vertical="center")
        c_adj.border        = brd()
        adj_row = row; row += 1
    else:
        adj_row = tot_row

    # ── Verdict row ───────────────────────────────────────────────────────────
    row = blank_row(row)

    ws.row_dimensions[row].height = 20
    # Verdict uses an Excel formula referencing the appropriate total cell
    score_ref = f"G{adj_row}"
    verdict_f = (
        f'=IF({score_ref}="","Score incomplete — fill qualitative tiers above",'
        f'IF({score_ref}>=80,"STRONG BUY",'
        f'IF({score_ref}>=65,"BUY",'
        f'IF({score_ref}>=50,"HOLD",'
        f'IF({score_ref}>=35,"REDUCE","SELL")))))'
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
    c_v = ws.cell(row=row, column=1, value=verdict_f)
    c_v.font      = Font(name="Arial", bold=True, size=12, color=C_WHITE)
    c_v.fill      = fll(C_SUMMARY_HD)
    c_v.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    # Scoring legend
    row = blank_row(row)
    ws.row_dimensions[row].height = 14
    row = merge_row(
        row,
        "SCORING GUIDE:  ≥80 STRONG BUY  |  65–79 BUY  |  50–64 HOLD  |  35–49 REDUCE  |  <35 SELL  "
        "  ||  Soft cap (non-banks only): −5pts per 1.0x of D/EBITDA above 3.0x; −4pts per 1.0x of EBIT/Int below 3.0x; min 40  "
        "  ||  Banks: gates exempt; P3 uses Equity/Assets (CET1 proxy) instead of D/EBITDA",
        bold=False, color="444444", bg="F4F8FB", size=8, indent=2
    )

    # ── Metrics dict for portfolio heatmap ────────────────────────────────────
    # auto_score = normalized 0-10 score from the 11 auto-scored criteria.
    # auto_score_raw = raw 0-87.5 value (used by callers to compute adj_score).
    # Raw max = 87.5 (Business Clarity 2.5 + Long-Term Potential 10.0 = 12.5 manual)
    # Uses continuous scores (0..SCORE_CAP) — no longer the discrete tier mapping.
    # Each weighted contribution = (score/10) * weight, mirroring Excel formula
    # `weight * score * 10` divided by 100 normalisation. SCORE_CAP=12 means a
    # criterion can earn up to 1.2 * its weight (rewards genuine outliers).
    _auto_criteria = [
        (tier_moat,      score_moat,      W["Moat"]),
        (tier_mgmt,      score_mgmt,      W["Mgmt"]),
        (tier_rev_cagr,  score_rev_cagr,  W["RevCAGR"]),
        (tier_fcf_ni,    score_fcf_ni,    W["FCFNI"]),
        (tier_cap_ret,   score_cap_ret,   W["CapRet"]),
        (tier_roic,      score_roic,      W["ROIC"]),
        (tier_leverage,  score_leverage,  W["Lev"]),     # D/EBITDA for non-banks; Equity/Assets for banks
        (tier_ebit_int,  score_ebit_int,  W["EBITInt"]),
        (tier_exec,      score_exec,      W["Exec"]),
        (tier_pe,        score_pe,        W["PE"]),
        (tier_pfcf,      score_pfcf,      W["PFCF"]),
    ]
    # Drop zero-weighted criteria so they don't even appear in the total
    _auto_criteria = [c for c in _auto_criteria if c[2] > 0]
    _active_weight = sum(w for _, _, w in _auto_criteria)   # total achievable weight (regime-adjusted)
    _scored = [(t, s, w) for t, s, w in _auto_criteria if t is not None and s is not None]
    _fg_valuation_gap = False  # initialise so it's always defined (used in conf_note below)
    if _scored:
        # Sum weighted scores. With SCORE_CAP=10, the maximum _raw_sum equals
        # the sum of all active weights (87.5 for a full non-bank scorecard).
        _raw_sum = sum((s / 10.0) * w for _, s, w in _scored)
        _scored_weight = sum(w for _, _, w in _scored)
        # Denominator normalization: if some criteria returned N/A (data gap,
        # not a regime zero), rescale to the full active weight pool so the
        # company isn't penalized for criteria it structurally cannot score on.
        # Sparse-data guardrail: if fewer than 50% of active criteria scored,
        # don't amplify — the surviving signals aren't representative enough
        # to extrapolate. Flag via low_data_confidence so callers can warn.
        #
        # F-G: Additional guard — when both P/E AND P/FCF tiers are None on a
        # standard (non-bank, non-EVS) company, suppress the rescale entirely.
        # This combination represents a data-fetch gap (FMP ratios API failed),
        # not a structural regime exclusion. Extrapolating 9 surviving criteria
        # to 100% coverage inflates the score by ~30% with no justification.
        # Flag LOW confidence so the report shows the data-gap caveat.
        _fg_valuation_gap = (
            tier_pe is None and tier_pfcf is None
            and not is_bank and not evs_regime
        )
        _low_data_confidence = (
            _fg_valuation_gap
            or ((_scored_weight < 0.5 * _active_weight) if _active_weight else True)
        )
        if (_scored_weight > 0 and _scored_weight < _active_weight
                and not _low_data_confidence):
            _raw_sum = _raw_sum * (_active_weight / _scored_weight)
        _auto_score_raw = round(_raw_sum, 1)
        if floor_cap is not None:
            _auto_score_raw = min(_auto_score_raw, floor_cap)
        # Normalize to 0-10 scale. Fixed 87.5 denominator regardless of regime so
        # scores are comparable across bank/standard/EVS weight regimes.
        # With SCORE_CAP=10, raw_sum is bounded at 87.5 by construction.
        _auto_score = round(min(10.0, _auto_score_raw / 87.5 * 10), 1)
    else:
        _auto_score_raw = None
        _auto_score = None
        _low_data_confidence = True

    # ── Score confidence flag ─────────────────────────────────────────────────
    _n_fcf_years = sum(
        1 for r in cf_data
        if (r.get("freeCashFlow") or
            (r.get("operatingCashFlow", 0) + r.get("capitalExpenditure", 0))) not in (None, 0)
    )
    if _low_data_confidence:
        _conf_level = "LOW"
        # F-G: distinguish data-fetch gap from genuine sparse-criteria case
        if _fg_valuation_gap:
            _conf_note = ("P/E and P/FCF data unavailable — FMP ratios fetch failed; "
                          "valuation criteria excluded, rescale suppressed")
        else:
            _conf_note = "fewer than 50% of scorecard criteria scored"
    elif _n_fcf_years < 3:
        _conf_level = "LOW"
        _conf_note  = f"only {_n_fcf_years} year(s) of FCF history"
    elif _n_analysts_eps == 0:
        _conf_level = "MEDIUM"
        _conf_note  = f"{_n_fcf_years} yr FCF history; no analyst EPS coverage"
    elif _n_fcf_years < 5 or _n_analysts_eps < 5:
        _conf_level = "MEDIUM"
        _conf_note  = f"{_n_fcf_years} yr FCF history; {_n_analysts_eps} analyst(s) covering"
    else:
        _conf_level = "HIGH"
        _conf_note  = f"{_n_fcf_years} yr FCF history; {_n_analysts_eps} analyst(s) covering"
    print(f"  Confidence: {_conf_level} ({_conf_note})")

    metrics = {
        "roic":          roic_latest,
        "rev_cagr":      rev_cagr,
        "fcf_ni":        fcf_ni_latest,
        "d_ebitda":      d_ebitda,        # None for banks (meaningless)
        "equity_assets": equity_assets,   # CET1 proxy; only set for banks
        "is_bank":       is_bank,
        "sector_bucket": _bucket,
        "auto_score":     _auto_score,        # normalized 0-10 for display
        "auto_score_raw": _auto_score_raw,    # raw 0-87.5 for adj_score computation in callers
        "floor_cap":     round(floor_cap / 87.5 * 10, 1) if floor_cap is not None else None,
        "low_data_confidence": _low_data_confidence,
        "scored_weight":       round(_scored_weight, 1) if _scored else 0.0,
        "active_weight":       round(_active_weight, 1),
        "pe_current":    pe_current,
        "pe_5yr_avg":         pe_5yr_avg,
        "pfcf_current":       trailing_pfcf,
        "pfcf_5yr_avg":       pfcf_5yr_avg,
        # SBC display metrics (Option B: no scoring impact, shown in report)
        "sbc_trailing_b":  (_sbc_raw / 1e9) if _sbc_raw else None,
        "fcf_ex_sbc_b":    (_fcf_ex_sbc / 1e9) if _fcf_ex_sbc else None,
        "sbc_pct_fcf":     _sbc_pct_fcf,
        "pfcf_adj":        (round(trailing_pfcf * (_fcf_raw_sbc / _fcf_ex_sbc), 1)
                            if (trailing_pfcf and _fcf_ex_sbc > 0 and _fcf_raw_sbc > 0)
                            else None),
        "ev_ebitda_5yr_avg":  ev_ebitda_5yr_avg,
        # All computed scorecard tiers — passed to report_bridge so HTML matches Excel exactly.
        # report_bridge must read ALL of these directly; it must NOT re-derive them.
        "tier_moat":      tier_moat,
        "tier_mgmt":      tier_mgmt,
        "tier_cap_ret":   tier_cap_ret,
        "tier_exec":      tier_exec,
        "tier_rev_cagr":  tier_rev_cagr,
        "tier_fcf_ni":    tier_fcf_ni,
        "tier_roic":      tier_roic,
        "tier_leverage":  tier_leverage,
        "tier_ebit_int":  tier_ebit_int,
        "tier_pe":        tier_pe,
        "tier_pfcf":      tier_pfcf,
        # Continuous scores (0..SCORE_CAP) for each auto-scored criterion.
        # report_bridge prefers these over tier-based lookup so HTML weighted
        # totals exactly match the Excel literal scores written in col F.
        "score_moat":      score_moat,
        "score_mgmt":      score_mgmt,
        "score_cap_ret":   score_cap_ret,
        "score_exec":      score_exec,
        "score_rev_cagr":  score_rev_cagr,
        "score_fcf_ni":    score_fcf_ni,
        "score_roic":      score_roic,
        "score_leverage":  score_leverage,
        "score_ebit_int":  score_ebit_int,
        "score_pe":        score_pe,
        "score_pfcf":      score_pfcf,
        # Regime-aware weight schema (#9). report_bridge uses these to compute
        # p1/p2/p3/p4 totals so the HTML weighted scores match the Excel sheet
        # under bank or EVS regime adjustments.
        "weights":         dict(W),
        "evs_regime":      evs_regime,
        # Forward multiples (for display in report)
        "forward_pe_val":          forward_pe_val,
        "forward_pfcf_val":        forward_pfcf_val,
        # TTM P/E from live quarterly IS (more accurate than FMP snapshot).
        # pe_current uses forward_pe_val when available (most actionable for investors),
        # falling back to ttm_pe. Both are surfaced so the report can show the source.
        "ttm_pe":                  trailing_pe,   # live TTM or FMP snapshot fallback
        "pe_source":               ("forward" if forward_pe_val else
                                    "ttm_live" if _ttm_ni_live else
                                    "fmp_snapshot"),
        # FCF yield vs 10-year treasury spread
        "fcf_yield":               _fcf_yield,
        "fcf_yield_spread":        _fcf_yield_spread,
        "rf_rate_scorecard":       _rf_rate_sc if _fcf_yield is not None else None,
        "fcf_spread_modifier":     _fcf_spread_modifier,
        # Earnings revision momentum
        "eps_revision_pct":        _eps_revision_pct,
        "eps_revision_dir":        _eps_revision_dir,
        "n_analysts_eps":          _n_analysts_eps,
        "rev_momentum_modifier":   _rev_momentum_modifier,
        # Score confidence
        "confidence_level":        _conf_level,
        "confidence_note":         _conf_note,
    }

    print("  Scorecard tab built.")
    return ws, metrics


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    ticker = input("Enter ticker symbol (e.g. AAPL, MSFT, NVDA): ").strip().upper()
    manual_rating_raw = input(
        "Enter S&P / Moody's credit rating (optional — press Enter to skip): "
    ).strip()
    # Normalise Moody's to S&P if needed; blank → None
    if manual_rating_raw:
        tok = manual_rating_raw.strip().split()[0].strip(".,;:()")
        manual_rating = MOODY_TO_SP.get(tok) or (tok.upper() if tok.upper() in VALID_SP_RATINGS else None)
        if not manual_rating:
            print(f"  Warning: '{manual_rating_raw}' not recognised — ignoring manual rating.")
    else:
        manual_rating = None
    print(f"\nFetching data for {ticker}...")

    try:
        is_data = fetch("income-statement",       ticker)[:YEARS][::-1]
        bs_data = fetch("balance-sheet-statement",ticker)[:YEARS][::-1]
        cf_data = fetch("cash-flow-statement",    ticker)[:YEARS][::-1]
    except ValueError as e:
        print(f"\nERROR: {e}")
        print("\nCommon fixes:")
        print("  1. Check API_KEY is correct at top of script")
        print("  2. Verify ticker — try AAPL to confirm API working")
        print("  3. Under Armour = UAA not UA")
        input("\nPress Enter to exit...")
        return

    years = [d.get("fiscalYear") or d.get("calendarYear") or d["date"][:4] for d in is_data]
    print(f"  Years: {years}")

    wb = Workbook()
    build_cover(wb, ticker, years, is_data)
    pl_refs = build_pl(wb, is_data, years, ticker)
    bs_refs = build_bs(wb, bs_data, years, ticker)
    cf_refs = build_cf(wb, cf_data, years, ticker)
    build_ratios(wb, is_data, bs_data, cf_data, years, ticker, pl_refs, bs_refs, cf_refs)
    build_segments(wb, ticker, years)
    wacc_refs = build_wacc(wb, ticker, is_data, bs_data, manual_rating)
    dcf_refs  = build_dcf(wb, ticker, is_data, bs_data, cf_data, years, pl_refs, bs_refs, wacc_refs, cf_refs=cf_refs)
    build_scorecard(wb, ticker, is_data, bs_data, cf_data, years)

    base  = f"{ticker}_FinancialModel_{years[-1]}"
    fname = f"{base}.xlsx"
    fpath = os.path.join(SCRIPT_DIR, fname)
    counter = 1
    while os.path.exists(fpath):
        fname = f"{base}_v{counter}.xlsx"
        fpath = os.path.join(SCRIPT_DIR, fname)
        counter += 1
    wb.save(fpath)
    print(f"\n  Saved: {fpath}")
    print("  Tabs: Cover | P&L | Balance Sheet | Cash Flow | Ratios & FCF | Segments | WACC | DCF | Scorecard")
    input("\nPress Enter to exit...")

if __name__ == "__main__":
    main()
